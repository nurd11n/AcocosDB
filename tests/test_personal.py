"""apps.personal — the Owner's PERSONAL (non-business) expense tracker.
Every test in the "isolation" section exists to prove a specific claim from
apps.personal.models's own docstring, not just to exercise the code path:
this app must be provably inert with respect to every business figure and
every business-facing report, not just conventionally so.
"""

from decimal import Decimal

import pytest
from django.contrib.auth.models import Group
from django.core.management import call_command
from django.utils import timezone

from apps.core.models import ExchangeRate
from apps.core.permissions import EDITOR, VIEWER
from apps.personal.models import PersonalExpense
from apps.personal.services import record_personal_expense

pytestmark = pytest.mark.django_db


def timezone_localdate():
    return timezone.localdate()


@pytest.fixture
def owner(client, django_user_model):
    call_command("setup_roles")
    user = django_user_model.objects.create_superuser("pe_owner", "o@e.com", "x" * 12)
    client.force_login(user)
    return user


# ---------------------------------------------------------------------------
# Basics: the rate freeze, the model, the DB constraint
# ---------------------------------------------------------------------------


def test_kgs_expense_freezes_rate_at_one():
    e = record_personal_expense(
        date=timezone_localdate(), tag=PersonalExpense.FOOD, amount=Decimal("1500"), currency="KGS"
    )
    assert e.rate_to_kgs == Decimal("1")
    assert e.amount_kgs == Decimal("1500")


def test_foreign_expense_converts_and_freezes_the_db_rate():
    today = timezone.localdate()
    ExchangeRate.objects.create(currency="USD", rate=Decimal("87.45"), date=today)
    e = record_personal_expense(
        date=today, tag=PersonalExpense.SHOPPING, amount=Decimal("10"), currency="USD"
    )
    assert e.rate_to_kgs == Decimal("87.45")
    assert e.amount_kgs == Decimal("874.50")

    ExchangeRate.objects.filter(currency="USD").update(rate=Decimal("99.00"))
    e.refresh_from_db()
    assert e.rate_to_kgs == Decimal("87.45"), "a filed row must not follow today's rate"


def test_db_rejects_a_som_row_with_a_bogus_rate_even_via_bulk_create():
    from django.db import IntegrityError, transaction

    with pytest.raises(IntegrityError), transaction.atomic():
        PersonalExpense.objects.bulk_create(
            [
                PersonalExpense(
                    date=timezone.localdate(),
                    tag=PersonalExpense.FOOD,
                    amount=Decimal("181000"),
                    currency="KGS",
                    rate_to_kgs=Decimal("88"),
                )
            ]
        )


def test_db_rejects_zero_or_negative_amount_and_rate():
    from django.db import IntegrityError, transaction

    today = timezone.localdate()
    with pytest.raises(IntegrityError), transaction.atomic():
        PersonalExpense.objects.create(
            date=today, tag=PersonalExpense.FOOD, amount=Decimal("0"), currency="KGS", rate_to_kgs=1
        )
    with pytest.raises(IntegrityError), transaction.atomic():
        PersonalExpense.objects.create(
            date=today,
            tag=PersonalExpense.FOOD,
            amount=Decimal("100"),
            currency="USD",
            rate_to_kgs=Decimal("0"),
        )


# ---------------------------------------------------------------------------
# (a) Zero foreign keys to any business model
# ---------------------------------------------------------------------------


def test_zero_foreign_keys_to_any_business_model():
    """Structural, not incidental — every field on PersonalExpense must be a
    plain scalar. Anything else would let a personal row reference a client,
    a product, a sale, or a contractor, which the model must never be able
    to do even by accident."""
    from django.db.models import ForeignKey

    fk_fields = [f for f in PersonalExpense._meta.get_fields() if isinstance(f, ForeignKey)]
    assert fk_fields == [], f"PersonalExpense has FK field(s): {fk_fields}"


def test_personal_app_not_in_business_apps_permission_map():
    """Deliberately absent from BUSINESS_MODEL_PERMISSIONS — same "safe
    default" as Campaign/BotContent/apps.manufacturing: Editor/Viewer get
    NOTHING here unless someone deliberately adds it."""
    from apps.core.permissions import BUSINESS_MODEL_PERMISSIONS

    assert "personal" not in BUSINESS_MODEL_PERMISSIONS


# ---------------------------------------------------------------------------
# (b) Never appears in send_daily_report (email + Telegram)
# ---------------------------------------------------------------------------


def test_personal_expense_never_appears_in_the_daily_report_xlsx():
    from apps.core.management.commands.send_daily_report import _build_xlsx
    from openpyxl import load_workbook
    import io

    record_personal_expense(
        date=timezone_localdate(),
        tag=PersonalExpense.ENTERTAINMENT,
        amount=Decimal("999999"),
        currency="KGS",
        description="UNIQUEPERSONALMARKER_zzz999",
    )

    wb = load_workbook(io.BytesIO(_build_xlsx()))
    assert set(wb.sheetnames) == {"Продажи", "Итоги", "Остаток", "Долги", "Заказы"}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                text = str(cell) if cell is not None else ""
                assert "UNIQUEPERSONALMARKER" not in text
                assert "999999" not in text


def test_personal_expense_never_appears_in_daily_report_csv():
    from apps.core.management.commands.send_daily_report import _build_csvs

    record_personal_expense(
        date=timezone_localdate(),
        tag=PersonalExpense.HEALTH,
        amount=Decimal("777777"),
        currency="KGS",
        description="UNIQUEPERSONALMARKER_csv",
    )
    csvs = _build_csvs()
    for name, content in csvs.items():
        text = content.decode("utf-8-sig")
        assert "UNIQUEPERSONALMARKER" not in text
        assert "777777" not in text


# ---------------------------------------------------------------------------
# (c) Never enters spent_kgs / net_cash / overhead_kgs / revenue / profit / COGS
# ---------------------------------------------------------------------------


def test_personal_expenses_never_move_any_business_figure():
    """Byte-identical (Decimal-equal, which for a fixed-scale DecimalField
    IS byte-identical) before vs after — the strongest form of this claim:
    not just "doesn't move much", but doesn't move AT ALL."""
    from apps.manufacturing.dashboard import net_cash, overhead_kgs, spent_kgs
    from apps.reports.dashboard import dashboard_data

    today = timezone_localdate()
    before_business_spent = spent_kgs(today, today)
    before_net_cash = net_cash(Decimal("0"), today, today)
    before_overhead = overhead_kgs(today, today)
    before_dash = dashboard_data("month")

    for tag in (
        PersonalExpense.FOOD,
        PersonalExpense.HOUSING,
        PersonalExpense.TRANSPORT,
        PersonalExpense.EDUCATION,
        PersonalExpense.HEALTH,
        PersonalExpense.SHOPPING,
        PersonalExpense.ENTERTAINMENT,
        PersonalExpense.SUBSCRIPTIONS,
        PersonalExpense.PERSONAL,
        PersonalExpense.BUSINESS,
        PersonalExpense.FEES,
    ):
        record_personal_expense(date=today, tag=tag, amount=Decimal("123456"), currency="KGS")

    after_business_spent = spent_kgs(today, today)
    after_net_cash = net_cash(Decimal("0"), today, today)
    after_overhead = overhead_kgs(today, today)
    after_dash = dashboard_data("month")

    assert after_business_spent == before_business_spent
    assert after_net_cash == before_net_cash
    assert after_overhead == before_overhead
    assert after_dash["metrics"]["revenue"] == before_dash["metrics"]["revenue"]
    assert after_dash["metrics"]["profit"] == before_dash["metrics"]["profit"]
    assert after_dash["metrics"]["received"] == before_dash["metrics"]["received"]
    assert after_dash["metrics"]["spent"] == before_dash["metrics"]["spent"]
    assert after_dash["metrics"]["net_cash"] == before_dash["metrics"]["net_cash"]
    assert after_dash["top_products"] == before_dash["top_products"], "COGS/profit-per-product"
    # The whole dict, not just the keys checked above — catches any future
    # key this test's author didn't think to name individually.
    assert after_dash == before_dash


# ---------------------------------------------------------------------------
# (d) Owner-only everywhere
# ---------------------------------------------------------------------------


def _mk(django_user_model, username, group):
    u = django_user_model.objects.create_user(username, password="x" * 12, is_staff=True)
    u.groups.add(Group.objects.get(name=group))
    return u


def test_editor_and_viewer_get_403_or_redirect_on_the_page(client, django_user_model):
    call_command("setup_roles")
    for username, group in (("pe_ed", EDITOR), ("pe_vw", VIEWER)):
        user = _mk(django_user_model, username, group)
        client.force_login(user)
        resp = client.get("/personal/")
        assert resp.status_code in (302, 403), f"{group} got {resp.status_code} on /personal/"
        client.logout()


def test_editor_and_viewer_get_403_or_redirect_on_the_export(client, django_user_model):
    call_command("setup_roles")
    for username, group in (("pe_ed2", EDITOR), ("pe_vw2", VIEWER)):
        user = _mk(django_user_model, username, group)
        client.force_login(user)
        resp = client.get("/personal/export/")
        assert resp.status_code in (302, 403)
        client.logout()


def test_anonymous_gets_redirected_not_200(client):
    assert client.get("/personal/").status_code != 200
    assert client.get("/personal/export/").status_code != 200


def test_owner_can_open_and_use_the_page(client, owner):
    assert client.get("/personal/").status_code == 200


def test_editor_cannot_add_via_panel_admin(client, django_user_model):
    call_command("setup_roles")
    editor = _mk(django_user_model, "pe_ed3", EDITOR)
    client.force_login(editor)
    resp = client.get("/panel/personal/personalexpense/add/")
    assert resp.status_code in (302, 403)


def test_owner_can_add_via_panel_admin_and_rate_is_frozen_not_typed(client, owner):
    body = client.get("/panel/personal/personalexpense/add/").content.decode()
    assert 'name="rate_to_kgs"' not in body, "a rate is never typed, at any stage"

    today = timezone_localdate()
    client.post(
        "/panel/personal/personalexpense/add/",
        {
            "date": today.isoformat(),
            "tag": PersonalExpense.FOOD,
            "amount": "1500",
            "currency": "KGS",
            "description": "",
            "_save": "",
        },
    )
    e = PersonalExpense.objects.get(amount=Decimal("1500"))
    assert e.rate_to_kgs == Decimal("1")


# ---------------------------------------------------------------------------
# Page behaviour: mirrors /manufacturing/expenses/'s own rules exactly
# ---------------------------------------------------------------------------


def test_page_has_a_back_link(client, owner):
    body = client.get("/personal/").content.decode()
    assert 'class="back-link"' in body


def test_currency_select_defaults_to_som_explicitly(client, owner):
    body = client.get("/personal/").content.decode()
    assert '<option value="KGS" selected>' in body


def test_foreign_expense_without_a_rate_saves_nothing(client, owner):
    assert not ExchangeRate.objects.filter(currency="USD").exists()
    resp = client.post(
        "/personal/",
        {
            "tag": PersonalExpense.SHOPPING,
            "amount": "10",
            "currency": "USD",
            "date": timezone_localdate().isoformat(),
        },
        follow=True,
    )
    assert PersonalExpense.objects.count() == 0
    assert "Нет курса" in resp.content.decode()


def test_list_shows_som_equivalent_for_a_foreign_row(client, owner):
    ExchangeRate.objects.create(currency="USD", rate=Decimal("87.45"), date=timezone_localdate())
    record_personal_expense(
        date=timezone_localdate(),
        tag=PersonalExpense.SHOPPING,
        amount=Decimal("10"),
        currency="USD",
    )
    body = client.get("/personal/").content.decode()
    assert "874,50" in body


def test_saving_confirms_both_figures_for_a_foreign_expense(client, owner):
    ExchangeRate.objects.create(currency="USD", rate=Decimal("87.45"), date=timezone_localdate())
    resp = client.post(
        "/personal/",
        {
            "tag": PersonalExpense.SHOPPING,
            "amount": "10",
            "currency": "USD",
            "date": timezone_localdate().isoformat(),
        },
        follow=True,
    )
    body = resp.content.decode()
    assert "874,50" in body and "≈" in body


def test_zero_amount_is_rejected(client, owner):
    client.post(
        "/personal/",
        {
            "tag": PersonalExpense.FOOD,
            "amount": "0",
            "currency": "KGS",
            "date": timezone_localdate().isoformat(),
        },
    )
    assert PersonalExpense.objects.count() == 0


def test_bad_tag_is_rejected(client, owner):
    client.post(
        "/personal/",
        {
            "tag": "not-a-real-tag",
            "amount": "100",
            "currency": "KGS",
            "date": timezone_localdate().isoformat(),
        },
    )
    assert PersonalExpense.objects.count() == 0


# ---------------------------------------------------------------------------
# The export
# ---------------------------------------------------------------------------


def test_export_is_a_separate_endpoint_and_contains_the_data(client, owner):
    record_personal_expense(
        date=timezone_localdate(),
        tag=PersonalExpense.FOOD,
        amount=Decimal("1500"),
        currency="KGS",
        description="EXPORTMARKER",
    )
    resp = client.get("/personal/export/")
    assert resp.status_code == 200
    assert resp["Content-Type"].startswith("text/csv")
    body = resp.content.decode("utf-8-sig")
    assert "EXPORTMARKER" in body
    assert "1500" in body


# ---------------------------------------------------------------------------
# Dashboard card
# ---------------------------------------------------------------------------


def test_dashboard_card_shows_tag_totals_and_net_cash_after_personal(client, owner):
    record_personal_expense(
        date=timezone_localdate(), tag=PersonalExpense.FOOD, amount=Decimal("1500"), currency="KGS"
    )
    body = client.get("/dashboard/?period=month").content.decode()
    assert "Личные расходы" in body
    assert "Еда" in body
    assert "Остаток после личных расходов" in body


def test_dashboard_card_stays_in_som_regardless_of_view_currency_toggle(client, owner):
    """The personal card is deliberately NOT walked by _convert_money — it
    must read the same in сом (its only real unit) no matter which cur=
    the rest of the dashboard is showing. A real USD rate is on record so
    the TOGGLE genuinely engages for the business figures around it —
    otherwise this would pass trivially (missing-rate already falls back
    to сом for everything)."""
    ExchangeRate.objects.create(currency="USD", rate=Decimal("87.45"), date=timezone.localdate())
    record_personal_expense(
        date=timezone_localdate(), tag=PersonalExpense.FOOD, amount=Decimal("1500"), currency="KGS"
    )
    kgs_body = client.get("/dashboard/?period=month&cur=KGS").content.decode()
    usd_body = client.get("/dashboard/?period=month&cur=USD").content.decode()
    # Both must show the SAME 1 500 сом figure for "Потрачено лично" — the
    # personal tile never gets converted, unlike the business ones around it
    # (which DO change between these two responses).
    assert "1\xa0500" in kgs_body
    assert "1\xa0500" in usd_body
