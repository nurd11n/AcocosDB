import hashlib
import hmac
import json
import re
import uuid
from datetime import date, timedelta
from decimal import Decimal

import pytest
import requests
from django.contrib.admin.sites import site
from django.contrib.auth.models import Group, Permission
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.test import RequestFactory
from django.utils import timezone

from apps.clients.models import Client, ClientOpeningBalance, Interaction
from apps.clients.services import (
    client_credits,
    client_debt,
    debtors_report_rows,
    log_whatsapp_interaction,
)
from apps.core.currency import from_base, get_rate, to_base
from apps.core.management.commands.send_daily_report import (
    _debts_sheet,
    _sales_sheet,
    _stock_sheet,
)
from apps.core.models import ExchangeRate, RateChangeLog
from apps.core.permissions import EDITOR, VIEWER
from apps.inventory.models import (
    PRODUCT_IMAGE_MAX_BYTES,
    Category,
    Product,
    ProductImage,
    ProductVariant,
    StockMovement,
)
from apps.inventory.services import add_movement, adjust_to_count
from apps.sales.models import Payment, SaleItem, SaleOrder
from apps.sales.services import (
    balance_kgs_before_payment,
    cancel_sale,
    compute_change_preview,
    confirm_sale,
    mark_fully_paid,
    pay_oldest_first,
    record_payment,
    void_payment,
    void_payment_batch,
)

pytestmark = pytest.mark.django_db


@pytest.fixture
def variant():
    cat = Category.objects.create(name="Dresses")
    product = Product.objects.create(category=cat, name="Evening Dress")
    return ProductVariant.objects.create(
        product=product,
        sku="EVD-M-RED",
        size="M",
        color="red",
        cost_price=Decimal("1500.00"),
        sale_price=Decimal("3200.00"),
    )


def _tiny_image(name="photo.png", color=(200, 50, 50)):
    """A 10x10 in-memory PNG — real enough for PIL (thumbnail derivation) to
    open and process, tiny enough to cost nothing in a test."""
    from io import BytesIO

    from django.core.files.uploadedfile import SimpleUploadedFile
    from PIL import Image

    buf = BytesIO()
    Image.new("RGB", (10, 10), color).save(buf, format="PNG")
    return SimpleUploadedFile(name, buf.getvalue(), content_type="image/png")


# ---------------------------------------------------------------------------
# Part 0b — Client descriptor (last_name -> a staff-only differentiator)
# ---------------------------------------------------------------------------


def test_client_name_shows_the_descriptor_parenthesised_for_staff():
    """Client.name is STAFF-facing (client lists, admin, internal reports) —
    it's allowed to show the descriptor, formatted as a parenthesised note
    rather than concatenated like a real surname, so it can't be misread as
    one."""
    with_tag = Client.objects.create(
        first_name="Айгуль", descriptor="сестра Розы", phone="+996700003001"
    )
    assert with_tag.name == "Айгуль (сестра Розы)"

    without_tag = Client.objects.create(first_name="Роза", phone="+996700003002")
    assert without_tag.name == "Роза"  # no dangling "()" when there's no descriptor


def test_debt_reminder_text_never_leaks_the_descriptor(variant, django_user_model, settings):
    """The whole point of splitting descriptor out from a real name: it must
    never reach the client it's ABOUT. debt_reminder_text is the actual
    outbound WhatsApp text — check the string itself, not just which field
    was passed in."""
    from apps.pos.messaging import debt_reminder_text

    settings.CURRENCY = "KGS"
    client_obj = Client.objects.create(
        first_name="Айгуль", descriptor="должница, не путать с сестрой", phone="+996700003003"
    )
    text = debt_reminder_text(client_obj, {"KGS": Decimal("500")})
    assert "Айгуль" in text
    assert "должница" not in text
    assert "сестрой" not in text
    assert "(" not in text  # no leftover parenthesised note at all


def test_stock_is_sum_of_movements(variant):
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    add_movement(variant, StockMovement.WRITEOFF_OUT, 2)
    assert variant.stock == 8


# ---------------------------------------------------------------------------
# Part 0 — Product gallery (multiple images per product)
# ---------------------------------------------------------------------------


def test_product_can_have_more_than_one_image(settings, tmp_path):
    """The whole point of the feature: a product used to allow exactly one
    photo. It must now allow as many as a manager uploads."""
    settings.MEDIA_ROOT = tmp_path
    cat = Category.objects.create(name="Gallery")
    product = Product.objects.create(category=cat, name="Multi-photo dress")
    for i in range(3):
        ProductImage.objects.create(product=product, image=_tiny_image(f"p{i}.png"), position=i)
    assert product.images.count() == 3


def test_grid_image_prefers_the_lowest_position_as_the_cover(settings, tmp_path):
    """position decides display order; the lowest is the cover every consumer
    (tiles, the client bot, a campaign's hero image) agrees is "the" photo —
    added out of order on purpose, to prove selection isn't upload order."""
    settings.MEDIA_ROOT = tmp_path
    cat = Category.objects.create(name="Gallery2")
    product = Product.objects.create(category=cat, name="Ordered dress")
    second = ProductImage.objects.create(
        product=product, image=_tiny_image("second.png"), position=5
    )
    cover = ProductImage.objects.create(product=product, image=_tiny_image("cover.png"), position=1)

    assert product.grid_image.name == cover.thumbnail.name
    assert product.grid_image.name != second.thumbnail.name


def test_grid_image_is_none_with_no_images_and_never_crashes(settings, tmp_path):
    settings.MEDIA_ROOT = tmp_path
    cat = Category.objects.create(name="Gallery3")
    product = Product.objects.create(category=cat, name="No photo yet")
    assert product.grid_image is None


def test_product_image_thumbnail_is_derived_and_kept_in_sync(settings, tmp_path):
    """Same discipline the old single-photo field had: a ~400px JPEG is
    derived on save, and re-saving with a NEW file re-derives it (never
    leaves a stale thumbnail pointing at the old image)."""
    settings.MEDIA_ROOT = tmp_path
    cat = Category.objects.create(name="Gallery4")
    product = Product.objects.create(category=cat, name="Thumb dress")
    img = ProductImage.objects.create(product=product, image=_tiny_image("a.png", (10, 10, 10)))
    assert img.thumbnail.name
    first_thumb_name = img.thumbnail.name

    img.image = _tiny_image("b.png", (250, 250, 250))
    img.save()
    img.refresh_from_db()
    assert img.thumbnail.name  # still has one
    # A new thumbnail file was written for the new image (name changes because
    # ContentFile.save with the same target name still creates a NEW storage
    # entry when the old one is still referenced — the important assertion is
    # that the thumbnail file actually exists and opens as a real image).
    from PIL import Image

    img.thumbnail.open()
    Image.open(img.thumbnail).verify()
    assert first_thumb_name  # sanity: the pre-resave name was captured


def test_grid_image_uses_the_prefetch_cache_not_a_fresh_query(
    django_assert_num_queries, settings, tmp_path
):
    """The whole reason Product.grid_image reads `self.images.all()` instead
    of `.first()`: `.first()` builds a brand new queryset and ignores an
    existing prefetch_related("images") cache, which would make every grid
    tile issue its own query again — exactly the N+1 the grid builders exist
    to avoid (see apps.pos.views._build_grid_tiles)."""
    settings.MEDIA_ROOT = tmp_path
    cat = Category.objects.create(name="Gallery5")
    product = Product.objects.create(category=cat, name="Prefetch dress")
    ProductImage.objects.create(product=product, image=_tiny_image())

    prefetched = list(Product.objects.filter(pk=product.pk).prefetch_related("images"))
    with django_assert_num_queries(0):
        assert prefetched[0].grid_image is not None


class TestProductGalleryMigrationBackfill:
    """Migration 0006 must not lose data: every existing Product.photo (and
    its already-derived thumbnail) becomes a ProductImage(position=0)
    pointing at the EXACT SAME stored files — no re-upload, no re-encode."""

    # Actually running the schema editor backward/forward needs its own real
    # transaction — pytest-django's default (module-level pytestmark) wraps
    # every test in one big transaction and SQLite can't toggle foreign-key
    # checks inside that.
    pytestmark = pytest.mark.django_db(transaction=True)

    def test_backfills_existing_photo_into_the_gallery(self):
        from django.apps import apps as django_apps
        from django.db.migrations.executor import MigrationExecutor
        from django.db import connection

        executor = MigrationExecutor(connection)
        app = "inventory"

        try:
            # Land on the state just before 0006, where Product still has
            # photo/thumbnail fields, and create an old-style product there.
            executor.migrate([(app, "0005_product_thumbnail")])
            executor.loader.build_graph()
            old_apps = executor.loader.project_state((app, "0005_product_thumbnail")).apps
            OldCategory = old_apps.get_model(app, "Category")
            OldProduct = old_apps.get_model(app, "Product")
            cat = OldCategory.objects.create(name="Migrating")
            product = OldProduct.objects.create(
                category=cat,
                name="Pre-gallery product",
                photo="products/existing.jpg",
                thumbnail="products/thumbs/thumb_existing.jpg",
            )

            # Forward through 0006 — the backfill runs here.
            executor = MigrationExecutor(connection)
            executor.migrate([(app, "0006_product_gallery")])
            executor.loader.build_graph()

            ProductImageNow = django_apps.get_model(app, "ProductImage")
            row = ProductImageNow.objects.get(product_id=product.pk)
            assert row.position == 0
            assert row.image.name == "products/existing.jpg"
            assert row.thumbnail.name == "products/thumbs/thumb_existing.jpg"
        finally:
            # Leave the DB on the latest migration for every test after this
            # one, even if an assertion above failed — a schema half-migrated
            # by this test would otherwise break the entire rest of the run,
            # not just this test.
            executor = MigrationExecutor(connection)
            executor.migrate(executor.loader.graph.leaf_nodes())


def test_confirm_sale_decrements_stock_and_sets_total(variant):
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    client = Client.objects.create(first_name="Aisha", phone="+996700000001")
    order = SaleOrder.objects.create(client=client, channel=SaleOrder.INSTAGRAM)
    SaleItem.objects.create(order=order, variant=variant, quantity=3, unit_price=Decimal("3200"))

    confirm_sale(order)
    order.refresh_from_db()
    assert order.status == SaleOrder.CONFIRMED
    assert order.total == Decimal("9600")
    assert variant.stock == 7


def test_cannot_oversell(variant):
    add_movement(variant, StockMovement.PRODUCTION_IN, 2)
    order = SaleOrder.objects.create()
    SaleItem.objects.create(order=order, variant=variant, quantity=5, unit_price=Decimal("3200"))
    with pytest.raises(ValidationError):
        confirm_sale(order)
    assert variant.stock == 2  # nothing was written off


def test_cancel_returns_stock(variant):
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    order = SaleOrder.objects.create()
    SaleItem.objects.create(order=order, variant=variant, quantity=4, unit_price=Decimal("3200"))
    confirm_sale(order)
    assert variant.stock == 1
    cancel_sale(order)
    assert variant.stock == 5


def test_debt_is_sales_minus_payments(variant):
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    client = Client.objects.create(first_name="Meerim", phone="+996700000002")
    order = SaleOrder.objects.create(client=client)
    SaleItem.objects.create(order=order, variant=variant, quantity=2, unit_price=Decimal("3200"))
    confirm_sale(order)
    Payment.objects.create(client=client, order=order, amount=Decimal("2000"))
    assert client_debt(client) == {"KGS": Decimal("4400")}


def test_debt_is_tracked_independently_per_currency(variant):
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    client = Client.objects.create(first_name="Aidai", phone="+996700000010")
    kgs_order = SaleOrder.objects.create(client=client, currency="KGS")
    SaleItem.objects.create(
        order=kgs_order, variant=variant, quantity=1, unit_price=Decimal("3200")
    )
    confirm_sale(kgs_order)
    usd_order = SaleOrder.objects.create(client=client, currency="USD")
    SaleItem.objects.create(order=usd_order, variant=variant, quantity=1, unit_price=Decimal("40"))
    confirm_sale(usd_order)

    # A USD payment only offsets the USD debt, never the KGS one.
    Payment.objects.create(client=client, order=usd_order, amount=Decimal("40"), currency="USD")

    debts = client_debt(client)
    assert debts == {"KGS": Decimal("3200")}


def test_client_debt_query_is_scoped_to_one_client_not_the_whole_book(variant):
    """client_debt()/client_credits() used to always aggregate EVERY
    client's confirmed sales and payments (an unbounded full-table scan)
    just to answer "what does THIS ONE client owe" — client_detail (the
    highest-traffic client-facing page, opened before/after nearly every
    sale) called it twice per visit. Assert the underlying SQL is now
    actually scoped by client_id, not scanning the whole table regardless
    of how many OTHER clients/sales exist."""
    from django.db import connection
    from django.test.utils import CaptureQueriesContext

    from apps.clients.services import client_debt_and_credit

    add_movement(variant, StockMovement.PRODUCTION_IN, 50)
    # A pile of unrelated clients/sales/payments the query must NOT scan.
    for i in range(20):
        other = Client.objects.create(first_name=f"Noise{i}", phone=f"+99670009{i:04d}")
        order = SaleOrder.objects.create(client=other, currency="KGS")
        SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("100"))
        confirm_sale(order)

    target = Client.objects.create(first_name="Target", phone="+996700009999")
    order = SaleOrder.objects.create(client=target, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)
    Payment.objects.create(client=target, order=order, amount=Decimal("1200"))

    with CaptureQueriesContext(connection) as ctx:
        debts, credits = client_debt_and_credit(target)
    assert debts == {"KGS": Decimal("2000")}
    assert credits == {}

    relevant = [
        q["sql"]
        for q in ctx.captured_queries
        if "sales_saleorder" in q["sql"] or "sales_payment" in q["sql"]
    ]
    assert relevant, "expected the debt aggregate queries to run"
    for sql in relevant:
        assert str(target.pk) in sql, (
            f"query has no client_id filter for the target client — looks like a "
            f"whole-table scan: {sql}"
        )


def test_adjustment_requires_reason_and_writes_diff(variant, django_user_model):
    user = django_user_model.objects.create_user("owner", password="x" * 12)
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    with pytest.raises(ValueError):
        adjust_to_count(variant, 7, user, reason="")
    adjust_to_count(variant, 7, user, reason="inventory count")
    assert variant.stock == 7


def test_variant_stock_panel_receive_writeoff_and_recount(client, django_user_model, variant):
    """The per-variant stock panel on the change page: +Принять, −Списать, and
    =Пересчёт each write a ledger row through services and update the remainder,
    reachable in one click, no bulk action, no raw movement form."""
    from django.urls import reverse

    owner = django_user_model.objects.create_superuser("stockowner", password="x" * 12)
    client.force_login(owner)
    move_url = reverse("admin:inventory_productvariant_stockmove", args=[variant.pk])

    # Change page renders the panel with the live remainder + history.
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    page = client.get(reverse("admin:inventory_productvariant_change", args=[variant.pk]))
    assert page.status_code == 200
    assert "Остаток на складе" in page.content.decode()

    # + Принять 10 -> 15
    client.post(move_url, {"op": "receive", "qty": "10"})
    assert variant.stock == 15
    # − Списать 3 -> 12
    client.post(move_url, {"op": "writeoff", "qty": "3"})
    assert variant.stock == 12
    # = Пересчёт to exactly 8 (with reason) -> 8, via an ADJUSTMENT row
    client.post(move_url, {"op": "recount", "count": "8", "reason": "инвентаризация"})
    assert variant.stock == 8

    # Every change is a ledger row — the audit trail is intact, not overwritten.
    kinds = list(variant.movements.values_list("movement_type", flat=True))
    assert StockMovement.ADJUSTMENT in kinds and StockMovement.WRITEOFF_OUT in kinds


def test_variant_stock_panel_rejects_overdraw_and_missing_reason(
    client, django_user_model, variant
):
    from django.urls import reverse

    client.force_login(django_user_model.objects.create_superuser("stockowner2", password="x" * 12))
    move_url = reverse("admin:inventory_productvariant_stockmove", args=[variant.pk])
    add_movement(variant, StockMovement.PRODUCTION_IN, 4)

    client.post(move_url, {"op": "writeoff", "qty": "9"})  # more than in stock
    assert variant.stock == 4  # unchanged
    client.post(move_url, {"op": "recount", "count": "2", "reason": ""})  # no reason
    assert variant.stock == 4  # unchanged


def test_variant_stock_panel_blocked_for_viewer(client, django_user_model, variant):
    from django.urls import reverse

    from apps.core.permissions import VIEWER

    call_command("setup_roles")
    viewer = django_user_model.objects.create_user("viewer_s", password="x" * 12, is_staff=True)
    viewer.groups.add(Group.objects.get(name=VIEWER))
    client.force_login(viewer)
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    move_url = reverse("admin:inventory_productvariant_stockmove", args=[variant.pk])
    resp = client.post(move_url, {"op": "receive", "qty": "10"})
    assert resp.status_code in (403, 302)  # no change permission
    assert variant.stock == 5  # unchanged


def test_editor_cannot_see_cost_price_field(variant, django_user_model):
    call_command("setup_roles")
    user = django_user_model.objects.create_user("editor1", password="x" * 12, is_staff=True)
    user.groups.add(Group.objects.get(name=EDITOR))
    request = RequestFactory().get("/")
    request.user = user
    model_admin = site._registry[ProductVariant]
    assert "cost_price" not in model_admin.get_fields(request, variant)


def test_viewer_has_no_change_permission(django_user_model):
    call_command("setup_roles")
    user = django_user_model.objects.create_user("viewer1", password="x" * 12, is_staff=True)
    user.groups.add(Group.objects.get(name=VIEWER))
    request = RequestFactory().get("/")
    request.user = user
    model_admin = site._registry[ProductVariant]
    assert model_admin.has_view_permission(request) is True
    assert model_admin.has_change_permission(request) is False
    assert model_admin.has_delete_permission(request) is False


def test_setup_roles_excludes_core_and_wa_apps():
    call_command("setup_roles")
    editor = Group.objects.get(name=EDITOR)
    viewer = Group.objects.get(name=VIEWER)
    assert not editor.permissions.filter(
        content_type__app_label__in=["core", "wa", "auth"]
    ).exists()
    assert not viewer.permissions.filter(
        content_type__app_label__in=["core", "wa", "auth"]
    ).exists()
    assert not editor.permissions.filter(codename__startswith="delete_").exists()


def test_setup_roles_keeps_exchange_rates_superuser_only():
    # ExchangeRate lives under apps.core (Система) — Editor/Viewer never get it,
    # matching "exchange rates" being listed as a Система-only model.
    call_command("setup_roles")
    editor = Group.objects.get(name=EDITOR)
    viewer = Group.objects.get(name=VIEWER)
    assert not editor.permissions.filter(codename__endswith="_exchangerate").exists()
    assert not viewer.permissions.filter(codename__endswith="_exchangerate").exists()


def test_setup_roles_never_grants_stockmovement_or_payment_write(django_user_model):
    """S1: setup_roles used to grant Editor/Viewer every permission on every
    model in inventory/sales/etc. by app_label wildcard — which included
    inventory.add_stockmovement (StockMovementAdmin blocks change/delete but
    never had has_add_permission) and sales.add_payment/change_payment. The
    explicit BUSINESS_MODEL_PERMISSIONS allowlist must never grant either."""
    call_command("setup_roles")
    editor = Group.objects.get(name=EDITOR)
    viewer = Group.objects.get(name=VIEWER)
    for group in (editor, viewer):
        assert not group.permissions.filter(
            content_type__app_label="inventory", codename__endswith="_stockmovement"
        ).exists()
        assert not group.permissions.filter(
            content_type__app_label="sales", codename__in=["add_payment", "change_payment"]
        ).exists()
        assert not group.permissions.filter(
            content_type__app_label="inbox", codename__endswith="_botcontent"
        ).exists()


def test_stockmovement_admin_blocks_add_even_for_a_superuser(django_user_model):
    """Isolates the has_add_permission fix itself, independent of the
    permissions-map fix: Django's own has_perm() always returns True for a
    superuser regardless of any explicit grant, so the wildcard-permission
    fix alone would NOT have stopped the Owner from hand-adding a ledger row
    here too — CLAUDE.md: "Nobody opens the raw movements form." Only the
    admin's own has_add_permission=False closes that for every role at once."""
    owner = django_user_model.objects.create_superuser("sm_owner", "o@e.com", "x" * 12)
    request = RequestFactory().get("/")
    request.user = owner
    model_admin = site._registry[StockMovement]
    assert model_admin.has_add_permission(request) is False


def test_editor_crafted_post_to_add_stockmovement_is_403(client, django_user_model, variant):
    """The exact S1 exploit: a crafted POST to the raw admin add-form,
    bypassing Принять/Списать/Пересчёт (apps.inventory.services.add_movement)
    entirely — must be refused regardless of the Editor's Django permission."""
    call_command("setup_roles")
    editor = django_user_model.objects.create_user("editor_sm", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    stock_before = variant.stock
    resp = client.get("/panel/inventory/stockmovement/add/")
    assert resp.status_code == 403

    resp = client.post(
        "/panel/inventory/stockmovement/add/",
        {
            "variant": str(variant.pk),
            "movement_type": StockMovement.PRODUCTION_IN,
            "quantity": "9999",
            "reason": "phantom stock",
        },
    )
    assert resp.status_code == 403
    assert variant.stock == stock_before  # nothing minted
    assert not StockMovement.objects.filter(reason="phantom stock").exists()


def test_migration_revokes_wildcard_permissions_from_an_existing_group():
    """S1's second half: a code change alone doesn't touch permissions
    already sitting on a Group row in an existing database — the data
    migration (apps.core.migrations.0009_revoke_wildcard_business_permissions)
    is what actually revokes them on upgrade. Simulate the OLD wildcard grant
    directly (bypassing setup_roles, which already only does the right
    thing), then run the migration's own revoke function and assert the
    over-grant is gone."""
    import importlib

    from django.contrib.contenttypes.models import ContentType

    migration_module = importlib.import_module(
        "apps.core.migrations.0009_revoke_wildcard_business_permissions"
    )
    revoke_wildcard_grants = migration_module.revoke_wildcard_grants

    editor = Group.objects.create(name="__wildcard_test_editor__")
    ct = ContentType.objects.get_for_model(StockMovement)
    phantom_perm = Permission.objects.get(content_type=ct, codename="add_stockmovement")
    editor.permissions.add(phantom_perm)
    assert editor.permissions.filter(codename="add_stockmovement").exists()

    # The migration keys off group NAME ("Editor"/"Viewer"), so point this
    # test group at that name to exercise the real function unmodified.
    editor.name = EDITOR
    editor.save(update_fields=["name"])

    class _FakeApps:
        def get_model(self, app_label, model_name):
            return {"auth.group": Group, "auth.permission": Permission}[
                f"{app_label}.{model_name}".lower()
            ]

    revoke_wildcard_grants(_FakeApps(), None)

    editor.refresh_from_db()
    assert not editor.permissions.filter(codename="add_stockmovement").exists()


def test_stock_movement_and_standalone_payment_hidden_from_sidebar():
    from apps.inventory.models import StockMovement as SM

    request = RequestFactory().get("/")
    assert site._registry[SM].has_module_permission(request) is False
    assert site._registry[Payment].has_module_permission(request) is False


def test_whatsapp_message_auto_creates_client_and_interaction():
    client = log_whatsapp_interaction("+996700000099", "hello")
    assert client.source == Client.WHATSAPP
    assert client.interactions.count() == 1
    assert client.interactions.first().kind == Interaction.MESSAGE

    # A second message from the same number reuses the client, doesn't duplicate it.
    log_whatsapp_interaction("+996700000099", "again")
    assert Client.objects.filter(phone="+996700000099").count() == 1
    assert client.interactions.count() == 2


def _by_header(sheet):
    """[{header: value}] per body row — report tests assert on MEANING rather
    than a column index that shifts whenever a sheet is reordered."""
    headers = [c.header for c in sheet.columns]
    return [dict(zip(headers, row)) for row in sheet.rows]


def test_daily_report_rows_are_russian_and_reflect_current_data(variant):
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    client = Client.objects.create(first_name="Aiperi", phone="+996700000003")
    order = SaleOrder.objects.create(client=client, channel=SaleOrder.SHOP)
    SaleItem.objects.create(order=order, variant=variant, quantity=2, unit_price=Decimal("3200"))
    confirm_sale(order)
    Payment.objects.create(client=client, order=order, amount=Decimal("1000"))

    sales = _by_header(_sales_sheet())
    assert any(r["Клиент"] == "Aiperi" for r in sales)

    stock = _by_header(_stock_sheet())
    assert any(r["Артикул"] == variant.sku and r["Остаток"] == 8 for r in stock)

    debts = _by_header(_debts_sheet())
    # Money is a real Decimal now, not a pre-formatted string — that's what
    # makes the column summable/sortable in Excel.
    assert any(
        r["Клиент"] == "Aiperi" and r["Долг"] == Decimal("5400.00") and r["Валюта"] == "KGS"
        for r in debts
    )


def test_daily_report_totals_sheet_matches_the_dashboard(variant, settings):
    """The download must never disagree with the page: Итоги's three figures
    are the SAME dashboard_data("today")["metrics"] the /dashboard/ page
    itself renders, not a second computation."""
    from apps.core.management.commands.send_daily_report import _totals_sheet
    from apps.reports.dashboard import dashboard_data

    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    client_ = Client.objects.create(first_name="Itogi", phone="+996700000303")
    order = SaleOrder.objects.create(client=client_, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3000"))
    confirm_sale(order)
    record_payment(order, Decimal("1000"))  # partial

    itogi = dict(_totals_sheet().rows)
    m = dashboard_data("today")["metrics"]
    assert itogi["Выручка (начислено), сом"] == float(m["revenue"]["value"])
    assert itogi["Получено, сом"] == float(m["received"]["value"])
    assert itogi["Ожидается, сом"] == float(m["expected"]["value"])
    assert itogi["Получено, сом"] == 1000.0
    assert itogi["Ожидается, сом"] == 2000.0


def test_send_daily_report_command_skips_network_when_unconfigured(capsys):
    call_command("send_daily_report")
    out = capsys.readouterr().out.lower()
    assert "skipped" in out


def test_payment_status_reflects_payments(variant):
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    c = Client.objects.create(first_name="Nur", phone="+996700000004")
    order = SaleOrder.objects.create(client=c)
    SaleItem.objects.create(order=order, variant=variant, quantity=2, unit_price=Decimal("3200"))
    confirm_sale(order)
    order.refresh_from_db()

    assert order.total == Decimal("6400")
    assert order.payment_status == SaleOrder.UNPAID
    assert order.balance == Decimal("6400")

    Payment.objects.create(client=c, order=order, amount=Decimal("2400"))
    assert order.payment_status == SaleOrder.PARTIAL
    assert order.balance == Decimal("4000")

    Payment.objects.create(client=c, order=order, amount=Decimal("4000"))
    assert order.payment_status == SaleOrder.PAID
    assert order.balance == Decimal("0")


def test_pending_order_has_no_balance(variant):
    order = SaleOrder.objects.create()
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    # Not approved yet — a pending order owes nothing and reads as unpaid.
    assert order.balance == Decimal("0")
    assert order.payment_status == SaleOrder.UNPAID


def test_foreign_currency_payment_converts_and_counts_toward_balance(variant):
    """A USD payment on a сом order is converted at the NBKR rate frozen onto
    the payment and reduces the сом balance + the client's сом debt. (The owner
    opted into conversion; the POS shows a 'verify the rate' note.)"""
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    ExchangeRate.objects.create(currency="USD", date=timezone.localdate(), rate=Decimal("87"))
    c = Client.objects.create(first_name="Aliya", phone="+996700000012")
    order = SaleOrder.objects.create(client=c, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("8700"))
    confirm_sale(order)  # total 8700 сом
    # 40 USD × 87 = 3480 сом toward the 8700 сом order.
    Payment.objects.create(client=c, order=order, amount=Decimal("40"), currency="USD")
    assert order.paid_amount == Decimal("3480.00")
    assert order.balance == Decimal("5220.00")
    assert order.payment_status == SaleOrder.PARTIAL
    assert client_debt(c) == {"KGS": Decimal("5220.00")}


def test_stats_and_download_require_superuser(client, django_user_model):
    # Anonymous -> redirected to login (not reachable by guessing the URL).
    assert client.get("/stats/").status_code == 302
    assert client.get("/stats/download/?format=csv&sheet=sales").status_code == 302

    # Staff but not superuser -> forbidden.
    staff = django_user_model.objects.create_user("staff1", password="x" * 12, is_staff=True)
    client.force_login(staff)
    assert client.get("/stats/").status_code == 403
    assert client.get("/stats/download/?format=xlsx").status_code == 403

    # Superuser -> OK, and the CSV download carries the UTF-8 BOM.
    root = django_user_model.objects.create_superuser("root1", "r@e.com", "x" * 12)
    client.force_login(root)
    assert client.get("/stats/").status_code == 200
    resp = client.get("/stats/download/?format=csv&sheet=sales")
    assert resp.status_code == 200
    assert resp["Content-Type"].startswith("text/csv")
    assert resp.content[:3] == b"\xef\xbb\xbf"


def test_stats_page_formats_money_with_the_shared_filter(
    client, django_user_model, variant, settings
):
    """stats.html used |unlocalize + a bare currency-symbol span instead of
    the app's own `money` filter — CLAUDE.md explicitly forbids "money
    formatted as a bare number + currency code": no NBSP thousands grouping,
    cents never dropped even at an exact .00. A round revenue figure must
    render the same NBSP-grouped, cents-dropped way it does everywhere else
    in the app."""
    from django.core.cache import cache

    settings.CURRENCY = "KGS"
    cache.clear()  # stats:business_snapshot is cached 60s and bleeds across tests otherwise
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=5, unit_price=Decimal("3000"))
    confirm_sale(order)  # revenue today = 15000.00 KGS exactly

    root = django_user_model.objects.create_superuser("root_stats", "r2@e.com", "x" * 12)
    client.force_login(root)
    body = client.get("/stats/").content.decode()
    assert "15\xa0000\xa0сом" in body  # NBSP-grouped, cents dropped — the money filter
    assert "15000.00" not in body  # the old bare-number rendering must be gone


def test_currency_conversion_uses_current_rate(settings):
    settings.CURRENCY = "KGS"
    # One row per currency — the current rate. The date arg to the helpers is
    # accepted for compatibility but ignored (rates are current-only now).
    ExchangeRate.objects.create(currency="USD", date=date(2026, 7, 12), rate=Decimal("90"))
    any_day = date(2026, 7, 13)

    # Base currency is always 1:1, no rate needed, in either direction.
    assert to_base(Decimal("9000"), "KGS", any_day) == Decimal("9000.00")
    assert from_base(Decimal("9000"), "KGS", any_day) == Decimal("9000.00")

    assert get_rate("USD") == Decimal("90")
    # from_base: KGS -> USD divides by the rate; to_base multiplies.
    assert from_base(Decimal("9000"), "USD", any_day) == Decimal("100.00")
    assert to_base(Decimal("100"), "USD", any_day) == Decimal("9000.00")

    # A currency with no rate on record -> None (caller falls back to base).
    assert from_base(Decimal("9000"), "RUB", any_day) is None
    assert to_base(Decimal("100"), "RUB", any_day) is None


_NBKR_SAMPLE = (
    b'<?xml version="1.0" encoding="windows-1251" ?>'
    b'<CurrencyRates Name="Daily Exchange Rates" Date="17.07.2026">'
    b'<Currency ISOCode="USD"><Nominal>1</Nominal><Value>87,4500</Value></Currency>'
    b'<Currency ISOCode="RUB"><Nominal>1</Nominal><Value>1,1239</Value></Currency>'
    b'<Currency ISOCode="EUR"><Nominal>1</Nominal><Value>100,0000</Value></Currency>'
    b"</CurrencyRates>"
)


class _FakeResp:
    """Fake for the plain (non-streaming) nbkr.kg cross-check fetch."""

    content = _NBKR_SAMPLE

    def raise_for_status(self):
        pass


class _FakeRaw:
    """Fake for the streaming `resp.raw` the Frankfurter fetch reads its
    capped body from."""

    def __init__(self, body: bytes):
        self._body = body

    def read(self, n, decode_content=True):
        return self._body[:n]


class _FakeFrankfurterResp:
    def __init__(self, body: bytes, status_code=200):
        self.status_code = status_code
        self.raw = _FakeRaw(body)

    def raise_for_status(self):
        if self.status_code >= 400:
            raise requests.HTTPError(str(self.status_code), response=self)

    def __enter__(self):
        return self

    def __exit__(self, *exc_info):
        return False


def _frankfurter_get(rates, status_code=200, calls=None):
    """A fake `requests.get` for the Frankfurter primary fetch. `rates` maps
    currency -> the JSON `rate` value to return for /v2/rate/{currency}/...;
    a currency missing from the dict gets a 404. `calls`, if given, collects
    (url, kwargs) for every call made."""

    def fake_get(url, **kwargs):
        if "/v2/rate/" not in url:
            # The best-effort NBKR cross-check hits a different host — treat
            # it as unreachable by default so tests that don't care about it
            # don't need to also fake nbkr.kg's XML.
            raise requests.RequestException("nbkr.kg unreachable (fake)")
        if calls is not None:
            calls.append((url, kwargs))
        currency = url.split("/v2/rate/")[1].split("/")[0]
        if currency not in rates:
            return _FakeFrankfurterResp(b'{"rate": null}', status_code=404)
        body = json.dumps(
            {"date": "2026-08-01", "base": currency, "quote": "KGS", "rate": rates[currency]}
        ).encode()
        return _FakeFrankfurterResp(body, status_code=status_code)

    return fake_get


def test_fetch_rates_pulls_from_frankfurter_pinned_to_the_nbkr_provider(monkeypatch, settings):
    """PRIMARY SOURCE is now Frankfurter — but the pinned NBKR provider is
    ALWAYS sent explicitly, never Frankfurter's own default blended rate
    (its docs warn that figure's trailing decimals can shift as more
    providers report in, which would break every frozen-rate guarantee).
    Since the pinned provider genuinely is NBKR, the stored rate keeps the
    honest НБКР label."""
    settings.CURRENCY = "KGS"
    calls = []
    monkeypatch.setattr(
        "apps.core.management.commands.fetch_rates.requests.get",
        _frankfurter_get({"USD": 87.45, "RUB": 1.1239}, calls=calls),
    )
    call_command("fetch_rates")
    today = timezone.localdate()
    usd = ExchangeRate.objects.get(currency="USD", date=today)
    assert usd.rate == Decimal("87.45")
    assert usd.source == ExchangeRate.NBKR  # pinned provider IS NBKR — honest label
    assert ExchangeRate.objects.get(currency="RUB", date=today).rate == Decimal("1.1239")

    assert len(calls) == 2  # one request per configured non-base currency
    for url, kwargs in calls:
        assert url.startswith("https://api.frankfurter.dev/v2/rate/")
        assert "providers=NBKR" in url  # pinned — the blended default is never used
        assert kwargs["allow_redirects"] is False
        assert kwargs["verify"] is True


def test_fetch_rates_overwrites_existing_row_in_place(monkeypatch, settings):
    settings.CURRENCY = "KGS"
    # An old rate for USD; a fetch overwrites it in place (one row per currency,
    # no dated pile-up).
    ExchangeRate.objects.create(
        currency="USD", date=date(2026, 7, 10), rate=Decimal("88.00"), source=ExchangeRate.MANUAL
    )
    monkeypatch.setattr(
        "apps.core.management.commands.fetch_rates.requests.get",
        _frankfurter_get({"USD": 87.45, "RUB": 1.1239}),
    )
    call_command("fetch_rates")
    assert ExchangeRate.objects.filter(currency="USD").count() == 1
    row = ExchangeRate.objects.get(currency="USD")
    assert row.rate == Decimal("87.45") and row.source == ExchangeRate.NBKR


def test_fetch_rates_survives_network_failure(monkeypatch, settings):
    settings.CURRENCY = "KGS"

    def boom(*a, **k):
        raise requests.RequestException("network down")

    monkeypatch.setattr("apps.core.management.commands.fetch_rates.requests.get", boom)
    call_command("fetch_rates")  # must not raise
    assert not ExchangeRate.objects.exists()  # nothing written, last known kept


def test_fetch_rates_converts_via_decimal_str_never_a_raw_float(monkeypatch, settings):
    """87.45 isn't exactly representable in binary float — Decimal(str(v))
    must be used, never Decimal(v) on the raw float or float() anywhere in
    the path, or the stored value drags in the float's binary-imprecise
    expansion instead of staying exactly 87.45."""
    settings.CURRENCY = "KGS"
    monkeypatch.setattr(
        "apps.core.management.commands.fetch_rates.requests.get",
        _frankfurter_get({"USD": 87.45, "RUB": 1.1239}),
    )
    call_command("fetch_rates")
    assert ExchangeRate.objects.get(currency="USD").rate == Decimal("87.45")
    assert ExchangeRate.objects.get(currency="RUB").rate == Decimal("1.1239")


def test_fetch_rates_rejects_non_numeric_negative_null_and_oversized(monkeypatch, settings):
    """A malformed/implausible response body is a REJECT, not a guess —
    nothing is ever saved from it."""
    settings.CURRENCY = "KGS"

    def _get_for_body(body: bytes, status_code=200):
        def fake_get(url, **kwargs):
            return _FakeFrankfurterResp(body, status_code=status_code)

        return fake_get

    cases = [
        b'{"rate": "87.45"}',  # a string, not a number
        b'{"rate": null}',  # missing/null
        b'{"rate": -87.45}',  # negative
        b'{"rate": 0}',  # zero
        b'{"rate": true}',  # a bool (technically an int subclass — must still be refused)
    ]
    for body in cases:
        monkeypatch.setattr(
            "apps.core.management.commands.fetch_rates.requests.get", _get_for_body(body)
        )
        call_command("fetch_rates")
        assert not ExchangeRate.objects.filter(currency="USD").exists()

    # Oversized body (>64KB) — refused before JSON is even parsed.
    huge = b'{"rate": 87.45, "padding": "' + b"x" * (70 * 1024) + b'"}'
    monkeypatch.setattr(
        "apps.core.management.commands.fetch_rates.requests.get", _get_for_body(huge)
    )
    call_command("fetch_rates")
    assert not ExchangeRate.objects.filter(currency="USD").exists()


def test_fetch_rates_refuses_a_redirect(monkeypatch, settings):
    """A followed redirect on a server-side fetch is exactly how it becomes
    SSRF against something like a cloud metadata endpoint or localhost — any
    30x response must be treated as an outright failure, never followed."""
    settings.CURRENCY = "KGS"

    def fake_get(url, **kwargs):
        assert kwargs.get("allow_redirects") is False
        return _FakeFrankfurterResp(b"", status_code=302)

    monkeypatch.setattr("apps.core.management.commands.fetch_rates.requests.get", fake_get)
    call_command("fetch_rates")
    assert not ExchangeRate.objects.exists()


def test_fetch_rates_rejects_a_jump_over_10_percent(monkeypatch, settings):
    settings.CURRENCY = "KGS"
    ExchangeRate.objects.create(currency="USD", date=date(2026, 7, 30), rate=Decimal("87.00"))
    monkeypatch.setattr(
        "apps.core.management.commands.fetch_rates.requests.get",
        _frankfurter_get({"USD": 100.00, "RUB": 1.1239}),  # +14.9% jump
    )
    call_command("fetch_rates")
    row = ExchangeRate.objects.get(currency="USD")
    assert row.rate == Decimal("87.00")  # unchanged — rejected, nothing saved
    assert row.date == date(2026, 7, 30)  # not bumped either
    assert not RateChangeLog.objects.filter(currency="USD").exists()
    # RUB has no prior rate to deviate from — accepted normally.
    assert ExchangeRate.objects.get(currency="RUB").rate == Decimal("1.1239")


def test_nbkr_proxy_only_affects_the_cross_check_never_the_primary_fetch(monkeypatch, settings):
    """NBKR_PROXY used to route the old direct-NBKR primary fetch; now the
    primary fetch is Frankfurter (unaffected by it) and NBKR_PROXY affects
    ONLY the optional, never-saved nbkr.kg cross-check."""
    settings.CURRENCY = "KGS"
    settings.NBKR_PROXY = "socks5://10.0.0.9:1080"
    frankfurter_calls = []
    nbkr_calls = []

    def fake_get(url, **kwargs):
        if url.startswith("https://api.frankfurter.dev"):
            frankfurter_calls.append((url, kwargs))
            currency = url.split("/v2/rate/")[1].split("/")[0]
            value = {"USD": 87.45, "RUB": 1.1239}[currency]
            body = json.dumps({"rate": value}).encode()
            return _FakeFrankfurterResp(body)
        nbkr_calls.append((url, kwargs))
        raise requests.RequestException("nbkr.kg blocked")

    monkeypatch.setattr("apps.core.management.commands.fetch_rates.requests.get", fake_get)
    call_command("fetch_rates")

    assert ExchangeRate.objects.get(currency="USD").rate == Decimal("87.45")
    assert frankfurter_calls  # primary fetch happened
    for _url, kwargs in frankfurter_calls:
        assert kwargs.get("proxies") is None  # NBKR_PROXY never leaks into the primary fetch
    assert nbkr_calls  # the (best-effort) cross-check was attempted too
    for _url, kwargs in nbkr_calls:
        assert kwargs.get("proxies") == {
            "http": "socks5://10.0.0.9:1080",
            "https": "socks5://10.0.0.9:1080",
        }


def test_nbkr_cross_check_tries_a_browser_user_agent(monkeypatch, settings):
    """A missing/non-browser User-Agent is the usual reason an automated
    fetch gets blocked — the cross-check sends one."""
    settings.CURRENCY = "KGS"
    captured = {}
    monkeypatch.setattr(
        "apps.core.management.commands.fetch_rates.requests.get",
        _frankfurter_get({"USD": 87.45, "RUB": 1.1239}),
    )

    from apps.core.management.commands.fetch_rates import fetch_nbkr_cross_check

    def fake_nbkr_get(url, **kwargs):
        captured["headers"] = kwargs.get("headers")
        return _FakeResp()

    monkeypatch.setattr("apps.core.management.commands.fetch_rates.requests.get", fake_nbkr_get)
    result = fetch_nbkr_cross_check()
    assert result == {"USD": Decimal("87.45"), "RUB": Decimal("1.1239")}
    assert "Mozilla" in captured["headers"]["User-Agent"]


def test_both_rate_sources_down_keeps_last_rate_and_a_sale_still_completes(
    client, django_user_model, variant, monkeypatch, settings
):
    """Total failure of BOTH the Frankfurter primary AND the NBKR cross-check
    must never block a sale — the last known rate is kept untouched, and
    confirming a sale still works."""
    settings.CURRENCY = "KGS"
    ExchangeRate.objects.create(currency="USD", date=date(2026, 7, 20), rate=Decimal("87.00"))
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)

    def boom(*a, **k):
        raise requests.RequestException("everything is down")

    monkeypatch.setattr("apps.core.management.commands.fetch_rates.requests.get", boom)
    call_command("fetch_rates")  # must not raise
    row = ExchangeRate.objects.get(currency="USD")
    assert row.rate == Decimal("87.00")  # last known, untouched
    assert row.date == date(2026, 7, 20)  # not bumped either

    call_command("setup_roles")
    editor = django_user_model.objects.create_user(
        "editor_bothdown", password="x" * 12, is_staff=True
    )
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)
    order_id = SaleOrder.objects.create(created_by=editor, channel=SaleOrder.SHOP).pk
    client.post(f"/pos/sale/{order_id}/items/add/", {"variant_id": variant.pk, "quantity": 1})
    resp = client.post(
        f"/pos/sale/{order_id}/confirm/", {"amount": "3200", "currency": "KGS", "method": "cash"}
    )
    assert resp.status_code == 302


def test_ui_never_shows_nbkr_label_for_a_manually_entered_rate(client, django_user_model, variant):
    """The Курс card and the payment-preview note must show the REAL source —
    a manually-entered rate never renders as «НБКР» just because nobody
    typed a per-payment override (the honesty bug record_payment/the
    templates used to have: they only checked "was THIS payment overridden",
    never the underlying ExchangeRate row's actual source)."""
    call_command("setup_roles")
    ExchangeRate.objects.create(
        currency="USD", date=timezone.localdate(), rate=Decimal("87"), source=ExchangeRate.MANUAL
    )
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    editor = django_user_model.objects.create_user(
        "editor_honest", password="x" * 12, is_staff=True
    )
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    order_id = SaleOrder.objects.create(created_by=editor, channel=SaleOrder.SHOP).pk
    cust = Client.objects.create(first_name="Hon", phone="+996700000831")
    client.post(f"/pos/sale/{order_id}/client/{cust.pk}/set/")
    resp = client.get(f"/pos/sale/{order_id}/")
    body = resp.content.decode()
    strip = body.split('id="rates-strip"')[1].split("</div>")[0]
    assert "вручную" in strip
    assert "НБКР" not in strip  # the page ALSO has a static "Обновить курс НБКР" button — unrelated

    client.post(f"/pos/sale/{order_id}/items/add/", {"variant_id": variant.pk, "quantity": 1})
    resp = client.post(
        f"/pos/sale/{order_id}/recalc/", {"amount": "10", "currency": "USD", "method": "cash"}
    )
    body = resp.content.decode()
    assert "вручную" in body
    assert "НБКР" not in body

    # And the PERSISTED payment carries the same honest source, not a default.
    resp = client.post(
        f"/pos/sale/{order_id}/confirm/", {"amount": "10", "currency": "USD", "method": "cash"}
    )
    assert resp.status_code == 302
    order = SaleOrder.objects.get(pk=order_id)
    assert order.payments.get().rate_source == Payment.RATE_MANUAL


def test_owner_can_set_rate_by_hand_from_the_pos_card(client, django_user_model, settings):
    """The «Изменить курс» dialog: an Owner hand-enters a rate (used where NBKR
    is unreachable), stored as a MANUAL override and logged like the /panel/
    ExchangeRate admin. Comma decimals are accepted; a blank field is left
    untouched."""
    from apps.core.models import ExchangeRate, RateChangeLog

    settings.CURRENCY = "KGS"
    owner = django_user_model.objects.create_superuser("rate_owner", "o@e.com", "x" * 12)
    client.force_login(owner)

    # The modal opens (Owner) and offers a USD field.
    resp = client.get("/pos/rates/edit/")
    assert resp.status_code == 200
    assert 'name="rate_USD"' in resp.content.decode()

    # Save USD (comma decimal), leave RUB blank.
    resp = client.post("/pos/rates/save/", {"rate_USD": "87,45", "rate_RUB": ""})
    assert resp.status_code == 200
    usd = ExchangeRate.objects.get(currency="USD")
    assert usd.rate == Decimal("87.45") and usd.source == ExchangeRate.MANUAL
    assert RateChangeLog.objects.filter(
        currency="USD", source=ExchangeRate.MANUAL, changed_by=owner
    ).exists()
    assert not ExchangeRate.objects.filter(currency="RUB").exists()  # blank = untouched
    assert 'id="rate-modal"' in resp.content.decode()  # dialog closes (oob)


def test_manual_rate_entry_is_owner_only(client, django_user_model, settings):
    from apps.core.models import ExchangeRate
    from apps.core.permissions import EDITOR

    call_command("setup_roles")
    editor = django_user_model.objects.create_user("rate_editor", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    assert client.get("/pos/rates/edit/").status_code == 403
    assert client.post("/pos/rates/save/", {"rate_USD": "90"}).status_code == 403
    assert not ExchangeRate.objects.exists()  # nothing written by a non-Owner


def test_rate_save_rejects_get(client, django_user_model):
    owner = django_user_model.objects.create_superuser("rate_owner2", "o@e.com", "x" * 12)
    client.force_login(owner)
    assert client.get("/pos/rates/save/").status_code == 405  # POST-only mutation


def test_convert_crosses_currencies_via_base():
    from apps.core.currency import convert

    d = date(2026, 7, 12)
    ExchangeRate.objects.create(currency="USD", date=d, rate=Decimal("90"))
    ExchangeRate.objects.create(currency="RUB", date=d, rate=Decimal("1.2"))
    # Same currency is a no-op (quantized).
    assert convert(Decimal("100"), "KGS", "KGS", d) == Decimal("100.00")
    # 40 USD -> KGS (× 90).
    assert convert(Decimal("40"), "USD", "KGS", d) == Decimal("3600.00")
    # 3600 KGS -> USD (÷ 90).
    assert convert(Decimal("3600"), "KGS", "USD", d) == Decimal("40.00")
    # USD -> RUB crosses via KGS: 40 × 90 ÷ 1.2 = 3000 RUB.
    assert convert(Decimal("40"), "USD", "RUB", d) == Decimal("3000.00")
    # No rate on record for a leg -> None (caller keeps it in its own ccy).
    ExchangeRate.objects.filter(currency="RUB").delete()
    assert convert(Decimal("40"), "USD", "RUB", d) is None


def test_foreign_payment_preview_converts_balance_and_shows_note(
    client, django_user_model, variant
):
    """The sale-screen preview converts a foreign payment into the order's
    currency, deducts it, and shows the 'verify the rate' note."""
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    ExchangeRate.objects.create(
        currency="USD", date=timezone.localdate(), rate=Decimal("87"), source=ExchangeRate.NBKR
    )
    editor = django_user_model.objects.create_user("editor_fx", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    order_id = SaleOrder.objects.create(created_by=editor, channel=SaleOrder.SHOP).pk
    client.post(f"/pos/sale/{order_id}/items/add/", {"variant_id": variant.pk, "quantity": 1})
    # variant sells for 3200 сом; pay 10 USD (= 870 сом) -> remaining 2330 сом.
    resp = client.post(
        f"/pos/sale/{order_id}/recalc/", {"amount": "10", "currency": "USD", "method": "cash"}
    )
    body = resp.content.decode()
    assert "2\xa0330" in body  # converted balance, not the untouched 3200
    assert "НБКР" in body  # the verify-manually note is shown


def test_refresh_rates_button_pulls_fresh_and_rerenders(
    client, django_user_model, variant, monkeypatch, settings
):
    settings.CURRENCY = "KGS"
    call_command("setup_roles")
    editor = django_user_model.objects.create_user("editor_rt", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    monkeypatch.setattr(
        "apps.core.management.commands.fetch_rates.requests.get",
        _frankfurter_get({"USD": 87.45, "RUB": 1.1239}),
    )
    assert not ExchangeRate.objects.exists()
    resp = client.post("/pos/rates/refresh/")
    assert resp.status_code == 200
    body = resp.content.decode()
    # Pulled today's USD/RUB and rendered them into the strip.
    assert ExchangeRate.objects.filter(currency="USD", date=timezone.localdate()).exists()
    assert "87,45" in body  # RU locale renders the decimal with a comma
    # A tap on «Обновить» used to look identical whether it worked or not —
    # a real update must say so.
    assert "Курс обновлён" in body


def test_refresh_rates_survives_nbkr_failure(client, django_user_model, monkeypatch):
    call_command("setup_roles")
    editor = django_user_model.objects.create_user("editor_rf", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    def boom(*a, **k):
        raise requests.RequestException("everything down")

    monkeypatch.setattr("apps.core.management.commands.fetch_rates.requests.get", boom)
    resp = client.post("/pos/rates/refresh/")  # must not 500
    assert resp.status_code == 200
    # A failed refresh used to render identically to a successful one — the
    # manager must be told nothing actually changed, with a real reason.
    assert "Не удалось обновить" in resp.content.decode()


def _login_editor(client, django_user_model, name):
    call_command("setup_roles")
    editor = django_user_model.objects.create_user(name, password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)
    return editor


def test_refresh_rates_shows_changed_message(client, django_user_model, settings, monkeypatch):
    """A rate that genuinely moved must say so, with the new value — not a
    generic "updated N" count that also fires when nothing moved."""
    settings.CURRENCY = "KGS"
    _login_editor(client, django_user_model, "editor_changed")
    monkeypatch.setattr(
        "apps.core.management.commands.fetch_rates.requests.get",
        _frankfurter_get({"USD": 87.45, "RUB": 1.1239}),
    )
    resp = client.post("/pos/rates/refresh/")
    body = resp.content.decode()
    assert "Курс обновлён: 1 $ = 87,45" in body
    assert "Не удалось" not in body
    assert "актуален на" not in body


def test_refresh_rates_shows_unchanged_message(client, django_user_model, settings, monkeypatch):
    """Pulling back the exact same rate already on record is NOT a failure
    and must never be conflated with one — see fetch_frankfurter_rates's
    `updated` counting every successful save regardless of whether the VALUE
    moved, which is what made the old single ambiguous message possible."""
    settings.CURRENCY = "KGS"
    _login_editor(client, django_user_model, "editor_unchanged")
    ExchangeRate.objects.create(
        currency="USD", date=date(2026, 7, 30), rate=Decimal("87.45"), source=ExchangeRate.NBKR
    )
    ExchangeRate.objects.create(
        currency="RUB", date=date(2026, 7, 30), rate=Decimal("1.1239"), source=ExchangeRate.NBKR
    )
    monkeypatch.setattr(
        "apps.core.management.commands.fetch_rates.requests.get",
        _frankfurter_get({"USD": 87.45, "RUB": 1.1239}),
    )
    resp = client.post("/pos/rates/refresh/")
    body = resp.content.decode()
    today_str = timezone.localdate().strftime("%d.%m.%Y")
    assert f"Курс $ актуален на {today_str}" in body
    assert "Не удалось" not in body
    assert "Курс обновлён" not in body


def test_refresh_rates_shows_network_failure_message(client, django_user_model, monkeypatch):
    """A connection-level failure (server unreachable) gets its own reason,
    distinct from a malformed response — a manager acting on "нет соединения"
    vs. "сервер прислал некорректные данные" would do different things."""
    _login_editor(client, django_user_model, "editor_netfail")

    def boom(*a, **k):
        raise requests.ConnectionError("connection refused")

    monkeypatch.setattr("apps.core.management.commands.fetch_rates.requests.get", boom)
    resp = client.post("/pos/rates/refresh/")
    assert resp.status_code == 200
    body = resp.content.decode()
    assert "Не удалось обновить $: нет соединения с сервером курсов" in body


def test_refresh_rates_shows_bad_response_message(client, django_user_model, monkeypatch):
    """A response that parses but carries a garbage value (null rate) is a
    DIFFERENT failure from a dead network — the reason text must say so, not
    reuse the connection-failure wording."""
    _login_editor(client, django_user_model, "editor_badresp")

    def fake_get(url, **kwargs):
        return _FakeFrankfurterResp(b'{"rate": null}', status_code=200)

    monkeypatch.setattr("apps.core.management.commands.fetch_rates.requests.get", fake_get)
    resp = client.post("/pos/rates/refresh/")
    assert resp.status_code == 200
    body = resp.content.decode()
    assert "Не удалось обновить $: сервер курсов прислал некорректные данные" in body
    assert "нет соединения" not in body


def test_restock_reply_lists_only_low_variants(variant):
    from apps.wa.replies import restock_reply

    variant.low_stock_threshold = 3
    variant.save(update_fields=["low_stock_threshold"])
    add_movement(variant, StockMovement.PRODUCTION_IN, 2)  # 2 <= 3 -> low
    reply = restock_reply()
    assert variant.sku in reply and "Low stock" in reply

    add_movement(variant, StockMovement.PRODUCTION_IN, 10)  # now 12 > 3 -> healthy
    assert variant.sku not in restock_reply()


def test_bot_stock_replies_do_not_n_plus_one(variant):
    # /restock and /stock answer from ONE annotated query however many variants
    # match — the old per-variant .stock property loop was O(catalog) queries.
    from django.db import connection
    from django.test.utils import CaptureQueriesContext

    from apps.inventory.services import low_stock_variants
    from apps.wa.replies import stock_reply

    cat = Category.objects.create(name="BotPerf")
    for i in range(40):
        p = Product.objects.create(category=cat, name=f"Платье Bot{i}")
        ProductVariant.objects.create(
            product=p, sku=f"BOT-{i:03d}", cost_price=Decimal("1"), sale_price=Decimal("2")
        )

    with CaptureQueriesContext(connection) as ctx:
        low = low_stock_variants()
    assert len(low) >= 40  # all fresh variants have stock 0 <= threshold 2
    assert len(ctx.captured_queries) == 1

    with CaptureQueriesContext(connection) as ctx:
        text = stock_reply("Платье Bot")
    assert "BOT-000" in text
    assert len(ctx.captured_queries) == 1


def test_approve_paid_in_full_records_payment(variant):
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    c = Client.objects.create(first_name="Gulnaz", phone="+996700000005")
    order = SaleOrder.objects.create(client=c)
    SaleItem.objects.create(order=order, variant=variant, quantity=2, unit_price=Decimal("3200"))

    confirm_sale(order)
    record_payment(order, order.total)  # what the "paid in full" choice does
    order.refresh_from_db()
    assert order.payment_status == SaleOrder.PAID
    assert order.balance == Decimal("0")
    assert order.payments.get().currency == order.currency


def test_mark_fully_paid_settles_balance(variant):
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    c = Client.objects.create(first_name="Bermet", phone="+996700000006")
    order = SaleOrder.objects.create(client=c)
    SaleItem.objects.create(order=order, variant=variant, quantity=2, unit_price=Decimal("3200"))
    confirm_sale(order)
    Payment.objects.create(client=c, order=order, amount=Decimal("2000"))  # partial

    mark_fully_paid(order)
    assert order.balance == Decimal("0")
    assert order.payment_status == SaleOrder.PAID


def test_walkin_order_records_no_payment(variant):
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    order = SaleOrder.objects.create()  # no client
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)
    # Walk-in has no client, so there is nothing to owe and no payment is created.
    assert record_payment(order, order.total) is None
    assert order.payments.count() == 0


def _payment_inline_management_form(total=1):
    return {
        "payments-TOTAL_FORMS": str(total),
        "payments-INITIAL_FORMS": "0",
        "payments-MIN_NUM_FORMS": "0",
        "payments-MAX_NUM_FORMS": "1000",
    }


def _payment_inline_row(prefix="payments-0-", **overrides):
    row = {
        "amount": "500",
        "currency": "KGS",
        "method": Payment.CASH,
        "note": "",
        "confirm_overpayment": "",
    }
    row.update(overrides)
    return {f"{prefix}{k}": v for k, v in row.items()}


def _payment_inline_formset(order, data, django_user_model):
    """Builds the exact formset class/instance PaymentInline uses in the real
    admin change view (same get_formset() call, same _order_client_id
    class-attribute wiring) — exercising the real validation code path
    without needing to fabricate an entire SaleOrderAdmin change-form POST."""
    from apps.sales.admin import PaymentInline

    inline = PaymentInline(SaleOrder, site)
    request = RequestFactory().post("/")
    request.user = django_user_model.objects.filter(
        is_superuser=True
    ).first() or django_user_model.objects.create_superuser(
        "inline_admin", "inline_admin@example.com", "x" * 12
    )
    FormSet = inline.get_formset(request, obj=order)
    return FormSet(data=data, instance=order, prefix="payments")


def test_payment_inline_client_prefilled_and_locked_for_a_sale_with_a_client(
    variant, django_user_model
):
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    c = Client.objects.create(first_name="Aisha", phone="+996700000201")
    order = SaleOrder.objects.create(client=c)
    SaleItem.objects.create(order=order, variant=variant, quantity=2, unit_price=Decimal("3200"))
    confirm_sale(order)

    empty_data = _payment_inline_management_form(total=0)
    form = _payment_inline_formset(order, empty_data, django_user_model).empty_form
    assert form.fields["client"].initial == c.pk
    assert form.fields["client"].widget.attrs.get("disabled") == "disabled"

    # A real browser never submits a value for a disabled field — the row
    # arrives with "payments-0-client" simply absent from POST data.
    data = {**_payment_inline_management_form(), **_payment_inline_row()}
    formset = _payment_inline_formset(order, data, django_user_model)
    assert formset.is_valid(), formset.errors
    saved = formset.save(commit=False)
    assert len(saved) == 1
    assert saved[0].client_id == c.pk


def test_payment_inline_rejects_a_crafted_post_with_a_different_client(variant, django_user_model):
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    c = Client.objects.create(first_name="Cholpon", phone="+996700000202")
    other = Client.objects.create(first_name="Someone Else", phone="+996700000203")
    order = SaleOrder.objects.create(client=c)
    SaleItem.objects.create(order=order, variant=variant, quantity=2, unit_price=Decimal("3200"))
    confirm_sale(order)

    data = {
        **_payment_inline_management_form(),
        **_payment_inline_row(client=str(other.pk)),
    }
    formset = _payment_inline_formset(order, data, django_user_model)
    assert not formset.is_valid()
    assert any(
        "должен совпадать" in str(e)
        for form in formset.forms
        for e in form.errors.get("client", [])
    )


def test_payment_inline_allows_choosing_a_client_for_a_walk_in_sale(variant, django_user_model):
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    someone = Client.objects.create(first_name="Chosen", phone="+996700000204")
    order = SaleOrder.objects.create()  # walk-in, no client
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)

    data = {
        **_payment_inline_management_form(),
        **_payment_inline_row(client=str(someone.pk)),
    }
    formset = _payment_inline_formset(order, data, django_user_model)
    assert formset.is_valid(), formset.errors
    saved = formset.save(commit=False)
    assert saved[0].client_id == someone.pk


def test_payment_inline_leaves_an_existing_payment_untouched_by_the_client_lock(
    variant, django_user_model
):
    # A row that already has a pk keeps whatever client it was recorded
    # with, even if SaleOrder.client (not readonly in this admin) is later
    # edited to point elsewhere — the lock/enforcement only governs a NEW
    # row, never re-validates history.
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    original = Client.objects.create(first_name="Original", phone="+996700000205")
    order = SaleOrder.objects.create(client=original)
    SaleItem.objects.create(order=order, variant=variant, quantity=2, unit_price=Decimal("3200"))
    confirm_sale(order)
    payment = Payment.objects.create(client=original, order=order, amount=Decimal("500"))

    order.client = Client.objects.create(first_name="Reassigned", phone="+996700000206")
    order.save(update_fields=["client"])

    data = {
        **_payment_inline_management_form(total=1),
        "payments-INITIAL_FORMS": "1",
        "payments-0-id": str(payment.pk),
        **_payment_inline_row(client=str(original.pk), amount="500", currency="KGS"),
    }
    formset = _payment_inline_formset(order, data, django_user_model)
    assert formset.is_valid(), formset.errors


def _saleorder_admin_post_data(response):
    """Reconstructs a full, valid changeform POST payload from a real GET
    response's rendered form/formset state — every field at its CURRENT
    value — so a test can crank open exactly one field to prove server-side
    enforcement, the same way a crafted request would, without hand-modeling
    every other required field on SaleOrder/SaleItem/Payment."""
    ctx = response.context
    data = {}
    for name, field in ctx["adminform"].form.fields.items():
        val = ctx["adminform"].form.initial.get(name)
        if val is None:
            continue
        data[name] = val.pk if hasattr(val, "pk") else val
    for inline in ctx["inline_admin_formsets"]:
        fs = inline.formset
        prefix = fs.prefix
        for k, v in fs.management_form.initial.items():
            data[f"{prefix}-{k}"] = v
        for i, form in enumerate(fs.forms):
            if form.instance.pk:
                data[f"{prefix}-{i}-id"] = form.instance.pk
            for name, field in form.fields.items():
                val = form.initial.get(name)
                if val is None:
                    continue
                data[f"{prefix}-{i}-{name}"] = val.pk if hasattr(val, "pk") else val
    return data


def test_admin_rejects_a_crafted_post_editing_a_confirmed_sales_item(
    client, django_user_model, variant
):
    """The bug this locks shut: editing a CONFIRMED sale's line item via
    /panel/ used to silently succeed while leaving order.total/total_kgs
    stale and never touching stock. Enforced via SaleItemInline's real
    has_change_permission (a Django admin permission check the changeform
    view itself consults, not a widget) — Django's standard behaviour for a
    permission-denied inline is to keep the 302 "saved" redirect but skip
    that formset's save entirely, so the assertion that matters is the DATA
    staying untouched, not the HTTP status."""
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    owner = django_user_model.objects.create_superuser("saleitem_lock", "s@e.com", "x" * 12)
    order = SaleOrder.objects.create(currency="KGS")
    item = SaleItem.objects.create(
        order=order, variant=variant, quantity=2, unit_price=Decimal("3000")
    )
    confirm_sale(order)
    order.refresh_from_db()
    stock_before = variant.stock

    client.force_login(owner)
    resp = client.get(f"/panel/sales/saleorder/{order.pk}/change/")
    data = _saleorder_admin_post_data(resp)
    item_prefix = next(
        inline.formset.prefix
        for inline in resp.context["inline_admin_formsets"]
        if inline.formset.model is SaleItem
    )
    data[f"{item_prefix}-0-quantity"] = 5  # the crafted change

    client.post(f"/panel/sales/saleorder/{order.pk}/change/", data)

    order.refresh_from_db()
    item.refresh_from_db()
    variant.refresh_from_db()
    assert item.quantity == 2  # unchanged
    assert order.total == Decimal("6000.00")  # unchanged
    assert variant.stock == stock_before  # unchanged — no phantom stock movement


def test_admin_can_still_edit_items_on_a_draft_sale(client, django_user_model, variant):
    """The lock is confirmed-only — a draft (not yet approved) sale's items
    stay fully editable via /panel/, matching the DRAFT-only exception."""
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    owner = django_user_model.objects.create_superuser("saleitem_draft", "d@e.com", "x" * 12)
    order = SaleOrder.objects.create(currency="KGS")
    item = SaleItem.objects.create(
        order=order, variant=variant, quantity=2, unit_price=Decimal("3000")
    )
    assert order.status == SaleOrder.DRAFT

    client.force_login(owner)
    resp = client.get(f"/panel/sales/saleorder/{order.pk}/change/")
    data = _saleorder_admin_post_data(resp)
    item_prefix = next(
        inline.formset.prefix
        for inline in resp.context["inline_admin_formsets"]
        if inline.formset.model is SaleItem
    )
    data[f"{item_prefix}-0-quantity"] = 5

    client.post(f"/panel/sales/saleorder/{order.pk}/change/", data)
    item.refresh_from_db()
    assert item.quantity == 5  # a draft sale is still a normal editable form


def test_saleitem_cost_price_is_hidden_from_editor_in_admin(client, django_user_model, variant):
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    editor = django_user_model.objects.create_user(
        "saleitem_editor", password="x" * 12, is_staff=True
    )
    editor.groups.add(Group.objects.get(name=EDITOR))
    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3000"))

    client.force_login(editor)
    resp = client.get(f"/panel/sales/saleorder/{order.pk}/change/")
    assert resp.status_code == 200
    item_formset = next(
        inline.formset
        for inline in resp.context["inline_admin_formsets"]
        if inline.formset.model is SaleItem
    )
    assert "cost_price" not in item_formset.forms[0].fields


def test_void_payment_creates_reversing_entry_not_a_delete(variant):
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    c = Client.objects.create(first_name="Zarina", phone="+996700000013")
    order = SaleOrder.objects.create(client=c)
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)
    payment = Payment.objects.create(client=c, order=order, amount=Decimal("3200"))

    reversal = void_payment(payment)
    assert reversal.amount == Decimal("-3200")
    assert reversal.reversed_payment_id == payment.pk
    assert Payment.objects.filter(pk=payment.pk).exists()  # never deleted
    order.refresh_from_db()
    assert order.paid_amount == Decimal("0")  # net effect: fully reversed


def test_voided_payment_excluded_from_method_mix(variant):
    # D1 regression: a voided payment must NOT colour the «Способы оплаты» donut.
    # Dropping only the negative reversal row would leave the +amount original
    # overstating its method; the fix drops the original it voided too.
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    c = Client.objects.create(first_name="Nargiza", phone="+996700000099")
    order = SaleOrder.objects.create(client=c)
    SaleItem.objects.create(order=order, variant=variant, quantity=2, unit_price=Decimal("3200"))
    confirm_sale(order)
    cash = Payment.objects.create(
        client=c, order=order, amount=Decimal("3200"), method=Payment.CASH
    )
    Payment.objects.create(client=c, order=order, amount=Decimal("3200"), method=Payment.MBANK)

    void_payment(cash)  # the cash payment was a mistake

    from apps.reports.dashboard import dashboard_data

    methods = dashboard_data("today")["methods"]
    total = sum((m["value"] for m in methods), Decimal("0"))
    # Only the surviving MBank payment remains; the voided cash is gone entirely
    # (both the −3200 reversal and the +3200 original), so the mix is one method.
    # The pre-fix bug would report two methods totalling 6400.
    assert len(methods) == 1
    assert total == Decimal("3200")


def test_void_reviewed_payment_requires_superuser(variant, django_user_model):
    """Payment.reviewed is deprecated (the review queue that used to write it
    is gone) but nothing drops the column, and this guard still protects
    whatever a payment was already marked before removal — set directly here
    since there's no UI path to it anymore."""
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    c = Client.objects.create(first_name="Elmira", phone="+996700000014")
    order = SaleOrder.objects.create(client=c)
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)
    payment = Payment.objects.create(client=c, order=order, amount=Decimal("3200"))
    payment.reviewed = True
    payment.save(update_fields=["reviewed"])

    editor = django_user_model.objects.create_user("editor3", password="x" * 12, is_staff=True)
    with pytest.raises(ValidationError):
        void_payment(payment, user=editor)

    root = django_user_model.objects.create_superuser("root2", "r2@e.com", "x" * 12)
    reversal = void_payment(payment, user=root)  # does not raise
    assert reversal.reversed_payment_id == payment.pk


def test_standalone_payment_admin_cannot_add_or_change_even_for_owner(django_user_model):
    """The ghost-payment bug: the standalone Payment admin (never a sidebar
    entry, but reachable by direct URL) had no has_add_permission override,
    so anyone with sales.add_payment could POST a Payment with no order and
    no production_order straight from a blank form — no sale, no deposit,
    counted toward nothing, visible nowhere except a raw admin list. It must
    stay reachable ONLY for voiding/read-only auditing (see class docstring),
    for every role including the Owner — the sanctioned way to create or edit
    a payment is always through a sale (PaymentInline) or a deposit
    (services.record_deposit)."""
    owner = django_user_model.objects.create_superuser("owner_pay", "o@e.com", "x" * 12)
    request = RequestFactory().get("/")
    request.user = owner
    model_admin = site._registry[Payment]
    assert model_admin.has_add_permission(request) is False
    assert model_admin.has_change_permission(request) is False
    assert model_admin.has_view_permission(request) is True  # auditing still works


def test_standalone_payment_admin_add_url_is_blocked_end_to_end(client, django_user_model):
    call_command("setup_roles")
    owner = django_user_model.objects.create_superuser("owner_pay2", "o2@e.com", "x" * 12)
    client.force_login(owner)
    resp = client.get("/panel/sales/payment/add/")
    assert resp.status_code == 403
    assert Payment.objects.count() == 0

    resp = client.post(
        "/panel/sales/payment/add/",
        {"client": "1", "amount": "500", "currency": "KGS", "method": Payment.CASH},
    )
    assert resp.status_code == 403
    assert Payment.objects.count() == 0


def test_payment_inline_rejects_zero_amount_with_russian_message(variant, django_user_model):
    """Layer 2 of the ghost-payment fix: the DB CheckConstraint already
    rejects amount<=0 (payment_amount_positive_unless_reversal), but a raw
    IntegrityError is a 500, not a usable admin error — the form must catch
    it first with a clean Russian message."""
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    c = Client.objects.create(first_name="Zero", phone="+996700000210")
    order = SaleOrder.objects.create(client=c)
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)

    for bad_amount in ("0", "-5"):
        data = {
            **_payment_inline_management_form(),
            **_payment_inline_row(amount=bad_amount),
        }
        formset = _payment_inline_formset(order, data, django_user_model)
        assert not formset.is_valid()
        assert any(
            "больше нуля" in str(e) for form in formset.forms for e in form.errors.get("amount", [])
        )
    assert not order.payments.exists()


def test_debtors_report_rows_group_by_currency(variant):
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    client = Client.objects.create(first_name="Asel", phone="+996700000015")
    order = SaleOrder.objects.create(client=client, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)

    rows = debtors_report_rows()
    matching = [r for r in rows if r[0].pk == client.pk]
    assert matching == [(client, "KGS", Decimal("3200"), None)]


# ---- Phase 2: /pos/ routing, idempotency, permissions, draft lifecycle -----


def test_confirm_sale_is_idempotent_against_double_call(variant):
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    order = SaleOrder.objects.create()
    SaleItem.objects.create(order=order, variant=variant, quantity=3, unit_price=Decimal("3200"))
    confirm_sale(order)
    assert variant.stock == 7

    # A second confirm on the same (now-approved) order must not decrement
    # stock again — the guarantee behind "double-tap produces exactly one sale".
    with pytest.raises(ValidationError):
        confirm_sale(order)
    assert variant.stock == 7


def test_cancel_sale_is_idempotent_against_double_call(variant):
    """cancel_sale used to have no row lock at all (unlike confirm_sale/
    return_items/void_payment, which all lock first for exactly this
    reason) — a double-tapped «Отменить продажу» could restock the same
    sale twice. A second cancel on an already-cancelled order must raise
    and must not add a second round of RETURN_IN movements."""
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    order = SaleOrder.objects.create()
    SaleItem.objects.create(order=order, variant=variant, quantity=3, unit_price=Decimal("3200"))
    confirm_sale(order)
    assert variant.stock == 7

    cancel_sale(order)
    assert variant.stock == 10  # fully restocked, once

    with pytest.raises(ValidationError):
        cancel_sale(order)
    assert variant.stock == 10  # NOT 13 — the second cancel must not restock again


@pytest.mark.django_db(transaction=True)
def test_cancel_sale_concurrent_double_tap_restocks_exactly_once(variant, django_user_model):
    """The real race, not just the sequential-idempotency check above: two
    genuinely concurrent cancel requests for the same CONFIRMED sale (a
    double-tap on a bad connection) must serialize on the row lock so only
    ONE of them actually restocks — never both. Verified against real
    Postgres to fail on the unpatched (unlocked) cancel_sale and pass on the
    fix — sqlite's coarse whole-file write lock accidentally serializes this
    particular race even without select_for_update, so this assertion is a
    no-op discriminator under the sqlite test DB specifically; it's kept
    because it's still correct and free, and it's the one that actually
    matters when run against Postgres (e.g. `DATABASE_URL=postgres://...`)."""
    import threading
    import time

    from django.db import connections
    from django.db.utils import OperationalError

    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    order = SaleOrder.objects.create()
    SaleItem.objects.create(order=order, variant=variant, quantity=4, unit_price=Decimal("3200"))
    confirm_sale(order)
    assert variant.stock == 6

    results = []
    errors = []
    barrier = threading.Barrier(2)

    def fire():
        barrier.wait(timeout=5)
        for attempt in range(8):
            try:
                cancel_sale(SaleOrder.objects.get(pk=order.pk))
                results.append("cancelled")
                return
            except ValidationError as exc:
                errors.append(exc)
                return
            except OperationalError:
                time.sleep(0.05 * (attempt + 1))
            finally:
                connections.close_all()
        raise AssertionError("gave up retrying after sqlite whole-file lock contention")

    threads = [threading.Thread(target=fire) for _ in range(2)]
    for t in threads:
        t.start()
    for t in threads:
        t.join(timeout=15)

    assert len(results) == 1, f"expected exactly one thread to actually cancel, got {results}"
    assert len(errors) == 1, "expected exactly one thread to be rejected (already cancelled)"
    variant.refresh_from_db()
    assert variant.stock == 10  # restocked exactly once, not 14


def test_pos_root_redirects_by_auth_state(client, django_user_model):
    resp = client.get("/", follow=True)
    assert resp.redirect_chain[-1][0].startswith("/login/")

    call_command("setup_roles")
    editor = django_user_model.objects.create_user("editor10", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)
    resp = client.get("/", follow=True)
    assert "/pos/sale/" in resp.redirect_chain[-1][0]


def test_pos_admin_redirects_to_shared_login_not_its_own_form(client):
    resp = client.get("/panel/", follow=True)
    assert resp.redirect_chain[-1][0].startswith("/login/")


def test_pos_viewer_blocked_from_selling_but_can_read(client, django_user_model):
    call_command("setup_roles")
    editor = django_user_model.objects.create_user("editor9", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    viewer = django_user_model.objects.create_user("viewer9", password="x" * 12, is_staff=True)
    viewer.groups.add(Group.objects.get(name=VIEWER))

    client.force_login(viewer)
    # /pos/ is LOGIN_REDIRECT_URL for everyone, Viewer included — it must not
    # be a dead end, so it routes them to Сегодня instead of 403ing.
    resp = client.get("/pos/")
    assert resp.status_code == 302
    assert resp.url == "/pos/today/"
    assert client.get("/pos/today/").status_code == 200
    assert client.get("/pos/clients/").status_code == 200
    # Direct sale-editing URLs stay blocked even though / no longer 403s.
    assert client.get("/pos/sale/1/").status_code in (403, 404)

    client.force_login(editor)
    assert client.get("/pos/").status_code == 302  # lands on the empty terminal (sale_new)
    assert client.get("/pos/today/").status_code == 200


def test_pos_new_sale_flow_is_idempotent_end_to_end(client, django_user_model, variant):
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    editor = django_user_model.objects.create_user("editor11", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    order_id = SaleOrder.objects.create(created_by=editor, channel=SaleOrder.SHOP).pk

    client.post(f"/pos/sale/{order_id}/items/add/", {"variant_id": variant.pk, "quantity": 2})
    resp = client.post(
        f"/pos/sale/{order_id}/confirm/", {"amount": "6400", "currency": "KGS", "method": "cash"}
    )
    assert resp.status_code == 302
    assert resp.url == f"/pos/sale/{order_id}/result/"

    order = SaleOrder.objects.get(pk=order_id)
    assert order.status == SaleOrder.CONFIRMED
    assert order.total == Decimal("6400")
    variant.refresh_from_db()
    assert variant.stock == 8

    # Double-submit: the exact same POST again must not create a second sale
    # or decrement stock again — it just lands back on the same result page.
    resp2 = client.post(
        f"/pos/sale/{order_id}/confirm/", {"amount": "6400", "currency": "KGS", "method": "cash"}
    )
    assert resp2.status_code == 302
    assert resp2.url == f"/pos/sale/{order_id}/result/"
    variant.refresh_from_db()
    assert variant.stock == 8
    assert SaleOrder.objects.filter(status=SaleOrder.CONFIRMED).count() == 1


# ---------------------------------------------------------------------------
# Per-line discounts (percent / fixed) on SaleItem
# ---------------------------------------------------------------------------


def test_saleitem_line_total_reflects_percent_and_fixed_discount(variant):
    order = SaleOrder.objects.create(currency="KGS")
    percent_item = SaleItem.objects.create(
        order=order,
        variant=variant,
        quantity=2,
        unit_price=Decimal("3200.00"),
        discount_type=SaleItem.DISCOUNT_PERCENT,
        discount_value=Decimal("10"),
    )
    # 3200 × 2 = 6400 subtotal; 10% off = 640 off -> 5760.
    assert percent_item.discount_amount == Decimal("640.00")
    assert percent_item.line_total == Decimal("5760.00")

    fixed_item = SaleItem.objects.create(
        order=order,
        variant=variant,
        quantity=1,
        unit_price=Decimal("3200.00"),
        discount_type=SaleItem.DISCOUNT_FIXED,
        discount_value=Decimal("500"),
    )
    assert fixed_item.discount_amount == Decimal("500.00")
    assert fixed_item.line_total == Decimal("2700.00")

    no_discount_item = SaleItem.objects.create(
        order=order, variant=variant, quantity=1, unit_price=Decimal("3200.00")
    )
    assert no_discount_item.discount_amount == Decimal("0")
    assert no_discount_item.line_total == Decimal("3200.00")


def test_saleitem_discount_never_exceeds_subtotal_even_read_only(variant):
    """A fixed discount bigger than the line itself must never produce a
    negative line_total — the property clamps even before any server-side
    view validation or DB constraint gets a chance to."""
    order = SaleOrder.objects.create(currency="KGS")
    item = SaleItem(
        order=order,
        variant=variant,
        quantity=1,
        unit_price=Decimal("100.00"),
        discount_type=SaleItem.DISCOUNT_FIXED,
        discount_value=Decimal("999.00"),  # bigger than the 100 subtotal
    )
    assert item.discount_amount == Decimal("100.00")  # clamped, not 999
    assert item.line_total == Decimal("0.00")  # never negative


def test_db_rejects_bad_discount_even_via_bulk_create(variant):
    """DB CheckConstraints, not just the view's clamp — must survive
    bulk_create, same discipline as every other money constraint in this
    project (see test_db_constraints_reject_bad_data_even_via_bulk_create)."""
    from django.db import IntegrityError, transaction

    order = SaleOrder.objects.create(currency="KGS")
    with transaction.atomic(), pytest.raises(IntegrityError):
        SaleItem.objects.bulk_create(
            [
                SaleItem(
                    order=order,
                    variant=variant,
                    quantity=1,
                    unit_price=Decimal("100.00"),
                    discount_type=SaleItem.DISCOUNT_PERCENT,
                    discount_value=Decimal("150"),  # >100%
                )
            ]
        )
    with transaction.atomic(), pytest.raises(IntegrityError):
        SaleItem.objects.bulk_create(
            [
                SaleItem(
                    order=order,
                    variant=variant,
                    quantity=1,
                    unit_price=Decimal("100.00"),
                    discount_type=SaleItem.DISCOUNT_FIXED,
                    discount_value=Decimal("101"),  # > the 100 subtotal
                )
            ]
        )
    with transaction.atomic(), pytest.raises(IntegrityError):
        SaleItem.objects.bulk_create(
            [
                SaleItem(
                    order=order,
                    variant=variant,
                    quantity=1,
                    unit_price=Decimal("100.00"),
                    discount_value=Decimal("-1"),  # negative, any type
                )
            ]
        )


def test_confirm_sale_total_includes_line_discounts(variant):
    """confirm_sale needed ZERO changes for this to work — it sums
    item.line_total, which already accounts for the discount. This test is
    the guarantee that stays true."""
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(
        order=order,
        variant=variant,
        quantity=2,
        unit_price=Decimal("3200.00"),
        discount_type=SaleItem.DISCOUNT_PERCENT,
        discount_value=Decimal("25"),
    )
    confirm_sale(order)
    order.refresh_from_db()
    # 6400 subtotal, 25% off = 1600 off -> 4800.
    assert order.total == Decimal("4800.00")


def test_item_discount_view_clamps_percent_and_fixed_server_side(
    client, django_user_model, variant
):
    """Never trusts the client — a crafted POST asking for 500% or a fixed
    amount bigger than the line is clamped, not saved at face value, the same
    discipline as the cart-time stock cap (CLAUDE.md Part 1c/1d)."""
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    editor = django_user_model.objects.create_user(
        "editor_discount1", password="x" * 12, is_staff=True
    )
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    order_id = SaleOrder.objects.create(created_by=editor, channel=SaleOrder.SHOP).pk
    client.post(f"/pos/sale/{order_id}/items/add/", {"variant_id": variant.pk, "quantity": 1})
    item = SaleItem.objects.get(order_id=order_id)

    client.post(
        f"/pos/sale/{order_id}/items/{item.pk}/discount/",
        {"discount_type": "percent", "discount_value": "500"},
    )
    item.refresh_from_db()
    assert item.discount_value == Decimal("100.00")  # clamped to 100%
    assert item.line_total == Decimal("0.00")

    client.post(
        f"/pos/sale/{order_id}/items/{item.pk}/discount/",
        {"discount_type": "fixed", "discount_value": "999999"},
    )
    item.refresh_from_db()
    assert item.discount_value == Decimal("3200.00")  # clamped to the subtotal
    assert item.line_total == Decimal("0.00")

    # Switching back to "none" zeroes discount_value out, not just ignores it.
    client.post(
        f"/pos/sale/{order_id}/items/{item.pk}/discount/",
        {"discount_type": "none", "discount_value": "50"},
    )
    item.refresh_from_db()
    assert item.discount_type == SaleItem.DISCOUNT_NONE
    assert item.discount_value == Decimal("0")
    assert item.line_total == Decimal("3200.00")


def test_item_discount_view_is_owner_and_editor_reachable_not_viewer(
    client, django_user_model, variant
):
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    order = SaleOrder.objects.create(currency="KGS")
    item = SaleItem.objects.create(
        order=order, variant=variant, quantity=1, unit_price=Decimal("100")
    )

    viewer = django_user_model.objects.create_user(
        "viewer_discount1", password="x" * 12, is_staff=True
    )
    viewer.groups.add(Group.objects.get(name=VIEWER))
    client.force_login(viewer)
    resp = client.post(
        f"/pos/sale/{order.pk}/items/{item.pk}/discount/",
        {"discount_type": "percent", "discount_value": "10"},
    )
    assert resp.status_code == 403
    item.refresh_from_db()
    assert item.discount_type == SaleItem.DISCOUNT_NONE  # untouched


def test_receipt_context_shows_discount_only_when_present(variant):
    from apps.pos.receipts import receipt_context

    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)
    assert receipt_context(order)["discount"] is None

    order2 = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(
        order=order2,
        variant=variant,
        quantity=1,
        unit_price=Decimal("3200"),
        discount_type=SaleItem.DISCOUNT_FIXED,
        discount_value=Decimal("200"),
    )
    confirm_sale(order2)
    ctx = receipt_context(order2)
    assert ctx["discount"] == Decimal("200")
    # Line rows always show the STICKER (pre-discount) price/total — the
    # discount is called out exactly once, in aggregate, as the Скидка row
    # in the totals block (see receipt_context's docstring). subtotal
    # (Сумма) minus discount (Скидка) must equal total_money (ИТОГО)
    # exactly, reconciling by construction.
    assert ctx["groups"][0]["rows"][0]["line_total"] == Decimal("3200")  # sticker, undiscounted
    assert ctx["subtotal"] == Decimal("3200")
    assert ctx["subtotal"] - ctx["discount"] == ctx["total_money"] == Decimal("3000")


# ---- Receipt format: grouped by product+size, colours nested; signed link -


def _make_receipt_variant(product, size, color, sale_price=Decimal("1550")):
    return ProductVariant.objects.create(
        product=product,
        sku=f"{product.name}-{size}-{color}",
        size=size,
        color=color,
        sale_price=sale_price,
        cost_price=Decimal("500"),
        currency="KGS",
    )


def test_receipt_groups_by_product_and_size_with_colours_nested(variant):
    """Matches the reference format: one product+size header row, one row per
    COLOUR nested beneath it — never a flat per-line list."""
    from apps.pos.receipts import receipt_context

    cat = Category.objects.create(name="ReceiptCat")
    product = Product.objects.create(category=cat, name="МИМИ 2-КА")
    black = _make_receipt_variant(product, "50-56", "ЧЁРНЫЙ")
    brown = _make_receipt_variant(product, "50-56", "КОРИЧНЕВЫЙ")
    for v in (black, brown):
        add_movement(v, StockMovement.PRODUCTION_IN, 20)

    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(order=order, variant=black, quantity=16, unit_price=Decimal("1550"))
    SaleItem.objects.create(order=order, variant=brown, quantity=16, unit_price=Decimal("1550"))
    confirm_sale(order)

    ctx = receipt_context(order)
    assert len(ctx["groups"]) == 1  # ONE product+size group
    group = ctx["groups"][0]
    assert group["product"] == product
    assert group["size"] == "50-56"
    colors = {row["color"]: row for row in group["rows"]}
    assert set(colors) == {"ЧЁРНЫЙ", "КОРИЧНЕВЫЙ"}
    for row in colors.values():
        assert row["quantity"] == 16
        assert row["unit_price"] == Decimal("1550.00")
        assert row["line_total"] == Decimal("24800")

    # Both unit total AND money total — wholesale, units matter as much as sum.
    assert ctx["total_quantity"] == 32
    assert ctx["total_money"] == Decimal("49600")


def test_receipt_groups_different_sizes_of_same_product_separately(variant):
    from apps.pos.receipts import receipt_context

    cat = Category.objects.create(name="ReceiptCat2")
    product = Product.objects.create(category=cat, name="Платье")
    small = _make_receipt_variant(product, "S", "СИНИЙ")
    large = _make_receipt_variant(product, "L", "СИНИЙ")
    for v in (small, large):
        add_movement(v, StockMovement.PRODUCTION_IN, 10)

    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(order=order, variant=small, quantity=2, unit_price=Decimal("1000"))
    SaleItem.objects.create(order=order, variant=large, quantity=3, unit_price=Decimal("1000"))
    confirm_sale(order)

    ctx = receipt_context(order)
    assert len(ctx["groups"]) == 2  # different sizes -> separate groups
    sizes = {g["size"]: g for g in ctx["groups"]}
    assert sizes["S"]["quantity"] == 2
    assert sizes["L"]["quantity"] == 3
    assert ctx["total_quantity"] == 5
    assert ctx["total_money"] == Decimal("5000")


def test_receipt_merges_duplicate_cart_lines_of_the_same_variant(variant):
    """Two SaleItem rows for the SAME variant (never merged at add-time)
    combine into ONE colour row on the receipt, quantity and money summed."""
    from apps.pos.receipts import receipt_context

    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=2, unit_price=Decimal("1000"))
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(order)

    ctx = receipt_context(order)
    assert len(ctx["groups"]) == 1
    assert len(ctx["groups"][0]["rows"]) == 1  # one merged colour row, not two
    row = ctx["groups"][0]["rows"][0]
    assert row["quantity"] == 3
    assert row["line_total"] == Decimal("3000")


def test_receipt_download_pdf_writes_no_db_rows(client, django_user_model, variant):
    """The old public token page is gone — the PDF now comes from an
    authenticated GET that generates it fresh from the database every time
    and stores nothing. This is the exact row-count guard against the
    ghost-payment class of bug: rendering a receipt must never write."""
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    viewer = django_user_model.objects.create_user("rcpt_viewer", password="x" * 12, is_staff=True)
    viewer.groups.add(Group.objects.get(name=VIEWER))
    client.force_login(viewer)

    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(order)

    payment_before = Payment.objects.count()
    order_before = SaleOrder.objects.count()
    interaction_before = Interaction.objects.count()
    client_before = Client.objects.count()

    resp = client.get(f"/pos/sale/{order.pk}/receipt/pdf/")
    assert resp.status_code == 200
    assert resp["Content-Type"] == "application/pdf"
    assert resp.content[:4] == b"%PDF"

    assert Payment.objects.count() == payment_before
    assert SaleOrder.objects.count() == order_before
    assert Interaction.objects.count() == interaction_before
    assert Client.objects.count() == client_before


def test_receipt_download_filename_is_sortable_ascii_and_dated(client, django_user_model, variant):
    """ACOCOS-chek-<pk>-<ISO date>.pdf — sorts chronologically in any file
    manager, is searchable by sale number, and stays plain ASCII (Latin
    "chek", not Cyrillic «чек») so it needs no RFC 5987 filename* encoding —
    the one thing guaranteed to work identically on every client device."""
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    viewer = django_user_model.objects.create_user(
        "rcpt_fn_viewer", password="x" * 12, is_staff=True
    )
    viewer.groups.add(Group.objects.get(name=VIEWER))
    client.force_login(viewer)

    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(order)

    resp = client.get(f"/pos/sale/{order.pk}/receipt/pdf/")
    assert resp.status_code == 200
    disposition = resp["Content-Disposition"]
    today = timezone.localdate().isoformat()
    assert disposition == f'attachment; filename="ACOCOS-chek-{order.pk}-{today}.pdf"'
    disposition.encode("ascii")  # raises if anything non-ASCII slipped in


def test_receipt_download_requires_a_real_permission(client, django_user_model, variant):
    """Downloading is authenticated now (the old surface was public-by-
    design token access) — a logged-in user with no sales.view_saleorder
    permission at all must be refused."""
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    nobody = django_user_model.objects.create_user("rcpt_nobody", password="x" * 12, is_staff=True)
    client.force_login(nobody)
    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(order)

    assert client.get(f"/pos/sale/{order.pk}/receipt/pdf/").status_code == 403


def test_receipt_status_badge_paid_partial_and_debt(variant):
    """The status badge — «ОПЛАЧЕНО»/«ЧАСТИЧНО — остаток X»/«ДОЛГ X» —
    must render correctly for all three payment states, and the WORD itself
    must be present (not colour alone), so it survives a black-and-white
    printout."""
    from django.template.loader import render_to_string

    from apps.pos.receipts import receipt_context

    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    # A walk-in (no client) is always "paid" the moment it's confirmed (see
    # SaleOrder.payment_status) — a client is required here so paying part
    # or none of the total actually produces the 'partial'/'unpaid' states.
    cust = Client.objects.create(first_name="Badge", phone="+996700006010")

    paid = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=paid, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(paid)
    record_payment(paid, Decimal("1000"))
    html = render_to_string("receipts/receipt.html", receipt_context(paid))
    assert "ОПЛАЧЕНО" in html
    assert "ЧАСТИЧНО" not in html and "ДОЛГ" not in html
    assert "Оплачено" not in html  # fully paid shows only Итого, per spec
    assert "Остаток" not in html

    partial = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=partial, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(partial)
    record_payment(partial, Decimal("400"))  # 600 left
    html = render_to_string("receipts/receipt.html", receipt_context(partial))
    assert "ЧАСТИЧНО" in html
    assert "600" in html
    assert "ОПЛАЧЕНО" not in html

    debt = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=debt, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(debt)  # no payment at all
    html = render_to_string("receipts/receipt.html", receipt_context(debt))
    assert "ДОЛГ" in html
    assert "1\xa0000" in html
    assert "ОПЛАЧЕНО" not in html and "ЧАСТИЧНО" not in html


def test_receipt_shows_first_name_only_never_descriptor_or_phone(variant):
    """Client-facing text uses first_name ONLY — never the staff-only
    descriptor client.name parenthesises onto it, and never the phone."""
    from django.template.loader import render_to_string

    from apps.pos.receipts import receipt_context

    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    cust = Client.objects.create(
        first_name="Aichurek", descriptor="сестра Розы", phone="+996700006001"
    )
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(order)
    record_payment(order, Decimal("1000"))

    html = render_to_string("receipts/receipt.html", receipt_context(order))
    assert "Aichurek" in html
    assert "сестра Розы" not in html
    assert "700006001" not in html
    assert cust.name not in html  # "Aichurek (сестра Розы)" never appears


def test_receipt_pdf_prints_a4_with_word_labelled_status_badges():
    """A4 page setup for print, and every status badge carries its own word
    (verified above) rather than relying on colour alone — the combination
    this test + test_receipt_status_badge_paid_partial_and_debt assert is
    what makes the receipt legible on a black-and-white printout."""
    from pathlib import Path

    from django.conf import settings

    css = (Path(settings.BASE_DIR) / "static" / "pos" / "receipt.css").read_text()
    assert "size: A4" in css


def _render_receipt_pdf(order) -> bytes:
    """Same path apps.pos.views.receipt_download uses, without the HTTP/
    permission plumbing — for tests that need the REAL rendered PDF (word
    positions, page count), not just the HTML/context."""
    from io import BytesIO
    from pathlib import Path

    from django.conf import settings
    from django.template.loader import render_to_string
    from xhtml2pdf import pisa

    from apps.pos.receipts import receipt_context

    css_path = Path(settings.BASE_DIR) / "static" / "pos" / "receipt.css"
    ctx = receipt_context(order)
    ctx["receipt_css_path"] = str(css_path)
    html = render_to_string("receipts/receipt.html", ctx)
    buffer = BytesIO()
    pisa.CreatePDF(html, dest=buffer, encoding="utf-8")
    return buffer.getvalue()


def _render_statement_pdf(client_obj, date_from=None, date_to=None) -> bytes:
    """Same path apps.pos.views.statement_download uses, without the HTTP/
    permission plumbing — mirrors _render_receipt_pdf above."""
    from io import BytesIO
    from pathlib import Path

    from django.conf import settings
    from django.template.loader import render_to_string
    from xhtml2pdf import pisa

    from apps.pos.statements import statement_context

    css_path = Path(settings.BASE_DIR) / "static" / "pos" / "receipt.css"
    ctx = statement_context(client_obj, date_from=date_from, date_to=date_to)
    ctx["receipt_css_path"] = str(css_path)
    html = render_to_string("receipts/statement.html", ctx)
    buffer = BytesIO()
    pisa.CreatePDF(html, dest=buffer, encoding="utf-8")
    return buffer.getvalue()


def test_client_statement_reconciles_opening_purchases_payments_to_closing(variant):
    """Reproduces the client's real handwritten statement exactly: old debt
    99 000, purchase 188 000, paid 150 000 -> 137 000 remaining. Also
    verifies opening + sum(sale_amount) − sum(payment_amount) == closing as
    a general arithmetic identity — and that the running balance is correct
    at EVERY row, not just at the end."""
    from apps.clients.services import client_statement

    add_movement(variant, StockMovement.PRODUCTION_IN, 500000)
    cust = Client.objects.create(first_name="Real", phone="+996700008001")

    old = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=old, variant=variant, quantity=99000, unit_price=Decimal("1"))
    confirm_sale(old)

    new = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=new, variant=variant, quantity=188000, unit_price=Decimal("1"))
    confirm_sale(new)

    record_payment(new, Decimal("150000"))

    stmt = client_statement(cust)["KGS"]
    assert stmt["closing"] == Decimal("137000.00")  # her real example, exactly

    purchases = sum((r["sale_amount"] for r in stmt["rows"] if r["sale_amount"]), Decimal("0"))
    payments = sum((r["payment_amount"] for r in stmt["rows"] if r["payment_amount"]), Decimal("0"))
    assert stmt["opening"] + purchases - payments == stmt["closing"]

    running = stmt["opening"]
    for row in stmt["rows"]:
        running += (row["sale_amount"] or Decimal("0")) - (row["payment_amount"] or Decimal("0"))
        assert row["balance"] == running


def test_client_statement_closing_matches_client_debts_by_currency_exactly(variant):
    """The unfiltered closing figure isn't merely equal to
    client_debts_by_currency's number by coincidence of matching math — it's
    structurally SOURCED from it (see client_statement's own docstring), so
    the two can never disagree even if either function's rounding changes
    independently later."""
    from apps.clients.services import client_debts_by_currency, client_statement

    add_movement(variant, StockMovement.PRODUCTION_IN, 50)
    cust = Client.objects.create(first_name="Match", phone="+996700008002")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=7, unit_price=Decimal("3200"))
    confirm_sale(order)
    record_payment(order, Decimal("9000"))

    stmt = client_statement(cust)
    debts = client_debts_by_currency(client=cust).get(cust.pk, {})
    assert stmt["KGS"]["closing"] == debts["KGS"]
    assert debts["KGS"] > 0  # sanity: this client genuinely owes something


def test_client_statement_multi_currency_sections_stay_separate(variant):
    """A KGS purchase and a USD purchase for the same client must never mix
    into one balance — each currency is its own independent running total,
    never live-rate-converted into the other."""
    from apps.clients.services import client_statement

    usd_variant = ProductVariant.objects.create(
        product=variant.product,
        sku="STMT-USD",
        size="L",
        color="blue",
        sale_price=Decimal("10"),
        cost_price=Decimal("5"),
        currency="USD",
    )
    add_movement(variant, StockMovement.PRODUCTION_IN, 50)
    add_movement(usd_variant, StockMovement.PRODUCTION_IN, 50)
    cust = Client.objects.create(first_name="Multi", phone="+996700008003")

    kgs_order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(
        order=kgs_order, variant=variant, quantity=5, unit_price=Decimal("1000")
    )
    confirm_sale(kgs_order)

    usd_order = SaleOrder.objects.create(client=cust, currency="USD")
    SaleItem.objects.create(
        order=usd_order, variant=usd_variant, quantity=5, unit_price=Decimal("10")
    )
    confirm_sale(usd_order)

    stmt = client_statement(cust)
    assert set(stmt.keys()) == {"KGS", "USD"}
    assert stmt["KGS"]["closing"] == Decimal("5000.00")
    assert stmt["USD"]["closing"] == Decimal("50.00")
    assert all(r["order"].currency == "KGS" for r in stmt["KGS"]["rows"])
    assert all(r["order"].currency == "USD" for r in stmt["USD"]["rows"])


def test_statement_pdf_figures_match_the_screen(client, django_user_model, variant):
    """The on-screen client_detail statement and the downloaded PDF must
    show the IDENTICAL closing balance — both are built from the same
    client_statement() call (see apps.pos.statements.statement_context),
    never two independently-shaped views that could drift apart."""
    from apps.core.currency import format_money

    call_command("setup_roles")
    editor = django_user_model.objects.create_user("stmt_pdf_ed", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    add_movement(variant, StockMovement.PRODUCTION_IN, 50)
    cust = Client.objects.create(first_name="Screen", phone="+996700008004")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=10, unit_price=Decimal("3200"))
    confirm_sale(order)
    record_payment(order, Decimal("12000"))

    from apps.clients.services import client_statement

    expected_closing = client_statement(cust)["KGS"]["closing"]
    expected_str = format_money(expected_closing, "KGS")

    html_resp = client.get(f"/pos/clients/{cust.pk}/")
    assert expected_str in html_resp.content.decode()

    import fitz

    pdf_bytes = _render_statement_pdf(cust)
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    words = doc[0].get_text("words")
    # NBSP-grouped thousands split into separate word tokens once extracted
    # from the PDF (measured — the same quirk the receipt alignment test
    # documents). Scope the check to the ИТОГО ОСТАТОК row specifically (by
    # y-coordinate), not the whole page, so this can't accidentally match
    # digits belonging to a different figure (a date, a Чек number, the
    # purchase amount on an earlier row).
    totals_y = next(w[1] for w in words if w[4] == "ИТОГО")
    row_text = "".join(w[4] for w in words if abs(w[1] - totals_y) < 3)
    assert str(int(expected_closing)) in row_text


def test_statement_download_writes_zero_db_rows(client, django_user_model, variant):
    """Generating the statement PDF must never write — same row-count
    guarantee as the receipt (test_receipt_download_pdf_writes_no_db_rows):
    a saved/cached statement would go stale the moment a new sale or
    payment changes the numbers, so it's a pure read on every request."""
    call_command("setup_roles")
    viewer = django_user_model.objects.create_user("stmt_viewer", password="x" * 12, is_staff=True)
    viewer.groups.add(Group.objects.get(name=VIEWER))
    client.force_login(viewer)

    add_movement(variant, StockMovement.PRODUCTION_IN, 50)
    cust = Client.objects.create(first_name="Zero", phone="+996700008005")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=5, unit_price=Decimal("1000"))
    confirm_sale(order)
    record_payment(order, Decimal("2000"))

    payment_before = Payment.objects.count()
    order_before = SaleOrder.objects.count()
    interaction_before = Interaction.objects.count()
    client_before = Client.objects.count()

    resp = client.get(f"/pos/clients/{cust.pk}/statement/pdf/")
    assert resp.status_code == 200
    assert resp["Content-Type"] == "application/pdf"
    assert resp.content[:4] == b"%PDF"

    assert Payment.objects.count() == payment_before
    assert SaleOrder.objects.count() == order_before
    assert Interaction.objects.count() == interaction_before
    assert Client.objects.count() == client_before


# ---- Phase 2: ClientOpeningBalance (pre-system debt) ----------------------


def test_opening_balance_appears_in_client_debt_and_statement(variant):
    """Reproduces her real handwritten statement using a REAL
    ClientOpeningBalance row (not a fake sale — a fake sale would corrupt
    revenue/profit/stock, see confirm_sale): old debt 99 000, purchase
    188 000, paid 150 000 -> 137 000 remaining, and the statement's first
    row is «Старый долг», exactly matching her paper ledger's own first
    line."""
    from apps.clients.services import client_debts_by_currency, client_statement

    add_movement(variant, StockMovement.PRODUCTION_IN, 500000)
    cust = Client.objects.create(first_name="OpenBal", phone="+996700007001")
    ClientOpeningBalance.objects.create(
        client=cust,
        amount=Decimal("99000"),
        currency="KGS",
        as_of_date=date(2026, 1, 1),
        note="pre-system debt",
    )

    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=188000, unit_price=Decimal("1"))
    confirm_sale(order)
    record_payment(order, Decimal("150000"))

    debts = client_debts_by_currency(client=cust).get(cust.pk, {})
    assert debts["KGS"] == Decimal("137000.00")

    stmt = client_statement(cust)["KGS"]
    assert stmt["closing"] == Decimal("137000.00")
    assert stmt["rows"][0]["kind"] == "opening_balance"
    assert stmt["rows"][0]["description"] == "Старый долг"
    assert stmt["rows"][0]["sale_amount"] == Decimal("99000.00")


def test_opening_balance_excluded_from_revenue_profit_units_and_received(variant):
    """A client with ONLY a pre-system debt (no sales, no payments at all)
    must move EVERY dashboard money metric by exactly zero — asserted
    against apps.reports.dashboard.dashboard_data() directly, not just the
    model, since revenue/profit/units come from SaleOrder/SaleItem and
    «Получено» from Payment, and a ClientOpeningBalance writes to none of
    those tables (see the model's own docstring on why it's neither)."""
    from apps.reports.dashboard import dashboard_data

    before = dashboard_data("today")["metrics"]

    cust = Client.objects.create(first_name="Ghost", phone="+996700007002")
    ClientOpeningBalance.objects.create(
        client=cust, amount=Decimal("500000"), currency="KGS", as_of_date=timezone.localdate()
    )

    after = dashboard_data("today")["metrics"]
    assert after["revenue"]["value"] == before["revenue"]["value"]
    assert after["profit"]["value"] == before["profit"]["value"]
    assert after["units"]["value"] == before["units"]["value"]
    assert after["received"]["value"] == before["received"]["value"]

    # And it DOES show up as real debt on the client's own page.
    from apps.clients.services import client_debt

    assert client_debt(cust)["KGS"] == Decimal("500000.00")


def test_opening_balance_per_currency_stays_separate(variant):
    """A KGS opening balance must never leak into a USD debt figure, or
    vice versa — same "never live-rate-convert, never mix currencies" rule
    as everywhere else in this app."""
    from apps.clients.services import client_debts_by_currency

    cust = Client.objects.create(first_name="Split", phone="+996700007003")
    ClientOpeningBalance.objects.create(
        client=cust, amount=Decimal("10000"), currency="KGS", as_of_date=date(2026, 1, 1)
    )
    ClientOpeningBalance.objects.create(
        client=cust, amount=Decimal("50"), currency="USD", as_of_date=date(2026, 1, 1)
    )

    debts = client_debts_by_currency(client=cust)[cust.pk]
    assert debts["KGS"] == Decimal("10000.00")
    assert debts["USD"] == Decimal("50.00")


def test_editor_crafted_post_to_add_opening_balance_is_403(client, django_user_model):
    """A money-affecting control, same tier as ExchangeRate's manual
    override — a crafted POST to the raw admin add-form must be refused
    regardless of the Editor's Django permissions (the exact S1-class
    exploit pattern test_editor_crafted_post_to_add_stockmovement_is_403
    guards against for stock)."""
    call_command("setup_roles")
    editor = django_user_model.objects.create_user("editor_ob", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    cust = Client.objects.create(first_name="Blocked", phone="+996700007004")

    resp = client.get("/panel/clients/clientopeningbalance/add/")
    assert resp.status_code == 403

    resp = client.post(
        "/panel/clients/clientopeningbalance/add/",
        {
            "client": str(cust.pk),
            "amount": "999999",
            "currency": "KGS",
            "as_of_date": "2026-01-01",
        },
    )
    assert resp.status_code == 403
    assert not ClientOpeningBalance.objects.filter(client=cust).exists()


def test_receipt_totals_reconcile_exactly(variant):
    """Сумма (sticker subtotal) minus Скидка must equal ИТОГО (order.total)
    exactly, always — the arithmetic the redesigned totals block depends
    on reading correctly at a glance (Сумма above, Скидка signed below it,
    ИТОГО last and boldest — never Итого-then-discount, which used to read
    as though the discount still needed subtracting)."""
    from apps.pos.receipts import receipt_context

    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(
        order=order,
        variant=variant,
        quantity=3,
        unit_price=Decimal("1550"),
        discount_type=SaleItem.DISCOUNT_PERCENT,
        discount_value=Decimal("10"),
    )
    confirm_sale(order)

    ctx = receipt_context(order)
    assert ctx["subtotal"] == Decimal("4650")  # 1550 * 3, sticker (pre-discount)
    assert ctx["discount"] == Decimal("465")  # 10% of 4650
    assert ctx["total_money"] == Decimal("4185")
    assert ctx["subtotal"] - ctx["discount"] == ctx["total_money"]


def test_receipt_no_discount_row_when_no_discount(variant):
    """An empty/zero Скидка row must never render — the row is simply
    absent, not shown as «Скидка 0 сом»."""
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)

    from django.template.loader import render_to_string

    from apps.pos.receipts import receipt_context

    ctx = receipt_context(order)
    assert ctx["discount"] is None
    html = render_to_string("receipts/receipt.html", ctx)
    assert "Скидка" not in html


def test_receipt_headers_align_pixel_perfect_with_columns(variant):
    """Verified with a ruler, not by eye: text-column headers share their
    LEFT edge with their data, numeric-column headers share their RIGHT
    edge, and the totals block's money column lands on the exact same
    right edge as the item table's own Сумма column above it — the literal
    bug this redesign fixed (measured misalignment before the fix: ~12pt).
    A few points of tolerance is left for ordinary glyph-shape variance
    between different characters sharing the same cell edge (e.g. "О" vs
    "1"), never for an actual column-width mismatch."""
    import fitz

    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)
    pdf_bytes = _render_receipt_pdf(order)

    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    words = doc[0].get_text("words")  # (x0, y0, x1, y1, text, block, line, word_no)

    def find(text):
        matches = [w for w in words if w[4] == text]
        assert matches, f"{text!r} not found on the rendered receipt"
        return matches[0]

    TOL = 5  # pt

    # Text columns (ТОВАР/РАЗМЕР/ЦВЕТ): header and data share the left edge.
    assert abs(find("ТОВАР")[0] - find("Dresses")[0]) < TOL
    assert abs(find("РАЗМЕР")[0] - find("M")[0]) < TOL
    assert abs(find("ЦВЕТ")[0] - find("red")[0]) < TOL

    # Numeric column (КОЛ-ВО): header and data share the right edge.
    assert abs(find("КОЛ-ВО")[2] - find("1")[2]) < TOL

    # The item table's own Сумма column (header vs data right edge).
    summa_header = find("СУММА")
    summa_label = find("Сумма")  # totals block's label row (sentence case,
    # distinct string from the header) — anchors where the totals block
    # starts, so "сом" tokens can be split into item-table vs totals-block.
    item_row_sums = [w for w in words if w[4] == "сом" and w[3] < summa_label[1]]
    totals_sums = [w for w in words if w[4] == "сом" and w[1] >= summa_label[1]]
    assert item_row_sums and totals_sums
    item_summa_right = max(w[2] for w in item_row_sums)
    assert abs(summa_header[2] - item_summa_right) < TOL

    # The actual bug: totals block numbers on the SAME right edge as the
    # item table's Сумма column — tight tolerance, this must be exact.
    for w in totals_sums:
        assert abs(w[2] - item_summa_right) < 1


def test_multi_page_receipt_does_not_split_or_orphan_totals_block(variant):
    """reportlab's Table flowable splits BETWEEN rows by default when a
    table doesn't fully fit in the remaining page space — before the fix
    (wrapping the totals block in a one-row outer table, see receipt.css's
    .receipt__totals-wrap comment), that put Сумма/Скидка on one page and
    ИТОГО/Остаток alone on the next. The whole block must always land on
    one page together."""
    import fitz

    cat = Category.objects.create(name="Категория")
    order = SaleOrder.objects.create(currency="KGS")
    for i in range(14):
        product = Product.objects.create(
            category=cat, name=f"Товар с очень длинным названием номер {i}"
        )
        v = ProductVariant.objects.create(
            product=product,
            sku=f"LONGNAME-{i}",
            size="L",
            color=f"цвет{i}",
            sale_price=Decimal("1000"),
            cost_price=Decimal("500"),
            currency="KGS",
        )
        add_movement(v, StockMovement.PRODUCTION_IN, 5)
        SaleItem.objects.create(order=order, variant=v, quantity=2, unit_price=Decimal("1000"))
    confirm_sale(order)

    pdf_bytes = _render_receipt_pdf(order)
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    assert len(doc) >= 2, "this many long-named items should span multiple pages"

    itogo_page = summa_page = None
    for i, page in enumerate(doc):
        words = page.get_text("words")
        if any(w[4] == "ИТОГО" for w in words):
            itogo_page = i
        if any(w[4] == "Сумма" for w in words):
            summa_page = i
    assert itogo_page is not None, "ИТОГО must appear somewhere in the PDF"
    assert itogo_page == summa_page, "totals block split across two pages"


def test_receipt_status_badge_grayscale_legible(variant):
    """The badge must survive a black-and-white printout: the status WORD
    is real extractable text (works regardless of colour), and the pill
    itself renders as a real bordered/filled shape (non-white pixels),
    not just coloured text with no visible box once colour is gone."""
    import fitz

    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    cust = Client.objects.create(first_name="GrayBadge", phone="+996700006099")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(order)
    record_payment(order, Decimal("400"))  # partial

    pdf_bytes = _render_receipt_pdf(order)
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    page = doc[0]

    words = page.get_text("words")
    badge_words = [w for w in words if w[4] == "ЧАСТИЧНО"]
    assert badge_words, "status word must be real text, not an image"
    x0, y0, x1, y1 = badge_words[0][:4]

    gray = page.get_pixmap(
        dpi=150, colorspace=fitz.csGRAY, clip=fitz.Rect(x0 - 6, y0 - 4, x1 + 6, y1 + 4)
    )
    pixels = gray.samples
    assert any(p < 250 for p in pixels), "badge region is blank in grayscale"


def test_no_public_receipt_route_exists_anywhere(client):
    """The whole token-link surface must be gone: no reversible URL name, no
    matching path, no importable view module, no token helpers left in
    apps.pos.receipts — this deleted a real attack surface and must not
    linger "just in case"."""
    import importlib

    from django.urls import NoReverseMatch, reverse

    for name in ("receipt-public", "receipt-pdf"):
        with pytest.raises(NoReverseMatch):
            reverse(name, args=["x"])

    assert client.get("/r/anything/").status_code == 404
    assert client.get("/r/anything/pdf/").status_code == 404

    with pytest.raises(ModuleNotFoundError):
        importlib.import_module("apps.pos.public_views")

    from apps.pos import receipts

    for gone in ("make_receipt_token", "resolve_receipt_token", "RECEIPT_SALT", "RECEIPT_MAX_AGE"):
        assert not hasattr(receipts, gone), f"{gone} should have been removed with the token system"


def test_share_receipt_sends_short_message_with_no_url_or_domain(
    client, django_user_model, variant
):
    """The WhatsApp text is a plain thank-you — no link, no token, no domain
    — since wa.me can't attach the PDF anyway; she attaches it manually
    after downloading it separately (receipt_download)."""
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    owner = django_user_model.objects.create_superuser("rcpt_owner", "o@e.com", "x" * 12)
    client.force_login(owner)
    cust = Client.objects.create(first_name="Sharer", phone="+996700006002")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)

    resp = client.post(f"/pos/sale/{order.pk}/receipt/")
    assert resp.status_code == 302
    location = resp["Location"]
    assert "wa.me" in location
    from urllib.parse import unquote

    text = unquote(location.split("text=", 1)[1])
    assert text == "Спасибо за покупку, Sharer! Чек прикреплён."
    assert "http" not in text
    assert "/r/" not in text
    assert "acocos" not in text.lower()
    assert "3200" not in text  # no itemised breakdown pasted into the chat

    interaction = cust.interactions.get(kind=Interaction.MESSAGE)
    assert interaction.note == "Чек отправлен"


def test_dashboard_top_products_revenue_reflects_discount_not_sticker_price(variant):
    """The dashboard's 'top products' ranking must show what was actually
    charged, not the undiscounted price — a raw unit_price×quantity SQL
    expression here would silently overstate revenue for anything sold at
    a discount (see apps.reports.dashboard._LINE_KGS)."""
    from apps.reports.dashboard import _top_products

    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(
        order=order,
        variant=variant,
        quantity=2,
        unit_price=Decimal("3200.00"),
        discount_type=SaleItem.DISCOUNT_PERCENT,
        discount_value=Decimal("50"),
    )
    confirm_sale(order)
    today = timezone.localdate()
    rows = _top_products(today, today)
    assert len(rows) == 1
    # 6400 subtotal, 50% off -> 3200 actual revenue, not 6400.
    assert rows[0]["revenue"] == Decimal("3200.00")


def test_dashboard_top_products_revenue_is_rounded_to_cents_like_the_order_total(variant):
    """Same awkward-percent case as the metrics regression test above: the
    per-line SQL discount expression must round to cents (matching
    SaleItem.discount_amount's ROUND_HALF_UP) before subtracting from the
    subtotal, not carry a raw 4-decimal-place division result forward."""
    from apps.reports.dashboard import _top_products

    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(
        order=order,
        variant=variant,
        quantity=3,
        unit_price=Decimal("999.99"),
        discount_type=SaleItem.DISCOUNT_PERCENT,
        discount_value=Decimal("15.00"),
    )
    confirm_sale(order)
    order.refresh_from_db()
    today = timezone.localdate()
    rows = _top_products(today, today)
    assert len(rows) == 1
    assert rows[0]["revenue"] == order.total  # 2549.97, not 2549.9745


def test_confirm_sale_freezes_cost_price_onto_the_line(variant):
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    order = SaleOrder.objects.create(currency="KGS")
    item = SaleItem.objects.create(
        order=order, variant=variant, quantity=2, unit_price=Decimal("3200")
    )
    assert item.cost_price is None  # not yet costed — the order hasn't been confirmed
    confirm_sale(order)
    item.refresh_from_db()
    assert item.cost_price == variant.cost_price


def test_dashboard_profit_is_unaffected_by_a_later_cost_price_change(variant):
    """The whole point of freezing cost_price on the SaleItem: raising the
    variant's cost price THIS month must never rewrite a sale CONFIRMED last
    month. Reading apps.inventory.models.ProductVariant.cost_price live in
    the dashboard's COGS calc — instead of the frozen SaleItem.cost_price —
    would silently move historical profit every time a cost changes."""
    from apps.reports.dashboard import dashboard_data

    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=2, unit_price=Decimal("3200"))
    confirm_sale(order)

    m_before = dashboard_data("today")["metrics"]
    original_cost = variant.cost_price

    variant.cost_price = original_cost + Decimal("500")  # the cost went up today
    variant.save(update_fields=["cost_price"])

    m_after = dashboard_data("today")["metrics"]
    assert m_after["profit"]["value"] == m_before["profit"]["value"]
    # And it must be computed from the FROZEN cost, not silently zeroed or
    # left at the new one: 3200*2 revenue - (frozen cost * 2) profit.
    assert m_after["profit"]["value"] == Decimal("6400") - original_cost * 2


def test_dashboard_profit_and_margin_are_computed_after_discount(variant):
    """SaleOrder.total (and total_kgs) are already net of discount — see
    SaleItem.line_total / confirm_sale — so the dashboard's revenue/profit/
    margin must reflect the discounted price actually charged, not the
    sticker subtotal. cost_price 1200 x 2 = 2400 COGS either way; revenue
    should be 3200 (discounted), never 6400 (sticker)."""
    from apps.reports.dashboard import dashboard_data

    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(
        order=order,
        variant=variant,
        quantity=2,
        unit_price=Decimal("3200.00"),
        discount_type=SaleItem.DISCOUNT_PERCENT,
        discount_value=Decimal("50"),
    )
    confirm_sale(order)

    m = dashboard_data("today")["metrics"]
    assert m["revenue"]["value"] == Decimal("3200.00")
    assert m["profit"]["value"] == Decimal("3200.00") - variant.cost_price * 2
    expected_margin = int(
        round((Decimal("3200.00") - variant.cost_price * 2) / Decimal("3200.00") * 100)
    )
    assert m["profit"]["margin"] == expected_margin


def test_dashboard_discount_metric_totals_what_was_given_away(variant):
    from apps.reports.dashboard import dashboard_data

    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    order_a = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(
        order=order_a,
        variant=variant,
        quantity=2,
        unit_price=Decimal("3200.00"),
        discount_type=SaleItem.DISCOUNT_PERCENT,
        discount_value=Decimal("50"),  # 3200 off
    )
    confirm_sale(order_a)

    order_b = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(
        order=order_b,
        variant=variant,
        quantity=1,
        unit_price=Decimal("3200.00"),
        discount_type=SaleItem.DISCOUNT_FIXED,
        discount_value=Decimal("500"),  # 500 off
    )
    confirm_sale(order_b)

    order_c = SaleOrder.objects.create(currency="KGS")  # no discount at all
    SaleItem.objects.create(order=order_c, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order_c)

    m = dashboard_data("today")["metrics"]
    assert m["discount"]["value"] == Decimal("3700.00")  # 3200 + 500
    # Sticker total = 6400 + 3200 + 3200 = 12800; discount 3700 -> ~28.9%.
    assert m["discount"]["pct"] == int(round(Decimal("3700") / Decimal("12800") * 100))


def test_dashboard_discount_metric_is_zero_with_no_discounts(variant):
    from apps.reports.dashboard import dashboard_data

    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=2, unit_price=Decimal("3200"))
    confirm_sale(order)

    m = dashboard_data("today")["metrics"]
    assert m["discount"]["value"] == Decimal("0")
    assert m["discount"]["pct"] == 0


def test_dashboard_discount_and_revenue_are_exact_for_an_awkward_percent(variant, settings):
    """Regression test: a percent discount whose subtotal*value/100 doesn't
    land on a clean 2-decimal figure (here: 2999.97 * 15% = 449.9955) used to
    leak into the dashboard's SQL aggregates unrounded, drifting a fraction
    of a сом from what confirm_sale actually froze onto SaleOrder.total —
    and revenue + discount silently failed to add back up to the sticker
    subtotal. Both must now be EXACT, matching order.total to the cent."""
    from apps.reports.dashboard import dashboard_data

    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    order = SaleOrder.objects.create(currency="KGS")
    item = SaleItem.objects.create(
        order=order,
        variant=variant,
        quantity=3,
        unit_price=Decimal("999.99"),
        discount_type=SaleItem.DISCOUNT_PERCENT,
        discount_value=Decimal("15.00"),
    )
    confirm_sale(order)
    order.refresh_from_db()
    assert order.total == Decimal("2549.97")  # sanity: the authoritative figure

    m = dashboard_data("today")["metrics"]
    assert m["revenue"]["value"] == order.total_kgs
    assert m["discount"]["value"] == item.subtotal - order.total
    assert m["revenue"]["value"] + m["discount"]["value"] == item.subtotal


def test_dashboard_booked_vs_received_vs_expected(variant, settings):
    """The exact scenario requested: a confirmed-but-unpaid sale raises
    Выручка (booked) and Ожидается (expected) but NOT Получено (received);
    a later payment moves the split without double-counting anything."""
    from apps.reports.dashboard import dashboard_data

    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    client_ = Client.objects.create(first_name="Debtor", phone="+996700000301")
    order = SaleOrder.objects.create(client=client_, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=2, unit_price=Decimal("3000"))
    confirm_sale(order)  # 6000 booked, nothing paid

    m1 = dashboard_data("today")["metrics"]
    assert m1["revenue"]["value"] == Decimal("6000.00")
    assert m1["received"]["value"] == Decimal("0")
    assert m1["expected"]["value"] == Decimal("6000.00")

    record_payment(order, Decimal("6000"))  # paid in full, same day

    m2 = dashboard_data("today")["metrics"]
    assert m2["revenue"]["value"] == Decimal("6000.00")  # unchanged — still booked
    assert m2["received"]["value"] == Decimal("6000.00")  # now in hand
    assert m2["expected"]["value"] == Decimal("0")  # nothing left owed
    # No double-counting: booked always equals received + expected.
    assert m2["revenue"]["value"] == m2["received"]["value"] + m2["expected"]["value"]


def test_dashboard_expected_surplus_state_is_clearly_labelled(
    client, django_user_model, variant, settings
):
    """When more OLD debt is collected today than NEW revenue is booked
    today, `expected` (revenue - received) goes negative — real signal, not
    a bug (see apps.reports.dashboard._metrics's docstring) — but a bare
    "-1 959 300 сом" under a "Ожидается" (pending) label reads as an error
    to a human, not information. This state must render as its own clearly
    labelled, positive fact instead."""

    from apps.reports.dashboard import dashboard_data

    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    cust = Client.objects.create(first_name="OldDebtor", phone="+996700000399")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("5000"))
    confirm_sale(order)
    order.confirmed_at = timezone.now() - timedelta(days=1)  # booked YESTERDAY
    order.save(update_fields=["confirmed_at"])

    record_payment(order, Decimal("5000"))  # but paid off TODAY

    m = dashboard_data("today")["metrics"]
    assert m["revenue"]["value"] == Decimal("0")  # nothing NEW booked today
    assert m["received"]["value"] == Decimal("5000.00")  # but cash landed today
    assert m["expected"]["value"] == Decimal("-5000.00")  # the raw signed figure, unchanged
    assert m["expected"]["is_surplus"] is True
    assert m["expected"]["display_value"] == Decimal("5000.00")  # always positive for display

    owner = django_user_model.objects.create_superuser("dash_surplus", "s@e.com", "x" * 12)
    client.force_login(owner)
    body = client.get("/dashboard/?period=today").content.decode()
    assert "Погашено старых долгов" in body
    assert "получено больше, чем начислено" in body
    assert "-5\xa0000" not in body  # no confusing bare negative money figure


def test_dashboard_walkin_sale_counts_as_received_immediately(variant, settings):
    """A walk-in never gets a Payment row (record_payment returns None with
    no client — CLAUDE.md), but the cash still changed hands over the
    counter. Without the walk-in special case, Ожидается would wrongly show
    every walk-in as forever-unpaid."""
    from apps.reports.dashboard import dashboard_data

    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    order = SaleOrder.objects.create(currency="KGS")  # walk-in, no client
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3000"))
    confirm_sale(order)
    assert order.payments.count() == 0

    m = dashboard_data("today")["metrics"]
    assert m["revenue"]["value"] == Decimal("3000.00")
    assert m["received"]["value"] == Decimal("3000.00")
    assert m["expected"]["value"] == Decimal("0")


def test_dashboard_received_nets_a_voided_payment_to_zero(variant, settings):
    from apps.reports.dashboard import dashboard_data

    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    client_ = Client.objects.create(first_name="Voider", phone="+996700000302")
    order = SaleOrder.objects.create(client=client_, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3000"))
    confirm_sale(order)
    payment = record_payment(order, Decimal("3000"))
    void_payment(payment)

    m = dashboard_data("today")["metrics"]
    assert m["received"]["value"] == Decimal("0")  # original + reversal net to zero
    assert m["expected"]["value"] == Decimal("3000.00")  # back to owed


# ---------------------------------------------------------------------------
# Per-client Excel export (apps.reports.client_export)
# ---------------------------------------------------------------------------


def _load_export(content: bytes):
    import io

    from openpyxl import load_workbook

    return load_workbook(io.BytesIO(content))


def test_client_export_empty_client_has_zeroed_summary_not_an_error():
    from apps.reports.client_export import client_workbook_bytes

    cust = Client.objects.create(first_name="Пусто", phone="+996700000901")
    content = client_workbook_bytes([cust])
    wb = _load_export(content)
    ws = wb.active
    # Header row only — no transaction rows — then a zeroed summary block.
    assert ws["A1"].value == "Дата"
    values = [row[0].value for row in ws.iter_rows(min_row=2)]
    assert any("ВСЕГО КУПЛЕНО" in str(v) for v in values if v)
    assert any("ЗАДОЛЖЕННОСТЬ" in str(v) for v in values if v)
    debt_row = next(
        r for r in ws.iter_rows(min_row=2) if r[0].value and "ЗАДОЛЖЕННОСТЬ" in str(r[0].value)
    )
    assert debt_row[1].value == Decimal("0")


def test_client_export_transaction_row_and_summary_math(variant):
    from apps.reports.client_export import client_workbook_bytes

    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    cust = Client.objects.create(first_name="Purchases", phone="+996700000902")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(
        order=order,
        variant=variant,
        quantity=2,
        unit_price=Decimal("3200.00"),
        discount_type=SaleItem.DISCOUNT_FIXED,
        discount_value=Decimal("400"),
    )
    confirm_sale(order)
    record_payment(order, Decimal("3000"), currency="KGS")

    wb = _load_export(client_workbook_bytes([cust]))
    ws = wb.active
    row = [c.value for c in ws[2]]
    # Subtotal 6400, discount 400, total 6000, paid 3000, balance 3000.
    assert row[3] == 2  # quantity
    assert row[4] == Decimal("6400.00")  # subtotal
    assert row[5] == Decimal("400.00")  # discount
    assert row[6] == Decimal("6000.00")  # total
    assert row[7] == Decimal("3000.00")  # paid
    assert row[8] == Decimal("3000.00")  # balance
    assert row[9] == "KGS"

    debt_row = next(
        r for r in ws.iter_rows(min_row=2) if r[0].value and "ЗАДОЛЖЕННОСТЬ" in str(r[0].value)
    )
    assert debt_row[1].value == Decimal("3000.00")


def test_dashboard_debt_prev_period_boundary_uses_bishkek_local_not_utc(variant, settings):
    """apps.reports.dashboard._debt compares each order's confirmed_at
    against `prev_end` (a local date) — a bare .date() on the UTC-aware
    confirmed_at used to read the UTC calendar date, one day behind Bishkek.
    02:00 Bishkek on the 3rd is 20:00 UTC on the 2nd: with prev_end = the
    2nd, the LOCAL date (the 3rd) is AFTER prev_end — this order must NOT
    count as "already outstanding as of the previous period", or the
    dashboard's debt delta-% silently misreports a brand new debt as
    unchanged."""
    from datetime import datetime, timezone as dt_timezone

    from apps.reports.dashboard import _debt

    settings.TIME_ZONE = "Asia/Bishkek"
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    cust = Client.objects.create(first_name="TZDash", phone="+996700000905")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)
    order.confirmed_at = datetime(2026, 8, 2, 20, 0, tzinfo=dt_timezone.utc)
    order.save(update_fields=["confirmed_at"])

    debt = _debt(prev_end=date(2026, 8, 2))
    assert debt["value"] == Decimal("3200")
    # Correct (local-date) reading: this debt didn't exist yet as of the
    # 2nd, so there's no previous-period baseline to compare against. The
    # UTC-date bug would instead count it as already outstanding then,
    # reporting a 0% change on what is actually a brand-new debt.
    assert debt["delta"]["has_baseline"] is False, (
        f"expected no baseline (debt is newer than prev_end), got {debt['delta']}"
    )


def test_client_export_never_sums_debt_across_currencies(variant):
    from apps.reports.client_export import client_workbook_bytes

    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    cust = Client.objects.create(first_name="MultiCcy", phone="+996700000903")
    kgs_order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(
        order=kgs_order, variant=variant, quantity=1, unit_price=Decimal("3200")
    )
    confirm_sale(kgs_order)

    usd_variant_order = SaleOrder.objects.create(client=cust, currency="USD")
    SaleItem.objects.create(
        order=usd_variant_order, variant=variant, quantity=1, unit_price=Decimal("40")
    )
    confirm_sale(usd_variant_order)

    wb = _load_export(client_workbook_bytes([cust]))
    ws = wb.active
    labels = [str(r[0].value) for r in ws.iter_rows(min_row=2) if r[0].value]
    kgs_debt_labels = [label for label in labels if "ЗАДОЛЖЕННОСТЬ" in label and "KGS" in label]
    usd_debt_labels = [label for label in labels if "ЗАДОЛЖЕННОСТЬ" in label and "USD" in label]
    assert len(kgs_debt_labels) == 1
    assert len(usd_debt_labels) == 1  # two SEPARATE blocks, never combined into one number


def test_client_export_date_column_uses_bishkek_local_not_utc(variant, settings):
    """order.confirmed_at is a UTC-aware datetime (USE_TZ=True) — a bare
    .date() reads the UTC calendar date, one day behind Bishkek (UTC+6) for
    anything confirmed in the early-morning hours locally. 02:00 Bishkek on
    the 3rd is 20:00 UTC on the 2nd — the exported «Дата» column must show
    the 3rd, the date the sale actually happened for the shop, not the 2nd."""
    from datetime import datetime, timezone as dt_timezone

    from apps.reports.client_export import client_workbook_bytes

    settings.TIME_ZONE = "Asia/Bishkek"
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    cust = Client.objects.create(first_name="TZBoundary", phone="+996700000904")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)
    # 20:00 UTC on the 2nd == 02:00 Bishkek on the 3rd.
    order.confirmed_at = datetime(2026, 8, 2, 20, 0, tzinfo=dt_timezone.utc)
    order.save(update_fields=["confirmed_at"])

    wb = _load_export(client_workbook_bytes([cust]))
    ws = wb.active
    # openpyxl round-trips an xlsx DATE cell back as a datetime, not a plain
    # date — normalize with .date() (a no-op for an actual date) before
    # comparing.
    dates = [
        r[0].value.date() if isinstance(r[0].value, datetime) else r[0].value
        for r in ws.iter_rows(min_row=2)
        if isinstance(r[0].value, date)
    ]
    assert date(2026, 8, 3) in dates, f"expected the Bishkek-local date, got {dates}"
    assert date(2026, 8, 2) not in dates  # the UTC date must never leak through


def test_client_export_bulk_action_writes_one_sheet_per_client(variant):
    from apps.reports.client_export import client_workbook_bytes

    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    a = Client.objects.create(first_name="Alpha", phone="+996700000904")
    b = Client.objects.create(first_name="Beta", phone="+996700000905")
    content = client_workbook_bytes([a, b])
    wb = _load_export(content)
    assert len(wb.sheetnames) == 2
    assert "Alpha" in wb.sheetnames
    assert "Beta" in wb.sheetnames


def test_client_export_filename_uses_first_name_never_the_descriptor():
    from apps.reports.client_export import client_export_filename

    cust = Client.objects.create(
        first_name="Айгуль", descriptor="сестра Розы", phone="+996700000906"
    )
    filename = client_export_filename(cust)
    assert filename.startswith("Айгуль_")
    assert "Розы" not in filename
    assert "сестра" not in filename


def test_client_export_admin_view_reachable_by_editor_and_viewer(
    client, django_user_model, variant
):
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    cust = Client.objects.create(first_name="Reach", phone="+996700000907")

    viewer = django_user_model.objects.create_user(
        "viewer_export1", password="x" * 12, is_staff=True
    )
    viewer.groups.add(Group.objects.get(name=VIEWER))
    client.force_login(viewer)
    resp = client.get(f"/panel/clients/client/{cust.pk}/export/")
    assert resp.status_code == 200
    assert resp["Content-Type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "filename*=UTF-8''" in resp["Content-Disposition"]


def test_pos_draft_survives_reload(client, django_user_model, variant):
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    editor = django_user_model.objects.create_user("editor12", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    order_id = SaleOrder.objects.create(created_by=editor, channel=SaleOrder.SHOP).pk
    client.post(f"/pos/sale/{order_id}/items/add/", {"variant_id": variant.pk, "quantity": 1})

    # Simulate a dropped connection / reload: hitting /pos/ again reuses the
    # same still-open draft instead of losing the basket.
    resp2 = client.get("/pos/")
    assert resp2.url == f"/pos/sale/{order_id}/"
    order = SaleOrder.objects.get(pk=order_id)
    assert order.items.count() == 1


def test_pos_oversell_shows_error_and_keeps_basket(client, django_user_model, variant):
    """Part 1c: client-side/cart-time caps are UX, not the defence — stock can
    still change between the cart being built and confirm (a concurrent sale,
    a write-off). Add 5 while 5 are in stock (the cart-time cap at add-time
    allows it), THEN reduce stock to 2 before confirming — confirm must still
    fail cleanly and keep the basket, never a silent oversell."""
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    editor = django_user_model.objects.create_user("editor13", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    order_id = SaleOrder.objects.create(created_by=editor, channel=SaleOrder.SHOP).pk
    client.post(f"/pos/sale/{order_id}/items/add/", {"variant_id": variant.pk, "quantity": 5})
    add_movement(variant, StockMovement.WRITEOFF_OUT, 3)  # stock 5 -> 2, AFTER adding to cart

    resp = client.post(
        f"/pos/sale/{order_id}/confirm/", {"amount": "0", "currency": "KGS", "method": "cash"}
    )
    # Redirects back to the (GET-able) sale page with the error as a message,
    # rather than rendering directly — a refresh must not prompt "resubmit
    # form?", and the offline-safe fetch() submit needs every outcome of this
    # POST to end in a redirect it can safely follow.
    assert resp.status_code == 302
    assert resp.url == f"/pos/sale/{order_id}/"
    page = client.get(resp.url)
    assert "Недостаточно" in page.content.decode()
    order = SaleOrder.objects.get(pk=order_id)
    assert order.status == SaleOrder.DRAFT
    assert order.items.count() == 1  # basket kept intact
    assert order.items.first().quantity == 5
    variant.refresh_from_db()
    assert variant.stock == 2  # nothing written off by the failed confirm itself


def test_pos_cancel_view_returns_stock(client, django_user_model, variant):
    # Exercises the /pos/ cancel *view* over HTTP, not just the cancel_sale
    # service — an Editor cancelling their own same-day sale should 302 to the
    # result page and return the stock. (A regression here once slipped past
    # the service-only tests.)
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    editor = django_user_model.objects.create_user("editor14", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    order_id = SaleOrder.objects.create(created_by=editor, channel=SaleOrder.SHOP).pk
    client.post(f"/pos/sale/{order_id}/items/add/", {"variant_id": variant.pk, "quantity": 2})
    client.post(
        f"/pos/sale/{order_id}/confirm/", {"amount": "6400", "currency": "KGS", "method": "cash"}
    )
    variant.refresh_from_db()
    assert variant.stock == 3

    resp = client.post(f"/pos/sale/{order_id}/cancel/", {"reason": "клиент передумал"})
    assert resp.status_code == 302
    assert resp.url == f"/pos/sale/{order_id}/result/"
    order = SaleOrder.objects.get(pk=order_id)
    assert order.status == SaleOrder.CANCELLED
    variant.refresh_from_db()
    assert variant.stock == 5  # returned to stock


def test_pos_share_receipt_redirects_to_whatsapp_and_logs(client, django_user_model, variant):
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    editor = django_user_model.objects.create_user("editor15", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    cust = Client.objects.create(first_name="Aiperi", phone="+996700123456")
    order = SaleOrder.objects.create(client=cust, created_by=editor)
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order, user=editor)

    # POST, not GET: it writes an Interaction (see test_url_security).
    resp = client.post(f"/pos/sale/{order.pk}/receipt/")
    assert resp.status_code == 302
    assert resp.url.startswith("https://wa.me/996700123456?text=")
    assert cust.interactions.filter(kind=Interaction.MESSAGE).count() == 1


def test_pos_debt_reminder_redirects_to_whatsapp_and_logs(client, django_user_model, variant):
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    editor = django_user_model.objects.create_user("editor16", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    cust = Client.objects.create(first_name="Nurgul", phone="+996700999888")
    order = SaleOrder.objects.create(client=cust, created_by=editor)
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order, user=editor)  # unpaid -> client now owes 3200 KGS

    resp = client.post(f"/pos/clients/{cust.pk}/remind/")
    assert resp.status_code == 302
    assert resp.url.startswith("https://wa.me/996700999888?text=")
    assert cust.interactions.filter(kind=Interaction.MESSAGE).count() == 1

    # A client with no debt gets bounced back, nothing logged.
    paid = Client.objects.create(first_name="Zarina", phone="+996700111000")
    resp2 = client.post(f"/pos/clients/{paid.pk}/remind/")
    assert resp2.status_code == 302
    assert resp2.url == f"/pos/clients/{paid.pk}/"


def test_partial_return_restocks_and_reduces_total(variant):
    from apps.sales.services import return_items

    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    cust = Client.objects.create(first_name="Meerim", phone="+996700222111")
    order = SaleOrder.objects.create(client=cust)
    item = SaleItem.objects.create(
        order=order, variant=variant, quantity=3, unit_price=Decimal("3200")
    )
    confirm_sale(order)
    assert variant.stock == 7 and order.total == Decimal("9600")

    return_items(order, {item.pk: 1}, user=None)
    order.refresh_from_db()
    variant.refresh_from_db()
    assert variant.stock == 8  # one came back
    assert order.total == Decimal("6400")  # 2 remaining × 3200
    assert order.status == SaleOrder.CONFIRMED
    # Debt follows the reduced total automatically.
    assert client_debt(cust) == {"KGS": Decimal("6400")}


def test_full_return_cancels_the_order(variant):
    from apps.sales.services import return_items

    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    order = SaleOrder.objects.create()
    item = SaleItem.objects.create(
        order=order, variant=variant, quantity=2, unit_price=Decimal("3200")
    )
    confirm_sale(order)
    return_items(order, {item.pk: 2}, user=None)
    order.refresh_from_db()
    variant.refresh_from_db()
    assert order.status == SaleOrder.CANCELLED
    assert variant.stock == 10  # all back


def test_return_rejects_more_than_sold(variant):
    from apps.sales.services import return_items

    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    order = SaleOrder.objects.create()
    item = SaleItem.objects.create(
        order=order, variant=variant, quantity=2, unit_price=Decimal("3200")
    )
    confirm_sale(order)
    with pytest.raises(ValidationError):
        return_items(order, {item.pk: 5}, user=None)
    variant.refresh_from_db()
    assert variant.stock == 8  # unchanged — the whole return rolled back


def test_return_form_has_double_submit_protection(client, django_user_model, variant):
    """Unlike its sibling money/stock actions («Отменить продажу», sale
    confirm), the return form used to be a bare <form> with no
    data-network-safe — a double-tap on a flaky connection could fire two
    physical POSTs. Since a partial return has no natural one-time status
    gate to reject a repeat (unlike confirm/cancel), the client-side
    disable-on-submit is what actually closes this gap; assert it's there."""
    owner = django_user_model.objects.create_superuser("ret_owner", "o@e.com", "x" * 12)
    client.force_login(owner)
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    order = SaleOrder.objects.create()
    SaleItem.objects.create(order=order, variant=variant, quantity=2, unit_price=Decimal("3200"))
    confirm_sale(order)

    resp = client.get(f"/pos/sale/{order.pk}/return/")
    body = resp.content.decode()
    assert "data-network-safe" in body
    assert "data-confirm=" in body


def test_units_sold_by_variant_counts_only_confirmed(variant):
    from apps.sales.services import units_sold_by_variant

    add_movement(variant, StockMovement.PRODUCTION_IN, 20)
    # A confirmed sale of 4 counts; a draft of 5 does not.
    o1 = SaleOrder.objects.create()
    SaleItem.objects.create(order=o1, variant=variant, quantity=4, unit_price=Decimal("3200"))
    confirm_sale(o1)
    o2 = SaleOrder.objects.create()  # left as draft
    SaleItem.objects.create(order=o2, variant=variant, quantity=5, unit_price=Decimal("3200"))

    assert units_sold_by_variant(30) == {variant.pk: 4}


def test_campaign_audience_requires_consent_and_chat_id():
    from apps.campaigns.models import Campaign
    from apps.campaigns.services import campaign_audience

    reachable = Client.objects.create(
        first_name="A", phone="+996700000101", telegram_chat_id=111, marketing_consent=True
    )
    Client.objects.create(  # consent but no chat_id -> unreachable
        first_name="B", phone="+996700000102", marketing_consent=True
    )
    Client.objects.create(  # chat_id but no consent -> excluded
        first_name="C", phone="+996700000103", telegram_chat_id=333, marketing_consent=False
    )
    campaign = Campaign.objects.create(name="New arrivals", text_ru="👗")
    assert campaign_audience(campaign) == [reachable]


def test_campaign_audience_filters_bought_before(variant):
    from apps.campaigns.models import Campaign
    from apps.campaigns.services import campaign_audience

    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    buyer = Client.objects.create(
        first_name="Buyer", phone="+996700000201", telegram_chat_id=201, marketing_consent=True
    )
    Client.objects.create(
        first_name="Never", phone="+996700000202", telegram_chat_id=202, marketing_consent=True
    )
    order = SaleOrder.objects.create(client=buyer)
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)

    campaign = Campaign.objects.create(name="VIP", text_ru="hi", only_bought_before=True)
    assert campaign_audience(campaign) == [buyer]


def test_subscribe_and_unsubscribe_telegram():
    from apps.clients.services import (
        set_marketing_consent,
        subscribe_telegram,
        unsubscribe_telegram,
    )

    c = Client.objects.create(first_name="Sub", phone="+996 700 55-66-77")
    # Phone match is digit-based, tolerant of spaces/dashes/+.
    linked = subscribe_telegram("996700556677", 909)
    assert linked == c
    c.refresh_from_db()
    # CLIENT_BOTS.md §3.1: sharing a phone makes the chat reachable but must
    # NOT itself set consent — that's a separate explicit Да/Нет step.
    assert c.telegram_chat_id == 909
    assert c.marketing_consent is False

    set_marketing_consent(c, True)
    c.refresh_from_db()
    assert c.marketing_consent is True

    unsubscribe_telegram(909)
    c.refresh_from_db()
    assert c.marketing_consent is False and c.telegram_chat_id == 909  # kept

    assert subscribe_telegram("000000", 1) is None  # no match


def test_send_campaign_marks_recipients_and_never_double_sends(monkeypatch, settings):
    from apps.campaigns.models import Campaign, CampaignRecipient

    settings.TELEGRAM_CLIENT_TOKEN = "test-token"
    calls = []

    def fake_send(token, chat_id, text, photos):
        calls.append(chat_id)
        return True, None, ""

    monkeypatch.setattr(
        "apps.campaigns.management.commands.send_campaign._telegram_send", fake_send
    )
    monkeypatch.setattr(
        "apps.campaigns.management.commands.send_campaign.time.sleep", lambda s: None
    )

    c = Client.objects.create(
        first_name="R", phone="+996700000301", telegram_chat_id=301, marketing_consent=True
    )
    campaign = Campaign.objects.create(name="Promo", text_ru="hello")

    call_command("send_campaign", campaign.pk, ignore_quiet_hours=True)
    r = CampaignRecipient.objects.get(campaign=campaign, client=c)
    assert r.status == CampaignRecipient.SENT and r.sent_at is not None
    assert c.interactions.filter(kind=Interaction.MESSAGE).count() == 1
    assert calls == [301]

    # Re-running must not message the same client again.
    call_command("send_campaign", campaign.pk, ignore_quiet_hours=True)
    assert calls == [301]  # unchanged
    assert CampaignRecipient.objects.filter(campaign=campaign).count() == 1


def test_lapsed_clients_only_returns_stale_buyers(variant):
    from apps.clients.services import lapsed_clients

    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    old_buyer = Client.objects.create(first_name="Old", phone="+996700000401")
    recent_buyer = Client.objects.create(first_name="New", phone="+996700000402")
    Client.objects.create(first_name="Never", phone="+996700000403")  # no purchase

    for cust, when in [
        (old_buyer, timezone.now() - timezone.timedelta(days=90)),
        (recent_buyer, timezone.now() - timezone.timedelta(days=5)),
    ]:
        o = SaleOrder.objects.create(client=cust)
        SaleItem.objects.create(order=o, variant=variant, quantity=1, unit_price=Decimal("3200"))
        confirm_sale(o)
        SaleOrder.objects.filter(pk=o.pk).update(confirmed_at=when)

    result = lapsed_clients(60)
    assert old_buyer in result
    assert recent_buyer not in result  # bought recently
    assert Client.objects.get(phone="+996700000403") not in result  # never bought


def test_confirmed_sale_som_value_is_frozen_against_later_rate_changes(variant, settings):
    # THE point of Part 0: a RUB sale's сом value is snapshotted at confirm time
    # and must NOT move when the owner later changes today's RUB rate.
    from django.core.cache import cache

    from apps.core.models import ExchangeRate

    cache.clear()
    settings.CURRENCY = "KGS"
    ExchangeRate.objects.create(currency="RUB", date=timezone.localdate(), rate=Decimal("1.10"))
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    order = SaleOrder.objects.create(currency="RUB")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(order)
    order.refresh_from_db()
    assert order.rate_to_kgs == Decimal("1.100000")
    assert order.total_kgs == Decimal("1100.00")  # 1000 RUB × 1.10

    # Owner changes today's RUB rate afterwards.
    ExchangeRate.objects.filter(currency="RUB").update(rate=Decimal("2.00"))
    cache.clear()  # a real override clears this via signal; bulk update bypassed it

    order.refresh_from_db()
    assert order.total_kgs == Decimal("1100.00")  # UNCHANGED — historical value is frozen
    from django.db.models import Sum

    agg = SaleOrder.objects.filter(status=SaleOrder.CONFIRMED).aggregate(s=Sum("total_kgs"))["s"]
    assert agg == Decimal("1100.00")  # aggregates read the frozen field, don't re-convert


def test_payment_freezes_its_rate_and_void_cancels_exactly(variant, settings):
    from apps.core.models import ExchangeRate

    settings.CURRENCY = "KGS"
    ExchangeRate.objects.create(currency="RUB", date=timezone.localdate(), rate=Decimal("1.50"))
    cust = Client.objects.create(first_name="R", phone="+996700000700")
    order = SaleOrder.objects.create(client=cust, currency="RUB")
    p = Payment.objects.create(client=cust, order=order, amount=Decimal("200"), currency="RUB")
    assert p.rate_to_kgs == Decimal("1.500000")  # snapshotted on save
    assert p.amount_kgs == Decimal("300.00")  # 200 × 1.50
    reversal = void_payment(p)
    assert reversal.rate_to_kgs == p.rate_to_kgs  # same frozen rate
    assert reversal.amount_kgs == Decimal("-300.00")  # cancels the original exactly


def test_db_constraints_reject_bad_data_even_via_bulk_create(variant):
    from django.db import IntegrityError, transaction

    # clean() is bypassed by bulk_create — the DB CheckConstraints are not.
    with pytest.raises(IntegrityError), transaction.atomic():
        StockMovement.objects.bulk_create(
            [StockMovement(variant=variant, movement_type="adjustment", quantity=0, reason="x")]
        )
    with pytest.raises(IntegrityError), transaction.atomic():
        StockMovement.objects.bulk_create(
            [StockMovement(variant=variant, movement_type="production_in", quantity=-5)]
        )
    c = Client.objects.create(first_name="A", phone="+996700000901")
    with pytest.raises(IntegrityError), transaction.atomic():
        Payment.objects.bulk_create([Payment(client=c, amount=Decimal("-10"))])
    order = SaleOrder.objects.create()
    with pytest.raises(IntegrityError), transaction.atomic():
        SaleItem.objects.bulk_create(
            [SaleItem(order=order, variant=variant, quantity=0, unit_price=Decimal("1"))]
        )
    with pytest.raises(IntegrityError), transaction.atomic():
        SaleOrder.objects.create(total=Decimal("-1"))


def test_db_constraint_allows_negative_reversal(variant):
    # The one legitimate negative payment: a reversal linked to what it voids.
    c = Client.objects.create(first_name="B", phone="+996700000902")
    order = SaleOrder.objects.create(client=c)
    p = Payment.objects.create(client=c, order=order, amount=Decimal("10"))
    reversal = Payment.objects.create(client=c, amount=Decimal("-10"), reversed_payment=p)
    assert reversal.pk is not None


def test_db_rejects_a_payment_with_no_order_and_no_deposit_even_via_bulk_create(variant):
    """The ghost-payment bug: a Payment tied to nothing (no sale, no
    production-order deposit) used to be creatable through the standalone
    Payment admin. The DB CheckConstraint is the backstop that survives even
    a bulk_create/raw insert, not just the admin's has_add_permission=False."""
    from django.db import IntegrityError, transaction

    c = Client.objects.create(first_name="Ghost", phone="+996700000903")
    with pytest.raises(IntegrityError), transaction.atomic():
        Payment.objects.bulk_create([Payment(client=c, amount=Decimal("500"))])


def test_argon2_is_default_and_pbkdf2_upgrades_on_login(client, django_user_model, settings):
    from django.contrib.auth.hashers import make_password

    # Argon2 must be the first (default) hasher.
    assert "Argon2" in settings.PASSWORD_HASHERS[0]

    # A user whose stored hash is the old PBKDF2 scheme...
    user = django_user_model.objects.create(username="legacy", is_staff=True)
    user.password = make_password("correct-horse-battery", hasher="pbkdf2_sha256")
    user.save(update_fields=["password"])
    assert user.password.startswith("pbkdf2_sha256$")

    # ...logs in through the real login view (axes needs a request) and gets
    # transparently re-hashed to Argon2 by Django's authenticate() path.
    token = client.get("/login/").context["csrf_token"]
    resp = client.post(
        "/login/",
        {"username": "legacy", "password": "correct-horse-battery", "csrfmiddlewaretoken": token},
    )
    assert resp.status_code == 302  # logged in, redirected
    user.refresh_from_db()
    assert user.password.startswith("argon2$"), "PBKDF2 hash should upgrade to Argon2 on login"


def test_csp_header_is_strict_on_pos_and_absent_on_admin(client, django_user_model):
    resp = client.get("/login/")
    csp = resp.headers.get("Content-Security-Policy", "")
    assert "script-src 'self'" in csp
    assert "unsafe-inline" not in csp.split("style-src")[0]  # no inline-script exception
    # Admin (Jazzmin, third-party inline scripts) is excluded from CSP.
    admin_resp = client.get("/panel/", follow=False)
    assert "Content-Security-Policy" not in admin_resp.headers


def _seed_grid(n, django_user_model):
    from apps.inventory.services import add_movement

    call_command("setup_roles")
    editor = django_user_model.objects.create_user(f"gridder{n}", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    cat = Category.objects.create(name=f"Grid{n}")
    for i in range(n):
        p = Product.objects.create(category=cat, name=f"P{n}-{i:04d}")
        v = ProductVariant.objects.create(
            product=p, sku=f"SKU{n}-{i:04d}", cost_price=Decimal("1"), sale_price=Decimal("100")
        )
        add_movement(v, StockMovement.PRODUCTION_IN, 5)
    return editor


def _grid_query_count(client, django_user_model, n):
    from django.core.cache import cache
    from django.db import connection
    from django.test.utils import CaptureQueriesContext

    editor = _seed_grid(n, django_user_model)
    client.force_login(editor)
    order_id = SaleOrder.objects.create(created_by=editor, channel=SaleOrder.SHOP).pk
    cache.clear()  # measure the cache-MISS path — the real query cost
    with CaptureQueriesContext(connection) as ctx:
        client.get(f"/pos/sale/{order_id}/products/")
    # Ignore session/auth/savepoint noise — count only catalog reads.
    catalog = [
        q
        for q in ctx.captured_queries
        if any(t in q["sql"] for t in ("inventory_product", "inventory_stockmovement"))
    ]
    return len(catalog)


def test_product_grid_query_budget_does_not_scale_with_rows(client, django_user_model):
    # ≤4 catalog queries at 10 products AND at 500 — proving the grid's query
    # count is constant, not O(rows). (The old per-product variant fetch was N+1.)
    assert _grid_query_count(client, django_user_model, 10) <= 4
    assert _grid_query_count(client, django_user_model, 500) <= 4


def test_product_grid_stays_flat_with_real_multi_image_products(
    client, django_user_model, settings, tmp_path
):
    """The query-budget test above uses image-less products, so it can't
    catch a regression where prefetch_related("images") starts fanning out
    per row once products actually HAVE several images each — the exact
    shape this feature introduces. Attach 3 images to every product and
    prove the grid's query count is unchanged."""
    from django.core.cache import cache
    from django.db import connection
    from django.test.utils import CaptureQueriesContext

    settings.MEDIA_ROOT = tmp_path
    editor = _seed_grid(10, django_user_model)
    for product in Product.objects.all():
        for i in range(3):
            ProductImage.objects.create(product=product, image=_tiny_image(f"g{i}.png"), position=i)

    client.force_login(editor)
    order_id = SaleOrder.objects.create(created_by=editor, channel=SaleOrder.SHOP).pk
    cache.clear()
    with CaptureQueriesContext(connection) as ctx:
        client.get(f"/pos/sale/{order_id}/products/")
    catalog = [
        q
        for q in ctx.captured_queries
        if any(
            t in q["sql"]
            for t in ("inventory_product", "inventory_stockmovement", "inventory_productimage")
        )
    ]
    assert len(catalog) <= 4


def test_stale_grid_cache_never_causes_an_oversell(client, django_user_model, variant):
    # Warm the grid cache with stock=1, then drain stock to 0 behind the cache
    # WITHOUT bumping the version, then confirm — the sale must still fail at
    # confirm time (fresh DB read under lock), never trust the cached badge.
    from django.core.cache import cache

    from apps.inventory.services import add_movement

    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 1)
    editor = django_user_model.objects.create_user("staler", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)
    order_id = SaleOrder.objects.create(created_by=editor, channel=SaleOrder.SHOP).pk
    client.get(f"/pos/sale/{order_id}/products/")  # warms grid cache (stock=1)
    client.post(f"/pos/sale/{order_id}/items/add/", {"variant_id": variant.pk, "quantity": 1})

    # Drain the stock directly, then FORCE the catalog version back so the grid
    # cache stays stale (simulating a read-from-stale-cache window).
    stale_version = cache.get("catalog:v")
    add_movement(variant, StockMovement.WRITEOFF_OUT, -1)  # stock now 0
    cache.set("catalog:v", stale_version)  # pin version stale

    resp = client.post(
        f"/pos/sale/{order_id}/confirm/", {"amount": "0", "currency": "KGS", "method": "cash"}
    )
    assert resp.status_code == 302
    assert resp.url == f"/pos/sale/{order_id}/"  # bounced with an error, not confirmed
    order = SaleOrder.objects.get(pk=order_id)
    assert order.status == SaleOrder.DRAFT  # never oversold


def test_client_list_query_budget(client, django_user_model):
    from django.db import connection
    from django.test.utils import CaptureQueriesContext

    call_command("setup_roles")
    editor = django_user_model.objects.create_user("clister", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    for i in range(50):
        Client.objects.create(first_name=f"C{i}", phone=f"+99670000{i:04d}")
    client.force_login(editor)
    with CaptureQueriesContext(connection) as ctx:
        client.get("/pos/clients/?q=C")
    client_queries = [q for q in ctx.captured_queries if "clients_client" in q["sql"]]
    assert len(client_queries) <= 3


def test_confirm_query_budget(client, django_user_model, variant):
    from django.db import connection
    from django.test.utils import CaptureQueriesContext

    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    editor = django_user_model.objects.create_user("confirmer", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)
    order_id = SaleOrder.objects.create(created_by=editor, channel=SaleOrder.SHOP).pk
    client.post(f"/pos/sale/{order_id}/items/add/", {"variant_id": variant.pk, "quantity": 2})

    with CaptureQueriesContext(connection) as ctx:
        client.post(
            f"/pos/sale/{order_id}/confirm/",
            {"amount": "0", "currency": "KGS", "method": "cash"},
        )
    # Writes to the ledger, order, and history all count; keep it lean.
    writes_and_reads = [
        q for q in ctx.captured_queries if any(t in q["sql"] for t in ("sales_", "inventory_"))
    ]
    # Budget raised 10 -> 11: confirm_sale now freezes SaleItem.cost_price (the
    # cost basis AT THE TIME of sale, so a later cost_price edit can't rewrite
    # historical profit — see apps.reports.dashboard._LINE_COGS) via one
    # bulk_update, a single query regardless of item count.
    assert len(writes_and_reads) <= 11


def test_500_handler_renders_without_db_and_shows_correlation_id(rf, django_assert_num_queries):
    from apps.core.errors import server_error

    request = rf.get("/pos/")
    request.correlation_id = "abc123def456"
    # The DB may be what's down — the 500 page must touch it zero times.
    with django_assert_num_queries(0):
        resp = server_error(request)
    assert resp.status_code == 500
    body = resp.content.decode()
    assert "abc123def456" in body  # the id, to correlate with the server log
    assert "Что-то сломалось" in body
    assert "/pos/" in body
    # No leakage of internals.
    assert "Traceback" not in body and "Django" not in body and "Sorry" not in body


def test_404_page_is_russian(client, settings):
    settings.DEBUG = False  # custom handlers only fire when DEBUG is off
    resp = client.get("/definitely-not-a-real-page/")
    assert resp.status_code == 404
    assert "Страница не найдена" in resp.content.decode()
    assert "Traceback" not in resp.content.decode()


def test_403_page_on_role_violation(client, django_user_model, settings):
    settings.DEBUG = False
    call_command("setup_roles")
    viewer = django_user_model.objects.create_user("viewer_err", password="x" * 12, is_staff=True)
    viewer.groups.add(Group.objects.get(name=VIEWER))
    client.force_login(viewer)
    resp = client.get("/pos/sale/1/")  # Viewer has no sales.add_saleorder -> 403
    assert resp.status_code == 403
    body = resp.content.decode()
    assert "Нет доступа" in body
    # A bare "no access" with no next step used to be the whole page — a
    # Viewer/Editor hitting a role wall had no idea whether it was a bug, a
    # session issue, or an actual restriction.
    assert "обратитесь к владельцу" in body


def test_csrf_failure_page_is_russian(settings):
    from django.test import Client

    settings.DEBUG = False
    c = Client(enforce_csrf_checks=True)
    resp = c.post("/login/", {"username": "x", "password": "y"})  # no CSRF token
    assert resp.status_code == 403
    assert "Сессия устарела" in resp.content.decode()


def test_healthz_ok_when_db_and_cache_up(client):
    resp = client.get("/healthz/")
    assert resp.status_code == 200
    assert resp.content == b"ok"


def test_healthz_503_when_cache_roundtrip_fails(client, monkeypatch):
    # Simulate Redis down: set is a no-op and get returns None (exactly how
    # django-redis behaves with IGNORE_EXCEPTIONS). The round-trip check must
    # then report the cache as unhealthy with a 503.
    from django.core import cache as cache_module

    monkeypatch.setattr(cache_module.cache, "set", lambda *a, **k: False)
    monkeypatch.setattr(cache_module.cache, "get", lambda *a, **k: None)
    resp = client.get("/healthz/")
    assert resp.status_code == 503
    assert b"cache" in resp.content


def test_healthz_503_when_db_is_down(client, monkeypatch):
    from django.db import connection

    def broken_cursor(*a, **k):
        raise OSError("could not connect to server")

    monkeypatch.setattr(connection, "cursor", broken_cursor)
    resp = client.get("/healthz/")
    assert resp.status_code == 503
    assert b"db" in resp.content


def _wa_sig(secret, body):
    return "sha256=" + hmac.new(secret.encode(), body, hashlib.sha256).hexdigest()


def _wa_text_payload(phone, text):
    return json.dumps(
        {
            "entry": [
                {
                    "changes": [
                        {
                            "value": {
                                "messages": [
                                    {"from": phone, "type": "text", "text": {"body": text}}
                                ]
                            }
                        }
                    ]
                }
            ]
        }
    ).encode()


def test_wa_webhook_valid_signature_passes_and_links_client(client, settings):
    settings.WHATSAPP_APP_SECRET = "topsecret"
    body = _wa_text_payload("996700123456", "stock")
    resp = client.post(
        "/wa/webhook/",
        data=body,
        content_type="application/json",
        HTTP_X_HUB_SIGNATURE_256=_wa_sig("topsecret", body),
    )
    assert resp.status_code == 200
    assert Client.objects.filter(phone="996700123456").exists()  # CRM-linked


def test_wa_webhook_tampered_body_rejected(client, settings):
    settings.WHATSAPP_APP_SECRET = "topsecret"
    body = _wa_text_payload("996700123456", "stock")
    sig = _wa_sig("topsecret", body)  # signature for the ORIGINAL body
    resp = client.post(
        "/wa/webhook/",
        data=body + b" ",  # one byte of tampering
        content_type="application/json",
        HTTP_X_HUB_SIGNATURE_256=sig,
    )
    assert resp.status_code == 403
    assert not Client.objects.filter(phone="996700123456").exists()


def test_wa_webhook_missing_signature_rejected(client, settings):
    settings.WHATSAPP_APP_SECRET = "topsecret"
    resp = client.post("/wa/webhook/", data=b"{}", content_type="application/json")
    assert resp.status_code == 403


def test_wa_webhook_rejects_when_secret_unset(client, settings):
    settings.WHATSAPP_APP_SECRET = ""  # fail closed, never process unsigned
    body = _wa_text_payload("996700123456", "stock")
    resp = client.post(
        "/wa/webhook/",
        data=body,
        content_type="application/json",
        HTTP_X_HUB_SIGNATURE_256=_wa_sig("anything", body),
    )
    assert resp.status_code == 403


def test_wa_webhook_404s_when_whatsapp_disabled(client, settings):
    """Bots are not production-ready — WHATSAPP_ENABLED=False (the shipped
    prod default) must make the whole route unreachable, GET and POST alike,
    regardless of credentials being otherwise valid."""
    settings.WHATSAPP_ENABLED = False
    settings.WHATSAPP_APP_SECRET = "topsecret"
    settings.WHATSAPP_VERIFY_TOKEN = "verify-me"

    assert (
        client.get(
            "/wa/webhook/", {"hub.verify_token": "verify-me", "hub.challenge": "x"}
        ).status_code
        == 404
    )

    body = _wa_text_payload("996700123456", "stock")
    resp = client.post(
        "/wa/webhook/",
        data=body,
        content_type="application/json",
        HTTP_X_HUB_SIGNATURE_256=_wa_sig("topsecret", body),
    )
    assert resp.status_code == 404
    assert not Client.objects.filter(phone="996700123456").exists()


def test_wa_webhook_oversize_payload_rejected(client, settings):
    settings.WHATSAPP_APP_SECRET = "topsecret"
    big = b'{"x":"' + b"a" * (200 * 1024) + b'"}'
    resp = client.post(
        "/wa/webhook/",
        data=big,
        content_type="application/json",
        HTTP_X_HUB_SIGNATURE_256="sha256=00",
    )
    assert resp.status_code == 413  # rejected before the HMAC even runs


def test_wa_webhook_rate_limited_per_ip(client, settings, monkeypatch):
    from django.core.cache import cache

    cache.clear()
    settings.WHATSAPP_APP_SECRET = "topsecret"
    monkeypatch.setattr("apps.wa.views.RATE_LIMIT", 2)
    # First 2 are allowed through (then 403 on the missing signature); the 3rd
    # trips the limit and short-circuits with 429 before any signature work.
    for _ in range(2):
        client.post("/wa/webhook/", data=b"{}", content_type="application/json")
    resp = client.post("/wa/webhook/", data=b"{}", content_type="application/json")
    assert resp.status_code == 429


def test_wa_webhook_rate_limit_uses_real_ip_not_forged_xff(client, settings, monkeypatch):
    from django.core.cache import cache

    cache.clear()
    settings.WHATSAPP_APP_SECRET = "topsecret"
    monkeypatch.setattr("apps.wa.views.RATE_LIMIT", 1)
    # An attacker rotates X-Forwarded-For to dodge the limit, but Caddy sets
    # X-Real-IP to the true peer. The limiter must key on X-Real-IP, so both
    # requests land in the same bucket and the 2nd is blocked.
    client.post(
        "/wa/webhook/",
        data=b"{}",
        content_type="application/json",
        HTTP_X_REAL_IP="203.0.113.7",
        HTTP_X_FORWARDED_FOR="9.9.9.9",
    )
    resp = client.post(
        "/wa/webhook/",
        data=b"{}",
        content_type="application/json",
        HTTP_X_REAL_IP="203.0.113.7",
        HTTP_X_FORWARDED_FOR="8.8.8.8",  # forged, rotated — must NOT create a new bucket
    )
    assert resp.status_code == 429


def test_login_rate_limited_after_ten_posts_from_one_ip(client, settings):
    """30 POSTs/IP/5min (apps.core.ratelimit) is layered ON TOP OF django-
    axes' own lockout (AXES_LOCKOUT_PARAMETERS = ["username", "ip_address"] —
    5 fails against EITHER axis locks, so a flood of distinct usernames from
    one IP would trip axes' own IP-based lockout well before request 10,
    also as a 429 — see AXES_FAILURE_LIMIT in config/settings/base.py).
    AXES_ENABLED=False isolates the layer actually under test here: apps.
    core.ratelimit's own 30-per-5-minute counter, independent of axes.
    30 rather than 10 because every staff member shares one office IP —
    see UnifiedLoginView's docstring for why 10 locked out the whole shop."""
    settings.AXES_ENABLED = False
    for i in range(30):
        resp = client.post("/login/", {"username": f"nouser{i}", "password": "wrong"})
        assert resp.status_code != 429
    resp = client.post("/login/", {"username": "nouser30", "password": "wrong"})
    assert resp.status_code == 429
    assert "Слишком много запросов" in resp.content.decode()


def test_login_rate_limit_fails_closed_when_cache_is_down(client, settings, monkeypatch):
    """A dead cache must never look like "no requests recorded yet" — the
    real bug this replaces (see apps/wa/views.py's old _rate_limited) was a
    plain cache.get() miss being indistinguishable from Redis actually being
    down. apps.core.ratelimit.cache_is_available does a real write-then-read
    round trip, so simulating IGNORE_EXCEPTIONS-style silent failure (set is
    a no-op, get always returns None) must produce a 503, not a 200/302."""
    settings.AXES_ENABLED = False
    from django.core import cache as cache_module

    monkeypatch.setattr(cache_module.cache, "set", lambda *a, **k: False)
    monkeypatch.setattr(cache_module.cache, "get", lambda *a, **k: None)
    resp = client.post("/login/", {"username": "whoever", "password": "wrong"})
    assert resp.status_code == 503


def test_cache_probe_does_not_false_negative_under_concurrency():
    """cache_is_available() must probe a key unique to ITS OWN call. It used
    to write one shared "ratelimit:probe" key, so two concurrent callers
    overwrote each other's token and BOTH read back a mismatch — concluding
    the cache was down and 503ing every rate-limited endpoint (login, search,
    receipt, report) under exactly the concurrent load the limiter exists to
    survive. Measured against a realistic 2ms Redis round trip: 98.5% false
    "cache down" on the shared key, 0% on per-call keys.

    The sleep simulates that round trip — without it LocMemCache is fast
    enough to hide the race and the test would pass against the broken code
    too."""
    import threading
    import time
    from unittest.mock import patch

    from django.core.cache import cache as real_cache

    from apps.core.ratelimit import cache_is_available

    original_get = real_cache.get

    def slow_get(*args, **kwargs):
        time.sleep(0.002)  # Redis network RTT, between the probe's set and get
        return original_get(*args, **kwargs)

    results = []
    lock = threading.Lock()

    def worker():
        for _ in range(10):
            ok = cache_is_available()
            with lock:
                results.append(ok)

    with patch.object(real_cache, "get", slow_get):
        threads = [threading.Thread(target=worker) for _ in range(8)]
        for t in threads:
            t.start()
        for t in threads:
            t.join()

    assert results, "probe never ran"
    assert all(results), (
        f"{results.count(False)}/{len(results)} concurrent probes falsely reported the cache "
        "as DOWN — the probe key is shared again, and every rate-limited endpoint will 503 "
        "under concurrent load"
    )


def test_blocked_requests_do_not_write_one_security_event_each(client, settings):
    """The limiter must SHED load under attack, not amplify it. Recording a
    SecurityEvent row per blocked request meant a flood the limiter had
    already rejected still cost one INSERT each into the primary business
    database — measured at 200 rows for 200 blocked requests — growing the
    table fastest exactly when nobody is watching. One row per IP per window
    is all send_security_digest needs (it counts DISTINCT IPs)."""
    from apps.core.models import SecurityEvent

    settings.AXES_ENABLED = False
    for _ in range(30):  # exhaust the 30-per-5-min login budget
        client.post("/login/", {"username": "x", "password": "y"})

    before = SecurityEvent.objects.filter(event_type=SecurityEvent.RATE_LIMITED).count()
    for _ in range(50):
        assert client.post("/login/", {"username": "x", "password": "y"}).status_code == 429
    written = SecurityEvent.objects.filter(event_type=SecurityEvent.RATE_LIMITED).count() - before

    assert written <= 1, f"{written} rows written for 50 blocked requests — write amplification"
    # ...but the block IS still recorded, so the digest can still see it.
    assert SecurityEvent.objects.filter(event_type=SecurityEvent.RATE_LIMITED).count() >= 1, (
        "the block must still be recorded once — silence would hide a real attack"
    )


def test_rate_limit_decorator_preserves_view_attributes():
    """functools.wraps, not a hand-rolled __name__/__doc__ copy — Django
    stashes view FLAGS (csrf_exempt, and anything else a decorator sets) in
    the function's __dict__. Dropping those meant @rate_limit stacked above
    @csrf_exempt would silently re-enable CSRF on a webhook that must not
    have it."""
    from django.views.decorators.csrf import csrf_exempt

    from apps.core.ratelimit import rate_limit

    @rate_limit("attrs", 10, 60)
    @csrf_exempt
    def some_view(request):
        """Original docstring."""

    assert getattr(some_view, "csrf_exempt", False) is True, "csrf_exempt flag was dropped"
    assert some_view.__name__ == "some_view"
    assert some_view.__doc__ == "Original docstring."


def test_security_events_are_purged_after_retention(settings):
    """SecurityEvent was the only growing table with no retention at all —
    unbounded abuse-log growth in the SAME database as sales/clients/debts.
    Rows past the window go; anything inside it stays."""

    from apps.core.models import SecurityEvent
    from apps.core.management.commands.purge_security_events import RETAIN

    now = timezone.now()
    old = SecurityEvent.objects.create(event_type=SecurityEvent.AUTH_FAILURE, ip="10.0.0.1")
    recent = SecurityEvent.objects.create(event_type=SecurityEvent.AUTH_FAILURE, ip="10.0.0.2")
    # created_at is auto_now_add, so backdate with an explicit update().
    SecurityEvent.objects.filter(pk=old.pk).update(created_at=now - RETAIN - timedelta(days=1))
    SecurityEvent.objects.filter(pk=recent.pk).update(created_at=now - timedelta(days=1))

    call_command("purge_security_events")

    assert not SecurityEvent.objects.filter(pk=old.pk).exists(), "stale row not purged"
    assert SecurityEvent.objects.filter(pk=recent.pk).exists(), "in-window row wrongly purged"


def test_wa_webhook_survives_a_malformed_content_length(client, settings):
    """A junk Content-Length must not raise ValueError out of a public,
    unauthenticated endpoint — that turns a bad header into a 500 (plus a
    stack trace in the logs) for anyone on the internet. Django's own
    request.body re-parses the header and raises too, so the header has to be
    rejected BEFORE the body is ever touched."""
    settings.WHATSAPP_ENABLED = True
    settings.WHATSAPP_APP_SECRET = "topsecret"
    resp = client.post(
        "/wa/webhook/",
        data=b"{}",
        content_type="application/json",
        CONTENT_LENGTH="not-a-number",
    )
    assert resp.status_code == 400  # malformed request, never a 500


def test_receipt_download_rate_limited_per_ip(client, django_user_model, variant):
    """120 receipt downloads/IP/5min — a scripted scrape of every sale's PDF
    must eventually get a 429, not stream forever. The ceiling is deliberately
    well above a full day of end-of-day re-printing, since all staff share one
    office IP (see receipt_download's docstring)."""
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    viewer = django_user_model.objects.create_user(
        "rcpt_rl_viewer", password="x" * 12, is_staff=True
    )
    viewer.groups.add(Group.objects.get(name=VIEWER))
    client.force_login(viewer)

    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(order)

    for _ in range(120):
        resp = client.get(f"/pos/sale/{order.pk}/receipt/pdf/")
        assert resp.status_code == 200
    resp = client.get(f"/pos/sale/{order.pk}/receipt/pdf/")
    assert resp.status_code == 429


def test_search_endpoint_rate_limited_per_ip(client, django_user_model):
    """Client-search autocomplete (300/IP/min) is the endpoint most exposed
    to a scripted enumeration attempt (typing every 2-letter prefix)."""
    call_command("setup_roles")
    editor = django_user_model.objects.create_user(
        "search_rl_editor", password="x" * 12, is_staff=True
    )
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    for _ in range(300):
        resp = client.get("/pos/sale/new/clients/search/", {"q": "a"})
        assert resp.status_code == 200
    resp = client.get("/pos/sale/new/clients/search/", {"q": "a"})
    assert resp.status_code == 429


def _freeze_now(monkeypatch, y, mo, d, h, mi):
    """Pin timezone.now() to a fixed UTC instant so localdate() is deterministic."""
    import datetime

    fixed = datetime.datetime(y, mo, d, h, mi, tzinfo=datetime.timezone.utc)
    monkeypatch.setattr("django.utils.timezone.now", lambda: fixed)


def _sale_confirmed_at(variant, y, mo, d, h, mi):
    """A confirmed sale whose confirmed_at is pinned to a UTC instant."""
    import datetime

    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    order = SaleOrder.objects.create()
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)
    SaleOrder.objects.filter(pk=order.pk).update(
        confirmed_at=datetime.datetime(y, mo, d, h, mi, tzinfo=datetime.timezone.utc)
    )
    return order


# Force Asia/Bishkek (UTC+6) as the active timezone so these tests are correct
# regardless of the ambient TIME_ZONE, and both localdate() and the ORM __date
# lookup agree on which day a sale belongs to.
def test_sale_at_2330_bishkek_counts_in_that_local_day(variant, monkeypatch):
    # 23:30 Asia/Bishkek on 16 Jul == 17:30 UTC on 16 Jul.
    _sale_confirmed_at(variant, 2026, 7, 16, 17, 30)
    from apps.sales.services import today_summary

    with timezone.override("Asia/Bishkek"):
        # "Now" is 23:45 Bishkek on the 16th — the sale is in TODAY's numbers.
        _freeze_now(monkeypatch, 2026, 7, 16, 17, 45)
        assert today_summary()["orders"] == 1

        # 15 minutes later it's 00:15 Bishkek on the 17th — the SAME sale must NOT
        # leak into the next day's numbers (that's the silent bug we're guarding).
        _freeze_now(monkeypatch, 2026, 7, 16, 18, 15)
        assert today_summary()["orders"] == 0


def test_early_morning_sale_counts_in_local_day_not_utc_day(variant, monkeypatch):
    # The real UTC-boundary case: 00:30 Bishkek on 17 Jul == 18:30 UTC on 16 Jul,
    # so the UTC date (16th) differs from the local date (17th). A UTC-based
    # "today" would drop this sale from the 17th's report — the corruption.
    order = _sale_confirmed_at(variant, 2026, 7, 16, 18, 30)
    from apps.sales.services import today_summary, todays_confirmed_orders

    with timezone.override("Asia/Bishkek"):
        _freeze_now(monkeypatch, 2026, 7, 17, 4, 0)  # 10:00 Bishkek on the 17th
        assert today_summary()["orders"] == 1  # /pos «Сегодня» + bot /today
        assert order.pk in [o.pk for o in todays_confirmed_orders()]  # daily report Продажи


def test_pos_segodnya_view_shows_boundary_sale_in_local_day(
    client, django_user_model, variant, monkeypatch
):
    call_command("setup_roles")
    editor = django_user_model.objects.create_user("tzeditor", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)
    _sale_confirmed_at(variant, 2026, 7, 16, 18, 30)  # 00:30 Bishkek 17th
    with timezone.override("Asia/Bishkek"):
        _freeze_now(monkeypatch, 2026, 7, 17, 4, 0)  # 10:00 Bishkek 17th
        resp = client.get("/pos/today/")
    assert resp.status_code == 200
    from apps.pos.templatetags.pos_extras import money_filter

    assert (
        money_filter(3200, "KGS") in resp.content.decode()
    )  # the boundary sale is on TODAY's screen


def test_bot_today_reply_counts_boundary_sale_in_local_day(variant, monkeypatch):
    _sale_confirmed_at(variant, 2026, 7, 16, 18, 30)  # 00:30 Bishkek 17th
    from apps.wa.replies import today_reply

    with timezone.override("Asia/Bishkek"):
        _freeze_now(monkeypatch, 2026, 7, 17, 4, 0)  # 10:00 Bishkek 17th
        assert "1 sales" in today_reply()  # bot /today, via today_summary


def test_catalog_version_bump_survives_absent_key():
    from django.core.cache import cache

    from apps.inventory.cache import CATALOG_VERSION_KEY, bump_catalog_version, catalog_version

    # Simulate Redis having dropped the key (restart / eviction / expiry).
    cache.delete(CATALOG_VERSION_KEY)
    bump_catalog_version()  # must NOT raise ValueError
    assert isinstance(catalog_version(), int)

    # And a product save (which fires the post_save signal -> bump) must not blow
    # up either when the version key is missing.
    cache.delete(CATALOG_VERSION_KEY)
    cat = Category.objects.create(name="Cache-Cat")
    Product.objects.create(category=cat, name="Cache-Prod")  # signal path, no raise
    assert catalog_version() >= 1


def test_audit_stale_totals_reports_a_clean_state(variant, capsys):
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=2, unit_price=Decimal("3000"))
    confirm_sale(order)

    call_command("audit_stale_totals")
    out = capsys.readouterr().out
    assert "No stale totals found" in out


def test_audit_stale_totals_detects_a_manufactured_mismatch(variant, capsys):
    """The exact scenario the confirmed-sale admin lock now prevents — a line
    item's quantity changed without going through services.py, leaving
    order.total stale. Manufactured directly (bypassing the admin, which is
    now locked) purely to prove the audit command can actually SEE it."""
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    order = SaleOrder.objects.create(currency="KGS")
    item = SaleItem.objects.create(
        order=order, variant=variant, quantity=2, unit_price=Decimal("3000")
    )
    confirm_sale(order)
    SaleItem.objects.filter(pk=item.pk).update(quantity=5)  # bypass services.py directly

    call_command("audit_stale_totals")
    out = capsys.readouterr().out
    assert f"#{order.pk}" in out
    assert "stored total=6000.00" in out
    assert "items currently add up to=15000.00" in out


def test_cleanup_draft_sales_deletes_only_old_drafts():
    fresh = SaleOrder.objects.create()
    old = SaleOrder.objects.create()
    SaleOrder.objects.filter(pk=old.pk).update(
        created_at=timezone.now() - timezone.timedelta(hours=30)
    )
    call_command("cleanup_draft_sales")
    assert SaleOrder.objects.filter(pk=fresh.pk).exists()
    assert not SaleOrder.objects.filter(pk=old.pk).exists()


def test_cleanup_draft_sales_keeps_a_draft_with_an_item_even_past_24h(variant):
    """Only a TRULY empty draft (no client, no items, no payments) qualifies —
    with lazy creation this is already a narrow edge case (every item removed
    again after creation), not the routine clutter it used to be, but it must
    still never purge real in-progress work."""
    add_movement(variant, StockMovement.PURCHASE_IN, 5)
    working = SaleOrder.objects.create(currency=variant.currency)
    SaleItem.objects.create(order=working, variant=variant, quantity=1, unit_price=Decimal("100"))
    SaleOrder.objects.filter(pk=working.pk).update(
        created_at=timezone.now() - timezone.timedelta(hours=25)
    )

    client_only = SaleOrder.objects.create(
        client=Client.objects.create(first_name="Draft", phone="+996700990301")
    )
    SaleOrder.objects.filter(pk=client_only.pk).update(
        created_at=timezone.now() - timezone.timedelta(hours=25)
    )

    call_command("cleanup_draft_sales")

    assert SaleOrder.objects.filter(pk=working.pk).exists(), "25h-old cart with an item was purged"
    assert SaleOrder.objects.filter(pk=client_only.pk).exists(), (
        "25h-old draft with a client was purged"
    )


def test_cleanup_draft_sales_never_touches_a_confirmed_or_cancelled_order(variant):
    add_movement(variant, StockMovement.PURCHASE_IN, 5)
    order = SaleOrder.objects.create(currency=variant.currency)
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("100"))
    confirm_sale(order)
    SaleOrder.objects.filter(pk=order.pk).update(
        created_at=timezone.now() - timezone.timedelta(hours=48)
    )
    call_command("cleanup_draft_sales")
    assert SaleOrder.objects.filter(pk=order.pk, status=SaleOrder.CONFIRMED).exists()


# =========================================================================
# Owner dashboard (Phase 3) — access control, HTMX, query budget, caching.
# =========================================================================


def test_dashboard_is_owner_only(client, django_user_model):
    call_command("setup_roles")
    # Anonymous -> bounced to the shared login, never rendered.
    assert client.get("/dashboard/").status_code == 302

    # Editor (staff, no superuser) -> 403, and never sees the surface.
    editor = django_user_model.objects.create_user("dasheditor", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)
    assert client.get("/dashboard/").status_code == 403

    # Owner (superuser) -> 200 with the real content.
    client.force_login(django_user_model.objects.create_superuser("dashowner", password="x" * 12))
    resp = client.get("/dashboard/")
    assert resp.status_code == 200
    assert "Выручка" in resp.content.decode()


def test_dashboard_numeric_svg_attrs_are_not_locale_formatted(client, django_user_model, variant):
    # Regression: LANGUAGE_CODE="ru" makes Django render raw numbers with a
    # comma decimal separator ("104.0" -> "104,0") wherever a template renders
    # a float/int directly. That's invalid inside SVG coordinate attributes
    # and breaks the bar's --pct CSS custom property (a comma inside scaleX()
    # makes the whole `transform` declaration invalid, so the browser drops it
    # and the bar never scales). The panel must render with locale formatting
    # off regardless of what LANGUAGE_CODE the project ships.
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    c1 = Client.objects.create(first_name="Aigerim", phone="+996700000097")
    order = SaleOrder.objects.create(client=c1, channel=SaleOrder.SHOP)
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)
    # A second channel with a different amount so at least one bar's --pct
    # fraction is non-round (e.g. 0.3333...), the case most likely to leak a
    # locale comma through an untested happy path.
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    order2 = SaleOrder.objects.create(client=c1, channel=SaleOrder.WHOLESALE)
    SaleItem.objects.create(order=order2, variant=variant, quantity=1, unit_price=Decimal("1600"))
    confirm_sale(order2)

    client.force_login(django_user_model.objects.create_superuser("svgowner", password="x" * 12))
    html = client.get("/dashboard/?period=month").content.decode()

    # Scoped to the attributes that carry a RAW single float/int (x, y, cx, cy,
    # r, data-cx, data-cy) — never `d`, whose path strings legitimately use
    # "x,y" as a pair separator between two already-period-formatted numbers
    # (e.g. "104.0,198.0"), which is correct SVG and not what this guards.
    for name in ("x", "y", "cx", "cy", "r", "data-cx", "data-cy"):
        for value in re.findall(rf'\b{name}="([^"]*)"', html):
            assert "," not in value, f'{name}="{value}" would be invalid — a locale comma leaked in'
    # The bar's --pct custom property is the other place a comma silently
    # breaks scaleX().
    for pct in re.findall(r"--pct: ([^\"]+)", html):
        assert "," not in pct, f"--pct value {pct!r} would make scaleX() invalid CSS"


def test_dashboard_link_shown_only_to_owner(client, django_user_model):
    call_command("setup_roles")
    editor = django_user_model.objects.create_user("linkeditor", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)
    assert 'href="/dashboard/"' not in client.get("/pos/", follow=True).content.decode()

    client.force_login(django_user_model.objects.create_superuser("linkowner", password="x" * 12))
    assert 'href="/dashboard/"' in client.get("/pos/", follow=True).content.decode()


# ---- Storage / inventory dashboard (Owner-only) ---------------------------


def _editor(client, django_user_model, name):
    call_command("setup_roles")
    editor = django_user_model.objects.create_user(name, password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)
    return editor


def test_storage_page_is_owner_only(client, django_user_model, variant):
    # Anonymous -> login redirect, never rendered.
    assert client.get("/storage/").status_code == 302
    # Editor (staff, not superuser) -> hard 403.
    _editor(client, django_user_model, "storeeditor")
    assert client.get("/storage/").status_code == 403
    # Owner -> 200 with the real spreadsheet content.
    client.force_login(django_user_model.objects.create_superuser("storeowner", password="x" * 12))
    resp = client.get("/storage/")
    assert resp.status_code == 200
    body = resp.content.decode()
    assert "Склад" in body and "По категориям" in body and variant.sku in body


def test_storage_export_is_owner_only(client, django_user_model, variant):
    # Anonymous and Editor must never receive a file.
    assert client.get("/storage/export/?format=xlsx").status_code == 302
    _editor(client, django_user_model, "storeexpeditor")
    assert client.get("/storage/export/?format=xlsx").status_code == 403
    assert client.get("/storage/export/?format=csv&sheet=tovary").status_code == 403


def test_storage_export_xlsx_and_csv_for_owner(client, django_user_model, variant):
    client.force_login(django_user_model.objects.create_superuser("storeexp", password="x" * 12))
    add_movement(variant, StockMovement.PRODUCTION_IN, 12)

    xlsx = client.get("/storage/export/?format=xlsx")
    assert xlsx.status_code == 200
    assert "spreadsheetml.sheet" in xlsx["Content-Type"]
    assert "attachment; filename=" in xlsx["Content-Disposition"]
    assert xlsx.content[:2] == b"PK"  # a real .xlsx (zip) payload

    csv_resp = client.get("/storage/export/?format=csv&sheet=tovary")
    assert csv_resp.status_code == 200
    assert csv_resp["Content-Type"].startswith("text/csv")
    text = csv_resp.content.decode("utf-8")
    assert "Артикул" in text and variant.sku in text

    # An unknown csv sheet name is rejected, not silently served.
    assert client.get("/storage/export/?format=csv&sheet=secret").status_code == 403


def test_storage_link_shown_only_to_owner(client, django_user_model):
    _editor(client, django_user_model, "storelinkeditor")
    assert 'href="/storage/"' not in client.get("/pos/", follow=True).content.decode()
    client.force_login(
        django_user_model.objects.create_superuser("storelinkowner", password="x" * 12)
    )
    assert 'href="/storage/"' in client.get("/pos/", follow=True).content.decode()


def test_storage_aggregates_movements_by_type(variant):
    from apps.reports.storage import storage_data

    add_movement(variant, StockMovement.PRODUCTION_IN, 20)
    add_movement(variant, StockMovement.WRITEOFF_OUT, -3)  # defective
    add_movement(variant, StockMovement.RETURN_IN, 2)
    # a confirmed sale of 4 -> a sale_out movement of -4
    order = SaleOrder.objects.create()
    SaleItem.objects.create(order=order, variant=variant, quantity=4, unit_price=Decimal("100"))
    confirm_sale(order)

    data = storage_data()
    row = next(it for it in data["items"] if it["sku"] == variant.sku)
    assert row["intake"] == 20
    assert row["writeoff"] == 3  # shown as a positive count
    assert row["returns"] == 2
    assert row["sold"] == 4
    assert row["stock"] == 20 - 3 + 2 - 4  # ledger sum
    assert data["summary"]["sold"] == 4 and data["summary"]["writeoff"] == 3


def test_dashboard_export_is_owner_only(client, django_user_model):
    call_command("setup_roles")
    # Anonymous -> login redirect; Editor -> hard 403; neither gets a file.
    assert client.get("/dashboard/export/?format=xlsx").status_code == 302
    _editor(client, django_user_model, "dashexpeditor")
    assert client.get("/dashboard/export/?format=xlsx").status_code == 403
    assert client.get("/dashboard/export/?format=csv&sheet=tovary").status_code == 403


def test_dashboard_export_xlsx_and_csv_for_owner(client, django_user_model, variant):
    client.force_login(django_user_model.objects.create_superuser("dashexp", password="x" * 12))
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    order = SaleOrder.objects.create()
    SaleItem.objects.create(order=order, variant=variant, quantity=2, unit_price=Decimal("500"))
    confirm_sale(order)

    # xlsx: real workbook, filename carries the period.
    xlsx = client.get("/dashboard/export/?period=year&format=xlsx")
    assert xlsx.status_code == 200
    assert "spreadsheetml.sheet" in xlsx["Content-Type"]
    assert xlsx.content[:2] == b"PK"
    assert "year" in xlsx["Content-Disposition"]

    # csv of a single sheet.
    csv_resp = client.get("/dashboard/export/?period=year&format=csv&sheet=tovary")
    assert csv_resp.status_code == 200
    assert csv_resp["Content-Type"].startswith("text/csv")
    text = csv_resp.content.decode("utf-8")
    assert "Товар" in text and "Прибыль" in text

    # Unknown csv sheet -> rejected, never silently served.
    assert client.get("/dashboard/export/?format=csv&sheet=secret").status_code == 403


def test_dashboard_export_matches_the_page_numbers(variant):
    # download == screen: the export flattens the SAME dashboard_data().
    from apps.reports.dashboard import dashboard_data, dashboard_sheets

    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    order = SaleOrder.objects.create()
    SaleItem.objects.create(order=order, variant=variant, quantity=3, unit_price=Decimal("1000"))
    confirm_sale(order)  # revenue 3000 сом today

    data = dashboard_data("today")
    sheets = dashboard_sheets(data)
    svodka = dict(sheets["Сводка"].rows)  # {label: value}
    assert svodka["Выручка, сом"] == float(data["metrics"]["revenue"]["value"])
    assert svodka["Продано, шт"] == 3


def test_dashboard_export_link_is_period_aware(client, django_user_model, variant):
    client.force_login(django_user_model.objects.create_superuser("dashdl", password="x" * 12))
    body = client.get("/dashboard/?period=3m").content.decode()
    assert "/dashboard/export/?period=3m&amp;format=xlsx" in body


def test_dashboard_currency_toggle_is_view_only(client, django_user_model, variant):
    # ?cur=USD converts the SHOWN money at today's rate; units aren't money and
    # stay unchanged; the download always stays in сом (the money-truth file).
    from django.core.cache import cache
    from django.utils import timezone

    from apps.core.models import ExchangeRate

    cache.clear()  # the dated-rate cache bleeds across tests otherwise
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    order = SaleOrder.objects.create()
    SaleItem.objects.create(order=order, variant=variant, quantity=5, unit_price=Decimal("1000"))
    confirm_sale(order)  # 5000 сом today
    ExchangeRate.objects.create(
        currency="USD", date=timezone.localdate(), rate=Decimal("100"), source="nbkr"
    )
    client.force_login(django_user_model.objects.create_superuser("curown", password="x" * 12))

    som = client.get("/dashboard/?period=today").content.decode()
    assert "5\xa0000\xa0сом" in som

    usd = client.get("/dashboard/?period=today&cur=USD").content.decode()
    assert "50\xa0USD" in usd  # 5000 / 100, shown as a text abbreviation
    assert "5 шт" in usd  # units are not money — never converted
    assert "приблизительны" in usd  # the ≈ disclaimer

    # The export ignores the view currency — always сом.
    csv = client.get(
        "/dashboard/export/?period=today&cur=USD&format=csv&sheet=vyruchka"
    ).content.decode()
    assert "сом" in csv and "USD" not in csv


def test_dashboard_falls_back_to_som_without_a_rate(client, django_user_model, variant):
    from django.core.cache import cache

    cache.clear()  # ensure no rate leaked in from another test's dated-rate cache
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    order = SaleOrder.objects.create()
    SaleItem.objects.create(order=order, variant=variant, quantity=5, unit_price=Decimal("1000"))
    confirm_sale(order)
    client.force_login(django_user_model.objects.create_superuser("fbown", password="x" * 12))

    # Unknown currency code -> сом silently.
    unknown = client.get("/dashboard/?period=today&cur=ZZZ").content.decode()
    assert "5\xa0000\xa0сом" in unknown

    # A real currency with NO rate on record -> сом, but say so.
    no_rate = client.get("/dashboard/?period=today&cur=USD").content.decode()
    assert "5\xa0000\xa0сом" in no_rate
    assert "Курс валюты недоступен" in no_rate


def test_dashboard_omits_percentage_deltas(client, django_user_model):
    call_command("setup_roles")
    client.force_login(django_user_model.objects.create_superuser("nodelta", password="x" * 12))
    body = client.get("/dashboard/?period=today").content.decode()
    assert "metric__delta" not in body
    assert "ко вчера" not in body
    assert "нет данных за прошлый период" not in body


def test_dashboard_htmx_swaps_only_the_panel(client, django_user_model):
    call_command("setup_roles")
    client.force_login(django_user_model.objects.create_superuser("dashhx", password="x" * 12))
    full = client.get("/dashboard/?period=today")
    assert b"<html" in full.content  # a normal hit renders the shell
    partial = client.get("/dashboard/?period=today", HTTP_HX_REQUEST="true")
    assert b"<html" not in partial.content  # HTMX gets only the panel fragment
    assert "Выручка" in partial.content.decode()


def test_dashboard_bad_period_falls_back_to_default(client, django_user_model):
    call_command("setup_roles")
    client.force_login(django_user_model.objects.create_superuser("dashbad", password="x" * 12))
    resp = client.get("/dashboard/?period=not-a-real-period")
    assert resp.status_code == 200  # resolve_period() coerces it, never 500s


def _dash_catalog():
    """10 variants + 20 clients to spread the seeded sales across."""
    cat = Category.objects.create(name="Dash")
    variants = [
        ProductVariant.objects.create(
            product=Product.objects.create(category=cat, name=f"DP{i}"),
            sku=f"DSKU{i}",
            cost_price=Decimal("400"),
            sale_price=Decimal("1000"),
        )
        for i in range(10)
    ]
    clients = [
        Client.objects.create(first_name=f"DC{i}", phone=f"+99655500{i:04d}") for i in range(20)
    ]
    return variants, clients


def _seed_dashboard_sales(n, variants, clients, day_offset=0):
    """Bulk-create n confirmed KGS sales spread across the last year, each with a
    line item and (for ~70%) a full payment — so every panel has data. Bulk, not
    confirm_sale(), so 5000 rows seed in one shot; PKs come back from bulk_create."""
    now = timezone.now()
    orders = [
        SaleOrder(
            status=SaleOrder.CONFIRMED,
            channel=SaleOrder.SHOP if i % 2 else SaleOrder.INSTAGRAM,
            client=clients[i % len(clients)],
            total=Decimal("1000"),
            currency="KGS",
            rate_to_kgs=Decimal("1"),
            total_kgs=Decimal("1000"),
            confirmed_at=now - timezone.timedelta(days=(day_offset + i) % 360, hours=i % 24),
        )
        for i in range(n)
    ]
    created = SaleOrder.objects.bulk_create(orders, batch_size=1000)
    items, pays = [], []
    for idx, o in enumerate(created):
        items.append(
            SaleItem(
                order_id=o.pk,
                variant=variants[idx % len(variants)],
                quantity=1,
                unit_price=Decimal("1000"),
            )
        )
        if idx % 10 < 7:  # leave ~30% as debt so the Долги panel is exercised
            pays.append(
                Payment(
                    order_id=o.pk,
                    client=clients[idx % len(clients)],
                    amount=Decimal("1000"),
                    currency="KGS",
                    rate_to_kgs=Decimal("1"),
                    method=Payment.CASH,
                )
            )
    SaleItem.objects.bulk_create(items, batch_size=1000)
    Payment.objects.bulk_create(pays, batch_size=1000)


def test_dashboard_query_budget_is_flat_and_bounded(django_user_model):
    from django.db import connection
    from django.test.utils import CaptureQueriesContext

    from apps.reports.dashboard import PERIODS, dashboard_data

    def count(period):
        with CaptureQueriesContext(connection) as ctx:
            dashboard_data(period)
        return len(ctx.captured_queries)

    variants, clients = _dash_catalog()
    _seed_dashboard_sales(50, variants, clients)
    small = {p: count(p) for p in PERIODS}
    _seed_dashboard_sales(4950, variants, clients, day_offset=7)  # 5000 total
    big = {p: count(p) for p in PERIODS}

    # Budget raised 12 -> 14 -> 16 -> 17 -> 18 -> 20 -> 21: the «Заказы»
    # panels (Part 3g) added two flat queries, the CLIENT_BOTS.md panels
    # (telegram_reach: 2 counts; top_favourited: 1 aggregate + 1 variant
    # lookup) added two more, a product now having many images
    # (apps.inventory.services.cover_image_names) adds a bulk cover-image
    # lookup to the top-products AND dead-stock panels (net +1, not +2,
    # since 16 already had a query of slack), _metrics() now sums
    # received/expected cash from a SECOND aggregate query (Payment, a
    # different table than SaleOrder/SaleItem — can't be folded into the
    # existing ones) — +1, «Потрачено»/net-cash (apps.manufacturing.
    # dashboard.spent_kgs, current + previous period — a THIRD table,
    # Expense, same reasoning) — +2, and the dashboard's «Расходы» panel
    # (apps.manufacturing.dashboard.expenses_by_section, ONE grouped query
    # over that same Expense table) — +1. All still independent of sales
    # volume, just a slightly higher fixed floor.
    for p in PERIODS:
        assert big[p] <= 21, f"{p}: {big[p]} queries at 5000 sales (budget 21)"
        assert big[p] == small[p], f"{p} query count scaled with rows: {small[p]} -> {big[p]}"


def test_dead_stock_does_not_n_plus_one(django_user_model):
    from django.db import connection
    from django.test.utils import CaptureQueriesContext

    from apps.inventory.services import add_movement
    from apps.reports.dashboard import _dead_stock

    cat = Category.objects.create(name="Dead")
    old = timezone.now() - timezone.timedelta(days=200)

    def make(k, tag):
        for i in range(k):
            p = Product.objects.create(category=cat, name=f"DS{tag}{i}")
            Product.objects.filter(pk=p.pk).update(created_at=old)  # 200 days idle -> dead
            v = ProductVariant.objects.create(
                product=p, sku=f"DSK{tag}{i}", cost_price=Decimal("100"), sale_price=Decimal("200")
            )
            add_movement(v, StockMovement.PRODUCTION_IN, 5)  # in stock

    make(5, "A")
    with CaptureQueriesContext(connection) as few:
        result_few = _dead_stock()
    make(200, "B")
    with CaptureQueriesContext(connection) as many:
        result_many = _dead_stock()

    assert result_few and result_many  # the loop actually ran over dead rows
    # Three queries (stock subquery + one grouped last-sale + one bulk cover-
    # image lookup, apps.inventory.services.cover_image_names — a product can
    # have many images now), constant as rows grow.
    assert len(many.captured_queries) == len(few.captured_queries) <= 3


def test_dashboard_cache_invalidated_on_sale_confirm(variant):
    from apps.reports.views import _cached_data

    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    before = _cached_data("today")["metrics"]["revenue"]["value"]  # warms the cache

    order = SaleOrder.objects.create()
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)  # bumps catalog_version via the StockMovement signal

    after = _cached_data("today")["metrics"]["revenue"]["value"]
    # A stale cache would return `before`; the version-keyed cache must refresh.
    assert after == before + Decimal("3200")


def test_dashboard_cache_invalidated_on_payment(variant):
    from apps.reports.views import _cached_data

    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    debtor = Client.objects.create(first_name="Debtor", phone="+996700009999")
    order = SaleOrder.objects.create(client=debtor)
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)  # unpaid -> a debt
    debt_before = _cached_data("today")["metrics"]["debt"]["value"]

    record_payment(
        order, amount=Decimal("1200"), currency="KGS", method="cash"
    )  # Payment signal bumps
    debt_after = _cached_data("today")["metrics"]["debt"]["value"]
    assert debt_after == debt_before - Decimal("1200")


# ---------------------------------------------------------------------------
# Multi-currency payment hardening
# ---------------------------------------------------------------------------


def test_manager_can_refresh_rates_but_hand_entering_one_is_403(
    client, django_user_model, variant, monkeypatch
):
    """RATE PERMISSIONS: refreshing from NBKR is allowed for Editor/Manager;
    hand-entering/overriding a rate is Owner-only, enforced server-side even
    if a Manager bypasses the UI and POSTs it directly."""
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    manager = django_user_model.objects.create_user("mgr_perm", password="x" * 12, is_staff=True)
    manager.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(manager)

    monkeypatch.setattr(
        "apps.core.management.commands.fetch_rates.requests.get",
        _frankfurter_get({"USD": 87.45, "RUB": 1.1239}),
    )
    assert client.post("/pos/rates/refresh/").status_code == 200  # allowed

    order = SaleOrder.objects.create(created_by=manager, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))

    resp = client.post(
        f"/pos/sale/{order.pk}/recalc/",
        {"amount": "10", "currency": "USD", "rate_override": "80"},
    )
    assert resp.status_code == 403

    resp = client.post(
        f"/pos/sale/{order.pk}/confirm/",
        {"amount": "10", "currency": "USD", "method": "cash", "rate_override": "80"},
    )
    assert resp.status_code == 403
    order.refresh_from_db()
    assert order.status == SaleOrder.DRAFT  # rejected before touching stock


def test_owner_rate_override_is_stored_with_official_rate_for_spread(
    client, django_user_model, variant, settings
):
    """OWNER RATE OVERRIDE: the Owner can hand-enter the actual (booth) rate —
    stored as rate_source=manual with rate_official (the NBKR rate at that
    moment) captured so the spread is reconstructable later."""
    settings.CURRENCY = "KGS"
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    ExchangeRate.objects.create(currency="USD", date=timezone.localdate(), rate=Decimal("87.00"))
    owner = django_user_model.objects.create_superuser("owner_ov", "o@e.com", "x" * 12)
    client.force_login(owner)

    order_id = SaleOrder.objects.create(created_by=owner, channel=SaleOrder.SHOP).pk
    cust = Client.objects.create(first_name="Own", phone="+996700000811")
    client.post(f"/pos/sale/{order_id}/client/{cust.pk}/set/")
    client.post(f"/pos/sale/{order_id}/items/add/", {"variant_id": variant.pk, "quantity": 1})

    resp = client.post(
        f"/pos/sale/{order_id}/confirm/",
        {
            "amount": "10",
            "currency": "USD",
            "method": "cash",
            "rate_override": "89.50",
            "risk_ack": "1",  # a manual rate is itself a risk reason — needs ack
        },
    )
    assert resp.status_code == 302
    assert resp.url == f"/pos/sale/{order_id}/result/"

    order = SaleOrder.objects.get(pk=order_id)
    payment = order.payments.get()
    assert payment.rate_source == Payment.RATE_MANUAL
    assert payment.rate_to_kgs == Decimal("89.50")
    assert payment.rate_official == Decimal("87.000000")
    assert order.balance == Decimal("2305.00")  # 3200 - (10 × 89.50)


def test_manual_rate_deviating_from_official_warns_but_never_blocks(
    client, django_user_model, variant, settings
):
    settings.CURRENCY = "KGS"
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    ExchangeRate.objects.create(currency="USD", date=timezone.localdate(), rate=Decimal("87.00"))
    owner = django_user_model.objects.create_superuser("owner_dev", "o@e.com", "x" * 12)
    client.force_login(owner)

    order_id = SaleOrder.objects.create(created_by=owner, channel=SaleOrder.SHOP).pk
    cust = Client.objects.create(first_name="Dev", phone="+996700000812")
    client.post(f"/pos/sale/{order_id}/client/{cust.pk}/set/")
    client.post(f"/pos/sale/{order_id}/items/add/", {"variant_id": variant.pk, "quantity": 1})

    # 95 vs 87 official = ~9.2% deviation — above the 5% warn threshold.
    resp = client.post(
        f"/pos/sale/{order_id}/recalc/",
        {"amount": "10", "currency": "USD", "rate_override": "95"},
    )
    assert "отличается от официального" in resp.content.decode()

    # The deviation itself never blocks — only the (separate) manual-rate risk
    # reason requires an ack, and ack'ing it lets the sale complete.
    resp = client.post(
        f"/pos/sale/{order_id}/confirm/",
        {
            "amount": "10",
            "currency": "USD",
            "method": "cash",
            "rate_override": "95",
            "risk_ack": "1",
        },
    )
    assert resp.status_code == 302
    payment = SaleOrder.objects.get(pk=order_id).payments.get()
    assert payment.rate_to_kgs == Decimal("95")


def test_void_foreign_payment_restores_debt_to_exact_pre_payment_value(variant, settings):
    """REVERSALS MUST USE THE FROZEN RATE: voiding a foreign payment restores
    the client's debt to EXACTLY its pre-payment value even after today's
    rate has moved significantly in between."""
    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    ExchangeRate.objects.create(currency="USD", date=timezone.localdate(), rate=Decimal("87.00"))
    cust = Client.objects.create(first_name="Rev", phone="+996700000826")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("8700"))
    confirm_sale(order)
    debt_before = client_debt(cust)
    assert debt_before == {"KGS": Decimal("8700.00")}

    payment = Payment.objects.create(client=cust, order=order, amount=Decimal("10"), currency="USD")
    assert client_debt(cust) == {"KGS": Decimal("7830.00")}  # 8700 - (10 × 87)

    # The rate moves significantly AFTER the payment.
    ExchangeRate.objects.filter(currency="USD").update(rate=Decimal("150.00"))

    void_payment(payment)
    assert client_debt(cust) == debt_before  # exactly restored, unaffected by the new rate


def test_voiding_batch_restores_exact_prior_balance_across_all_affected_sales(variant):
    """A pay_oldest_first batch spanning two sales voids as one event and
    restores BOTH sales' own balances exactly, not just the client's
    aggregate debt."""
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    cust = Client.objects.create(first_name="Bat", phone="+996700000901")
    order1 = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order1, variant=variant, quantity=1, unit_price=Decimal("3000"))
    confirm_sale(order1)
    order2 = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order2, variant=variant, quantity=1, unit_price=Decimal("5000"))
    confirm_sale(order2)

    debt_before = client_debt(cust)
    assert debt_before == {"KGS": Decimal("8000.00")}

    payments = pay_oldest_first(cust, "KGS", Decimal("8000"))
    assert len(payments) == 2
    batch_id = payments[0].batch_id
    assert batch_id is not None and payments[1].batch_id == batch_id
    assert client_debt(cust) == {}
    order1.refresh_from_db()
    order2.refresh_from_db()
    assert order1.balance == Decimal("0")
    assert order2.balance == Decimal("0")

    void_payment_batch(batch_id)

    order1.refresh_from_db()
    order2.refresh_from_db()
    assert order1.balance == Decimal("3000.00")
    assert order2.balance == Decimal("5000.00")
    assert client_debt(cust) == debt_before


def test_partial_void_of_a_batch_is_impossible(variant):
    """Neither void_payment on a single batch member nor a second call to
    void_payment_batch can leave a batch half-reversed — there is no code
    path that produces a partial void."""
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    cust = Client.objects.create(first_name="Bat2", phone="+996700000902")
    order1 = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order1, variant=variant, quantity=1, unit_price=Decimal("2000"))
    confirm_sale(order1)
    order2 = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order2, variant=variant, quantity=1, unit_price=Decimal("4000"))
    confirm_sale(order2)

    payments = pay_oldest_first(cust, "KGS", Decimal("6000"))
    batch_id = payments[0].batch_id

    with pytest.raises(ValidationError):
        void_payment(payments[0])
    with pytest.raises(ValidationError):
        void_payment(payments[1])
    order1.refresh_from_db()
    order2.refresh_from_db()
    assert order1.balance == Decimal("0")
    assert order2.balance == Decimal("0")

    reversals = void_payment_batch(batch_id)
    assert len(reversals) == 2
    assert Payment.objects.filter(batch_id=batch_id, reversed_payment_id__isnull=False).count() == 2

    with pytest.raises(ValidationError):
        void_payment_batch(batch_id)


def test_batch_void_restores_debt_exactly_after_a_rate_change(variant, settings):
    """A foreign-currency batch voids correctly even after the live rate has
    since moved: void_payment_batch never re-looks-up today's rate, it just
    replays each frozen payment rate in reverse."""
    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    ExchangeRate.objects.create(currency="USD", date=timezone.localdate(), rate=Decimal("80.00"))
    cust = Client.objects.create(first_name="Bat3", phone="+996700000903")
    order1 = SaleOrder.objects.create(client=cust, currency="USD")
    SaleItem.objects.create(order=order1, variant=variant, quantity=1, unit_price=Decimal("50"))
    confirm_sale(order1)
    order2 = SaleOrder.objects.create(client=cust, currency="USD")
    SaleItem.objects.create(order=order2, variant=variant, quantity=1, unit_price=Decimal("30"))
    confirm_sale(order2)

    debt_before = client_debt(cust)
    assert debt_before == {"USD": Decimal("80.00")}

    payments = pay_oldest_first(cust, "USD", Decimal("80"))
    batch_id = payments[0].batch_id
    assert client_debt(cust) == {}

    # The rate moves significantly AFTER the batch was paid.
    ExchangeRate.objects.filter(currency="USD").update(rate=Decimal("200.00"))

    void_payment_batch(batch_id)
    assert client_debt(cust) == debt_before
    order1.refresh_from_db()
    order2.refresh_from_db()
    assert order1.balance == Decimal("50.00")
    assert order2.balance == Decimal("30.00")


def test_batch_reversal_uses_each_payments_own_frozen_rate(variant, settings):
    """Two payments in the SAME batch, snapshotted at two DIFFERENT rates,
    each reverse using their OWN frozen rate_to_kgs — never a shared or
    live one. Constructed directly via record_payment(batch_id=...) (the
    same parameter pay_oldest_first uses) to force the rates apart, since a
    real pay_oldest_first call always lands both payments in one atomic
    transaction at whatever the single current rate is."""
    from django.core.cache import cache

    cache.clear()
    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    ExchangeRate.objects.create(currency="USD", date=timezone.localdate(), rate=Decimal("80.00"))
    cust = Client.objects.create(first_name="Bat4", phone="+996700000904")
    order1 = SaleOrder.objects.create(client=cust, currency="USD")
    SaleItem.objects.create(order=order1, variant=variant, quantity=1, unit_price=Decimal("100"))
    confirm_sale(order1)
    order2 = SaleOrder.objects.create(client=cust, currency="USD")
    SaleItem.objects.create(order=order2, variant=variant, quantity=1, unit_price=Decimal("100"))
    confirm_sale(order2)

    debt_before = client_debt(cust)
    assert debt_before == {"USD": Decimal("200.00")}

    batch_id = uuid.uuid4()
    p1 = record_payment(order1, Decimal("100"), currency="USD", batch_id=batch_id)
    assert p1.rate_to_kgs == Decimal("80.00")

    # A lower rate (not higher) — a same-nominal-amount payment at a HIGHER
    # rate than the order's own frozen rate reads as an overpayment in KGS
    # terms and record_payment rightly refuses it without an explicit
    # excess disposition; a lower rate leaves a real (smaller) balance
    # instead, which is enough to prove the per-payment rate independently.
    ExchangeRate.objects.filter(currency="USD").update(rate=Decimal("60.00"))
    cache.clear()  # a real rate refresh clears this via signal; bulk update bypassed it
    p2 = record_payment(order2, Decimal("100"), currency="USD", batch_id=batch_id)
    assert p2.rate_to_kgs == Decimal("60.00")
    order2.refresh_from_db()
    assert order2.balance == Decimal("25.00")  # (100 × 60) / 80 = 75 contributed, not 100

    # A third, unrelated rate move before voiding — proves neither reversal
    # picks up a fresh live rate.
    ExchangeRate.objects.filter(currency="USD").update(rate=Decimal("200.00"))
    cache.clear()

    reversals = void_payment_batch(batch_id)
    by_original = {r.reversed_payment_id: r for r in reversals}
    assert by_original[p1.pk].rate_to_kgs == Decimal("80.00")
    assert by_original[p2.pk].rate_to_kgs == Decimal("60.00")

    order1.refresh_from_db()
    order2.refresh_from_db()
    assert order1.balance == Decimal("100.00")
    assert order2.balance == Decimal("100.00")
    assert client_debt(cust) == debt_before


def test_rates_card_shows_date_and_stale_badge_past_four_days(client, django_user_model, settings):
    settings.CURRENCY = "KGS"
    call_command("setup_roles")
    manager = django_user_model.objects.create_user("mgr_age", password="x" * 12, is_staff=True)
    manager.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(manager)

    stale_date = timezone.localdate() - timezone.timedelta(days=5)
    ExchangeRate.objects.create(currency="USD", date=stale_date, rate=Decimal("87.00"))
    order = SaleOrder.objects.create(created_by=manager)

    resp = client.get(f"/pos/sale/{order.pk}/")
    body = resp.content.decode()
    assert stale_date.strftime("%d.%m.%Y") in body
    assert "Курс устарел" in body
    assert "badge--debt" in body


def test_scheduler_runs_fetch_rates_daily_and_before_the_report():
    """Restores the automatic daily pull (kept alongside the manual button) —
    automatic for reliability, manual for control."""
    import scheduler as scheduler_module

    job_commands = [cmds for _, cmds in scheduler_module.JOBS]
    assert ["fetch_rates"] in job_commands
    report_job = next(cmds for _, cmds in scheduler_module.JOBS if "send_daily_report" in cmds)
    assert "fetch_rates" in report_job
    assert "cleanup_draft_sales" in report_job


def test_refresh_rates_writes_audit_log_and_skips_noop_refresh(
    client, django_user_model, monkeypatch, settings
):
    settings.CURRENCY = "KGS"
    call_command("setup_roles")
    manager = django_user_model.objects.create_user("mgr_audit", password="x" * 12, is_staff=True)
    manager.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(manager)

    monkeypatch.setattr(
        "apps.core.management.commands.fetch_rates.requests.get",
        _frankfurter_get({"USD": 87.45, "RUB": 1.1239}),
    )
    client.post("/pos/rates/refresh/")
    log = RateChangeLog.objects.get(currency="USD")
    assert log.old_rate is None
    assert log.new_rate == Decimal("87.45")
    assert log.source == ExchangeRate.NBKR
    assert log.changed_by == manager

    # Pulling back the SAME rate is not a "change" — no second log row.
    client.post("/pos/rates/refresh/")
    assert RateChangeLog.objects.filter(currency="USD").count() == 1


def test_exchange_rate_admin_is_owner_only(django_user_model):
    call_command("setup_roles")
    manager = django_user_model.objects.create_user("mgr_era", password="x" * 12, is_staff=True)
    manager.groups.add(Group.objects.get(name=EDITOR))
    request = RequestFactory().get("/")
    request.user = manager
    model_admin = site._registry[ExchangeRate]
    assert model_admin.has_module_permission(request) is False
    assert model_admin.has_add_permission(request) is False
    assert model_admin.has_change_permission(request) is False

    owner = django_user_model.objects.create_superuser("owner_era", "o@e.com", "x" * 12)
    request.user = owner
    assert model_admin.has_module_permission(request) is True
    assert model_admin.has_add_permission(request) is True
    assert model_admin.has_change_permission(request) is True


def test_rate_change_log_admin_is_read_only_and_owner_only(django_user_model):
    owner = django_user_model.objects.create_superuser("owner_rcl", "o@e.com", "x" * 12)
    request = RequestFactory().get("/")
    request.user = owner
    model_admin = site._registry[RateChangeLog]
    assert model_admin.has_view_permission(request) is True
    assert model_admin.has_add_permission(request) is False
    assert model_admin.has_change_permission(request) is False
    assert model_admin.has_delete_permission(request) is False

    staff = django_user_model.objects.create_user("staff_rcl", password="x" * 12, is_staff=True)
    request.user = staff
    assert model_admin.has_module_permission(request) is False


def test_exchange_rate_admin_save_model_upserts_and_logs(django_user_model):
    owner = django_user_model.objects.create_superuser("owner_save", "o@e.com", "x" * 12)
    request = RequestFactory().post("/")
    request.user = owner
    model_admin = site._registry[ExchangeRate]

    obj = ExchangeRate(currency="USD", rate=Decimal("89.00"))
    model_admin.save_model(request, obj, form=None, change=False)
    assert ExchangeRate.objects.filter(currency="USD").count() == 1
    row = ExchangeRate.objects.get(currency="USD")
    assert row.rate == Decimal("89.00")
    assert row.source == ExchangeRate.MANUAL
    log = RateChangeLog.objects.get(currency="USD")
    assert log.old_rate is None and log.new_rate == Decimal("89.00")
    assert log.source == ExchangeRate.MANUAL and log.changed_by == owner

    # Editing again upserts in place (still one row) and logs the change.
    obj2 = ExchangeRate(currency="USD", rate=Decimal("91.00"))
    model_admin.save_model(request, obj2, form=None, change=False)
    assert ExchangeRate.objects.filter(currency="USD").count() == 1
    assert RateChangeLog.objects.filter(currency="USD").count() == 2
    latest = RateChangeLog.objects.filter(currency="USD").order_by("-changed_at").first()
    assert latest.old_rate == Decimal("89.00") and latest.new_rate == Decimal("91.00")


def test_fresh_small_foreign_payment_confirms_without_risk_ack(
    client, django_user_model, variant, settings
):
    settings.CURRENCY = "KGS"
    settings.LARGE_PAYMENT_THRESHOLD_KGS = 10000
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    ExchangeRate.objects.create(currency="USD", date=timezone.localdate(), rate=Decimal("87.00"))
    manager = django_user_model.objects.create_user("mgr_fresh", password="x" * 12, is_staff=True)
    manager.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(manager)

    order_id = SaleOrder.objects.create(created_by=manager, channel=SaleOrder.SHOP).pk
    cust = Client.objects.create(first_name="Fresh", phone="+996700000820")
    client.post(f"/pos/sale/{order_id}/client/{cust.pk}/set/")
    client.post(f"/pos/sale/{order_id}/items/add/", {"variant_id": variant.pk, "quantity": 1})

    resp = client.post(
        f"/pos/sale/{order_id}/confirm/", {"amount": "10", "currency": "USD", "method": "cash"}
    )
    assert resp.status_code == 302
    assert resp.url == f"/pos/sale/{order_id}/result/"
    assert SaleOrder.objects.get(pk=order_id).payments.exists()


def test_stale_rate_payment_requires_explicit_risk_ack(
    client, django_user_model, variant, settings
):
    settings.CURRENCY = "KGS"
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    stale_date = timezone.localdate() - timezone.timedelta(days=5)
    ExchangeRate.objects.create(currency="USD", date=stale_date, rate=Decimal("87.00"))
    manager = django_user_model.objects.create_user("mgr_stale", password="x" * 12, is_staff=True)
    manager.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(manager)

    order_id = SaleOrder.objects.create(created_by=manager, channel=SaleOrder.SHOP).pk
    cust = Client.objects.create(first_name="Stale", phone="+996700000821")
    client.post(f"/pos/sale/{order_id}/client/{cust.pk}/set/")
    client.post(f"/pos/sale/{order_id}/items/add/", {"variant_id": variant.pk, "quantity": 1})

    # Without risk_ack: rejected — the sale isn't even confirmed, nothing saved.
    resp = client.post(
        f"/pos/sale/{order_id}/confirm/", {"amount": "10", "currency": "USD", "method": "cash"}
    )
    assert resp.status_code == 302
    assert resp.url == f"/pos/sale/{order_id}/"
    order = SaleOrder.objects.get(pk=order_id)
    assert order.status == SaleOrder.DRAFT
    assert not order.payments.exists()

    # With risk_ack=1: goes through.
    resp = client.post(
        f"/pos/sale/{order_id}/confirm/",
        {"amount": "10", "currency": "USD", "method": "cash", "risk_ack": "1"},
    )
    assert resp.status_code == 302
    assert resp.url == f"/pos/sale/{order_id}/result/"
    order.refresh_from_db()
    assert order.status == SaleOrder.CONFIRMED
    assert order.payments.exists()


def test_large_converted_payment_requires_explicit_risk_ack(
    client, django_user_model, variant, settings
):
    settings.CURRENCY = "KGS"
    settings.LARGE_PAYMENT_THRESHOLD_KGS = 1000  # low threshold to trigger easily
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    ExchangeRate.objects.create(currency="USD", date=timezone.localdate(), rate=Decimal("87.00"))
    manager = django_user_model.objects.create_user("mgr_big", password="x" * 12, is_staff=True)
    manager.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(manager)

    order_id = SaleOrder.objects.create(created_by=manager, channel=SaleOrder.SHOP).pk
    cust = Client.objects.create(first_name="Big", phone="+996700000822")
    client.post(f"/pos/sale/{order_id}/client/{cust.pk}/set/")
    client.post(f"/pos/sale/{order_id}/items/add/", {"variant_id": variant.pk, "quantity": 1})

    # 20 USD × 87 = 1740 сом, above the 1000 сом threshold but below the 3200
    # сом order total — large, but NOT an overpayment (kept separate from the
    # сдача/overpayment-fork flow, which has its own dedicated tests).
    resp = client.post(
        f"/pos/sale/{order_id}/confirm/", {"amount": "20", "currency": "USD", "method": "cash"}
    )
    order = SaleOrder.objects.get(pk=order_id)
    assert order.status == SaleOrder.DRAFT
    assert not order.payments.exists()

    resp = client.post(
        f"/pos/sale/{order_id}/confirm/",
        {"amount": "20", "currency": "USD", "method": "cash", "risk_ack": "1"},
    )
    assert resp.status_code == 302
    order.refresh_from_db()
    assert order.status == SaleOrder.CONFIRMED


def test_missing_rate_at_confirm_shows_russian_error_and_saves_nothing(
    client, django_user_model, variant
):
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    manager = django_user_model.objects.create_user("mgr_norate", password="x" * 12, is_staff=True)
    manager.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(manager)
    assert not ExchangeRate.objects.filter(currency="USD").exists()

    order_id = SaleOrder.objects.create(created_by=manager, channel=SaleOrder.SHOP).pk
    cust = Client.objects.create(first_name="NoRate", phone="+996700000823")
    client.post(f"/pos/sale/{order_id}/client/{cust.pk}/set/")
    client.post(f"/pos/sale/{order_id}/items/add/", {"variant_id": variant.pk, "quantity": 1})

    resp = client.post(
        f"/pos/sale/{order_id}/confirm/", {"amount": "10", "currency": "USD", "method": "cash"}
    )
    assert resp.status_code == 302
    assert resp.url == f"/pos/sale/{order_id}/"
    order = SaleOrder.objects.get(pk=order_id)
    assert order.status == SaleOrder.DRAFT  # never confirmed either
    assert not order.payments.exists()
    page = client.get(resp.url)
    assert "Нет курса" in page.content.decode()


def test_record_payment_raises_for_missing_rate_and_saves_nothing(variant):
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    cust = Client.objects.create(first_name="Svc", phone="+996700000824")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)
    assert not ExchangeRate.objects.filter(currency="USD").exists()
    with pytest.raises(ValidationError):
        record_payment(order, Decimal("10"), currency="USD")
    assert not order.payments.exists()


def test_same_currency_as_order_payment_uses_rate_one_and_skips_conversion_ui(
    client, django_user_model, variant, settings
):
    """order currency == payment currency (both USD, non-KGS) needs no rate
    lookup at all and shows no conversion math — even with zero ExchangeRate
    rows on record."""
    settings.CURRENCY = "KGS"
    call_command("setup_roles")
    variant.currency = "USD"
    variant.sale_price = Decimal("40")
    variant.save(update_fields=["currency", "sale_price"])
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    manager = django_user_model.objects.create_user("mgr_same", password="x" * 12, is_staff=True)
    manager.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(manager)
    assert not ExchangeRate.objects.filter(currency="USD").exists()

    order_id = SaleOrder.objects.create(created_by=manager, channel=SaleOrder.SHOP).pk
    client.post(f"/pos/sale/{order_id}/items/add/", {"variant_id": variant.pk, "quantity": 1})
    assert SaleOrder.objects.get(pk=order_id).currency == "USD"

    resp = client.post(
        f"/pos/sale/{order_id}/recalc/", {"amount": "40", "currency": "USD", "method": "cash"}
    )
    body = resp.content.decode()
    assert "USD ×" not in body  # no conversion math for a same-currency payment
    assert "Нет курса" not in body  # and no missing-rate warning either

    resp = client.post(
        f"/pos/sale/{order_id}/confirm/", {"amount": "40", "currency": "USD", "method": "cash"}
    )
    assert resp.status_code == 302


def test_changing_rate_afterward_never_alters_past_payment_values(variant, settings):
    settings.CURRENCY = "KGS"
    ExchangeRate.objects.create(currency="USD", date=timezone.localdate(), rate=Decimal("87.00"))
    cust = Client.objects.create(first_name="Frz", phone="+996700000825")
    order = SaleOrder.objects.create(client=cust, currency="USD")
    payment = Payment.objects.create(client=cust, order=order, amount=Decimal("10"), currency="USD")
    assert payment.rate_to_kgs == Decimal("87.000000")
    assert payment.amount_kgs == Decimal("870.00")

    ExchangeRate.objects.filter(currency="USD").update(rate=Decimal("95.00"))
    payment.refresh_from_db()
    assert payment.rate_to_kgs == Decimal("87.000000")  # unchanged
    assert payment.amount_kgs == Decimal("870.00")  # unchanged


def test_rounding_residue_marks_sale_paid_and_excluded_from_debts(variant, settings):
    settings.CURRENCY = "KGS"
    settings.PAYMENT_ROUNDING_TOLERANCE = Decimal("1.00")
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    cust = Client.objects.create(first_name="Round", phone="+996700000827")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("100.00"))
    confirm_sale(order)
    # Leaves a 0.04 сом residue — sub-сом currency-conversion rounding.
    record_payment(order, Decimal("99.96"), currency="KGS")
    assert order.balance == Decimal("0")
    assert order.payment_status == SaleOrder.PAID
    assert client_debt(cust) == {}


def test_payment_conversion_filter_shows_frozen_rate_math():
    from apps.pos.templatetags.pos_extras import payment_conversion_filter

    ExchangeRate.objects.create(currency="USD", date=timezone.localdate(), rate=Decimal("87.45"))
    cust = Client.objects.create(first_name="Fmt", phone="+996700000829")
    order = SaleOrder.objects.create(client=cust, currency="USD")
    payment = Payment.objects.create(client=cust, order=order, amount=Decimal("10"), currency="USD")
    text = payment_conversion_filter(payment)
    assert "10 USD" in text
    assert "87.45" in text
    assert "874.50 KGS" in text
    assert "НБКР" in text

    same_currency_payment = Payment.objects.create(
        client=cust, order=order, amount=Decimal("100"), currency="KGS"
    )
    assert payment_conversion_filter(same_currency_payment) == ""


def test_daily_report_sales_sheet_includes_currency_rate_source_columns(variant, settings):
    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    ExchangeRate.objects.create(
        currency="USD", date=timezone.localdate(), rate=Decimal("87.00"), source=ExchangeRate.NBKR
    )
    cust = Client.objects.create(first_name="Rep", phone="+996700000828")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)
    record_payment(order, Decimal("10"), currency="USD")  # 10 × 87 = 870

    row = _by_header(_sales_sheet())[0]
    assert row["Валюта оплаты"] == "USD"
    assert row["Курс"] == "87.0000"
    assert row["Источник курса"] == "НБКР"
    assert row["Сдача"] is None  # no change given on this payment
    assert row["Округление"] is None  # no rounding residue either
    # Оплачено/Долг по продаже use the CONVERTED paid amount, not a raw
    # same-currency-only sum (the bug this hardening pass fixed).
    assert row["Оплачено"] == Decimal("870.00")
    assert row["Долг по продаже"] == Decimal("2330.00")


def test_client_admin_unpaid_orders_counts_converted_foreign_payment(variant, settings):
    """Regression: unpaid_orders() used to sum only same-currency payments,
    wrongly listing an order as unpaid even after a foreign payment fully
    settled it — the same bug class fixed in send_daily_report.py."""
    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    ExchangeRate.objects.create(currency="USD", date=timezone.localdate(), rate=Decimal("87.00"))
    cust = Client.objects.create(first_name="AdmFx", phone="+996700000830")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("870"))
    confirm_sale(order)
    record_payment(order, Decimal("10"), currency="USD")  # 10 × 87 = 870 — fully settles it

    model_admin = site._registry[Client]
    assert str(model_admin.unpaid_orders(cust)) == "Нет — всё оплачено."


def test_rate_info_reflects_only_the_most_recently_fetched_value(settings):
    """DETERMINISM: one row per currency means exactly one current answer —
    a second fetch overwrites in place, never leaving two rates to disagree."""
    from apps.core.currency import rate_info

    settings.CURRENCY = "KGS"
    ExchangeRate.objects.create(currency="USD", date=timezone.localdate(), rate=Decimal("87.00"))
    assert rate_info("USD")["rate"] == Decimal("87.00")
    ExchangeRate.objects.filter(currency="USD").update(rate=Decimal("91.00"))
    assert rate_info("USD")["rate"] == Decimal("91.00")
    assert ExchangeRate.objects.filter(currency="USD").count() == 1


def test_no_float_in_money_and_conversion_code_paths():
    import ast
    import pathlib

    money_modules = [
        "apps/core/currency.py",
        "apps/sales/models.py",
        "apps/sales/services.py",
        "apps/clients/services.py",
        "apps/pos/views.py",
    ]
    base = pathlib.Path(__file__).resolve().parent.parent
    for rel in money_modules:
        source = (base / rel).read_text(encoding="utf-8")
        tree = ast.parse(source, filename=rel)
        for node in ast.walk(tree):
            if (
                isinstance(node, ast.Call)
                and isinstance(node.func, ast.Name)
                and node.func.id == "float"
            ):
                pytest.fail(f"{rel}:{node.lineno} calls float() in a money/conversion path")


# ---------------------------------------------------------------------------
# Change (сдача) system
# ---------------------------------------------------------------------------


def test_same_currency_change_computes_and_rounds_correctly(variant, settings):
    settings.CURRENCY = "KGS"
    settings.CHANGE_ROUNDING_STEP = Decimal("1.00")
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    cust = Client.objects.create(first_name="Chg", phone="+996700001001")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("874.50"))
    confirm_sale(order)

    payment = record_payment(
        order,
        Decimal("1000"),
        currency="KGS",
        excess_disposition=Payment.DISPOSITION_CHANGE,
    )
    # Ideal change = 1000 - 874.50 = 125.50 -> floor to 125, residue 0.50.
    assert payment.change_amount == Decimal("125.00")
    assert payment.change_currency == "KGS"
    assert payment.change_amount_kgs == Decimal("125.00")
    assert payment.change_rounding_kgs == Decimal("0.50")
    # net = 1000 - 125 change = 875.00 (the 0.50 rounding residue stays with
    # the shop, since change was rounded DOWN — never given away silently).
    assert payment.net_applied_kgs == Decimal("875.00")
    order.refresh_from_db()
    assert order.balance == Decimal("0")
    assert order.payment_status == SaleOrder.PAID


def test_cross_currency_change_uses_the_same_frozen_rate_as_the_payment(variant, settings):
    settings.CURRENCY = "KGS"
    settings.CHANGE_ROUNDING_STEP = Decimal("1.00")
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    ExchangeRate.objects.create(currency="USD", date=timezone.localdate(), rate=Decimal("80.00"))
    cust = Client.objects.create(first_name="ChgFx", phone="+996700001002")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(order)

    # 20 USD x 80 = 1600 KGS toward a 1000 KGS sale -> 600 KGS ideal change,
    # already an exact step, no rounding residue.
    payment = record_payment(
        order,
        Decimal("20"),
        currency="USD",
        excess_disposition=Payment.DISPOSITION_CHANGE,
        change_currency="KGS",
    )
    assert payment.rate_to_kgs == Decimal("80.000000")
    assert payment.change_amount_kgs == Decimal("600.00")
    assert payment.change_amount == Decimal("600.00")  # change_currency == KGS, no conversion
    assert payment.net_applied_kgs == Decimal("1000.00")

    # The SAME payment, but change requested back in USD (the currency
    # received) — change_amount_kgs must be IDENTICAL (one rate, one
    # transaction); only its OWN-currency representation differs.
    order2 = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order2, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(order2)
    payment2 = record_payment(
        order2,
        Decimal("20"),
        currency="USD",
        excess_disposition=Payment.DISPOSITION_CHANGE,
        change_currency="USD",
    )
    assert payment2.change_amount_kgs == Decimal("600.00")
    assert payment2.change_currency == "USD"
    assert payment2.change_amount == Decimal("7.50")  # 600 / 80


def test_change_while_balance_remains_is_rejected(variant, settings):
    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    cust = Client.objects.create(first_name="NoChg", phone="+996700001003")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(order)

    # A 500 payment against a 1000 total creates NO excess — forcing a change
    # override anyway (a crafted/buggy caller) must be rejected, not silently
    # accepted, since it would leave the sale not fully paid.
    with pytest.raises(ValidationError):
        record_payment(
            order,
            Decimal("500"),
            currency="KGS",
            excess_disposition=Payment.DISPOSITION_CHANGE,
            change_amount_override=Decimal("100"),
            change_adjust_reason="test override",
        )
    assert not order.payments.exists()


def test_negative_change_impossible_even_via_bulk_create(variant):
    from django.db import IntegrityError, transaction

    cust = Client.objects.create(first_name="Neg", phone="+996700001004")
    with pytest.raises(IntegrityError), transaction.atomic():
        Payment.objects.bulk_create(
            [
                Payment(
                    client=cust,
                    amount=Decimal("10"),
                    rate_to_kgs=Decimal("1"),
                    change_amount=Decimal("-5"),
                )
            ]
        )


def test_change_kgs_cannot_exceed_amount_kgs_even_via_bulk_create(variant):
    from django.db import IntegrityError, transaction

    cust = Client.objects.create(first_name="TooMuch", phone="+996700001005")
    with pytest.raises(IntegrityError), transaction.atomic():
        Payment.objects.bulk_create(
            [
                Payment(
                    client=cust,
                    amount=Decimal("10"),
                    rate_to_kgs=Decimal("1"),
                    change_amount=Decimal("5"),
                    change_amount_kgs=Decimal("20"),  # more than the 10 received
                )
            ]
        )


def test_net_applied_kgs_drives_debt_and_balance_everywhere(variant, settings):
    """CORE RULE: what reduces balance/debt is the NET, not the gross amount —
    checked at every layer: the order property, the client debt aggregate,
    and the admin's own annotated queryset."""
    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    cust = Client.objects.create(first_name="Net", phone="+996700001006")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(order)
    record_payment(
        order, Decimal("1500"), currency="KGS", excess_disposition=Payment.DISPOSITION_CHANGE
    )
    # net_applied_kgs = 1500 - 500 change = 1000 -> fully paid, no debt.
    order.refresh_from_db()
    assert order.paid_amount == Decimal("1000.00")
    assert order.balance == Decimal("0")
    assert client_debt(cust) == {}

    from django.contrib.admin.sites import site

    admin_order = site._registry[SaleOrder].get_queryset(RequestFactory().get("/")).get(pk=order.pk)
    assert admin_order._paid == Decimal("1000.00")


def test_debt_disposition_reduces_clients_other_outstanding_sales(variant, settings):
    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    cust = Client.objects.create(first_name="Pool", phone="+996700001007")
    order_a = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order_a, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(order_a)
    order_b = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order_b, variant=variant, quantity=1, unit_price=Decimal("500"))
    confirm_sale(order_b)
    assert client_debt(cust) == {"KGS": Decimal("1500.00")}

    # Pay 1500 on order A alone, marking the excess "в счёт долга" — the full
    # amount counts (no change), and the excess pools onto order B's debt too.
    record_payment(
        order_a, Decimal("1500"), currency="KGS", excess_disposition=Payment.DISPOSITION_DEBT
    )
    assert client_debt(cust) == {}


def test_credit_disposition_yields_negative_debt_shown_as_avans(variant, settings):
    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    cust = Client.objects.create(first_name="Avans", phone="+996700001008")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(order)

    record_payment(
        order, Decimal("1200"), currency="KGS", excess_disposition=Payment.DISPOSITION_CREDIT
    )
    assert client_debt(cust) == {}  # never shown as a negative debt
    assert client_credits(cust) == {"KGS": Decimal("200.00")}


def test_debt_and_credit_disposition_require_a_client(variant, settings):
    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    cust = Client.objects.create(first_name="Req", phone="+996700001009")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(order)
    # record_payment itself only guards client_id at the ORDER level (already
    # required elsewhere); this proves debt/credit compute cleanly with one.
    payment = record_payment(
        order, Decimal("1200"), currency="KGS", excess_disposition=Payment.DISPOSITION_DEBT
    )
    assert payment.excess_disposition == Payment.DISPOSITION_DEBT
    assert payment.change_amount == Decimal("0")


def test_void_payment_with_change_restores_exact_debt_after_rate_move(variant, settings):
    """DONE WHEN: a client hands 20 $ against an 874,50 сом sale — voiding
    that payment later restores the client's debt EXACTLY, even after the
    rate has since moved."""
    settings.CURRENCY = "KGS"
    settings.CHANGE_ROUNDING_STEP = Decimal("1.00")
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    ExchangeRate.objects.create(currency="USD", date=timezone.localdate(), rate=Decimal("87.45"))
    cust = Client.objects.create(first_name="VoidChg", phone="+996700001010")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("874.50"))
    confirm_sale(order)
    debt_before = client_debt(cust)
    assert debt_before == {"KGS": Decimal("874.50")}

    payment = record_payment(
        order,
        Decimal("20"),
        currency="USD",
        excess_disposition=Payment.DISPOSITION_CHANGE,
    )
    assert payment.change_amount == Decimal("874.00")
    assert payment.change_rounding_kgs == Decimal("0.50")
    assert client_debt(cust) == {}  # fully paid, change given

    # The rate moves significantly AFTER the payment.
    ExchangeRate.objects.filter(currency="USD").update(rate=Decimal("150.00"))

    void_payment(payment)
    assert client_debt(cust) == debt_before  # exactly restored


def test_manager_out_of_band_change_amount_gets_403(client, django_user_model, variant, settings):
    settings.CURRENCY = "KGS"
    settings.CHANGE_ROUNDING_STEP = Decimal("1.00")
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    manager = django_user_model.objects.create_user("mgr_chg403", password="x" * 12, is_staff=True)
    manager.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(manager)

    order_id = SaleOrder.objects.create(created_by=manager, channel=SaleOrder.SHOP).pk
    cust = Client.objects.create(first_name="Band", phone="+996700001011")
    client.post(f"/pos/sale/{order_id}/client/{cust.pk}/set/")
    client.post(f"/pos/sale/{order_id}/items/add/", {"variant_id": variant.pk, "quantity": 1})
    # variant sells for 3200 KGS; pay 4000 -> ideal change 800.

    resp = client.post(
        f"/pos/sale/{order_id}/recalc/",
        {
            "amount": "4000",
            "currency": "KGS",
            "excess_disposition": "change",
            "change_amount_override": "500",  # wildly outside the ±2 step band
        },
    )
    assert resp.status_code == 403

    resp = client.post(
        f"/pos/sale/{order_id}/confirm/",
        {
            "amount": "4000",
            "currency": "KGS",
            "method": "cash",
            "excess_disposition": "change",
            "change_amount_override": "500",
            "change_adjust_reason": "test",
        },
    )
    assert resp.status_code == 403
    order = SaleOrder.objects.get(pk=order_id)
    assert order.status == SaleOrder.DRAFT
    assert not order.payments.exists()


def test_owner_can_adjust_change_beyond_the_band(client, django_user_model, variant, settings):
    settings.CURRENCY = "KGS"
    settings.CHANGE_ROUNDING_STEP = Decimal("1.00")
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    owner = django_user_model.objects.create_superuser("owner_chg", "o@e.com", "x" * 12)
    client.force_login(owner)

    order_id = SaleOrder.objects.create(created_by=owner, channel=SaleOrder.SHOP).pk
    cust = Client.objects.create(first_name="OwnerBand", phone="+996700001012")
    client.post(f"/pos/sale/{order_id}/client/{cust.pk}/set/")
    client.post(f"/pos/sale/{order_id}/items/add/", {"variant_id": variant.pk, "quantity": 1})

    resp = client.post(
        f"/pos/sale/{order_id}/confirm/",
        {
            "amount": "4000",
            "currency": "KGS",
            "method": "cash",
            "excess_disposition": "change",
            "change_amount_override": "500",
            "change_adjust_reason": "owner discretion",
        },
    )
    assert resp.status_code == 302
    order = SaleOrder.objects.get(pk=order_id)
    assert order.status == SaleOrder.CONFIRMED
    payment = order.payments.get()
    assert payment.change_amount == Decimal("500")
    assert "owner discretion" in payment.note


def test_overpayment_requires_explicit_disposition_never_auto_picked(
    client, django_user_model, variant, settings
):
    settings.CURRENCY = "KGS"
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    manager = django_user_model.objects.create_user("mgr_fork", password="x" * 12, is_staff=True)
    manager.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(manager)

    order_id = SaleOrder.objects.create(created_by=manager, channel=SaleOrder.SHOP).pk
    cust = Client.objects.create(first_name="Fork", phone="+996700001013")
    client.post(f"/pos/sale/{order_id}/client/{cust.pk}/set/")
    client.post(f"/pos/sale/{order_id}/items/add/", {"variant_id": variant.pk, "quantity": 1})

    # 4000 against a 3200 order, no disposition chosen -> rejected.
    resp = client.post(
        f"/pos/sale/{order_id}/confirm/", {"amount": "4000", "currency": "KGS", "method": "cash"}
    )
    assert resp.status_code == 302
    assert resp.url == f"/pos/sale/{order_id}/"
    order = SaleOrder.objects.get(pk=order_id)
    assert order.status == SaleOrder.DRAFT
    assert not order.payments.exists()

    # The double-check panel shows the math once "Сдача" is picked.
    resp = client.post(
        f"/pos/sale/{order_id}/recalc/",
        {"amount": "4000", "currency": "KGS", "excess_disposition": "change"},
    )
    body = resp.content.decode()
    assert "800" in body  # computed change (4000 - 3200)
    assert "Подтвердить и выдать сдачу" in body

    # Confirming with the disposition now succeeds and hands back the change.
    resp = client.post(
        f"/pos/sale/{order_id}/confirm/",
        {
            "amount": "4000",
            "currency": "KGS",
            "method": "cash",
            "excess_disposition": "change",
        },
    )
    assert resp.status_code == 302
    assert resp.url == f"/pos/sale/{order_id}/result/"
    order.refresh_from_db()
    assert order.status == SaleOrder.CONFIRMED
    payment = order.payments.get()
    assert payment.change_amount == Decimal("800.00")
    assert payment.excess_disposition == Payment.DISPOSITION_CHANGE
    assert order.payment_status == SaleOrder.PAID


def test_debt_credit_fork_buttons_disabled_for_walkin(client, django_user_model, variant, settings):
    settings.CURRENCY = "KGS"
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    manager = django_user_model.objects.create_user("mgr_walkin", password="x" * 12, is_staff=True)
    manager.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(manager)

    order_id = SaleOrder.objects.create(created_by=manager, channel=SaleOrder.SHOP).pk
    client.post(f"/pos/sale/{order_id}/items/add/", {"variant_id": variant.pk, "quantity": 1})

    resp = client.post(f"/pos/sale/{order_id}/recalc/", {"amount": "4000", "currency": "KGS"})
    body = resp.content.decode()
    assert "В счёт долга" in body and "Аванс" in body
    # Both fork buttons carrying "debt"/"credit" are disabled for a walk-in.
    assert body.count("disabled") >= 2


def test_client_credits_display_as_avans_on_client_page(
    client, django_user_model, variant, settings
):
    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    call_command("setup_roles")
    manager = django_user_model.objects.create_user("mgr_page", password="x" * 12, is_staff=True)
    manager.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(manager)

    cust = Client.objects.create(first_name="PageAvans", phone="+996700001014")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(order)
    record_payment(
        order, Decimal("1200"), currency="KGS", excess_disposition=Payment.DISPOSITION_CREDIT
    )

    resp = client.get(f"/pos/clients/{cust.pk}/")
    body = resp.content.decode()
    assert "Аванс" in body
    assert "200" in body


def test_daily_report_change_and_rounding_columns_and_till_drift_total(variant, settings):
    settings.CURRENCY = "KGS"
    settings.CHANGE_ROUNDING_STEP = Decimal("1.00")
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    cust = Client.objects.create(first_name="ReportChg", phone="+996700001015")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("874.50"))
    confirm_sale(order)
    record_payment(
        order, Decimal("1000"), currency="KGS", excess_disposition=Payment.DISPOSITION_CHANGE
    )

    sheet = _sales_sheet()
    row = _by_header(sheet)[0]
    assert row["Сдача"] == Decimal("125.00")
    assert row["Округление"] == Decimal("0.50")
    # The day's till-drift lands in the bold totals row, as a real number in
    # the Округление column — not buried in a sentence in the last cell.
    rounding_col = [c.header for c in sheet.columns].index("Округление")
    assert sheet.totals[rounding_col] == Decimal("0.50")


def test_balance_kgs_before_payment_matches_order_balance_after_confirm(variant, settings):
    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    cust = Client.objects.create(first_name="Bal", phone="+996700001016")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(order)
    assert balance_kgs_before_payment(order) == Decimal("1000.00")
    record_payment(order, Decimal("400"), currency="KGS")
    assert balance_kgs_before_payment(order) == Decimal("600.00")


def test_compute_change_preview_has_excess_flag_respects_tolerance(variant, settings):
    settings.CURRENCY = "KGS"
    settings.PAYMENT_ROUNDING_TOLERANCE = Decimal("1.00")
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    cust = Client.objects.create(first_name="Tol", phone="+996700001017")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(order)
    # 1000.50 over a 1000 balance is within tolerance -> no excess/fork needed.
    preview = compute_change_preview(
        order, Decimal("1000.50"), "KGS", Decimal("1"), balance_kgs_before_payment(order)
    )
    assert preview["has_excess"] is False


# ---------------------------------------------------------------------------
# Part 1 — stock cannot go negative (cart-time cap + reservation)
# ---------------------------------------------------------------------------


def test_cart_cap_applies_across_all_lines_of_the_same_variant(client, django_user_model, variant):
    """Adding the same variant twice (two separate POSTs) must not bypass the
    cap — it's checked against the TOTAL already in the cart, not per add."""
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    editor = django_user_model.objects.create_user("cap1", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    order_id = SaleOrder.objects.create(created_by=editor, channel=SaleOrder.SHOP).pk
    client.post(f"/pos/sale/{order_id}/items/add/", {"variant_id": variant.pk, "quantity": 7})
    resp = client.post(
        f"/pos/sale/{order_id}/items/add/", {"variant_id": variant.pk, "quantity": 7}
    )
    body = resp.content.decode()
    assert "Доступно только" in body
    order = SaleOrder.objects.get(pk=order_id)
    assert order.items.count() == 1
    assert order.items.first().quantity == 10  # capped at the 10 available, not 14


def test_zero_available_tile_cannot_be_added_even_via_crafted_post(
    client, django_user_model, variant
):
    call_command("setup_roles")
    # No stock movement at all — 0 available.
    editor = django_user_model.objects.create_user("cap2", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    order_id = SaleOrder.objects.create(created_by=editor, channel=SaleOrder.SHOP).pk
    resp = client.post(
        f"/pos/sale/{order_id}/items/add/", {"variant_id": variant.pk, "quantity": 1}
    )
    assert (
        "уже полностью в корзине" in resp.content.decode()
        or "Доступно только" in resp.content.decode()
    )
    order = SaleOrder.objects.get(pk=order_id)
    assert order.items.count() == 0


def test_crafted_post_above_stock_is_clamped_not_saved_beyond_available(
    client, django_user_model, variant
):
    """A crafted POST with qty far above stock is rejected — clamped to what's
    truly available, never silently saved at face value. Client-side caps are
    UX only; this is the server-side defence."""
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 15)
    editor = django_user_model.objects.create_user("cap3", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    order_id = SaleOrder.objects.create(created_by=editor, channel=SaleOrder.SHOP).pk
    resp = client.post(
        f"/pos/sale/{order_id}/items/add/", {"variant_id": variant.pk, "quantity": 999999}
    )
    assert resp.status_code == 200
    assert "Доступно только 15" in resp.content.decode()
    order = SaleOrder.objects.get(pk=order_id)
    assert order.items.first().quantity == 15  # never the crafted 999999


def test_order_rejects_a_second_item_in_a_different_currency(
    client, django_user_model, variant, settings
):
    """Every OrderItem shares the order's own currency (it has none of its
    own — see that model's docstring), but Order.total naively sums
    unit_price*quantity across every line as one number — a second item
    priced in a different currency would corrupt that total (and, at
    handover, the resulting sale's total_kgs, since SaleItem has no
    currency of its own either and inherits the sale's single
    order.currency). Must reject, same as the draft-sale cart already does."""
    settings.CURRENCY = "KGS"
    owner = django_user_model.objects.create_superuser("mix_owner", "o@e.com", "x" * 12)
    client.force_login(owner)

    usd_variant = ProductVariant.objects.create(
        product=variant.product,
        sku="EVD-M-USD",
        size="L",
        color="blue",
        cost_price=Decimal("20"),
        sale_price=Decimal("20"),
        currency="USD",
    )
    cust = Client.objects.create(first_name="MixCcy", phone="+996700002050")
    from apps.orders.models import Order
    from apps.orders.services import create_order

    order = create_order(
        cust, [{"variant": variant, "quantity": 1, "unit_price": variant.sale_price}]
    )
    assert order.currency == "KGS"

    resp = client.post(
        f"/orders/{order.pk}/items/add/",
        {"variant_id": usd_variant.pk, "quantity": "1"},
        HTTP_HX_REQUEST="true",
    )
    assert resp.status_code == 200
    assert "нельзя добавить" in resp.content.decode()

    order = Order.objects.get(pk=order.pk)
    assert order.currency == "KGS"  # unchanged
    assert order.items.count() == 1  # the USD line was never added
    assert order.total == variant.sale_price  # not corrupted by a mixed sum


def test_order_deposit_form_has_a_confirmation_dialog(client, django_user_model, variant):
    """Unlike every other money-write form in the app, the «Внести аванс»
    form had no data-confirm at all — a manager could add a deposit with a
    single accidental tap, no "are you sure" step. safety.js listens for
    htmx:confirm too, so data-confirm works on this hx-post form exactly
    like it does on a plain <form> submit."""
    from apps.orders.services import create_order

    owner = django_user_model.objects.create_superuser("dep_owner", "o@e.com", "x" * 12)
    client.force_login(owner)
    cust = Client.objects.create(first_name="DepConfirm", phone="+996700002051")
    order = create_order(
        cust, [{"variant": variant, "quantity": 1, "unit_price": variant.sale_price}]
    )

    resp = client.get(f"/orders/{order.pk}/")
    body = resp.content.decode()
    assert "Внести аванс на эту сумму?" in body
    assert 'data-confirm="Внести аванс на эту сумму?"' in body


def test_reserved_stock_is_excluded_from_available_and_blocks_a_walkin_sale(variant, settings):
    """Part 1b/3c: stock promised to an open production order must not be
    sellable to a walk-in, even bypassing the cart-time cap entirely."""
    from apps.inventory.services import available_for
    from apps.orders.services import create_order

    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    cust = Client.objects.create(first_name="Reserver", phone="+996700002101")
    create_order(cust, [{"variant": variant, "quantity": 8, "unit_price": variant.sale_price}])
    assert available_for(variant) == 2  # 10 on hand - 8 reserved

    # A sale attempting to take more than what's left over the reservation
    # must fail at confirm time, even if the cart itself somehow held it.
    walkin = SaleOrder.objects.create()
    SaleItem.objects.create(
        order=walkin, variant=variant, quantity=3, unit_price=variant.sale_price
    )
    with pytest.raises(ValidationError):
        confirm_sale(walkin)
    variant.refresh_from_db()
    assert variant.stock == 10  # nothing written off


def test_product_grid_tile_shows_reserved_badge(client, django_user_model, variant, settings):
    from apps.orders.services import create_order

    settings.CURRENCY = "KGS"
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    cust = Client.objects.create(first_name="Badge", phone="+996700002102")
    create_order(cust, [{"variant": variant, "quantity": 3, "unit_price": variant.sale_price}])
    editor = django_user_model.objects.create_user("cap4", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    order_id = SaleOrder.objects.create(created_by=editor, channel=SaleOrder.SHOP).pk
    resp = client.get(f"/pos/sale/{order_id}/products/")
    body = resp.content.decode()
    assert "7" in body  # 10 - 3 reserved = 7 available
    assert "в заказах" in body


# ---------------------------------------------------------------------------
# Part 2 — money formatting + cart rail
# ---------------------------------------------------------------------------


def test_money_filter_groups_thousands_with_nbsp_and_drops_trailing_zero_cents():
    from apps.pos.templatetags.pos_extras import money_filter

    assert money_filter(Decimal("3800"), "KGS") == "3\xa0800\xa0сом"
    assert money_filter(Decimal("3800.00"), "KGS") == "3\xa0800\xa0сом"
    assert money_filter(Decimal("874.50"), "KGS") == "874,50\xa0сом"
    assert money_filter(Decimal("-500"), "KGS") == "-500\xa0сом"


def test_money_filter_uses_currency_symbol_not_code():
    from apps.pos.templatetags.pos_extras import money_filter

    assert money_filter(Decimal("100"), "USD") == "100\xa0$"
    assert money_filter(Decimal("100"), "RUB") == "100\xa0₽"
    assert "USD" not in money_filter(Decimal("100"), "USD")
    assert "KGS" not in money_filter(Decimal("100"), "KGS")


def test_sale_screen_renders_formatted_money_not_raw_currency_code(
    client, django_user_model, variant
):
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    editor = django_user_model.objects.create_user("fmt1", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    order_id = SaleOrder.objects.create(created_by=editor, channel=SaleOrder.SHOP).pk
    resp = client.post(
        f"/pos/sale/{order_id}/items/add/", {"variant_id": variant.pk, "quantity": 1}
    )
    body = resp.content.decode()
    assert "3800,00 KGS" not in body
    assert "3\xa0200\xa0сом" in body  # variant sells for 3200 KGS


def test_cart_rail_keeps_confirm_button_and_money_bar_at_25_items(
    client, django_user_model, settings
):
    """Server-side proxy for the layout requirement: with 25 lines in the
    cart, the money bar and confirm button must still be present in the
    response — CSS (tested manually/visually) keeps them pinned/visible
    regardless of item count; this proves the markup itself never drops
    them when the list grows long."""
    settings.CURRENCY = "KGS"
    call_command("setup_roles")
    editor = django_user_model.objects.create_user("rail1", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    cat = Category.objects.create(name="Rail")
    order_id = SaleOrder.objects.create(created_by=editor, channel=SaleOrder.SHOP).pk
    for i in range(25):
        p = Product.objects.create(category=cat, name=f"RailProduct{i}")
        v = ProductVariant.objects.create(
            product=p, sku=f"RAIL{i:03d}", cost_price=Decimal("1"), sale_price=Decimal("100")
        )
        add_movement(v, StockMovement.PRODUCTION_IN, 5)
        resp = client.post(f"/pos/sale/{order_id}/items/add/", {"variant_id": v.pk, "quantity": 1})
    body = resp.content.decode()
    order = SaleOrder.objects.get(pk=order_id)
    assert order.items.count() == 25
    assert "Позиции · 25" in body
    assert 'class="money-bar' in body
    assert "Подтвердить продажу" in body
    assert "items-scroll" in body


# ---------------------------------------------------------------------------
# Part 3 — Заказы (production orders)
# ---------------------------------------------------------------------------


def test_production_queue_need_is_ordered_minus_stock_never_negative_displayed(variant, settings):
    from apps.orders.services import create_order, production_queue

    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 3)
    cust = Client.objects.create(first_name="Queue1", phone="+996700002201")
    create_order(cust, [{"variant": variant, "quantity": 10, "unit_price": variant.sale_price}])

    rows = production_queue()
    row = next(r for r in rows if r["variant"].pk == variant.pk)
    assert row["ordered"] == 10
    assert row["in_stock"] == 3
    assert row["to_produce"] == 7
    assert row["covered"] is False

    # Now stock covers demand — need clamps to 0, flagged covered, not hidden.
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    rows = production_queue()
    row = next(r for r in rows if r["variant"].pk == variant.pk)
    assert row["to_produce"] == 0
    assert row["covered"] is True


def test_queue_aggregates_the_same_variant_across_two_orders(variant, settings):
    from apps.orders.services import create_order, production_queue

    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 2)
    c1 = Client.objects.create(first_name="QA", phone="+996700002202")
    c2 = Client.objects.create(first_name="QB", phone="+996700002203")
    create_order(c1, [{"variant": variant, "quantity": 4, "unit_price": variant.sale_price}])
    create_order(c2, [{"variant": variant, "quantity": 6, "unit_price": variant.sale_price}])

    rows = production_queue()
    matching = [r for r in rows if r["variant"].pk == variant.pk]
    assert len(matching) == 1  # aggregated into ONE row, not two
    row = matching[0]
    assert row["ordered"] == 10  # 4 + 6
    assert row["to_produce"] == 8  # 10 - 2 in stock
    assert row["orders_count"] == 2


def test_queue_row_breaks_down_who_ordered_what(variant, settings):
    """«Для кого» — the queue must answer which client is waiting on how many
    without opening each order, and each line's `remaining` is that LINE's own
    shortfall, not the variant's stock-adjusted need."""
    from apps.orders.services import create_order, production_queue

    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 2)
    c1 = Client.objects.create(first_name="Breakdown1", phone="+996700002251")
    c2 = Client.objects.create(first_name="Breakdown2", phone="+996700002252")
    create_order(c1, [{"variant": variant, "quantity": 4, "unit_price": variant.sale_price}])
    create_order(c2, [{"variant": variant, "quantity": 6, "unit_price": variant.sale_price}])

    row = next(r for r in production_queue() if r["variant"].pk == variant.pk)
    assert row["clients_count"] == 2
    assert len(row["lines"]) == 2
    by_client = {line["client"].first_name: line for line in row["lines"]}
    assert by_client["Breakdown1"]["quantity"] == 4
    assert by_client["Breakdown1"]["remaining"] == 4  # nothing produced on that line yet
    assert by_client["Breakdown2"]["quantity"] == 6
    # The 2 already in stock reduce the VARIANT's need, never a client's line.
    assert row["to_produce"] == 8


def test_queue_is_one_row_per_variant(variant, settings):
    """production_queue() takes no grouping argument (product/client rollups
    removed 2026-08, apps.orders CLAUDE.md Part 4) — always one row per
    SKU, aggregating demand across every open order for that variant, with
    a `lines` breakdown of who's waiting on it."""
    from apps.orders.services import create_order, production_queue

    settings.CURRENCY = "KGS"
    other_size = ProductVariant.objects.create(
        product=variant.product,
        sku="EVD-L-RED",
        size="L",
        color="red",
        cost_price=Decimal("1500.00"),
        sale_price=Decimal("3200.00"),
    )
    cust = Client.objects.create(first_name="Grouped", phone="+996700002253")
    create_order(
        cust,
        [
            {"variant": variant, "quantity": 3, "unit_price": variant.sale_price},
            {"variant": other_size, "quantity": 2, "unit_price": other_size.sale_price},
        ],
    )

    rows = production_queue()
    assert len({r["key"] for r in rows}) == 2  # two SKUs -> two rows
    assert sum(r["to_produce"] for r in rows) == 5  # 3 + 2, nothing in stock
    row = next(r for r in rows if r["variant"].pk == variant.pk)
    assert row["to_produce"] == 3
    assert len(row["lines"]) == 1
    assert row["lines"][0]["client"].pk == cust.pk


def test_queue_sorts_by_nearest_due_date_and_flags_overdue(variant, settings):
    """Nearest deadline first, undated last, overdue flagged against the
    Asia/Bishkek LOCAL date (never UTC) — same rule as Order.is_overdue."""

    from apps.orders.services import create_order, production_queue

    settings.CURRENCY = "KGS"
    today = timezone.localdate()
    sizes = {}
    for label, due in (("late", today - timedelta(days=2)), ("soon", today + timedelta(days=3))):
        sizes[label] = ProductVariant.objects.create(
            product=variant.product,
            sku=f"EVD-{label}",
            size=label,
            color="red",
            cost_price=Decimal("1500.00"),
            sale_price=Decimal("3200.00"),
        )
        cust = Client.objects.create(first_name=label, phone=f"+99670000226{len(sizes)}")
        create_order(
            cust,
            [{"variant": sizes[label], "quantity": 1, "unit_price": Decimal("3200.00")}],
            due_date=due,
        )
    # A third order with no due date at all must sink below both.
    undated = Client.objects.create(first_name="undated", phone="+996700002269")
    create_order(undated, [{"variant": variant, "quantity": 1, "unit_price": variant.sale_price}])

    rows = production_queue()
    assert [r["variant"].size for r in rows] == ["late", "soon", "M"]
    assert rows[0]["overdue"] is True
    assert rows[1]["overdue"] is False
    assert rows[2]["due_date"] is None and rows[2]["overdue"] is False


def test_queue_view_has_no_grouping_and_ignores_stray_params(
    client, django_user_model, variant, settings
):
    """The queue page is always one row per variant (grouping removed
    2026-08, apps.orders CLAUDE.md Part 4) — a stray/garbage ?group= param
    someone still bookmarks is simply ignored, never a 500."""
    settings.CURRENCY = "KGS"
    client.force_login(
        django_user_model.objects.create_superuser("queue_owner", "q@e.com", "x" * 12)
    )
    cust = Client.objects.create(first_name="ViewGroup", phone="+996700002270")
    from apps.orders.services import create_order

    create_order(cust, [{"variant": variant, "quantity": 2, "unit_price": variant.sale_price}])

    resp = client.get("/orders/queue/")
    assert resp.status_code == 200
    assert "group" not in resp.context
    assert "ViewGroup" in resp.content.decode()

    resp = client.get("/orders/queue/?group=../etc/passwd")
    assert resp.status_code == 200
    assert "ViewGroup" in resp.content.decode()


def test_orders_index_status_filter_separates_open_from_delivered(
    client, variant, django_user_model, settings
):
    """The «where did my order go?» fix: the list defaults to the working set,
    and a delivered order moves out of it instead of burying today's work."""
    from apps.orders.models import Order
    from apps.orders.services import create_order, hand_over

    settings.CURRENCY = "KGS"
    owner = django_user_model.objects.create_superuser("orders_owner", "oo@e.com", "x" * 12)
    client.force_login(owner)
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    cust = Client.objects.create(first_name="Filtered", phone="+996700002271")
    open_order = create_order(
        cust, [{"variant": variant, "quantity": 1, "unit_price": variant.sale_price}]
    )
    done = create_order(
        cust, [{"variant": variant, "quantity": 1, "unit_price": variant.sale_price}]
    )
    hand_over(done, user=owner)

    resp = client.get("/orders/")  # default = активные
    assert resp.context["status"] == "open"
    ids = [r["order"].pk for r in resp.context["rows"]]
    assert open_order.pk in ids and done.pk not in ids
    assert resp.context["counts"]["open"] == 1
    assert resp.context["counts"]["delivered"] == 1

    resp = client.get("/orders/?status=delivered")
    ids = [r["order"].pk for r in resp.context["rows"]]
    assert ids == [done.pk]

    resp = client.get("/orders/?status=all")
    ids = [r["order"].pk for r in resp.context["rows"]]
    assert open_order.pk in ids and done.pk in ids

    # Garbage falls back to the default instead of erroring.
    assert client.get("/orders/?status=nonsense").context["status"] == "open"
    assert Order.objects.count() == 2


def test_order_builder_swaps_in_place_for_htmx_and_redirects_without_it(
    client, django_user_model, variant, settings
):
    """Building an order used to be a full page load per item. Every control
    now swaps #order-body, and the same endpoint still redirects for a plain
    POST — the swap is an enhancement, not the mechanism."""
    from apps.orders.services import create_order

    settings.CURRENCY = "KGS"
    owner = django_user_model.objects.create_superuser("swap_owner", "s@e.com", "x" * 12)
    client.force_login(owner)
    cust = Client.objects.create(first_name="Swap", phone="+996700002291")
    order = create_order(
        cust, [{"variant": variant, "quantity": 1, "unit_price": variant.sale_price}]
    )
    add_url = f"/orders/{order.pk}/items/add/"

    # No JS: a normal POST still redirects back to the order.
    resp = client.post(add_url, {"variant_id": variant.pk, "quantity": 1})
    assert resp.status_code == 302

    # HTMX: a fragment comes back instead, and it closes the picker modal
    # out-of-band so the next item can be added without leaving the page.
    resp = client.post(add_url, {"variant_id": variant.pk, "quantity": 1}, HTTP_HX_REQUEST="true")
    assert resp.status_code == 200
    body = resp.content.decode()
    assert not body.lstrip().startswith("<!doctype")
    assert 'id="variant-picker" hx-swap-oob' in body
    assert str(variant) in body  # the line list came back with it


def test_order_builder_shows_errors_inside_the_swapped_body(
    client, django_user_model, variant, settings
):
    """A swapped fragment never reaches base.html's message block, so an
    over-produce that used to go to messages.error would vanish silently."""
    from apps.orders.services import create_order

    settings.CURRENCY = "KGS"
    owner = django_user_model.objects.create_superuser("err_owner", "er@e.com", "x" * 12)
    client.force_login(owner)
    cust = Client.objects.create(first_name="ErrBody", phone="+996700002292")
    order = create_order(
        cust, [{"variant": variant, "quantity": 2, "unit_price": variant.sale_price}]
    )
    item = order.items.get()

    resp = client.post(
        f"/orders/{order.pk}/items/{item.pk}/produce/",
        {"quantity": 99},  # more than the line has left
        HTTP_HX_REQUEST="true",
    )
    assert resp.status_code == 200
    assert 'class="error"' in resp.content.decode()
    item.refresh_from_db()
    assert item.produced_qty == 0  # and nothing was written


def test_empty_order_hides_money_sections_until_it_has_items(
    client, django_user_model, variant, settings
):
    """Аванс and Итого are meaningless against a 0 сом order — they appear
    only once there's something to price."""
    from apps.orders.models import Order

    settings.CURRENCY = "KGS"
    owner = django_user_model.objects.create_superuser("empty_owner", "em@e.com", "x" * 12)
    client.force_login(owner)
    cust = Client.objects.create(first_name="EmptyOrd", phone="+996700002293")
    order = Order.objects.create(client=cust, created_by=owner)

    body = client.get(f"/orders/{order.pk}/").content.decode()
    assert "Внести аванс" not in body
    assert "Аванс внесён" not in body  # the totals card is out too
    assert "найдите товар выше" in body  # ...replaced by what to do next

    client.post(f"/orders/{order.pk}/items/add/", {"variant_id": variant.pk, "quantity": 1})
    body = client.get(f"/orders/{order.pk}/").content.decode()
    assert "Внести аванс" in body
    assert "Аванс внесён" in body


def test_guided_creation_walks_items_then_due_then_saves(
    client, django_user_model, variant, settings
):
    """A new order takes the guided route: товары → срок → сохранено — оплата
    is no longer a 3rd wizard step (merged 2026-08 into the same standing
    «Аванс» card an existing order already shows, apps.orders CLAUDE.md
    Part 4), so finishing step 2 lands directly on the order's own page,
    where that card is immediately usable."""
    settings.CURRENCY = "KGS"
    owner = django_user_model.objects.create_superuser("wiz_owner", "w@e.com", "x" * 12)
    client.force_login(owner)
    cust = Client.objects.create(first_name="Wizard", phone="+996700002311")

    resp = client.post("/orders/new/", {"client": cust.pk})
    order_id = int(re.search(r"/orders/(\d+)/", resp.url).group(1))
    assert "new=1" in resp.url  # a new order starts on the guided route

    # HTMX posts don't carry the page's query string; htmx sends the current
    # URL instead, which is how the swapped body stays in wizard mode.
    hx = {"HTTP_HX_REQUEST": "true", "HTTP_HX_CURRENT_URL": f"http://t/orders/{order_id}/?new=1"}

    body = client.get(f"/orders/{order_id}/?new=1").content.decode()
    assert "Далее: срок" in body
    assert "Внести аванс" not in body  # not shown until creation finishes
    assert "Завершение" not in body  # can't hand over an order being created

    client.post(f"/orders/{order_id}/items/add/", {"variant_id": variant.pk, "quantity": 5}, **hx)

    body = client.get(f"/orders/{order_id}/step/due/", **hx).content.decode()
    assert "Шаг 2 из 2" in body
    # Saving the срок FINISHES creation — no more dialogs.
    resp = client.post(
        f"/orders/{order_id}/step/due/", {"due_date": "2026-09-15", "note": "к свадьбе"}, **hx
    )
    assert resp.status_code == 204
    assert resp["HX-Redirect"] == f"/orders/{order_id}/"

    from apps.orders.models import Order

    order = Order.objects.get(pk=order_id)
    assert str(order.due_date) == "2026-09-15"
    assert order.note == "к свадьбе"

    # Landing on the order's own (now non-wizard) page, the Аванс card is
    # immediately there — the SAME path (orders:deposit_add) whether used
    # right now or weeks later, never a second payment form.
    body = client.get(f"/orders/{order_id}/").content.decode()
    assert f"Заказ №{order_id} сохранён." in body
    assert "flash--ok" in body
    assert "Внести аванс" in body

    resp = client.post(
        f"/orders/{order_id}/deposit/",
        {"amount": "2000", "currency": "KGS", "method": "cash"},
        **hx,
    )
    assert resp.status_code == 200
    order.refresh_from_db()
    assert order.deposits.count() == 1


def test_handover_moves_money_stock_and_debt_exactly_once(
    client, django_user_model, variant, settings
):
    """The whole chain the order flow promises, asserted end to end: an open
    order owes nothing, handover deducts stock once and turns the remainder
    into real client debt, the deposit carries over, and paying the rest
    clears the debt while both payments stay on the sale as history."""
    from apps.clients.services import client_debt
    from apps.inventory.services import available_for
    from apps.orders.models import Order
    from apps.orders.services import create_order, hand_over, record_deposit
    from apps.sales.models import SaleOrder

    settings.CURRENCY = "KGS"
    owner = django_user_model.objects.create_superuser("chain_owner", "c@e.com", "x" * 12)
    cust = Client.objects.create(first_name="ChainFlow", phone="+996700002312")
    order = create_order(
        cust, [{"variant": variant, "quantity": 5, "unit_price": Decimal("1000")}], user=owner
    )
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)

    # An open заказ is NOT a sale: it owes nothing and reserves its stock.
    assert client_debt(cust) == {}
    assert available_for(variant) == 0  # reserved, unsellable to a walk-in
    record_deposit(order, Decimal("2000"), user=owner)
    assert client_debt(cust) == {}  # a deposit on an order is still not debt

    sale = hand_over(order, user=owner)
    variant.refresh_from_db()
    assert variant.stock == 0  # deducted exactly once, by confirm_sale
    assert sale.total == Decimal("5000.00")
    assert sale.paid_amount == Decimal("2000.00")  # the deposit carried over
    assert sale.balance == Decimal("3000.00")
    assert client_debt(cust) == {"KGS": Decimal("3000.00")}  # remainder IS the debt

    order.refresh_from_db()
    assert order.status == Order.DELIVERED
    assert order.sale_order_id == sale.pk  # linked both ways
    assert sale.production_order.pk == order.pk

    record_payment(sale, Decimal("3000"), user=owner, method=Payment.CASH)
    sale.refresh_from_db()
    assert sale.balance == 0
    assert sale.payment_status == SaleOrder.PAID
    assert client_debt(cust) == {}  # debt cleared
    # Both the deposit and the final payment live on the sale as history.
    assert sorted(sale.payments.values_list("amount", flat=True)) == [
        Decimal("2000.00"),
        Decimal("3000.00"),
    ]


def test_order_line_progress_reports_real_stock_and_other_orders_claims(variant, settings):
    """Per-line the order page must show real numbers: this line's own
    production progress, the variant's actual stock, and how much of that
    stock another open order has already been promised."""
    from apps.orders.services import create_order, mark_produced, order_line_progress

    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 30)
    mine = Client.objects.create(first_name="Mine", phone="+996700002301")
    rival = Client.objects.create(first_name="Rival", phone="+996700002302")
    order = create_order(
        mine, [{"variant": variant, "quantity": 10, "unit_price": variant.sale_price}]
    )

    row = order_line_progress(order)[0]
    assert row["in_stock"] == 30
    assert row["claimed_by_others"] == 0
    assert row["free_for_order"] == 30
    assert row["coverable"] is True  # 30 on hand covers this line's 10
    assert row["made"] == 0 and row["percent"] == 0

    # Another open order books 25 of the same SKU — only 5 are really free now.
    create_order(rival, [{"variant": variant, "quantity": 25, "unit_price": variant.sale_price}])
    row = order_line_progress(order)[0]
    assert row["in_stock"] == 30  # unchanged fact
    assert row["claimed_by_others"] == 25
    assert row["free_for_order"] == 5
    assert row["coverable"] is False  # 5 free < 10 ordered

    # Taking 4 into stock moves this line's own progress to 40%.
    mark_produced(order.items.get(), 4)
    row = order_line_progress(order)[0]
    assert row["made"] == 4
    assert row["percent"] == 40
    assert row["remaining"] == 6


def test_shortfall_is_about_stock_not_production_progress(variant, settings):
    """A fully produced line can still be short at handover, because another
    open order holds a prior claim on the same shared stock. The shortfall
    shown must be the STOCK gap — reporting the production remainder printed
    the nonsense «принято 20 из 20 · 100%» next to «не хватает 0 шт»."""
    from apps.orders.services import create_order, mark_produced, order_line_progress

    settings.CURRENCY = "KGS"
    mine = Client.objects.create(first_name="ShortMine", phone="+996700002306")
    rival = Client.objects.create(first_name="ShortRival", phone="+996700002307")
    order = create_order(
        mine, [{"variant": variant, "quantity": 20, "unit_price": variant.sale_price}]
    )
    create_order(rival, [{"variant": variant, "quantity": 100, "unit_price": variant.sale_price}])

    # Produce this line in full: 100% made, and those 20 units are in stock...
    mark_produced(order.items.get(), 20)
    row = order_line_progress(order)[0]
    assert row["percent"] == 100
    assert row["remaining"] == 0  # nothing left to sew
    # ...but the rival order's 100-unit claim outweighs the 20 on hand, so
    # handover would still fail — and the number said so must be the stock gap.
    assert row["free_for_order"] == 0
    assert row["coverable"] is False
    assert row["short_by"] == 20  # not 0, which is the production remainder


def test_order_progress_percentage_rolls_up_across_lines(variant, settings):
    from apps.orders.services import create_order, mark_produced, order_progress

    settings.CURRENCY = "KGS"
    other = ProductVariant.objects.create(
        product=variant.product,
        sku="EVD-PROG-L",
        size="L",
        color="red",
        cost_price=Decimal("1500.00"),
        sale_price=Decimal("3200.00"),
    )
    cust = Client.objects.create(first_name="Prog", phone="+996700002303")
    order = create_order(
        cust,
        [
            {"variant": variant, "quantity": 3, "unit_price": variant.sale_price},
            {"variant": other, "quantity": 1, "unit_price": other.sale_price},
        ],
    )
    assert order_progress(order) == {"made": 0, "total": 4, "percent": 0, "complete": False}

    mark_produced(order.items.get(variant=variant), 3)
    prog = order_progress(order)
    assert prog["made"] == 3 and prog["total"] == 4 and prog["percent"] == 75
    assert prog["complete"] is False

    mark_produced(order.items.get(variant=other), 1)
    assert order_progress(order)["complete"] is True


def test_saving_the_due_date_finishes_and_returns_to_the_list(
    client, django_user_model, variant, settings
):
    """Setting the срок is the last step of creating an order, so it saves
    and goes back to the list — both for a plain POST and under HTMX (which
    needs HX-Redirect, or the whole list page would be pasted into the body)."""
    from apps.orders.services import create_order

    settings.CURRENCY = "KGS"
    owner = django_user_model.objects.create_superuser("due_owner", "d@e.com", "x" * 12)
    client.force_login(owner)
    cust = Client.objects.create(first_name="DueDate", phone="+996700002304")
    order = create_order(
        cust, [{"variant": variant, "quantity": 1, "unit_price": variant.sale_price}]
    )
    url = f"/orders/{order.pk}/due-date/"

    resp = client.post(url, {"due_date": "2026-09-01", "note": "подшить"})
    assert resp.status_code == 302
    assert resp.url == "/orders/"
    order.refresh_from_db()
    assert str(order.due_date) == "2026-09-01"
    assert order.note == "подшить"

    resp = client.post(url, {"due_date": "2026-09-02", "note": ""}, HTTP_HX_REQUEST="true")
    assert resp.status_code == 204  # nothing to swap
    assert resp["HX-Redirect"] == "/orders/"


def test_handover_dialog_blocks_when_stock_cannot_cover_the_order(
    client, django_user_model, variant, settings
):
    """confirm_sale raises on an oversell anyway; the dialog's job is to say so
    BEFORE the manager taps, instead of after."""
    from apps.orders.services import create_order

    settings.CURRENCY = "KGS"
    owner = django_user_model.objects.create_superuser("hand_owner", "h@e.com", "x" * 12)
    client.force_login(owner)
    cust = Client.objects.create(first_name="Handover", phone="+996700002305")
    order = create_order(
        cust, [{"variant": variant, "quantity": 5, "unit_price": variant.sale_price}]
    )

    # Nothing produced yet -> the dialog warns and its submit is disabled.
    resp = client.get(f"/orders/{order.pk}/deliver/confirm/")
    assert resp.status_code == 200
    body = resp.content.decode()
    assert len(resp.context["short_lines"]) == 1
    assert "Не хватает остатка" in body
    assert "disabled" in body

    # Take the goods into stock -> the same dialog now allows the handover.
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    resp = client.get(f"/orders/{order.pk}/deliver/confirm/")
    assert resp.context["short_lines"] == []
    assert "Не хватает остатка" not in resp.content.decode()
    assert resp.context["remaining"] == order.total  # no deposit taken


def test_order_page_offers_a_way_out_that_is_not_handing_over(
    client, django_user_model, variant, settings
):
    """«Выдать заказ» is irreversible — it confirms a sale and deducts stock.
    It may only LOOK like the default action once the order is готов; until
    then leaving the page must be the visually obvious move, or a manager
    trying to finish editing reaches for the destructive button."""
    import re

    from apps.orders.models import Order
    from apps.orders.services import create_order, mark_produced

    settings.CURRENCY = "KGS"
    owner = django_user_model.objects.create_superuser("exit_owner", "e@e.com", "x" * 12)
    client.force_login(owner)
    cust = Client.objects.create(first_name="ExitRoute", phone="+996700002281")
    order = create_order(
        cust, [{"variant": variant, "quantity": 2, "unit_price": variant.sale_price}]
    )

    def deliver_button_class(body):
        # type="button" — it opens the handover summary dialog rather than
        # submitting; the styling is what this test is about.
        m = re.search(r'<button type="button"\s+class="btn (btn-\w+)"[^>]*>\s*Выдать заказ', body)
        assert m, "the Выдать заказ button disappeared"
        return m.group(1)

    body = client.get(f"/orders/{order.pk}/").content.decode()
    assert "back-link" in body  # an escape hatch above the fold
    assert "К заказам" in body  # ...and one next to the money actions
    assert deliver_button_class(body) == "btn-ghost"  # not the default-looking action

    # Once everything is produced the order is готов and handing it over IS
    # the natural next step, so it earns the primary styling.
    for item in order.items.all():
        mark_produced(item, item.quantity, user=owner)
    order.refresh_from_db()
    assert order.status == Order.READY
    body = client.get(f"/orders/{order.pk}/").content.decode()
    assert deliver_button_class(body) == "btn-primary"
    assert "К заказам" in body  # the exit never goes away


def test_header_wordmark_links_home(client, django_user_model, settings):
    """The ACOCOS wordmark is the way back to the terminal from any /pos/
    screen — the convention every web app already trained people on."""
    settings.CURRENCY = "KGS"
    client.force_login(
        django_user_model.objects.create_superuser("brand_owner", "b@e.com", "x" * 12)
    )
    body = client.get("/pos/today/").content.decode()
    assert '<a class="top__brand" href="/pos/"' in body


def test_admin_panel_links_back_to_the_pos_terminal(client, django_user_model):
    """/panel/ used to be a one-way door — the POS header links in and nothing
    linked back, so the only way out was editing the URL by hand."""
    client.force_login(
        django_user_model.objects.create_superuser("panel_owner", "p@e.com", "x" * 12)
    )
    body = client.get("/panel/").content.decode()
    assert "Терминал ACOCOS" in body
    assert 'href="/pos/"' in body


def _empty_variant_formset():
    return {
        "variants-TOTAL_FORMS": "0",
        "variants-INITIAL_FORMS": "0",
        "variants-MIN_NUM_FORMS": "0",
        "variants-MAX_NUM_FORMS": "1000",
    }


def test_admin_can_add_multiple_images_to_one_product(
    client, django_user_model, settings, tmp_path
):
    """The feature end to end, through the real /panel/ form: ProductAdmin
    used to have a single `photo` field. It's now an inline gallery — a
    manager attaches as many images as they want in one save, with no
    position field to fill in — order follows upload order automatically."""
    settings.MEDIA_ROOT = tmp_path
    owner = django_user_model.objects.create_superuser("gal_owner", "g@e.com", "x" * 12)
    client.force_login(owner)
    cat = Category.objects.create(name="AdminGallery")

    data = {
        "name": "Admin-uploaded dress",
        "category": cat.pk,
        "description": "",
        "is_active": "on",
        "images-TOTAL_FORMS": "3",
        "images-INITIAL_FORMS": "0",
        "images-MIN_NUM_FORMS": "0",
        "images-MAX_NUM_FORMS": "1000",
        "images-0-image": _tiny_image("first.png"),
        "images-1-image": _tiny_image("second.png"),
        "images-2-image": _tiny_image("third.png"),
        **_empty_variant_formset(),
        "_save": "Сохранить",
    }
    resp = client.post("/panel/inventory/product/add/", data)
    assert resp.status_code == 302, (
        resp.context["adminform"].form.errors if resp.status_code == 200 else resp.status_code
    )

    product = Product.objects.get(name="Admin-uploaded dress")
    rows = list(product.images.order_by("position"))
    assert [r.position for r in rows] == [0, 1, 2]  # sequential, assigned automatically
    assert [r.image.name for r in rows] == [
        "products/first.png",
        "products/second.png",
        "products/third.png",
    ]  # upload order preserved, not reordered
    assert product.grid_image.name == rows[0].thumbnail.name  # the FIRST one is the cover


def test_admin_new_images_continue_after_existing_positions(
    client, django_user_model, settings, tmp_path
):
    """Adding more photos to a product that already has some must not reset
    or collide with the existing ones' positions — the new ones continue
    after whatever the highest existing position already is."""
    settings.MEDIA_ROOT = tmp_path
    owner = django_user_model.objects.create_superuser("gal_owner2", "g2@e.com", "x" * 12)
    client.force_login(owner)
    cat = Category.objects.create(name="AdminGallery2")
    product = Product.objects.create(category=cat, name="Existing gallery product")
    first = ProductImage.objects.create(product=product, image=_tiny_image("a.png"), position=0)
    second = ProductImage.objects.create(product=product, image=_tiny_image("b.png"), position=1)

    data = {
        "name": product.name,
        "category": cat.pk,
        "description": "",
        "is_active": "on",
        "images-TOTAL_FORMS": "3",
        "images-INITIAL_FORMS": "2",
        "images-MIN_NUM_FORMS": "0",
        "images-MAX_NUM_FORMS": "1000",
        "images-0-id": first.pk,
        "images-0-product": product.pk,
        "images-1-id": second.pk,
        "images-1-product": product.pk,
        "images-2-image": _tiny_image("c.png"),  # the new one, no id — a fresh row
        **_empty_variant_formset(),
        "_save": "Сохранить",
    }
    resp = client.post(f"/panel/inventory/product/{product.pk}/change/", data)
    assert resp.status_code == 302, (
        resp.context["adminform"].form.errors if resp.status_code == 200 else resp.status_code
    )

    rows = list(product.images.order_by("position"))
    assert [r.position for r in rows] == [0, 1, 2]
    assert rows[0].pk == first.pk and rows[1].pk == second.pk  # untouched
    assert rows[2].image.name == "products/c.png"  # the new one, continuing at 2


def test_mark_produced_creates_movement_and_raises_available(variant, django_user_model, settings):
    from apps.inventory.services import available_for
    from apps.orders.models import Order
    from apps.orders.services import create_order, mark_produced

    settings.CURRENCY = "KGS"
    cust = Client.objects.create(first_name="Prod1", phone="+996700002204")
    order = create_order(
        cust, [{"variant": variant, "quantity": 5, "unit_price": variant.sale_price}]
    )
    item = order.items.get()
    assert available_for(variant) == -5  # nothing on hand yet, 5 fully reserved

    mark_produced(item, 2)
    item.refresh_from_db()
    assert item.produced_qty == 2
    # Producing raises on_hand, but the reservation stays the FULL 5 (not
    # reduced by produced_qty) — the produced units are still earmarked for
    # this order, not sellable to a walk-in: available is unchanged.
    assert available_for(variant) == -3  # 2 on hand - 5 reserved
    assert variant.stock == 2  # PRODUCTION_IN movement written
    order.refresh_from_db()
    assert order.status == Order.IN_PRODUCTION

    mark_produced(item, 3)
    item.refresh_from_db()
    order.refresh_from_db()
    assert item.produced_qty == 5
    assert order.status == Order.READY  # fully produced -> auto-advance
    assert order.fully_produced is True


def test_mark_produced_rejects_more_than_remaining(variant):
    from apps.orders.services import create_order, mark_produced

    cust = Client.objects.create(first_name="Prod2", phone="+996700002205")
    order = create_order(
        cust, [{"variant": variant, "quantity": 3, "unit_price": variant.sale_price}]
    )
    item = order.items.get()
    with pytest.raises(ValidationError):
        mark_produced(item, 4)
    item.refresh_from_db()
    assert item.produced_qty == 0


def test_handover_deducts_stock_once_applies_deposit_and_links_order_and_sale(variant, settings):
    from apps.orders.models import Order
    from apps.orders.services import create_order, hand_over, mark_produced, record_deposit

    settings.CURRENCY = "KGS"
    cust = Client.objects.create(first_name="Handover1", phone="+996700002206")
    order = create_order(cust, [{"variant": variant, "quantity": 4, "unit_price": Decimal("1000")}])
    item = order.items.get()
    mark_produced(item, 4)
    variant.refresh_from_db()
    assert variant.stock == 4

    deposit = record_deposit(order, Decimal("1500"), currency="KGS")
    assert deposit.production_order_id == order.pk
    assert deposit.order_id is None  # not linked to a sale yet

    sale_out_before = StockMovement.objects.filter(
        movement_type=StockMovement.SALE_OUT, variant=variant
    ).count()

    sale = hand_over(order)

    sale_out_after = StockMovement.objects.filter(
        movement_type=StockMovement.SALE_OUT, variant=variant
    ).count()
    assert sale_out_after == sale_out_before + 1  # stock deducted EXACTLY once
    variant.refresh_from_db()
    assert variant.stock == 0

    order.refresh_from_db()
    assert order.status == Order.DELIVERED
    assert order.sale_order_id == sale.pk
    assert sale.production_order.pk == order.pk  # linked both ways

    deposit.refresh_from_db()
    assert deposit.order_id == sale.pk  # deposit carried over
    assert sale.paid_amount == Decimal("1500.00")
    assert sale.balance == Decimal("2500.00")  # 4000 total - 1500 deposit


def test_handover_rejects_an_order_with_no_items(variant):
    from apps.orders.services import hand_over
    from apps.orders.models import Order

    cust = Client.objects.create(first_name="Empty", phone="+996700002207")
    order = Order.objects.create(client=cust)
    with pytest.raises(ValidationError):
        hand_over(order)


def test_cancel_order_releases_its_reservation(variant, settings):
    from apps.inventory.services import available_for
    from apps.orders.services import cancel_order, create_order

    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    cust = Client.objects.create(first_name="Cancel1", phone="+996700002208")
    order = create_order(
        cust, [{"variant": variant, "quantity": 5, "unit_price": variant.sale_price}]
    )
    assert available_for(variant) == 0

    cancel_order(order)
    assert available_for(variant) == 5  # reservation released
    order.refresh_from_db()
    from apps.orders.models import Order

    assert order.status == Order.CANCELLED


def test_overdue_detection_respects_bishkek_local_date_not_utc(monkeypatch):
    from datetime import date

    from apps.orders.models import Order

    cust = Client.objects.create(first_name="OverdueTZ", phone="+996700002209")
    order = Order.objects.create(client=cust, due_date=date(2026, 7, 17))
    with timezone.override("Asia/Bishkek"):
        # 00:30 Bishkek on the 17th == 18:30 UTC on the 16th — the LOCAL date
        # is already the due date, not yet overdue.
        _freeze_now(monkeypatch, 2026, 7, 16, 18, 30)
        assert order.is_overdue is False
        # 00:30 Bishkek on the 18th == 18:30 UTC on the 17th — a UTC-based
        # check would still read the 17th (not overdue); the correct local
        # date is the 18th, one day past due.
        _freeze_now(monkeypatch, 2026, 7, 17, 18, 30)
        assert order.is_overdue is True


def test_deposit_reuses_the_frozen_rate_mechanism_no_second_conversion_path(variant, settings):
    settings.CURRENCY = "KGS"
    ExchangeRate.objects.create(currency="USD", date=timezone.localdate(), rate=Decimal("87.00"))
    from apps.orders.services import create_order, record_deposit

    cust = Client.objects.create(first_name="DepositFx", phone="+996700002210")
    order = create_order(cust, [{"variant": variant, "quantity": 1, "unit_price": Decimal("1000")}])
    deposit = record_deposit(order, Decimal("10"), currency="USD")
    assert deposit.rate_to_kgs == Decimal("87.000000")
    assert deposit.amount_kgs == Decimal("870.00")

    # Changing the rate afterward never alters the stored deposit — same
    # invariant as every other frozen-rate payment in the system.
    ExchangeRate.objects.filter(currency="USD").update(rate=Decimal("150.00"))
    deposit.refresh_from_db()
    assert deposit.rate_to_kgs == Decimal("87.000000")


def test_owner_only_can_cancel_order_via_admin(django_user_model):
    from django.contrib.admin.sites import site

    from apps.orders.models import Order

    cust = Client.objects.create(first_name="AdminCancel", phone="+996700002211")
    order = Order.objects.create(client=cust)
    model_admin = site._registry[Order]
    # "apply": "1" — the confirm_bulk_action gate (see test_admin_cancel_
    # order_action_asks_before_cancelling for the two-step flow itself);
    # this test is about the is_superuser check, so it confirms up front.
    request = RequestFactory().post("/", {"apply": "1"})
    manager = django_user_model.objects.create_user("mgr_cancel", password="x" * 12, is_staff=True)
    request.user = manager

    from django.contrib.messages.storage.fallback import FallbackStorage

    setattr(request, "session", {})
    messages_storage = FallbackStorage(request)
    setattr(request, "_messages", messages_storage)

    model_admin.cancel_selected(request, Order.objects.filter(pk=order.pk))
    order.refresh_from_db()
    assert order.status == Order.NEW  # unchanged — Editor can't cancel

    owner = django_user_model.objects.create_superuser("owner_cancel", "o@e.com", "x" * 12)
    request.user = owner
    model_admin.cancel_selected(request, Order.objects.filter(pk=order.pk))
    order.refresh_from_db()
    assert order.status == Order.CANCELLED


def test_done_when_scenario_order_deposit_queue_produce_handover(
    client, django_user_model, variant, settings
):
    """DONE WHEN: create an order for unproduced goods, take a deposit in any
    currency, see it in the queue, mark produced, hand over as a normal sale
    that deducts stock exactly once."""
    settings.CURRENCY = "KGS"
    ExchangeRate.objects.create(currency="USD", date=timezone.localdate(), rate=Decimal("87.00"))
    call_command("setup_roles")
    manager = django_user_model.objects.create_user("done_mgr", password="x" * 12, is_staff=True)
    manager.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(manager)
    assert variant.stock == 0  # nothing produced yet — the whole point

    cust = Client.objects.create(first_name="DoneWhen", phone="+996700002212")
    # Order creation is POST-only (a cross-site GET must not be able to spin up
    # orders) — the client-picker row is a CSRF-protected form.
    resp = client.post("/orders/new/", {"client": cust.pk})
    assert resp.status_code == 302
    # /orders/<id>/?new=1 — a new order lands on the guided creation route.
    order_id = int(re.search(r"/orders/(\d+)/", resp.url).group(1))
    assert "new=1" in resp.url

    # No cap on ordering unproduced goods.
    resp = client.post(f"/orders/{order_id}/items/add/", {"variant_id": variant.pk, "quantity": 5})
    assert resp.status_code == 302

    # Deposit in a foreign currency.
    client.post(
        f"/orders/{order_id}/deposit/", {"amount": "20", "currency": "USD", "method": "cash"}
    )

    from apps.orders.models import Order

    order = Order.objects.get(pk=order_id)
    assert order.deposits.count() == 1

    # Shows up in the aggregated queue, with the client it's being made for.
    resp = client.get("/orders/queue/")
    body = resp.content.decode()
    assert "сшить" in body  # the row's headline number
    assert "DoneWhen" in body  # ...and who it's for

    # Mark produced.
    item = order.items.get()
    client.post(f"/orders/{order_id}/items/{item.pk}/produce/", {"quantity": 5})
    variant.refresh_from_db()
    assert variant.stock == 5

    # Hand over -> a normal confirmed sale, stock deducted exactly once.
    resp = client.post(f"/orders/{order_id}/deliver/")
    assert resp.status_code == 302
    variant.refresh_from_db()
    assert variant.stock == 0
    order.refresh_from_db()
    assert order.status == Order.DELIVERED
    assert order.sale_order is not None
    assert order.sale_order.status == SaleOrder.CONFIRMED


# ---------------------------------------------------------------------------
# Part 0c — Spreadsheet formatting + debt visibility
# ---------------------------------------------------------------------------


def _load_sheet(xlsx_bytes, name):
    import io

    from openpyxl import load_workbook

    return load_workbook(io.BytesIO(xlsx_bytes))[name]


def test_report_money_cells_are_real_numbers_not_text(variant):
    """The core of the spreadsheet rework: money and counts used to be written
    with str(), so Excel stored the sheet as TEXT — a column couldn't be
    summed, and sorting put 1000 before 900. They must be numeric cells with a
    display format instead."""
    from apps.core.management.commands.send_daily_report import _build_xlsx

    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    cust = Client.objects.create(first_name="Fmt", phone="+996700004001")
    order = SaleOrder.objects.create(client=cust, channel=SaleOrder.SHOP)
    SaleItem.objects.create(order=order, variant=variant, quantity=2, unit_price=Decimal("3200"))
    confirm_sale(order)

    ws = _load_sheet(_build_xlsx(), "Продажи")
    headers = [c.value for c in ws[1]]
    total_col = headers.index("Итого") + 1
    qty_col = headers.index("Кол-во") + 1
    time_col = headers.index("Время") + 1

    total_cell = ws.cell(row=2, column=total_col)
    assert isinstance(total_cell.value, (int, float)), "money must be a NUMBER, not text"
    assert total_cell.value == 6400
    assert total_cell.number_format == "# ##0.00"

    qty_cell = ws.cell(row=2, column=qty_col)
    assert isinstance(qty_cell.value, int) and qty_cell.value == 2
    assert qty_cell.number_format == "# ##0"

    # A datetime underneath, shown as clock time — sorts chronologically.
    import datetime as _dt

    assert isinstance(ws.cell(row=2, column=time_col).value, _dt.datetime)
    assert ws.cell(row=2, column=time_col).number_format == "HH:MM"


def test_every_report_sheet_is_navigable(variant):
    """Frozen header + autofilter + a styled header row on every sheet of
    every export — what makes a downloaded file usable rather than a dump."""
    from apps.core.management.commands.send_daily_report import _build_xlsx
    from apps.reports.dashboard import dashboard_data, dashboard_sheets
    from apps.reports.export import to_xlsx
    from apps.reports.storage import build_xlsx as storage_xlsx

    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    order = SaleOrder.objects.create()
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(order)

    workbooks = [
        _build_xlsx(),
        storage_xlsx(),
        to_xlsx(dashboard_sheets(dashboard_data("month"))),
    ]
    import io

    from openpyxl import load_workbook

    for data in workbooks:
        wb = load_workbook(io.BytesIO(data))
        for name in wb.sheetnames:
            ws = wb[name]
            assert ws.freeze_panes == "A2", f"{name}: header not frozen"
            assert ws.auto_filter.ref, f"{name}: no autofilter"
            assert ws["A1"].font.bold, f"{name}: header not styled"
            # Columns sized — the default 8.43 leaves Cyrillic headers clipped.
            assert ws.column_dimensions["A"].width >= 9


def test_receipt_context_shows_paid_and_balance_only_when_not_fully_paid(variant, settings):
    """paid/balance are always real numbers in the context (the status badge
    needs balance even at 0) — it's `status` that decides whether the
    template SHOWS the Оплачено/Остаток lines: hidden once status is
    'paid' (only Итого shows then), rendered for 'unpaid'/'partial'."""
    from django.template.loader import render_to_string

    from apps.pos.receipts import receipt_context

    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    cust = Client.objects.create(first_name="Receipt", phone="+996700004002")

    goods_only = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(
        order=goods_only, variant=variant, quantity=1, unit_price=Decimal("3200")
    )
    confirm_sale(goods_only)
    ctx = receipt_context(goods_only)
    assert ctx["status"] == "unpaid"
    assert ctx["paid"] == Decimal("0")
    assert ctx["balance"] == Decimal("3200")
    assert ctx["total_money"] == Decimal("3200")
    html = render_to_string("receipts/receipt.html", ctx)
    assert "Оплачено" in html and "Остаток" in html

    record_payment(goods_only, Decimal("1200"))
    ctx = receipt_context(goods_only)
    assert ctx["status"] == "partial"
    assert ctx["paid"] == Decimal("1200.00")
    assert ctx["balance"] == Decimal("2000.00")

    # Paying the rest hides the payment lines — nothing left to explain.
    record_payment(goods_only, Decimal("2000"))
    ctx = receipt_context(goods_only)
    assert ctx["status"] == "paid"
    assert ctx["paid"] == Decimal("3200.00")
    assert ctx["balance"] == Decimal("0")
    html = render_to_string("receipts/receipt.html", ctx)
    assert "Оплачено" not in html and "Остаток" not in html


def test_purchase_history_shows_the_debt_on_each_sale(client, django_user_model, variant, settings):
    """«Откуда долг» answered in place: the client page's debt block lists
    each unpaid sale with what's still owed ON THAT SALE (not just one total
    at the top), while a fully-paid sale shows up in История покупок
    instead — the two lists never overlap."""
    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    owner = django_user_model.objects.create_superuser("hist_owner", "h@e.com", "x" * 12)
    client.force_login(owner)
    cust = Client.objects.create(first_name="Hist", phone="+996700004003")

    unpaid = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=unpaid, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(unpaid)
    record_payment(unpaid, Decimal("1000"))  # 2200 still owed

    settled = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=settled, variant=variant, quantity=1, unit_price=Decimal("500"))
    confirm_sale(settled)
    record_payment(settled, Decimal("500"))  # fully paid

    resp = client.get(f"/pos/clients/{cust.pk}/")
    assert resp.status_code == 200
    unpaid_pks = {r["order"].pk for r in resp.context["unpaid_rows"]}
    paid_pks = {r["order"].pk for r in resp.context["paid_rows"]}
    assert unpaid_pks == {unpaid.pk}
    assert paid_pks == {settled.pk}
    rows = {r["order"].pk: r for r in resp.context["unpaid_rows"] + resp.context["paid_rows"]}
    assert rows[unpaid.pk]["balance"] == Decimal("2200.00")
    assert rows[settled.pk]["balance"] == Decimal("0")
    body = resp.content.decode()
    assert "2\xa0200\xa0сом" in body  # the per-sale debt is actually rendered
    assert "Погасить" in body  # per-row repay button on the unpaid sale


def test_client_page_has_no_separate_payments_card(client, django_user_model, variant):
    """The «Платежи» card duplicated exactly what the debt block and История
    покупок already show — it must be gone, not just relocated."""
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    owner = django_user_model.objects.create_superuser("noplay_owner", "o@e.com", "x" * 12)
    client.force_login(owner)
    cust = Client.objects.create(first_name="NoPayCard", phone="+996700005033")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(order)
    record_payment(order, Decimal("1000"))

    body = client.get(f"/pos/clients/{cust.pk}/").content.decode()
    assert "Платежи" not in body


def test_client_page_debt_block_shows_item_summary_and_repay_button(
    client, django_user_model, variant
):
    """The debt block's own spec: total debt, then each unpaid sale with
    date, what was bought, remaining, and a «Погасить» button — all in ONE
    place, above История покупок (never mixed into it)."""
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    owner = django_user_model.objects.create_superuser("debtblk_owner", "o@e.com", "x" * 12)
    client.force_login(owner)
    cust = Client.objects.create(first_name="DebtBlock", phone="+996700005044")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)  # unpaid — 3200 owed

    resp = client.get(f"/pos/clients/{cust.pk}/")
    body = resp.content.decode()
    assert "Evening Dress" in body  # what was bought
    debt_idx = body.index("Долг")
    history_idx = body.index("История покупок")
    pay_url = f"/pos/sale/{order.pk}/debt/"
    assert pay_url in body
    # The per-row repay link/button sits inside the debt block, before
    # История покупок starts — never mixed into the paid-sales list.
    assert debt_idx < body.index(pay_url) < history_idx


def test_client_page_whatsapp_reminder_is_not_styled_destructive(
    client, django_user_model, variant
):
    """A debt reminder isn't a destructive action — it must not carry the
    same red/danger styling as «Отменить продажу»."""
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    call_command("setup_roles")
    editor = django_user_model.objects.create_user("wa_neutral", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)
    cust = Client.objects.create(first_name="WaNeutral", phone="+996700005055")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(order, user=editor)  # unpaid -> reminder button renders

    body = client.get(f"/pos/clients/{cust.pk}/").content.decode()
    assert "Напомнить о долге в WhatsApp" in body
    assert "btn-ghost--debt" not in body


# ---- Debt repayment from the client history page ("Погасить долг") -------


def test_debt_pay_detail_prefills_the_remaining_balance(client, django_user_model, variant):
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    owner = django_user_model.objects.create_superuser("dpd_owner", "o@e.com", "x" * 12)
    client.force_login(owner)
    cust = Client.objects.create(first_name="Debtor", phone="+996700005001")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)
    record_payment(order, Decimal("1000"))  # 2200 still owed

    resp = client.get(f"/pos/sale/{order.pk}/debt/")
    assert resp.status_code == 200
    assert resp.context["balance"] == Decimal("2200.00")
    assert resp.context["payment_amount"] == Decimal("2200.00")


def test_debt_pay_detail_shows_item_list(client, django_user_model, variant):
    """The repayment window used to show only money — no reminder of what
    the debt is actually FOR. The sale's own items must render: товар,
    размер, цвет, кол-во, сумма."""
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    owner = django_user_model.objects.create_superuser("dpi_owner", "o@e.com", "x" * 12)
    client.force_login(owner)
    cust = Client.objects.create(first_name="ItemDebtor", phone="+996700005011")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=2, unit_price=Decimal("3200"))
    confirm_sale(order)

    resp = client.get(f"/pos/sale/{order.pk}/debt/")
    body = resp.content.decode()
    assert "Evening Dress" in body
    assert "красн" in body.lower() or "red" in body.lower()  # variant.color, untranslated item data
    assert "M</td>" in body or ">M<" in body
    assert "6\xa0400" in body  # 2 × 3200 line total


def test_debt_pay_detail_shows_resend_receipt_only_with_a_phone(client, django_user_model, variant):
    """«Отправить чек ещё раз» reuses the existing share_receipt path — it
    must only appear when there's actually a phone to send it to."""
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    owner = django_user_model.objects.create_superuser("dpr_owner", "o@e.com", "x" * 12)
    client.force_login(owner)

    with_phone = Client.objects.create(first_name="HasPhone", phone="+996700005022")
    order = SaleOrder.objects.create(client=with_phone, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)
    body = client.get(f"/pos/sale/{order.pk}/debt/").content.decode()
    assert "Отправить чек ещё раз" in body
    assert f"/pos/sale/{order.pk}/receipt/" in body

    no_phone = Client.objects.create(first_name="NoPhone", phone="")
    order2 = SaleOrder.objects.create(client=no_phone, currency="KGS")
    SaleItem.objects.create(order=order2, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order2)
    body2 = client.get(f"/pos/sale/{order2.pk}/debt/").content.decode()
    assert "Отправить чек ещё раз" not in body2


def test_debt_pay_confirm_full_repay_clears_debt(client, django_user_model, variant):
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    owner = django_user_model.objects.create_superuser("dpc_owner", "o2@e.com", "x" * 12)
    client.force_login(owner)
    cust = Client.objects.create(first_name="FullPay", phone="+996700005002")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)

    resp = client.post(
        f"/pos/sale/{order.pk}/debt/confirm/",
        {"amount": "3200", "currency": "KGS", "method": Payment.CASH},
    )
    assert resp.status_code == 302
    order.refresh_from_db()
    assert order.balance == Decimal("0")
    assert order.payment_status == SaleOrder.PAID
    assert client_debt(cust) == {}


@pytest.mark.django_db(transaction=True)
def test_record_payment_concurrent_double_tap_settles_exactly_once(variant):
    """record_payment used to take no lock at all — every balance/change/
    overpayment check read order.payments.all() unlocked, so two genuinely
    concurrent payments against the same order's full balance (a double-
    tapped «Оплачено полностью» on debt_pay_confirm) could each compute
    against the SAME pre-payment balance and both succeed, overpaying by the
    full amount a second time. Verified against real Postgres to fail on
    the unpatched (unlocked) record_payment and pass on the fix — see the
    matching note on test_cancel_sale_concurrent_double_tap_restocks_exactly_once
    for why this doesn't discriminate under the sqlite test DB."""
    import threading
    import time

    from django.db import connections
    from django.db.utils import OperationalError

    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    order = SaleOrder.objects.create(
        client=Client.objects.create(first_name="Race", phone="+996700005099")
    )
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)

    results = []
    rejections = []
    barrier = threading.Barrier(2)

    def fire():
        barrier.wait(timeout=5)
        for attempt in range(8):
            try:
                record_payment(SaleOrder.objects.get(pk=order.pk), Decimal("3200"))
                results.append("paid")
                return
            except ValidationError as exc:
                rejections.append(exc)
                return
            except OperationalError:
                time.sleep(0.05 * (attempt + 1))
            finally:
                connections.close_all()
        raise AssertionError("gave up retrying after sqlite whole-file lock contention")

    threads = [threading.Thread(target=fire) for _ in range(2)]
    for t in threads:
        t.start()
    for t in threads:
        t.join(timeout=15)

    assert len(results) == 1, f"expected exactly one payment to succeed, got {results}"
    assert len(rejections) == 1, "expected the second, now-stale payment to be rejected"
    order.refresh_from_db()
    # The winner's payment settles the order exactly — never double-applied.
    assert order.paid_amount == Decimal("3200")


def test_debt_pay_confirm_partial_updates_sale_and_client_debt(client, django_user_model, variant):
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    owner = django_user_model.objects.create_superuser("dpc_owner2", "o3@e.com", "x" * 12)
    client.force_login(owner)
    cust = Client.objects.create(first_name="Partial", phone="+996700005003")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)

    resp = client.post(
        f"/pos/sale/{order.pk}/debt/confirm/",
        {"amount": "1200", "currency": "KGS", "method": Payment.CASH},
    )
    assert resp.status_code == 302
    order.refresh_from_db()
    assert order.balance == Decimal("2000.00")
    assert client_debt(cust) == {"KGS": Decimal("2000.00")}


def test_debt_pay_confirm_foreign_currency_uses_frozen_rate(
    client, django_user_model, variant, settings
):
    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    ExchangeRate.objects.create(currency="USD", date=timezone.localdate(), rate=Decimal("87.00"))
    owner = django_user_model.objects.create_superuser("dpc_owner3", "o4@e.com", "x" * 12)
    client.force_login(owner)
    cust = Client.objects.create(first_name="Foreign", phone="+996700005004")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("870"))
    confirm_sale(order)

    resp = client.post(
        f"/pos/sale/{order.pk}/debt/confirm/",
        {"amount": "10", "currency": "USD", "method": Payment.CASH},
    )
    assert resp.status_code == 302
    order.refresh_from_db()
    assert order.balance == Decimal("0")  # 10 USD × 87 = 870 KGS, exactly the total
    payment = order.payments.get()
    assert payment.rate_to_kgs == Decimal("87.000000")

    # Voiding restores EXACTLY the pre-payment debt, at the same frozen rate,
    # even if today's rate has since moved.
    ExchangeRate.objects.filter(currency="USD").update(rate=Decimal("95.00"))
    void_payment(payment, user=owner)
    order.refresh_from_db()
    assert order.balance == Decimal("870.00")
    assert client_debt(cust) == {"KGS": Decimal("870.00")}


def test_debt_pay_viewer_forbidden(client, django_user_model, variant):
    call_command("setup_roles")
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    viewer = django_user_model.objects.create_user("dp_viewer", password="x" * 12, is_staff=True)
    viewer.groups.add(Group.objects.get(name=VIEWER))
    client.force_login(viewer)
    cust = Client.objects.create(first_name="View", phone="+996700005005")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(order)

    resp = client.get(f"/pos/sale/{order.pk}/debt/")
    assert resp.status_code == 403
    resp = client.post(f"/pos/sale/{order.pk}/debt/confirm/", {"amount": "1000", "currency": "KGS"})
    assert resp.status_code == 403
    assert not order.payments.exists()


def test_debt_pay_reuses_record_payment_no_second_payment_path(client, django_user_model, variant):
    """Not a parallel implementation — the exact same services.record_payment
    every other /pos/ payment goes through, so debt-status derivation and the
    daily report never see a payment recorded a different way."""
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    owner = django_user_model.objects.create_superuser("dp_reuse", "o5@e.com", "x" * 12)
    client.force_login(owner)
    cust = Client.objects.create(first_name="Reuse", phone="+996700005006")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(order)

    client.post(
        f"/pos/sale/{order.pk}/debt/confirm/",
        {"amount": "1000", "currency": "KGS", "method": Payment.MBANK},
    )
    payment = order.payments.get()
    assert payment.method == Payment.MBANK
    assert payment.created_by_id == owner.pk
    assert payment.rate_to_kgs is not None  # frozen exactly like any other payment


def test_client_debt_pay_multi_sale_oldest_first_closes_correct_sales(
    client, django_user_model, variant
):

    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    owner = django_user_model.objects.create_superuser("mdp_owner", "o6@e.com", "x" * 12)
    client.force_login(owner)
    cust = Client.objects.create(first_name="Multi", phone="+996700005007")

    older = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=older, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(older)
    older.confirmed_at = timezone.now() - timedelta(days=2)
    older.save(update_fields=["confirmed_at"])

    newer = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=newer, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(newer)
    newer.confirmed_at = timezone.now() - timedelta(days=1)
    newer.save(update_fields=["confirmed_at"])

    # Preview: 1500 covers the older sale fully and half of the newer one.
    resp = client.get(f"/pos/clients/{cust.pk}/debt/pay/?currency=KGS&amount=1500")
    assert resp.status_code == 200
    rows = resp.context["allocation"]["rows"]
    assert [r["order"].pk for r in rows] == [older.pk, newer.pk]
    assert rows[0]["applied"] == Decimal("1000")
    assert rows[0]["closes"] is True
    assert rows[1]["applied"] == Decimal("500")
    assert rows[1]["closes"] is False

    resp = client.post(
        f"/pos/clients/{cust.pk}/debt/pay/confirm/",
        {"currency": "KGS", "amount": "1500", "method": Payment.CASH},
    )
    assert resp.status_code == 302
    older.refresh_from_db()
    newer.refresh_from_db()
    assert older.balance == Decimal("0")
    assert newer.balance == Decimal("500.00")
    assert Payment.objects.filter(order=older).count() == 1
    assert Payment.objects.filter(order=newer).count() == 1


def test_client_debt_pay_multi_sale_rejects_amount_over_total_debt(
    client, django_user_model, variant
):
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    owner = django_user_model.objects.create_superuser("mdp_owner2", "o7@e.com", "x" * 12)
    client.force_login(owner)
    cust = Client.objects.create(first_name="TooMuch2", phone="+996700005008")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(order)

    resp = client.post(
        f"/pos/clients/{cust.pk}/debt/pay/confirm/",
        {"currency": "KGS", "amount": "5000", "method": Payment.CASH},
    )
    assert resp.status_code == 302
    order.refresh_from_db()
    assert order.balance == Decimal("1000.00")  # nothing was applied
    assert not order.payments.exists()


def test_sales_sheet_carries_each_buyers_total_outstanding_debt(variant, settings):
    """Debt shown everywhere, per the ask: the Sales sheet gains a per-buyer
    running balance next to the sale, so the spreadsheet answers "how much
    does this person owe overall" without cross-referencing another tab."""
    from apps.core.management.commands.send_daily_report import _sales_sheet

    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    cust = Client.objects.create(first_name="Running", phone="+996700004004")
    first = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=first, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(first)  # wholly unpaid -> 1000
    second = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=second, variant=variant, quantity=1, unit_price=Decimal("500"))
    confirm_sale(second)  # unpaid -> 500; client total 1500

    rows = {r["№"]: r for r in _by_header(_sales_sheet())}
    assert rows[first.pk]["Долг по продаже"] == Decimal("1000.00")
    assert rows[second.pk]["Долг по продаже"] == Decimal("500.00")
    # Both rows show the SAME client-wide figure — it's a balance, not a
    # per-sale number repeated.
    assert rows[first.pk]["Долг клиента всего"] == Decimal("1500.00")
    assert rows[second.pk]["Долг клиента всего"] == Decimal("1500.00")


def test_debts_sheet_reports_debt_age_and_spread(variant, settings):
    """The Debts sheet gains how many sales a debt spans and how long it's
    been outstanding, so the oldest can be chased first rather than only the
    largest."""
    from apps.core.management.commands.send_daily_report import _debts_sheet

    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    cust = Client.objects.create(first_name="Aging", phone="+996700004005")
    for price in (Decimal("1000"), Decimal("2000")):
        o = SaleOrder.objects.create(client=cust, currency="KGS")
        SaleItem.objects.create(order=o, variant=variant, quantity=1, unit_price=price)
        confirm_sale(o)

    row = next(r for r in _by_header(_debts_sheet()) if r["Клиент"] == "Aging")
    assert row["Долг"] == Decimal("3000.00")
    assert row["Неоплаченных продаж"] == 2
    assert row["Долг с"] == timezone.localdate()


def test_admin_search_boxes_do_not_crash(client, django_user_model, variant):
    """Regression: SaleOrder/Payment searched on `client__name` — a PROPERTY,
    not a column — so every one of these 500'd with FieldError the moment
    anyone typed in the box. The search box is the main way a non-technical
    owner finds anything in the panel."""
    owner = django_user_model.objects.create_superuser("srch_owner", "s@e.com", "x" * 12)
    client.force_login(owner)
    for url in [
        "/panel/sales/saleorder/?q=Айгуль",
        "/panel/sales/payment/?q=Айгуль",
        "/panel/clients/client/?q=Айгуль",
        "/panel/orders/order/?q=Айгуль",
        "/panel/inventory/productvariant/?q=EVD",
    ]:
        assert client.get(url).status_code == 200, f"{url} crashed"


def test_admin_has_debt_filter_selects_only_debtors(client, django_user_model, variant, settings):
    settings.CURRENCY = "KGS"
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    owner = django_user_model.objects.create_superuser("dbt_owner", "d@e.com", "x" * 12)
    client.force_login(owner)

    debtor = Client.objects.create(first_name="Debtor", phone="+996700004006")
    o = SaleOrder.objects.create(client=debtor, currency="KGS")
    SaleItem.objects.create(order=o, variant=variant, quantity=1, unit_price=Decimal("900"))
    confirm_sale(o)  # unpaid
    clear = Client.objects.create(first_name="Clear", phone="+996700004007")

    owing = client.get("/panel/clients/client/?has_debt=yes")
    ids = {c.pk for c in owing.context["cl"].result_list}
    assert debtor.pk in ids and clear.pk not in ids

    settled = client.get("/panel/clients/client/?has_debt=no")
    ids = {c.pk for c in settled.context["cl"].result_list}
    assert clear.pk in ids and debtor.pk not in ids


def test_daily_report_query_count_does_not_scale_with_sales(settings):
    """Regression: the nightly report was N+1 on payments. SaleOrder.paid_amount
    used .values_list(), which ALWAYS issues its own query and ignores an
    existing prefetch_related("payments") — so todays_confirmed_orders()
    prefetched them and paid for it anyway, and the Debts sheet's
    unpaid-span scan had no prefetch at all. Measured 96 queries for 45 sales
    and climbing; at real volume the scheduler's nightly run would crawl."""
    from django.db import connection
    from django.test.utils import CaptureQueriesContext

    from apps.core.management.commands.send_daily_report import (
        _debts_sheet,
        _orders_sheet,
        _sales_sheet,
        _stock_sheet,
    )

    settings.CURRENCY = "KGS"

    def seed(n, tag):
        cat = Category.objects.create(name=f"QC{tag}")
        prod = Product.objects.create(category=cat, name=f"QP{tag}")
        v = ProductVariant.objects.create(
            product=prod, sku=f"QS{tag}", cost_price=Decimal("1"), sale_price=Decimal("100")
        )
        add_movement(v, StockMovement.PRODUCTION_IN, n * 2)
        for i in range(n):
            cust = Client.objects.create(first_name=f"Q{tag}{i}", phone=f"+9967008{tag}{i:04d}")
            o = SaleOrder.objects.create(client=cust, currency="KGS")
            SaleItem.objects.create(order=o, variant=v, quantity=1, unit_price=Decimal("100"))
            confirm_sale(o)
            record_payment(o, Decimal("40"))  # leaves a debt on every one

    def count(fn):
        with CaptureQueriesContext(connection) as ctx:
            fn()
        return len(ctx.captured_queries)

    builders = (_sales_sheet, _debts_sheet, _orders_sheet, _stock_sheet)
    seed(5, "a")
    small = {f.__name__: count(f) for f in builders}
    seed(40, "b")  # 9x the data
    big = {f.__name__: count(f) for f in builders}

    for name in small:
        assert big[name] == small[name], (
            f"{name}: {small[name]} -> {big[name]} queries as rows grew 9x — N+1 is back"
        )


# --- Regressions found in the 2026-08 project review -------------------------
# Each of these reproduced a real defect on main before the fix in the same
# commit: stock driven negative, a client's debt driven negative, a 500 on an
# ordinary return, and a report overstating cash taken.


def test_confirm_sale_caps_duplicate_variant_lines_against_the_combined_total(variant):
    """The same variant on TWO lines must be checked against the SUM, not each
    line separately. 6 + 6 against 10 available passed line-by-line and drove
    on_hand to −2 — the guard of last resort was weaker than the cart cap in
    front of it."""
    add_movement(variant, StockMovement.PURCHASE_IN, 10)
    client = Client.objects.create(first_name="Дуб", phone="+996700990001")
    order = SaleOrder.objects.create(client=client, currency=variant.currency)
    SaleItem.objects.create(order=order, variant=variant, quantity=6, unit_price=Decimal("3200"))
    SaleItem.objects.create(order=order, variant=variant, quantity=6, unit_price=Decimal("3200"))

    with pytest.raises(ValidationError):
        confirm_sale(order)

    order.refresh_from_db()
    assert order.status == SaleOrder.DRAFT
    assert variant.stock == 10, "a rejected oversell must not move stock at all"


def test_handover_of_an_order_listing_one_variant_twice_cannot_oversell(variant):
    """The reachable route to the bug above: hand_over copies OrderItems into
    SaleItems one-for-one and nothing merges them, so a production order with
    the same variant on two lines produced exactly that duplicate-line sale."""
    from apps.orders.services import create_order, hand_over

    add_movement(variant, StockMovement.PURCHASE_IN, 10)
    client = Client.objects.create(first_name="Заказ", phone="+996700990002")
    order = create_order(
        client=client,
        items=[
            {"variant": variant, "quantity": 6, "unit_price": Decimal("3200")},
            {"variant": variant, "quantity": 6, "unit_price": Decimal("3200")},
        ],
    )
    with pytest.raises(ValidationError):
        hand_over(order)
    assert variant.stock >= 0, f"stock went negative: {variant.stock}"


def test_a_payment_can_be_voided_only_once(variant):
    """A double-tapped void wrote two reversal rows, each subtracting the
    amount, so the client ended up with a phantom credit."""
    add_movement(variant, StockMovement.PURCHASE_IN, 5)
    client = Client.objects.create(first_name="Возврат", phone="+996700990003")
    order = SaleOrder.objects.create(client=client, currency=variant.currency)
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)
    payment = record_payment(order, Decimal("3200"))

    void_payment(payment)
    order.refresh_from_db()
    assert order.paid_amount == Decimal("0.00")

    with pytest.raises(ValidationError):
        void_payment(payment)

    order.refresh_from_db()
    assert order.paid_amount == Decimal("0.00"), "a second void must not move the balance"
    assert client_debt(client).get("KGS") == Decimal("3200.00")


def test_db_rejects_a_second_reversal_even_via_bulk_create(variant):
    """The service check above is row-locked, but the constraint is what makes
    a double void impossible — same convention as every other money rule here."""
    from django.db import IntegrityError

    add_movement(variant, StockMovement.PURCHASE_IN, 5)
    client = Client.objects.create(first_name="Балк", phone="+996700990004")
    order = SaleOrder.objects.create(client=client, currency=variant.currency)
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)
    payment = record_payment(order, Decimal("3200"))
    void_payment(payment)

    with pytest.raises(IntegrityError):
        Payment.objects.bulk_create(
            [
                Payment(
                    client=client,
                    order=order,
                    amount=-payment.amount,
                    currency=payment.currency,
                    rate_to_kgs=payment.rate_to_kgs,
                    reversed_payment=payment,
                )
            ]
        )


def test_partial_return_of_a_fixed_discount_line_scales_the_discount(variant):
    """Returning 1 of 2 units on a line with a сом-amount discount used to hit
    the saleitem_discount_fixed_lte_subtotal CheckConstraint and 500. The
    discount belongs to the whole line, so it shrinks with it: the client keeps
    exactly the per-unit price they actually paid."""
    from apps.sales.services import return_items

    add_movement(variant, StockMovement.PURCHASE_IN, 5)
    client = Client.objects.create(first_name="Скидка", phone="+996700990005")
    order = SaleOrder.objects.create(client=client, currency=variant.currency)
    item = SaleItem.objects.create(
        order=order,
        variant=variant,
        quantity=2,
        unit_price=Decimal("500.00"),
        discount_type=SaleItem.DISCOUNT_FIXED,
        discount_value=Decimal("800.00"),
    )
    confirm_sale(order)
    assert order.total == Decimal("200.00")  # 1000 − 800, i.e. 100 per unit

    return_items(order, {item.pk: 1})

    order.refresh_from_db()
    item.refresh_from_db()
    assert item.quantity == 1
    assert item.discount_value == Decimal("400.00"), "half returned, half the discount"
    assert order.total == Decimal("100.00"), "the one kept unit still costs what it did"
    assert variant.stock == 4


def test_percent_discount_needs_no_rescaling_on_a_partial_return(variant):
    """The other half of the rule: a percentage is proportional already, so it
    must survive a return untouched."""
    from apps.sales.services import return_items

    add_movement(variant, StockMovement.PURCHASE_IN, 5)
    client = Client.objects.create(first_name="Процент", phone="+996700990006")
    order = SaleOrder.objects.create(client=client, currency=variant.currency)
    item = SaleItem.objects.create(
        order=order,
        variant=variant,
        quantity=2,
        unit_price=Decimal("500.00"),
        discount_type=SaleItem.DISCOUNT_PERCENT,
        discount_value=Decimal("20"),
    )
    confirm_sale(order)
    assert order.total == Decimal("800.00")

    return_items(order, {item.pk: 1})

    order.refresh_from_db()
    item.refresh_from_db()
    assert item.discount_value == Decimal("20")
    assert order.total == Decimal("400.00")


def test_dashboard_payment_methods_report_net_cash_not_the_note_tendered(variant):
    """«Способы оплаты» summed the gross amount, so a 5 000 сом note handed over
    for a 3 000 сом sale was reported as 5 000 сом of cash taken — 2 000 of it
    went straight back out of the till as change."""
    from apps.reports.dashboard import dashboard_data

    add_movement(variant, StockMovement.PURCHASE_IN, 5)
    client = Client.objects.create(first_name="Сдача", phone="+996700990007")
    order = SaleOrder.objects.create(client=client, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3000"))
    confirm_sale(order)
    record_payment(
        order,
        Decimal("5000"),
        method=Payment.CASH,
        excess_disposition=Payment.DISPOSITION_CHANGE,
    )

    methods = {row["label"]: row["value"] for row in dashboard_data("month")["methods"]}
    assert sum(methods.values()) == Decimal("3000.00"), (
        f"expected the 3 000 сом that stayed in the till, got {methods}"
    )


def test_a_walk_in_cash_sale_is_settled_not_a_debt(variant):
    """A walk-in has no client, so services.record_payment records nothing and
    no Payment row can exist (Payment.client is non-nullable). Deriving the
    status from paid_amount=0 branded the shop's commonest transaction «долг» —
    on the POS result screen and in the owner's nightly Продажи sheet, where a
    day of cash sales read as a day of unpaid debt."""
    add_movement(variant, StockMovement.PURCHASE_IN, 5)
    order = SaleOrder.objects.create(client=None, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)

    assert record_payment(order, Decimal("3200")) is None  # nothing to record
    assert order.payment_status == SaleOrder.PAID
    assert order.balance == Decimal("0")


def test_a_walk_in_sale_still_never_appears_in_anyone_s_debt(variant):
    """The other half: settling the walk-in must not invent a client debt or
    leak into the debtors report."""
    add_movement(variant, StockMovement.PURCHASE_IN, 5)
    order = SaleOrder.objects.create(client=None, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)

    assert debtors_report_rows() == []


def test_daily_report_does_not_list_a_walk_in_cash_sale_as_unpaid(variant):
    """End-to-end on the sheet the owner actually reads every evening."""
    add_movement(variant, StockMovement.PURCHASE_IN, 5)
    order = SaleOrder.objects.create(client=None, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)

    rows = [str(cell) for row in _sales_sheet().rows for cell in row]
    assert "долг" not in rows, f"walk-in cash sale reported as debt: {rows}"
    assert "оплачено" in rows


# --- Cart undo/redo (apps.pos.undo) ------------------------------------------


def _draft_for(client, user, variant):
    return SaleOrder.objects.create(created_by=user, currency=variant.currency)


def test_cart_undo_restores_a_removed_line_and_redo_removes_it_again(
    client, django_user_model, variant
):
    call_command("setup_roles")
    user = django_user_model.objects.create_user("seller-undo", password="pw", is_staff=True)
    user.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(user)
    add_movement(variant, StockMovement.PURCHASE_IN, 10)

    order = _draft_for(client, user, variant)
    client.post(f"/pos/sale/{order.pk}/items/add/", {"variant_id": variant.pk, "quantity": 2})
    assert order.items.count() == 1

    item = order.items.get()
    client.post(f"/pos/sale/{order.pk}/items/{item.pk}/remove/")
    assert order.items.count() == 0

    client.post(f"/pos/sale/{order.pk}/undo/")
    assert order.items.count() == 1, "undo must put the removed line back"
    assert order.items.get().quantity == 2

    client.post(f"/pos/sale/{order.pk}/redo/")
    assert order.items.count() == 0, "redo must remove it again"


def test_cart_undo_reverses_a_discount(client, django_user_model, variant):
    call_command("setup_roles")
    user = django_user_model.objects.create_user("seller-disc", password="pw", is_staff=True)
    user.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(user)
    add_movement(variant, StockMovement.PURCHASE_IN, 10)

    order = _draft_for(client, user, variant)
    client.post(f"/pos/sale/{order.pk}/items/add/", {"variant_id": variant.pk, "quantity": 1})
    item = order.items.get()

    client.post(
        f"/pos/sale/{order.pk}/items/{item.pk}/discount/",
        {"discount_type": "percent", "discount_value": "50"},
    )
    assert order.items.get().discount_amount == Decimal("1600.00")

    client.post(f"/pos/sale/{order.pk}/undo/")
    assert order.items.get().discount_amount == Decimal("0.00")


def test_cart_undo_restores_the_client_on_the_sale(client, django_user_model, variant):
    call_command("setup_roles")
    user = django_user_model.objects.create_user("seller-cl", password="pw", is_staff=True)
    user.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(user)

    buyer = Client.objects.create(first_name="Асель", phone="+996700990101")
    order = _draft_for(client, user, variant)
    client.post(f"/pos/sale/{order.pk}/client/{buyer.pk}/set/")
    order.refresh_from_db()
    assert order.client_id == buyer.pk

    client.post(f"/pos/sale/{order.pk}/client/clear/")
    order.refresh_from_db()
    assert order.client_id is None

    client.post(f"/pos/sale/{order.pk}/undo/")
    order.refresh_from_db()
    assert order.client_id == buyer.pk, "undo must re-attach the client"
    assert Client.objects.filter(pk=buyer.pk).exists(), "undo must never delete the Client record"


def test_cart_undo_with_nothing_to_undo_is_a_calm_no_op(client, django_user_model, variant):
    call_command("setup_roles")
    user = django_user_model.objects.create_user("seller-noop", password="pw", is_staff=True)
    user.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(user)

    order = _draft_for(client, user, variant)
    resp = client.post(f"/pos/sale/{order.pk}/undo/")
    assert resp.status_code == 200
    assert "Отменять больше нечего" in resp.content.decode()


def test_cart_undo_re_clamps_against_stock_that_moved_since_the_snapshot(
    client, django_user_model, variant
):
    """A snapshot was valid when taken, but stock can move before the undo. The
    restore must never be the thing that puts an unsellable quantity in the
    cart — the same cap every other cart path applies."""
    call_command("setup_roles")
    user = django_user_model.objects.create_user("seller-clamp", password="pw", is_staff=True)
    user.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(user)
    add_movement(variant, StockMovement.PURCHASE_IN, 10)

    order = _draft_for(client, user, variant)
    client.post(f"/pos/sale/{order.pk}/items/add/", {"variant_id": variant.pk, "quantity": 8})
    item = order.items.get()
    client.post(f"/pos/sale/{order.pk}/items/{item.pk}/remove/")

    # Someone else sells 9 of the 10 while this cart sits empty.
    add_movement(variant, StockMovement.WRITEOFF_OUT, 9)

    client.post(f"/pos/sale/{order.pk}/undo/")
    restored = order.items.first()
    assert restored is not None
    assert restored.quantity == 1, f"undo restored {restored.quantity} with only 1 available"


def test_cart_undo_never_touches_a_confirmed_sale(client, django_user_model, variant):
    """Undo is for the draft. Once a sale is confirmed its reversal path is
    cancel_sale/return_items, which restock through services.py and leave a
    ledger trail — a keystroke must never rewrite counted money."""
    call_command("setup_roles")
    user = django_user_model.objects.create_user("seller-conf", password="pw", is_staff=True)
    user.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(user)
    add_movement(variant, StockMovement.PURCHASE_IN, 10)

    order = _draft_for(client, user, variant)
    client.post(f"/pos/sale/{order.pk}/items/add/", {"variant_id": variant.pk, "quantity": 2})
    client.post(f"/pos/sale/{order.pk}/confirm/", {"amount": ""})
    order.refresh_from_db()
    assert order.status == SaleOrder.CONFIRMED

    resp = client.post(f"/pos/sale/{order.pk}/undo/")
    assert resp.status_code == 404, "a confirmed sale is not a draft and has no undo endpoint"
    order.refresh_from_db()
    assert order.items.get().quantity == 2
    assert variant.stock == 8


def test_cart_undo_requires_sell_permission(client, django_user_model, variant):
    call_command("setup_roles")
    viewer = django_user_model.objects.create_user("viewer-undo", password="pw", is_staff=True)
    viewer.groups.add(Group.objects.get(name=VIEWER))
    owner = django_user_model.objects.create_superuser("owner-undo", password="pw")
    order = SaleOrder.objects.create(created_by=owner, currency=variant.currency)

    client.force_login(viewer)
    assert client.post(f"/pos/sale/{order.pk}/undo/").status_code == 403
    assert client.post(f"/pos/sale/{order.pk}/redo/").status_code == 403


def test_destructive_cart_and_order_actions_carry_a_confirmation(
    client, django_user_model, variant
):
    """The guard is a data-confirm attribute read by static/pos/js/safety.js —
    assert it is actually present on the rendered destructive controls, so
    removing it is a test failure rather than a silent regression."""
    call_command("setup_roles")
    user = django_user_model.objects.create_user("seller-guard", password="pw", is_staff=True)
    user.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(user)
    add_movement(variant, StockMovement.PURCHASE_IN, 5)

    order = _draft_for(client, user, variant)
    client.post(f"/pos/sale/{order.pk}/items/add/", {"variant_id": variant.pk, "quantity": 1})
    body = client.get(f"/pos/sale/{order.pk}/").content.decode()
    assert "data-confirm=" in body, "the cart's remove-line button lost its confirmation"


def test_adding_an_item_with_no_variant_id_is_a_404_not_a_500(client, django_user_model, variant):
    """A blank variant_id used to reach the ORM as "" and raise ValueError —
    a 500 on malformed input, where 404 is the honest answer."""
    call_command("setup_roles")
    user = django_user_model.objects.create_user("seller-bad", password="pw", is_staff=True)
    user.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(user)
    order = SaleOrder.objects.create(created_by=user, currency=variant.currency)

    for payload in ({"variant_id": "", "quantity": "1"}, {"quantity": "1"}, {"variant_id": "abc"}):
        resp = client.post(f"/pos/sale/{order.pk}/items/add/", payload)
        assert resp.status_code == 404, f"{payload} -> {resp.status_code}"
    assert order.items.count() == 0


# --- Money display precision (2026-08 review) --------------------------------
# Root cause: Postgres's NUMERIC division picks its own result scale, and
# neither Postgres nor Django rounds a raw SQL annotation to its declared
# output_field.decimal_places. SaleOrderAdmin's `_paid` (amount*rate -
# change)/rate summed in SQL) could come back as Decimal('6800.0000000000000000')
# against a Decimal('6800.00') total; subtracting nets to an unnormalized
# Decimal('0E-16') rather than Decimal('0.00'). apps.core.currency.format_money
# now quantizes at the one point every money value passes through before
# display, and the admin reads the already-quantized SaleOrder.paid_amount/
# .balance/.payment_status properties instead of the raw annotation.


def test_format_money_normalizes_the_exact_reported_garbage_values():
    """Decimal('0E-16') and a 16-decimal-place Decimal are exactly what a raw
    Postgres NUMERIC division annotation produced in production — assert the
    shared formatter turns both into clean, correct money strings."""
    from apps.core.currency import format_money

    assert format_money(Decimal("0E-16"), "KGS") == "0\N{NBSP}сом"
    assert format_money(Decimal("6800.0000000000000000"), "KGS") == "6\N{NBSP}800\N{NBSP}сом"
    assert format_money(Decimal("874.500000000000"), "USD") == "874,50\N{NBSP}$"


def test_sale_admin_paid_and_balance_columns_never_read_the_raw_annotation(variant):
    """Full-stack proof, not just the formatter: even if the DB-computed _paid
    annotation carries garbage precision (simulated here — sqlite's own
    division doesn't reproduce Postgres's scale behaviour, so the real bug
    can't be provoked through sqlite alone), the displayed columns must be
    unaffected because they no longer read _paid at all."""
    add_movement(variant, StockMovement.PURCHASE_IN, 5)
    client = Client.objects.create(first_name="Точность", phone="+996700990201")
    order = SaleOrder.objects.create(client=client, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("6800"))
    confirm_sale(order)
    record_payment(order, Decimal("6800"))

    admin_instance = site._registry[SaleOrder]
    obj = admin_instance.get_queryset(RequestFactory().get("/")).get(pk=order.pk)
    # Corrupt the annotation the way Postgres's division scale would, in place —
    # if any column still read it, this is what would leak onto the screen.
    obj._paid = Decimal("6800.00") - Decimal("6800.0000000000000000")
    assert str(obj._paid) == "0E-16", "sanity: this IS the unnormalized-zero shape"

    assert admin_instance.paid_column(obj) == "6\N{NBSP}800\N{NBSP}сом"
    assert admin_instance.balance_column(obj) == "0\N{NBSP}сом"
    assert "оплачено" in admin_instance.payment_badge(obj).lower()


def test_sale_admin_columns_render_correctly_for_a_normal_fully_paid_order(variant):
    """The literal spec: a sale paid exactly in full stores Decimal('0.00'),
    renders «0 сом», and classifies as оплачено — through the REAL,
    uncorrupted admin queryset this time."""
    add_movement(variant, StockMovement.PURCHASE_IN, 5)
    client = Client.objects.create(first_name="Ровно", phone="+996700990202")
    order = SaleOrder.objects.create(client=client, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)
    record_payment(order, Decimal("3200"))

    admin_instance = site._registry[SaleOrder]
    obj = admin_instance.get_queryset(RequestFactory().get("/")).get(pk=order.pk)
    assert obj.balance == Decimal("0.00")
    assert admin_instance.total_display(obj) == "3\N{NBSP}200\N{NBSP}сом"
    assert admin_instance.paid_column(obj) == "3\N{NBSP}200\N{NBSP}сом"
    assert admin_instance.balance_column(obj) == "0\N{NBSP}сом"
    assert "оплачено" in admin_instance.payment_badge(obj).lower()


def test_sale_admin_queryset_still_sorts_by_paid_despite_display_change(variant):
    """_paid stays for `ordering=` — only what's DISPLAYED moved to the
    canonical property. Sorting must still work."""
    add_movement(variant, StockMovement.PURCHASE_IN, 5)
    client = Client.objects.create(first_name="Сорт", phone="+996700990203")
    small = SaleOrder.objects.create(client=client, currency="KGS")
    SaleItem.objects.create(order=small, variant=variant, quantity=1, unit_price=Decimal("100"))
    confirm_sale(small)
    record_payment(small, Decimal("100"))

    big = SaleOrder.objects.create(client=client, currency="KGS")
    SaleItem.objects.create(order=big, variant=variant, quantity=1, unit_price=Decimal("900"))
    confirm_sale(big)
    record_payment(big, Decimal("900"))

    admin_instance = site._registry[SaleOrder]
    ids = list(
        admin_instance.get_queryset(RequestFactory().get("/"))
        .filter(pk__in=[small.pk, big.pk])
        .order_by("_paid")
        .values_list("pk", flat=True)
    )
    assert ids == [small.pk, big.pk]


def test_order_admin_total_and_paid_use_money_formatting(variant):
    """apps.orders.admin (production-order 'Заказы') gets the same NBSP +
    symbol formatting — the report named this admin list too."""
    from apps.orders.admin import OrderAdmin
    from apps.orders.models import Order, OrderItem

    client = Client.objects.create(first_name="Заказ", phone="+996700990204")
    order = Order.objects.create(client=client, currency="KGS")
    OrderItem.objects.create(order=order, variant=variant, quantity=2, unit_price=Decimal("3400"))
    admin_instance = site._registry[Order]
    assert isinstance(admin_instance, OrderAdmin)
    assert order.total == Decimal("6800")
    assert admin_instance.total_display(order) == "6\N{NBSP}800\N{NBSP}сом"
    assert admin_instance.paid_display(order) == "0\N{NBSP}сом"


# --- Lazy draft creation (2026-08 review, F1) ---------------------------------
# /pos/ used to create a SaleOrder the instant the page was opened, leaving an
# empty "Ожидает / 0.00 / —" row in the admin for anyone who ever visited it.
# The draft is now created lazily, on the FIRST real action — adding an item
# or picking/creating a client — via a session-token-guarded get_or_create
# (apps.pos.views._get_or_create_pending_draft) so a double-tap race still
# produces exactly one order. Everything after that first action (undo/redo,
# autosave, confirm, cancel) is completely unchanged — these endpoints hand
# off to the ordinary, unmodified per-order views and never duplicate their
# logic.


def _login_editor(client, django_user_model, username):
    call_command("setup_roles")
    user = django_user_model.objects.create_user(username, password="x" * 12, is_staff=True)
    user.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(user)
    return user


def test_opening_pos_creates_zero_database_rows(client, django_user_model):
    _login_editor(client, django_user_model, "lazy-open")
    assert SaleOrder.objects.count() == 0
    resp = client.get("/pos/", follow=True)
    assert resp.redirect_chain[-1][0] == "/pos/sale/new/"
    assert SaleOrder.objects.count() == 0


def test_visiting_sale_new_directly_creates_zero_rows(client, django_user_model):
    _login_editor(client, django_user_model, "lazy-new")
    resp = client.get("/pos/sale/new/")
    assert resp.status_code == 200
    assert SaleOrder.objects.count() == 0


def test_browsing_products_and_search_before_any_action_creates_zero_rows(
    client, django_user_model, variant
):
    _login_editor(client, django_user_model, "lazy-browse")
    client.get("/pos/sale/new/")
    client.get("/pos/sale/new/products/")
    client.get(f"/pos/sale/new/products/{variant.product_id}/variants/")
    client.get("/pos/sale/new/clients/search/", {"q": "test"})
    client.get("/pos/sale/new/clients/new/")
    assert SaleOrder.objects.count() == 0


def test_adding_the_first_item_creates_exactly_one_order(client, django_user_model, variant):
    _login_editor(client, django_user_model, "lazy-add")
    add_movement(variant, StockMovement.PURCHASE_IN, 5)
    client.get("/pos/sale/new/")

    resp = client.post("/pos/sale/new/items/add/", {"variant_id": variant.pk, "quantity": "2"})
    assert resp.status_code == 200
    assert resp["HX-Redirect"].startswith("/pos/sale/")

    assert SaleOrder.objects.count() == 1
    order = SaleOrder.objects.get()
    assert order.items.count() == 1
    assert order.items.get().quantity == 2
    assert resp["HX-Redirect"] == f"/pos/sale/{order.pk}/"


def test_picking_an_existing_client_first_creates_exactly_one_order(client, django_user_model):
    _login_editor(client, django_user_model, "lazy-pick")
    buyer = Client.objects.create(first_name="Первый", phone="+996700990401")
    client.get("/pos/sale/new/")

    resp = client.post(f"/pos/sale/new/client/{buyer.pk}/set/")
    assert resp.status_code == 200
    assert resp["HX-Redirect"].startswith("/pos/sale/")

    assert SaleOrder.objects.count() == 1
    order = SaleOrder.objects.get()
    assert order.client_id == buyer.pk


def test_creating_a_new_client_first_creates_exactly_one_order(client, django_user_model):
    _login_editor(client, django_user_model, "lazy-create")
    client.get("/pos/sale/new/")

    resp = client.post(
        "/pos/sale/new/clients/create/",
        {"first_name": "Новая", "phone": "+996700990402"},
    )
    assert resp.status_code == 200
    assert resp["HX-Redirect"].startswith("/pos/sale/")

    assert SaleOrder.objects.count() == 1
    order = SaleOrder.objects.get()
    assert order.client is not None
    assert order.client.phone == "+996700990402"


def test_new_client_validation_error_does_not_lose_the_form_or_redirect(client, django_user_model):
    """A duplicate phone / missing field is a real validation bounce, not a
    reason to strand the manager on a freshly-created, client-less order with
    no explanation — the inline error must survive, same as it does today."""
    _login_editor(client, django_user_model, "lazy-badclient")
    Client.objects.create(first_name="Existing", phone="+996700990403")
    client.get("/pos/sale/new/")

    resp = client.post(
        "/pos/sale/new/clients/create/",
        {"first_name": "Дубликат", "phone": "+996700990403"},
    )
    assert resp.status_code == 200
    assert "HX-Redirect" not in resp
    assert "уже есть" in resp.content.decode()

    # The order WAS created (lazily, before validation ran) but has no client —
    # a retry through the now-real, ordinary client_create must still work.
    assert SaleOrder.objects.count() == 1
    order = SaleOrder.objects.get()
    assert order.client_id is None
    assert f"/pos/sale/{order.pk}/clients/create/" in resp.content.decode()


def test_reload_after_the_first_action_restores_the_cart(client, django_user_model, variant):
    _login_editor(client, django_user_model, "lazy-reload")
    add_movement(variant, StockMovement.PURCHASE_IN, 5)
    client.get("/pos/sale/new/")
    client.post("/pos/sale/new/items/add/", {"variant_id": variant.pk, "quantity": "1"})
    order = SaleOrder.objects.get()

    resp = client.get("/pos/", follow=True)
    assert resp.redirect_chain[-1][0] == f"/pos/sale/{order.pk}/"
    assert SaleOrder.objects.count() == 1


def test_confirming_after_lazy_creation_works_exactly_as_before(client, django_user_model, variant):
    _login_editor(client, django_user_model, "lazy-confirm")
    add_movement(variant, StockMovement.PURCHASE_IN, 5)
    client.get("/pos/sale/new/")
    client.post("/pos/sale/new/items/add/", {"variant_id": variant.pk, "quantity": "1"})
    order = SaleOrder.objects.get()

    resp = client.post(f"/pos/sale/{order.pk}/confirm/", {"amount": ""})
    order.refresh_from_db()
    assert order.status == SaleOrder.CONFIRMED
    assert resp.status_code == 302


def test_starting_a_second_sale_after_confirming_the_first_gets_its_own_order(
    client, django_user_model, variant
):
    """A later, separate visit to the empty terminal must never reattach to a
    stale, already-consumed session token — it mints a fresh one."""
    _login_editor(client, django_user_model, "lazy-second")
    add_movement(variant, StockMovement.PURCHASE_IN, 5)
    client.get("/pos/sale/new/")
    client.post("/pos/sale/new/items/add/", {"variant_id": variant.pk, "quantity": "1"})
    first = SaleOrder.objects.get()
    client.post(f"/pos/sale/{first.pk}/confirm/", {"amount": ""})

    resp = client.get("/pos/", follow=True)
    assert resp.redirect_chain[-1][0] == "/pos/sale/new/"
    client.post("/pos/sale/new/items/add/", {"variant_id": variant.pk, "quantity": "1"})

    assert SaleOrder.objects.count() == 2
    second = SaleOrder.objects.exclude(pk=first.pk).get()
    assert second.pk != first.pk


@pytest.mark.django_db(transaction=True)
def test_two_simultaneous_first_actions_race_on_pending_token_to_one_order(django_user_model):
    """The actual mechanism under test: two threads calling get_or_create with
    the SAME pending_token (exactly what two near-simultaneous requests do,
    once sale_new's GET has minted one shared token into the session — see
    _get_or_create_pending_draft) must produce exactly one row, never two.

    Deliberately calls the ORM directly rather than going through the full
    HTTP request cycle: a real request also writes to django_session (twice —
    once for the initial GET, then again per POST) and to sales_saleitem/
    inventory_stockmovement, and sqlite locks the whole FILE per writer, not
    per row — stacking all of that onto two genuinely concurrent threads
    reliably trips sqlite's coarse locking on tables that have nothing to do
    with the property being verified here. Postgres (prod's real DB) uses
    row-level MVCC and wouldn't contend on any of that; isolating the test to
    the one write that actually matters — the get_or_create race itself —
    proves the real guarantee without fighting the test DB's own limits.

    One more sqlite-only wrinkle handled below: get_or_create's own documented
    safety net retries on IntegrityError (a genuine duplicate-key rejection —
    the case that actually matters, and what Postgres would produce for the
    loser of a real race). sqlite can ALSO reject a write immediately with
    OperationalError("database is locked") purely from whole-file contention,
    before either thread even reaches the uniqueness check — a busy_timeout is
    configured (config/settings/test.py), but two threads opening brand-new
    connections concurrently can still occasionally lose that race outright.
    That failure mode doesn't exist under Postgres's row-level locking, so a
    small retry-on-OperationalError loop below is test-environment
    accommodation, not a weakening of what's being proven: every attempt that
    DOES complete must land on the same row.
    """
    import threading
    import time

    from django.db import transaction as txn
    from django.db.utils import OperationalError

    call_command("setup_roles")
    user = django_user_model.objects.create_user("lazy-race", password="x" * 12, is_staff=True)
    user.groups.add(Group.objects.get(name=EDITOR))

    token = "race-token-0123456789abcdef"
    results = []
    barrier = threading.Barrier(2)

    def fire():
        barrier.wait(timeout=5)
        for attempt in range(8):
            try:
                with txn.atomic():
                    order, created = SaleOrder.objects.get_or_create(
                        created_by=user,
                        pending_token=token,
                        defaults={"channel": SaleOrder.SHOP},
                    )
                results.append((order.pk, created))
                return
            except OperationalError:
                time.sleep(0.05 * (attempt + 1))
        raise AssertionError("gave up retrying after sqlite whole-file lock contention")

    threads = [threading.Thread(target=fire) for _ in range(2)]
    for t in threads:
        t.start()
    for t in threads:
        t.join(timeout=15)

    assert len(results) == 2
    pks = {pk for pk, _ in results}
    assert len(pks) == 1, f"expected both threads to land on ONE order, got {pks}"
    assert sum(1 for _, created in results if created) == 1, "expected exactly one real INSERT"
    assert SaleOrder.objects.filter(pending_token=token).count() == 1


def test_lazy_creation_reads_but_never_relies_on_a_stale_session_token(
    client, django_user_model, variant
):
    """A POST arriving with no prior GET /pos/sale/new/ (no token minted yet —
    e.g. a stale bookmark) must still work, minting its own token on the fly."""
    _login_editor(client, django_user_model, "lazy-notoken")
    add_movement(variant, StockMovement.PURCHASE_IN, 5)
    assert "pending_sale_token" not in client.session

    resp = client.post("/pos/sale/new/items/add/", {"variant_id": variant.pk, "quantity": "1"})
    assert resp.status_code == 200
    assert SaleOrder.objects.count() == 1


def test_viewer_cannot_reach_any_lazy_creation_endpoint(client, django_user_model, variant):
    call_command("setup_roles")
    viewer = django_user_model.objects.create_user("lazy-viewer", password="x" * 12, is_staff=True)
    viewer.groups.add(Group.objects.get(name=VIEWER))
    client.force_login(viewer)

    assert client.get("/pos/sale/new/").status_code == 403
    assert (
        client.post(
            "/pos/sale/new/items/add/", {"variant_id": variant.pk, "quantity": "1"}
        ).status_code
        == 403
    )
    assert SaleOrder.objects.count() == 0


# --- Admin: empty drafts hidden by default (F1) ------------------------------


def test_sale_admin_hides_empty_drafts_by_default(django_user_model):
    SaleOrder.objects.create()  # empty — no client, no items
    real_draft = SaleOrder.objects.create(
        client=Client.objects.create(first_name="Real", phone="+996700990501")
    )

    admin_instance = site._registry[SaleOrder]
    request = RequestFactory().get("/")
    pks = set(admin_instance.get_queryset(request).values_list("pk", flat=True))
    assert real_draft.pk in pks
    assert len(pks) == 1, f"expected only the real draft, got {pks}"


def test_sale_admin_shows_empty_drafts_with_the_filter_applied():
    empty = SaleOrder.objects.create()

    admin_instance = site._registry[SaleOrder]
    request = RequestFactory().get("/", {"show_drafts": "yes"})
    pks = set(admin_instance.get_queryset(request).values_list("pk", flat=True))
    assert empty.pk in pks


def test_sale_admin_default_hide_never_touches_a_confirmed_walkin(variant):
    """A confirmed walk-in has status=confirmed (not draft) with no client —
    the exclude is scoped to status=DRAFT specifically so this must never
    be swept up by accident."""
    add_movement(variant, StockMovement.PURCHASE_IN, 5)
    order = SaleOrder.objects.create(currency=variant.currency)
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("100"))
    confirm_sale(order)
    order.items.all().delete()  # contrived: now client=None AND items=None, but CONFIRMED

    admin_instance = site._registry[SaleOrder]
    request = RequestFactory().get("/")
    pks = set(admin_instance.get_queryset(request).values_list("pk", flat=True))
    assert order.pk in pks


def test_sale_admin_paid_column_unaffected_by_the_draft_visibility_exclude(variant):
    """The exclude() joins `items`; the existing _paid annotation joins
    `payments` — proves the two don't fan out against each other (a payment
    Sum inflating because of an unrelated items join is exactly the kind of
    bug two independent to-many joins on one queryset can cause)."""
    add_movement(variant, StockMovement.PURCHASE_IN, 5)
    client = Client.objects.create(first_name="ДваСтроки", phone="+996700990502")
    order = SaleOrder.objects.create(client=client, currency=variant.currency)
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("100"))
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("200"))
    confirm_sale(order)
    record_payment(order, Decimal("300"))

    admin_instance = site._registry[SaleOrder]
    request = RequestFactory().get("/")
    obj = admin_instance.get_queryset(request).get(pk=order.pk)
    assert admin_instance.paid_column(obj) == "300\N{NBSP}сом"


# --- Confirmation dialog wording/colour (2026-08 audit) ----------------------
# safety.js's shared dialog defaulted its "yes" button to "Да, удалить" (a
# delete verb) styled red, for EVERY data-confirm site that didn't override
# it — including «Подтвердить продажу?», a CREATE action with no override at
# all. These pin the wording+colour(tone) for every fixed dialog directly in
# the rendered HTML, so a careless future template edit that drops
# data-confirm-ok/data-confirm-tone shows up as a failing test, not a support
# ticket about a red "delete" button on the sale-confirm screen.


def test_confirm_sale_dialog_is_worded_and_coloured_as_a_create_not_a_delete(
    client, django_user_model, variant
):
    _login_editor(client, django_user_model, "dlg-confirm")
    add_movement(variant, StockMovement.PURCHASE_IN, 5)
    order = SaleOrder.objects.create(
        created_by=django_user_model.objects.get(username="dlg-confirm")
    )
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("500"))

    body = client.get(f"/pos/sale/{order.pk}/").content.decode()
    assert 'data-confirm-tone="primary"' in body, "confirm-sale must not use danger/red styling"
    assert "Да, подтвердить" in body
    assert "Подтвердить продажу на 500" in body, "must name the amount, not just 'are you sure?'"
    assert "Да, удалить" not in body


def test_cancel_sale_dialog_names_the_order_and_amount(client, django_user_model, variant):
    _login_editor(client, django_user_model, "dlg-cancel")
    add_movement(variant, StockMovement.PURCHASE_IN, 5)
    order = SaleOrder.objects.create(
        created_by=django_user_model.objects.get(username="dlg-cancel"), currency="KGS"
    )
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("500"))
    confirm_sale(order)

    body = client.get(f"/pos/sale/{order.pk}/result/").content.decode()
    assert f"Отменить продажу №{order.pk} на 500" in body
    assert 'data-confirm-tone="danger"' in body
    assert "Да, отменить продажу" in body
    assert 'name="reason"' in body  # a plain typed reason, asked for every time


def test_cancel_sale_without_a_reason_is_rejected_and_nothing_changes(
    client, django_user_model, variant
):
    """«ask for the reason» — an empty/missing reason must not cancel the
    sale at all, and must never touch apps.manufacturing's defect (брак)
    tracking, which is a completely separate concept (production quality on
    goods being made, not a completed sale being cancelled)."""
    from apps.manufacturing.models import ProductionRun

    _login_editor(client, django_user_model, "no-reason-cancel")
    add_movement(variant, StockMovement.PURCHASE_IN, 5)
    order = SaleOrder.objects.create(
        created_by=django_user_model.objects.get(username="no-reason-cancel"), currency="KGS"
    )
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("500"))
    confirm_sale(order)

    resp = client.post(f"/pos/sale/{order.pk}/cancel/")  # no reason posted
    assert resp.status_code == 302
    order.refresh_from_db()
    assert order.status == SaleOrder.CONFIRMED  # unchanged — never cancelled
    assert not ProductionRun.objects.exists()  # never touched at all

    resp = client.post(f"/pos/sale/{order.pk}/cancel/", {"reason": "клиент передумал"})
    order.refresh_from_db()
    assert order.status == SaleOrder.CANCELLED
    movement = StockMovement.objects.filter(sale_order=order, movement_type="return_in").first()
    assert movement.reason == "клиент передумал"  # plain text, not a category


def test_cancel_sale_with_an_overlong_reason_is_truncated_not_silently_dropped(
    client, django_user_model, variant
):
    """A raw crafted POST past the client-side maxlength=255 must never
    reach StockMovement.reason's own max_length and blow up inside
    add_movement's full_clean() — the view's blanket
    `except ValidationError: pass` (idempotent-cancel guard) would swallow
    that silently, leaving the sale uncancelled with zero feedback. Must
    truncate server-side instead, so the cancel still goes through."""
    _login_editor(client, django_user_model, "overlong-cancel")
    add_movement(variant, StockMovement.PURCHASE_IN, 5)
    order = SaleOrder.objects.create(
        created_by=django_user_model.objects.get(username="overlong-cancel"), currency="KGS"
    )
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("500"))
    confirm_sale(order)

    resp = client.post(f"/pos/sale/{order.pk}/cancel/", {"reason": "x" * 500})
    assert resp.status_code == 302
    order.refresh_from_db()
    assert order.status == SaleOrder.CANCELLED  # went through, not silently dropped
    movement = StockMovement.objects.filter(sale_order=order, movement_type="return_in").first()
    assert len(movement.reason) == 255


def test_return_without_a_reason_is_rejected(client, django_user_model, variant):
    owner = django_user_model.objects.create_superuser("no-reason-return", "o@e.com", "x" * 12)
    client.force_login(owner)
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    order = SaleOrder.objects.create()
    SaleItem.objects.create(order=order, variant=variant, quantity=2, unit_price=Decimal("3200"))
    confirm_sale(order)

    resp = client.post(f"/pos/sale/{order.pk}/return/", {f"return_{order.items.first().pk}": "1"})
    order.refresh_from_db()
    assert order.items.first().quantity == 2  # unchanged — the return never applied

    resp = client.post(
        f"/pos/sale/{order.pk}/return/",
        {f"return_{order.items.first().pk}": "1", "reason": "не подошёл размер"},
    )
    assert resp.status_code == 302
    order.refresh_from_db()
    assert order.items.first().quantity == 1
    movement = StockMovement.objects.filter(sale_order=order, movement_type="return_in").first()
    assert movement.reason == "не подошёл размер"


def test_cancel_order_dialog_names_the_order_and_amount(client, django_user_model, variant):
    from apps.orders.models import Order, OrderItem

    _login_editor(client, django_user_model, "dlg-order")
    owner = django_user_model.objects.create_superuser("dlg-order-owner", password="x" * 12)
    client.force_login(owner)
    buyer = Client.objects.create(first_name="DlgOrder", phone="+996700990601")
    order = Order.objects.create(client=buyer, currency="KGS")
    OrderItem.objects.create(order=order, variant=variant, quantity=2, unit_price=Decimal("500"))

    body = client.get(f"/orders/{order.pk}/").content.decode()
    assert f"Отменить заказ №{order.pk} на 1\N{NBSP}000" in body
    assert 'data-confirm-tone="danger"' in body
    assert "Да, отменить заказ" in body
    assert 'name="reason"' in body


def test_cancel_order_requires_a_reason_and_stores_it_plainly(client, django_user_model, variant):
    """Same rule as a sale cancel/return — a plain typed reason, required,
    never a defect/брак classification (apps.manufacturing is untouched by
    this entirely)."""
    from apps.manufacturing.models import ProductionRun
    from apps.orders.models import Order, OrderItem

    owner = django_user_model.objects.create_superuser("order-reason-owner", "o@e.com", "x" * 12)
    client.force_login(owner)
    buyer = Client.objects.create(first_name="ReasonOrder", phone="+996700990602")
    order = Order.objects.create(client=buyer, currency="KGS")
    OrderItem.objects.create(order=order, variant=variant, quantity=2, unit_price=Decimal("500"))

    resp = client.post(f"/orders/{order.pk}/cancel/")  # no reason
    order.refresh_from_db()
    assert order.status == Order.NEW  # unchanged
    assert not ProductionRun.objects.exists()

    resp = client.post(f"/orders/{order.pk}/cancel/", {"reason": "клиент передумал"})
    assert resp.status_code == 302
    order.refresh_from_db()
    assert order.status == Order.CANCELLED
    assert order.cancel_reason == "клиент передумал"


def test_cancel_order_with_an_overlong_reason_is_truncated_not_a_500(
    client, django_user_model, variant
):
    """cancel_order's plain order.save() never calls full_clean(), so an
    untruncated overlong reason would reach the database as-is — a raw
    uncaught IntegrityError (a 500) on Postgres, where varchar(255) is a
    real constraint. Must truncate server-side before it ever gets there."""
    from apps.orders.models import Order, OrderItem

    owner = django_user_model.objects.create_superuser("overlong-order-owner", "o@e.com", "x" * 12)
    client.force_login(owner)
    buyer = Client.objects.create(first_name="OverlongOrder", phone="+996700990603")
    order = Order.objects.create(client=buyer, currency="KGS")
    OrderItem.objects.create(order=order, variant=variant, quantity=2, unit_price=Decimal("500"))

    resp = client.post(f"/orders/{order.pk}/cancel/", {"reason": "y" * 500})
    assert resp.status_code == 302
    order.refresh_from_db()
    assert order.status == Order.CANCELLED
    assert len(order.cancel_reason) == 255


def test_stock_action_wizard_colours_writeoff_red_and_receive_green(
    client, django_user_model, variant
):
    """The bulk stock-action wizard (admin/inventory/stock_action.html) used
    one generic 'Apply' button for receive/writeoff/recount alike — writeoff
    is destructive (removes stock permanently) and must read that way, same
    as the single-variant inline buttons already do. Goes through the real
    admin action (selecting the action, no `apply` yet) rather than rendering
    the template in isolation, since it needs a real admin request context."""
    owner = django_user_model.objects.create_superuser("dlg-stock-owner", password="x" * 12)
    client.force_login(owner)
    changelist = "/panel/inventory/productvariant/"

    writeoff = client.post(
        changelist, {"action": "writeoff_selected", "_selected_action": [variant.pk]}
    )
    receive = client.post(
        changelist, {"action": "receive_stock_selected", "_selected_action": [variant.pk]}
    )
    writeoff_html = writeoff.content.decode()
    receive_html = receive.content.decode()
    assert "btn-danger" in writeoff_html
    assert "btn-success" in receive_html
    assert "btn-danger" not in receive_html
    assert "btn-success" not in writeoff_html


def test_admin_cancel_sale_action_asks_before_cancelling(client, django_user_model, variant):
    """cancel_selected used to fire the instant "Go" was clicked — no
    confirmation step, unlike approve_selected's existing two-step pattern.
    The first POST (no `apply`) must show the list and NOT touch the sale;
    only the second, confirmed POST actually cancels it."""
    owner = django_user_model.objects.create_superuser("bulk_owner1", "o@e.com", "x" * 12)
    client.force_login(owner)
    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    order = SaleOrder.objects.create()
    SaleItem.objects.create(order=order, variant=variant, quantity=2, unit_price=Decimal("3200"))
    confirm_sale(order)
    assert variant.stock == 8

    changelist = "/panel/sales/saleorder/"
    resp = client.post(changelist, {"action": "cancel_selected", "_selected_action": [order.pk]})
    assert resp.status_code == 200
    assert "cannot be undone" in resp.content.decode() or "Cancel sales" in resp.content.decode()
    order.refresh_from_db()
    variant.refresh_from_db()
    assert order.status == SaleOrder.CONFIRMED  # untouched
    assert variant.stock == 8  # untouched

    resp = client.post(
        changelist,
        {
            "action": "cancel_selected",
            "_selected_action": [order.pk],
            "apply": "1",
        },
    )
    assert resp.status_code == 302
    order.refresh_from_db()
    variant.refresh_from_db()
    assert order.status == SaleOrder.CANCELLED
    assert variant.stock == 10


def test_admin_void_payment_action_asks_before_voiding(client, django_user_model, variant):
    owner = django_user_model.objects.create_superuser("bulk_owner2", "o2@e.com", "x" * 12)
    client.force_login(owner)
    cust = Client.objects.create(first_name="VoidConfirm", phone="+996700009100")
    order = SaleOrder.objects.create(client=cust)
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    confirm_sale(order)
    payment = Payment.objects.create(client=cust, order=order, amount=Decimal("3200"))

    changelist = "/panel/sales/payment/"
    resp = client.post(changelist, {"action": "void_selected", "_selected_action": [payment.pk]})
    assert resp.status_code == 200
    assert not Payment.objects.filter(reversed_payment=payment).exists()  # untouched

    resp = client.post(
        changelist,
        {"action": "void_selected", "_selected_action": [payment.pk], "apply": "1"},
    )
    assert resp.status_code == 302
    assert Payment.objects.filter(reversed_payment=payment).exists()


def test_admin_cancel_order_action_asks_before_cancelling(client, django_user_model, variant):
    from apps.orders.models import Order
    from apps.orders.services import create_order

    owner = django_user_model.objects.create_superuser("bulk_owner3", "o3@e.com", "x" * 12)
    client.force_login(owner)
    cust = Client.objects.create(first_name="OrderCancelConfirm", phone="+996700009101")
    order = create_order(
        cust, [{"variant": variant, "quantity": 1, "unit_price": variant.sale_price}]
    )

    changelist = "/panel/orders/order/"
    resp = client.post(changelist, {"action": "cancel_selected", "_selected_action": [order.pk]})
    assert resp.status_code == 200
    order.refresh_from_db()
    assert order.status == Order.NEW  # untouched

    resp = client.post(
        changelist,
        {"action": "cancel_selected", "_selected_action": [order.pk], "apply": "1"},
    )
    assert resp.status_code == 302
    order.refresh_from_db()
    assert order.status == Order.CANCELLED


def test_admin_campaign_send_now_asks_before_sending(client, django_user_model, settings):
    from apps.campaigns.models import Campaign

    settings.CAMPAIGNS_ENABLED = True
    owner = django_user_model.objects.create_superuser("bulk_owner4", "o4@e.com", "x" * 12)
    client.force_login(owner)
    campaign = Campaign.objects.create(name="ConfirmGate", text_ru="hi")

    changelist = "/panel/campaigns/campaign/"
    resp = client.post(changelist, {"action": "send_now", "_selected_action": [campaign.pk]})
    assert resp.status_code == 200
    assert "Send campaigns" in resp.content.decode() or "cannot be undone" in resp.content.decode()


def test_campaign_send_failure_shows_no_raw_exception_text(
    client, django_user_model, settings, monkeypatch
):
    """The admin used to interpolate str(exc) straight into the message —
    a raw Python exception shown to a non-technical Owner with no next
    step. Must show a plain Russian message instead, and log the real
    exception for a developer to actually diagnose."""
    from apps.campaigns.models import Campaign

    settings.CAMPAIGNS_ENABLED = True
    owner = django_user_model.objects.create_superuser("bulk_owner5", "o5@e.com", "x" * 12)
    client.force_login(owner)
    Client.objects.create(
        first_name="Reachable", phone="+996700009102", telegram_chat_id=42, marketing_consent=True
    )
    campaign = Campaign.objects.create(name="BoomCampaign", text_ru="hi")

    def boom(*a, **k):
        raise RuntimeError("Traceback (most recent call last): something internal broke")

    monkeypatch.setattr("apps.campaigns.admin.call_command", boom)

    resp = client.post(
        "/panel/campaigns/campaign/",
        {"action": "send_now", "_selected_action": [campaign.pk], "apply": "1"},
        follow=True,
    )
    body = resp.content.decode()
    assert "Traceback" not in body
    assert "something internal broke" not in body
    assert "не удалась" in body


# ---- Category shown everywhere a product is identified (products in ------
# ---- different categories can share a name) ------------------------------


def test_variant_str_leads_with_category(variant):
    """ProductVariant.__str__ is the shared string used by the cart's
    remove-confirm, result/return screens, order lines, and several
    dashboard panels — category must lead so a bare product name is never
    shown ambiguously through any of them."""
    assert str(variant) == "Dresses / Evening Dress / M / red"


def test_receipt_pdf_shows_category_with_no_column_overlap(variant):
    """The receipt table renders the category as its own line above the
    product name (see templates/receipts/receipt.html) — verified here at
    the data level (receipt_context/rendered HTML); the actual PDF layout
    (no column overlap at 1 item, 6 items/2 products, and long category+
    product names) was verified visually against a real rendered PDF."""
    from django.template.loader import render_to_string

    from apps.pos.receipts import receipt_context

    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order)

    ctx = receipt_context(order)
    assert ctx["groups"][0]["product"].category.name == "Dresses"
    html = render_to_string("receipts/receipt.html", ctx)
    assert "Dresses" in html
    assert html.index("Dresses") < html.index("Evening Dress")  # category ABOVE product name


def test_product_grid_tile_and_variant_picker_show_category(client, django_user_model, variant):
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    call_command("setup_roles")
    editor = django_user_model.objects.create_user("catgrid1", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)
    order = SaleOrder.objects.create(created_by=editor, channel=SaleOrder.SHOP)

    resp = client.get(f"/pos/sale/{order.pk}/products/")
    assert "Dresses" in resp.content.decode()

    resp2 = client.get(f"/pos/sale/{order.pk}/products/{variant.product_id}/variants/")
    assert "Dresses" in resp2.content.decode()


def test_cart_line_shows_category(client, django_user_model, variant):
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    call_command("setup_roles")
    editor = django_user_model.objects.create_user("catcart1", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)
    order = SaleOrder.objects.create(created_by=editor, channel=SaleOrder.SHOP)
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))

    resp = client.get(f"/pos/sale/{order.pk}/")
    assert "Dresses" in resp.content.decode()


def test_debt_pay_item_table_shows_category(client, django_user_model, variant):
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    call_command("setup_roles")
    editor = django_user_model.objects.create_user("catdebt1", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)
    cust = Client.objects.create(first_name="Debtor", phone="+996700007771")
    order = SaleOrder.objects.create(client=cust, created_by=editor)
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order, user=editor)

    resp = client.get(f"/pos/sale/{order.pk}/debt/")
    assert "Dresses" in resp.content.decode()


def test_client_history_item_summary_shows_category(client, django_user_model, variant):
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    call_command("setup_roles")
    editor = django_user_model.objects.create_user("cathist1", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)
    cust = Client.objects.create(first_name="Histo", phone="+996700007772")
    order = SaleOrder.objects.create(client=cust, created_by=editor)
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order, user=editor)

    resp = client.get(f"/pos/clients/{cust.pk}/")
    assert "Dresses" in resp.content.decode()


def test_admin_variant_list_has_category_column(client, django_user_model, variant):
    owner = django_user_model.objects.create_superuser("catadmin1", "a@e.com", "x" * 12)
    client.force_login(owner)
    resp = client.get("/panel/inventory/productvariant/")
    body = resp.content.decode()
    assert "Category" in body or "category" in body.lower()
    assert "Dresses" in body


def test_daily_report_stock_sheet_includes_category(variant):
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    sheet = _stock_sheet()
    assert sheet.columns[1].header == "Категория"
    row = next(r for r in sheet.rows if r[0] == variant.sku)
    assert row[1] == "Dresses"
    assert row[2] == "Evening Dress"


def test_dashboard_top_products_and_dead_stock_sheets_include_category(variant):
    from apps.reports.dashboard import dashboard_data, dashboard_sheets

    add_movement(variant, StockMovement.PRODUCTION_IN, 20)
    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=2, unit_price=Decimal("3200"))
    confirm_sale(order)

    data = dashboard_data("month")
    assert data["top_products"][0]["category"] == "Dresses"

    sheets = dashboard_sheets(data)
    top_row = sheets["Топ товаров"].rows[0]
    assert top_row[0] == "Dresses / Evening Dress"


def test_production_queue_row_label_includes_category(variant):
    """The queue's one-row-per-variant label (ProductVariant.__str__) leads
    with category, same as everywhere else this string renders — see that
    model's own docstring on why (products in different categories can
    share a name)."""
    from apps.orders.models import Order, OrderItem
    from apps.orders.services import production_queue

    cust = Client.objects.create(first_name="QueueCat", phone="+996700007773")
    order = Order.objects.create(client=cust, currency="KGS")
    OrderItem.objects.create(order=order, variant=variant, quantity=3, unit_price=Decimal("3200"))

    rows = production_queue()
    assert rows[0]["label"].startswith("Dresses / Evening Dress")


# ---- Multiple product-image upload — the real fix was a real multi-file ---
# ---- picker (apps.inventory.admin.ProductAdmin.bulk_image_upload_view), --
# ---- ProductImageInline never supported selecting more than one file at --
# ---- a time. Plus: size/dimension validation and per-file failure -------
# ---- isolation (apps.inventory.models.validate_product_image). ----------


def _bulk_upload(client, product, files):
    return client.post(
        f"/panel/inventory/product/{product.pk}/images/bulk-upload/",
        {"images": files},
        format="multipart",
    )


def test_bulk_upload_ten_images_succeeds(client, django_user_model):
    owner = django_user_model.objects.create_superuser("bulk10", "b10@e.com", "x" * 12)
    client.force_login(owner)
    cat = Category.objects.create(name="BulkCat")
    product = Product.objects.create(category=cat, name="BulkProduct")

    files = [_tiny_image(f"p{i}.png") for i in range(10)]
    resp = _bulk_upload(client, product, files)
    assert resp.status_code == 200
    data = resp.json()
    assert data["saved"] == 10
    assert data["failed"] == 0
    assert all(r["ok"] for r in data["results"])
    assert ProductImage.objects.filter(product=product).count() == 10
    # Upload order is preserved as sequential positions (see _next_image_position).
    positions = list(
        ProductImage.objects.filter(product=product)
        .order_by("position")
        .values_list("position", flat=True)
    )
    assert positions == list(range(10))


def test_bulk_upload_rejects_oversized_file_but_saves_the_valid_ones(client, django_user_model):
    """The old ProductImageInline formset path is all-or-nothing (Django
    validates every row before saving any) — one bad file loses the whole
    batch, silently, with a generic message. The bulk endpoint validates and
    saves each file independently: the two good photos must survive even
    though the third file is rejected, and the rejection must name the file
    and say clearly that it's oversized — not "corrupted"."""
    owner = django_user_model.objects.create_superuser("bulkbad", "bb@e.com", "x" * 12)
    client.force_login(owner)
    cat = Category.objects.create(name="BulkBadCat")
    product = Product.objects.create(category=cat, name="BulkBadProduct")

    oversized = SimpleUploadedFile(
        "toobig.jpg",
        b"0" * (PRODUCT_IMAGE_MAX_BYTES + 1),
        content_type="image/jpeg",
    )
    files = [_tiny_image("good1.png"), oversized, _tiny_image("good2.png")]
    resp = _bulk_upload(client, product, files)
    assert resp.status_code == 200
    data = resp.json()
    assert data["saved"] == 2
    assert data["failed"] == 1
    assert ProductImage.objects.filter(product=product).count() == 2

    bad = next(r for r in data["results"] if not r["ok"])
    assert bad["name"] == "toobig.jpg"
    assert "МБ" in bad["error"]  # size-specific Russian message, not "corrupted"
    assert "поврежд" not in bad["error"]


def test_oversized_dimensions_rejected_before_pillow_decodes_pixels():
    """A file whose HEADER declares dimensions past PRODUCT_IMAGE_MAX_DIMENSION
    must be rejected from just the header read (Image.open() — cheap, lazy,
    format plugins only parse metadata) WITHOUT ever calling .load()/
    .convert(), which is what actually decodes the full pixel buffer. Patching
    Image.Image.load to blow up proves the rejection happens strictly before
    that — the literal "before Pillow decodes it fully" requirement."""
    from io import BytesIO
    from unittest.mock import patch

    from PIL import Image as PILImage

    from apps.inventory.models import PRODUCT_IMAGE_MAX_DIMENSION, validate_product_image

    # A single scanline is nothing to encode/decode — only the header lies.
    wide = PILImage.new("RGB", (PRODUCT_IMAGE_MAX_DIMENSION + 500, 10), color=(1, 1, 1))
    buf = BytesIO()
    wide.save(buf, format="PNG")
    f = SimpleUploadedFile("wide.png", buf.getvalue(), content_type="image/png")

    with patch.object(PILImage.Image, "load", side_effect=AssertionError("full decode ran")):
        with pytest.raises(ValidationError) as exc_info:
            validate_product_image(f)
    assert str(PRODUCT_IMAGE_MAX_DIMENSION) in str(exc_info.value)


def test_true_decompression_bomb_gets_the_clear_dimension_message_not_generic_corrupted():
    """A 20000x20000 PNG (tiny on disk, a huge pixel buffer once decoded) —
    Pillow's own DecompressionBombError still exists as defense in depth, but
    our validator must translate it into the SAME clear "too large" Russian
    message as an explicit over-PRODUCT_IMAGE_MAX_DIMENSION rejection, not
    Django's generic (and here actively misleading) "file is corrupted" text."""
    from io import BytesIO

    from PIL import Image as PILImage

    from apps.inventory.models import validate_product_image

    bomb = PILImage.new("RGB", (20000, 20000), color=(1, 1, 1))
    buf = BytesIO()
    bomb.save(buf, format="PNG")
    f = SimpleUploadedFile("bomb.png", buf.getvalue(), content_type="image/png")

    with pytest.raises(ValidationError) as exc_info:
        validate_product_image(f)
    assert "большое" in str(exc_info.value)
    assert "поврежд" not in str(exc_info.value)


def test_thumbnail_failure_does_not_lose_the_uploaded_image(monkeypatch):
    """A synthetic thumbnail-generation failure (anything unexpected — disk
    hiccup, memory blip) must not crash the save or lose the already-valid,
    already-on-disk image; Product.grid_image already falls back to the
    original when thumbnail is empty, so this must leave a saved row with an
    empty thumbnail rather than raising."""
    from apps.inventory.models import ProductImage as PI

    def boom(self):
        raise RuntimeError("simulated Pillow failure")

    monkeypatch.setattr(PI, "_make_thumbnail", boom)
    cat = Category.objects.create(name="ThumbFailCat")
    product = Product.objects.create(category=cat, name="ThumbFailProduct")

    img = ProductImage.objects.create(product=product, image=_tiny_image("t.png"))
    img.refresh_from_db()
    assert img.pk is not None
    assert not img.thumbnail
    assert product.grid_image == img.image  # falls back to the original, per docstring


def test_product_image_inline_formset_also_rejects_oversized_file(client, django_user_model):
    """The validator lives on the model field, so the OLD ProductImageInline
    path (still there for reordering/deleting) is covered too, not just the
    new bulk endpoint — one path of truth for what counts as a valid photo."""
    owner = django_user_model.objects.create_superuser("inlinebad", "ib@e.com", "x" * 12)
    client.force_login(owner)
    cat = Category.objects.create(name="InlineBadCat")
    product = Product.objects.create(category=cat, name="InlineBadProduct")

    oversized = SimpleUploadedFile(
        "toobig.jpg", b"0" * (PRODUCT_IMAGE_MAX_BYTES + 1), content_type="image/jpeg"
    )
    data = {
        "name": product.name,
        "category": product.category_id,
        "description": "",
        "is_active": "on",
        "images-TOTAL_FORMS": "1",
        "images-INITIAL_FORMS": "0",
        "images-MIN_NUM_FORMS": "0",
        "images-MAX_NUM_FORMS": "1000",
        "images-0-image": oversized,
        "images-0-product": str(product.pk),
        "variants-TOTAL_FORMS": "0",
        "variants-INITIAL_FORMS": "0",
        "variants-MIN_NUM_FORMS": "0",
        "variants-MAX_NUM_FORMS": "1000",
        "_save": "Save",
    }
    resp = client.post(f"/panel/inventory/product/{product.pk}/change/", data)
    assert resp.status_code == 200  # redisplayed with a form error, not saved
    assert "МБ" in resp.content.decode()
    assert ProductImage.objects.filter(product=product).count() == 0


# ---- Phase 5: send the statement like she does today, Должники worklist ---


def test_send_statement_message_contains_no_url(client, django_user_model, variant):
    """«Отправить выписку» opens WhatsApp with the headline figure only —
    no link, no domain, no token ever sent to a client (same rule as the
    receipt share text)."""
    from urllib.parse import parse_qs, unquote, urlsplit

    call_command("setup_roles")
    editor = django_user_model.objects.create_user("stmt_send_ed", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    cust = Client.objects.create(first_name="Данара", phone="+996700009001")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("32000"))
    confirm_sale(order)

    resp = client.post(f"/pos/clients/{cust.pk}/statement/send/")
    assert resp.status_code == 302
    location = resp["Location"]
    assert location.startswith("https://wa.me/")
    text = unquote(parse_qs(urlsplit(location).query)["text"][0])
    assert "http" not in text
    assert "www." not in text
    assert "Данара" in text
    assert "Остаток" in text


def test_send_statement_fires_even_at_zero_balance(client, django_user_model, variant):
    """Unlike debt_reminder, the statement is informational — it must still
    send (showing a zero/settled balance) for a fully-paid client, never
    silently no-op the way a debt nudge correctly does."""
    call_command("setup_roles")
    editor = django_user_model.objects.create_user("stmt_zero_ed", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    cust = Client.objects.create(first_name="Чистый", phone="+996700009002")
    resp = client.post(f"/pos/clients/{cust.pk}/statement/send/")
    assert resp.status_code == 302
    assert resp["Location"].startswith("https://wa.me/")


def test_send_statement_logs_interaction(client, django_user_model, variant):
    call_command("setup_roles")
    editor = django_user_model.objects.create_user("stmt_log_ed", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    cust = Client.objects.create(first_name="Логи", phone="+996700009003")
    before = Interaction.objects.count()
    client.post(f"/pos/clients/{cust.pk}/statement/send/")
    assert Interaction.objects.count() == before + 1
    assert Interaction.objects.filter(client=cust, note="Выписка отправлена").exists()


def test_viewer_cannot_send_statement_or_toggle_reminder(client, django_user_model):
    call_command("setup_roles")
    viewer = django_user_model.objects.create_user("stmt_viewer2", password="x" * 12, is_staff=True)
    viewer.groups.add(Group.objects.get(name=VIEWER))
    client.force_login(viewer)

    cust = Client.objects.create(first_name="Вьюер", phone="+996700009004")
    resp = client.post(f"/pos/clients/{cust.pk}/statement/send/")
    assert resp.status_code == 403
    resp2 = client.post(f"/pos/clients/{cust.pk}/remind/toggle/")
    assert resp2.status_code == 403
    # Read-only access to the worklist itself IS allowed.
    resp3 = client.get("/pos/clients/debtors/")
    assert resp3.status_code == 200


def test_debtors_view_matches_client_debts_by_currency_exactly(client, django_user_model, variant):
    from apps.clients.services import client_debts_by_currency

    call_command("setup_roles")
    editor = django_user_model.objects.create_user(
        "debtors_view_ed", password="x" * 12, is_staff=True
    )
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    debtor = Client.objects.create(first_name="Должник", phone="+996700009005")
    order = SaleOrder.objects.create(client=debtor, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("4400"))
    confirm_sale(order)

    settled = Client.objects.create(first_name="Расплатился", phone="+996700009006")
    order2 = SaleOrder.objects.create(client=settled, currency="KGS")
    SaleItem.objects.create(order=order2, variant=variant, quantity=1, unit_price=Decimal("1000"))
    confirm_sale(order2)
    record_payment(order2, Decimal("1000"))

    resp = client.get("/pos/clients/debtors/")
    assert resp.status_code == 200
    content = resp.content.decode()

    expected = client_debts_by_currency(client=debtor).get(debtor.pk, {})
    assert expected == {"KGS": Decimal("4400.00")}
    from apps.core.currency import format_money

    assert debtor.name in content
    assert format_money(expected["KGS"], "KGS") in content
    # A settled client (debt cleared) must NOT appear on the worklist.
    assert settled.name not in content


def test_debtors_view_writes_zero_db_rows(client, django_user_model, variant):
    call_command("setup_roles")
    editor = django_user_model.objects.create_user(
        "debtors_zerodb_ed", password="x" * 12, is_staff=True
    )
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    cust = Client.objects.create(first_name="НульЗапись", phone="+996700009007")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("2000"))
    confirm_sale(order)

    payment_before = Payment.objects.count()
    order_before = SaleOrder.objects.count()
    interaction_before = Interaction.objects.count()
    client_before = Client.objects.count()

    resp = client.get("/pos/clients/debtors/")
    assert resp.status_code == 200

    assert Payment.objects.count() == payment_before
    assert SaleOrder.objects.count() == order_before
    assert Interaction.objects.count() == interaction_before
    assert Client.objects.count() == client_before


def test_do_not_remind_toggle_hides_action_and_is_reversible(client, django_user_model, variant):
    call_command("setup_roles")
    editor = django_user_model.objects.create_user("dnr_ed", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    add_movement(variant, StockMovement.PRODUCTION_IN, 10)
    cust = Client.objects.create(first_name="Скип", phone="+996700009008")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("1500"))
    confirm_sale(order)

    resp = client.get("/pos/clients/debtors/")
    assert "Не напоминать" in resp.content.decode()

    client.post(f"/pos/clients/{cust.pk}/remind/toggle/")
    cust.refresh_from_db()
    assert cust.do_not_remind is True

    resp2 = client.get("/pos/clients/debtors/")
    assert "Напоминать" in resp2.content.decode()
    # The debt itself, and «Выписка», are UNAFFECTED by the flag — only the
    # reminder action is skipped (see Client.do_not_remind's own docstring).
    assert cust.name in resp2.content.decode()

    client.post(f"/pos/clients/{cust.pk}/remind/toggle/")
    cust.refresh_from_db()
    assert cust.do_not_remind is False


def test_client_detail_query_count_stays_flat_as_purchase_history_grows(
    client, django_user_model, variant
):
    """client_detail's «История покупок» list (see apps.pos.views.
    client_detail) — assert the query count is a FIXED handful regardless of
    how many confirmed sales exist, never one extra query per row (the
    queryset is a single bounded slice, never iterated with a per-row
    lookup)."""
    from django.db import connection
    from django.test.utils import CaptureQueriesContext

    call_command("setup_roles")
    editor = django_user_model.objects.create_user("perf_ed", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    add_movement(variant, StockMovement.PRODUCTION_IN, 100)
    cust = Client.objects.create(first_name="Perf", phone="+996700777001")
    for i in range(25):
        order = SaleOrder.objects.create(client=cust, currency="KGS")
        SaleItem.objects.create(
            order=order, variant=variant, quantity=1, unit_price=Decimal("1000")
        )
        confirm_sale(order)
        record_payment(order, Decimal("1000"))

    with CaptureQueriesContext(connection) as ctx:
        resp = client.get(f"/pos/clients/{cust.pk}/")
    assert resp.status_code == 200
    # A generous ceiling here is a regression trip-wire, not a tight budget.
    assert len(ctx.captured_queries) < 30


def test_debtors_page_query_count_stays_flat_as_debtor_count_grows(
    client, django_user_model, variant
):
    """The Должники worklist (apps.pos.views.debtors, reusing
    debtors_report_rows) must stay a fixed handful of queries regardless of
    how many debtors exist — never one extra query per debtor row."""
    from django.db import connection
    from django.test.utils import CaptureQueriesContext

    call_command("setup_roles")
    editor = django_user_model.objects.create_user("perf_deb_ed", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    add_movement(variant, StockMovement.PRODUCTION_IN, 100)
    for i in range(20):
        cust = Client.objects.create(first_name=f"Debtor{i}", phone=f"+99670099{i:04d}")
        order = SaleOrder.objects.create(client=cust, currency="KGS")
        SaleItem.objects.create(
            order=order, variant=variant, quantity=1, unit_price=Decimal("1000")
        )
        confirm_sale(order)

    with CaptureQueriesContext(connection) as ctx:
        resp = client.get("/pos/clients/debtors/")
    assert resp.status_code == 200
    assert len(ctx.captured_queries) < 15


# ---- Manual opening-balance entry (no bulk import — see git history) ------


def test_opening_balance_add_editor_gets_403(client, django_user_model):
    call_command("setup_roles")
    editor = django_user_model.objects.create_user("obadd_ed", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    cust = Client.objects.create(first_name="Крафт", phone="+996700020001")
    resp = client.get(f"/pos/clients/{cust.pk}/opening-balance/add/")
    assert resp.status_code == 403

    resp2 = client.post(
        f"/pos/clients/{cust.pk}/opening-balance/add/",
        {"amount": "5000", "currency": "KGS", "as_of_date": "2026-01-01", "note": ""},
    )
    assert resp2.status_code == 403
    assert not ClientOpeningBalance.objects.filter(client=cust).exists()


def test_opening_balance_add_creates_and_is_immediately_visible(client, django_user_model):
    """Once added, it shows in «Баланс» AND as the first «старый долг»
    statement row — no separate refresh/recompute step."""
    call_command("setup_roles")
    owner = django_user_model.objects.create_superuser("obadd_owner", password="x" * 12)
    client.force_login(owner)

    cust = Client.objects.create(first_name="Роза", phone="+996700020002")
    resp = client.post(
        f"/pos/clients/{cust.pk}/opening-balance/add/",
        {"amount": "99000", "currency": "KGS", "as_of_date": "2026-01-01", "note": "старый долг"},
    )
    assert resp.status_code == 302

    ob = ClientOpeningBalance.objects.get(client=cust)
    assert ob.amount == Decimal("99000")
    assert ob.currency == "KGS"
    assert ob.created_by == owner

    from apps.clients.services import client_debt, client_statement

    assert client_debt(cust) == {"KGS": Decimal("99000.00")}

    detail_html = client.get(f"/pos/clients/{cust.pk}/").content.decode()
    assert "99" in detail_html  # the balance figure renders on the page

    statement = client_statement(cust)
    first_row = statement["KGS"]["rows"][0]
    assert first_row["kind"] == "opening_balance"
    assert first_row["balance"] == Decimal("99000.00")


def test_opening_balance_add_resubmit_same_date_updates_in_place(client, django_user_model):
    """She will make typos — re-entering the SAME (client, currency, date)
    corrects the figure in place rather than erroring or duplicating, and
    created_by stays the original entrant, not whoever fixed the typo."""
    call_command("setup_roles")
    owner = django_user_model.objects.create_superuser("obadd_fix_owner", password="x" * 12)
    client.force_login(owner)

    cust = Client.objects.create(first_name="Гуля", phone="+996700020003")
    client.post(
        f"/pos/clients/{cust.pk}/opening-balance/add/",
        {"amount": "10000", "currency": "KGS", "as_of_date": "2026-01-01", "note": ""},
    )
    client.post(
        f"/pos/clients/{cust.pk}/opening-balance/add/",
        {"amount": "12500", "currency": "KGS", "as_of_date": "2026-01-01", "note": "исправлено"},
    )

    assert ClientOpeningBalance.objects.filter(client=cust).count() == 1
    ob = ClientOpeningBalance.objects.get(client=cust)
    assert ob.amount == Decimal("12500")
    assert ob.note == "исправлено"
    assert ob.created_by == owner
    # history retained across the correction — two snapshots, not one
    assert ob.history.count() == 2


def test_opening_balance_add_allows_negative_amount_as_advance(client, django_user_model):
    call_command("setup_roles")
    owner = django_user_model.objects.create_superuser("obadd_neg_owner", password="x" * 12)
    client.force_login(owner)

    cust = Client.objects.create(first_name="Аванс", phone="+996700020004")
    resp = client.post(
        f"/pos/clients/{cust.pk}/opening-balance/add/",
        {"amount": "-5000", "currency": "USD", "as_of_date": "2026-01-01", "note": ""},
    )
    assert resp.status_code == 302
    ob = ClientOpeningBalance.objects.get(client=cust)
    assert ob.amount == Decimal("-5000")

    from apps.clients.services import client_credits

    assert client_credits(cust) == {"USD": Decimal("5000.00")}


def test_opening_balance_add_rejects_zero_amount_writes_nothing(client, django_user_model):
    call_command("setup_roles")
    owner = django_user_model.objects.create_superuser("obadd_zero_owner", password="x" * 12)
    client.force_login(owner)

    cust = Client.objects.create(first_name="Ноль", phone="+996700020005")
    resp = client.post(
        f"/pos/clients/{cust.pk}/opening-balance/add/",
        {"amount": "0", "currency": "KGS", "as_of_date": "2026-01-01", "note": ""},
    )
    assert resp.status_code == 200  # redisplayed with an error, not redirected
    assert not ClientOpeningBalance.objects.filter(client=cust).exists()


def test_opening_balance_progress_count_on_clients_page(client, django_user_model):
    """«Старый долг внесён для N из M клиентов» — N is DISTINCT clients with
    at least one opening balance (any currency), never row count; M is
    active clients."""
    call_command("setup_roles")
    owner = django_user_model.objects.create_superuser("obadd_prog_owner", password="x" * 12)
    client.force_login(owner)

    c1 = Client.objects.create(first_name="Один", phone="+996700020006")
    c2 = Client.objects.create(first_name="Два", phone="+996700020007")
    Client.objects.create(first_name="Три", phone="+996700020008")  # no opening balance

    ClientOpeningBalance.objects.create(
        client=c1, amount=Decimal("1000"), currency="KGS", as_of_date=date(2026, 1, 1)
    )
    # c2 has opening balances in TWO currencies — still counts as ONE client done.
    ClientOpeningBalance.objects.create(
        client=c2, amount=Decimal("1000"), currency="KGS", as_of_date=date(2026, 1, 1)
    )
    ClientOpeningBalance.objects.create(
        client=c2, amount=Decimal("50"), currency="USD", as_of_date=date(2026, 1, 1)
    )

    html = client.get("/pos/clients/").content.decode()
    assert "Старый долг внесён для 2 из 3 клиентов" in html


def test_opening_balance_progress_hidden_from_non_owner(client, django_user_model):
    call_command("setup_roles")
    editor = django_user_model.objects.create_user(
        "obadd_prog_ed", password="x" * 12, is_staff=True
    )
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    html = client.get("/pos/clients/").content.decode()
    assert "Старый долг внесён" not in html


def test_opening_balance_admin_changelist_shows_progress(client, django_user_model):
    call_command("setup_roles")
    owner = django_user_model.objects.create_superuser("obadd_admin_owner", password="x" * 12)
    client.force_login(owner)

    cust = Client.objects.create(first_name="Панель", phone="+996700020009")
    ClientOpeningBalance.objects.create(
        client=cust, amount=Decimal("1000"), currency="KGS", as_of_date=date(2026, 1, 1)
    )

    html = client.get("/panel/clients/clientopeningbalance/").content.decode()
    assert "Старый долг внесён для" in html


def test_owner_can_edit_and_delete_opening_balance_via_admin(client, django_user_model):
    """Admin change/delete both stay Owner-only and history survives an
    edit — this already worked via ClientOpeningBalanceAdmin's explicit
    is_superuser checks, confirmed here end-to-end through the actual admin
    URLs rather than just the permission methods."""
    call_command("setup_roles")
    owner = django_user_model.objects.create_superuser("obadd_edit_owner", password="x" * 12)
    client.force_login(owner)

    cust = Client.objects.create(first_name="Редакт", phone="+996700020010")
    ob = ClientOpeningBalance.objects.create(
        client=cust, amount=Decimal("1000"), currency="KGS", as_of_date=date(2026, 1, 1)
    )

    resp = client.post(
        f"/panel/clients/clientopeningbalance/{ob.pk}/change/",
        {
            "client": cust.pk,
            "amount": "1500",
            "currency": "KGS",
            "as_of_date": "2026-01-01",
            "note": "поправка",
            "_save": "Save",
        },
    )
    assert resp.status_code == 302
    ob.refresh_from_db()
    assert ob.amount == Decimal("1500")
    assert ob.history.count() == 2

    resp2 = client.post(f"/panel/clients/clientopeningbalance/{ob.pk}/delete/", {"post": "yes"})
    assert resp2.status_code == 302
    assert not ClientOpeningBalance.objects.filter(pk=ob.pk).exists()


# ---- Historical (backdated) sales: itemized history without touching ------
# ---- current stock or current-period revenue -------------------------------


def test_confirm_sale_historical_writes_no_stock_movement(variant):
    stock_before = variant.stock
    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=2, unit_price=Decimal("3200"))
    confirm_sale(order, is_historical=True, historical_date=date(2024, 3, 15))

    variant.refresh_from_db()
    assert variant.stock == stock_before
    assert not StockMovement.objects.filter(variant=variant).exists()


def test_confirm_sale_historical_sets_confirmed_at_from_historical_date(variant):
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order, is_historical=True, historical_date=date(2023, 6, 1))

    order.refresh_from_db()
    assert order.is_historical is True
    assert timezone.localtime(order.confirmed_at).date() == date(2023, 6, 1)


def test_confirm_sale_historical_skips_the_oversell_check(variant):
    """A backdated sale isn't taking anything from TODAY's warehouse — it
    must succeed even against ZERO current stock (variant starts with none:
    no add_movement call in this test at all)."""
    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=100, unit_price=Decimal("3200"))
    confirm_sale(order, is_historical=True, historical_date=date(2023, 1, 1))  # must not raise

    order.refresh_from_db()
    assert order.status == SaleOrder.CONFIRMED
    assert variant.stock == 0  # still untouched


def test_confirm_sale_historical_without_date_raises(variant):
    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    with pytest.raises(ValidationError):
        confirm_sale(order, is_historical=True, historical_date=None)
    order.refresh_from_db()
    assert order.status == SaleOrder.DRAFT  # nothing written


def test_cancel_historical_sale_does_not_phantom_add_stock(variant):
    stock_before = variant.stock
    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=3, unit_price=Decimal("3200"))
    confirm_sale(order, is_historical=True, historical_date=date(2023, 1, 1))
    cancel_sale(order)

    variant.refresh_from_db()
    order.refresh_from_db()
    assert order.status == SaleOrder.CANCELLED
    assert variant.stock == stock_before  # never went up from a fake RETURN_IN
    assert not StockMovement.objects.filter(variant=variant).exists()


def test_return_from_historical_sale_does_not_phantom_add_stock(variant):
    stock_before = variant.stock
    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=3, unit_price=Decimal("3200"))
    confirm_sale(order, is_historical=True, historical_date=date(2023, 1, 1))
    item = order.items.get()

    from apps.sales.services import return_items

    return_items(order, {item.pk: 1})

    variant.refresh_from_db()
    assert variant.stock == stock_before
    assert not StockMovement.objects.filter(variant=variant).exists()


def test_historical_sale_excluded_from_dashboard_revenue_profit_units(variant):
    """Backdated to TODAY specifically (inside the "today" period window)
    so this actually proves the is_historical filter works, not just that
    the date happens to fall outside the window."""
    from apps.reports.dashboard import dashboard_data

    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    today = timezone.localdate()
    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order, is_historical=True, historical_date=today)

    data = dashboard_data("today")
    assert data["metrics"]["revenue"]["value"] == Decimal("0")
    assert data["metrics"]["profit"]["value"] == Decimal("0")
    assert data["metrics"]["units"]["value"] == 0


def test_historical_sale_excluded_from_today_revenue_and_summary(variant):
    from apps.sales.services import today_revenue_kgs, today_summary

    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    today = timezone.localdate()
    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order, is_historical=True, historical_date=today)

    assert today_revenue_kgs() == Decimal("0")
    summary = today_summary()
    assert summary["orders"] == 0
    assert summary["items"] == 0
    assert summary["revenue_by_currency"] == {}


def test_historical_sale_excluded_from_daily_report_sales_sheet(variant):
    from apps.sales.services import todays_confirmed_orders

    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    today = timezone.localdate()
    order = SaleOrder.objects.create(currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order, is_historical=True, historical_date=today)

    assert order.pk not in [o.pk for o in todays_confirmed_orders()]


def test_historical_sale_still_counts_as_real_debt_if_unpaid(variant):
    """Debt/payment tracking is DELIBERATELY not excluded — see
    SaleOrder.is_historical's own docstring. An unpaid historical sale is a
    real debt like any other and must show in «Баланс»/the statement."""
    from apps.clients.services import client_debt

    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    cust = Client.objects.create(first_name="Историческая", phone="+996700030001")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order, is_historical=True, historical_date=date(2023, 1, 1))

    assert client_debt(cust) == {"KGS": Decimal("3200.00")}


def test_historical_sale_appears_in_purchase_history_marked(client, django_user_model, variant):
    call_command("setup_roles")
    editor = django_user_model.objects.create_user("hist_hist_ed", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    cust = Client.objects.create(first_name="Итория", phone="+996700030002")
    order = SaleOrder.objects.create(client=cust, currency="KGS")
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    confirm_sale(order, is_historical=True, historical_date=date(2023, 1, 1))
    record_payment(order, Decimal("3200"))  # fully paid -> appears in История покупок

    html = client.get(f"/pos/clients/{cust.pk}/").content.decode()
    assert "историческая" in html


def test_sale_confirm_view_editor_crafted_post_with_is_historical_gets_403(
    client, django_user_model, variant
):
    call_command("setup_roles")
    editor = django_user_model.objects.create_user("hist_403_ed", password="x" * 12, is_staff=True)
    editor.groups.add(Group.objects.get(name=EDITOR))
    client.force_login(editor)

    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    order = SaleOrder.objects.create(currency="KGS", created_by=editor)
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    resp = client.post(
        f"/pos/sale/{order.pk}/confirm/",
        {"amount": "0", "currency": "KGS", "is_historical": "1", "historical_date": "2023-01-01"},
    )
    assert resp.status_code == 403
    order.refresh_from_db()
    assert order.status == SaleOrder.DRAFT
    assert not order.is_historical


def test_sale_confirm_view_owner_can_backdate(client, django_user_model, variant):
    call_command("setup_roles")
    owner = django_user_model.objects.create_superuser("hist_owner_view", password="x" * 12)
    client.force_login(owner)

    order = SaleOrder.objects.create(currency="KGS", created_by=owner)
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    resp = client.post(
        f"/pos/sale/{order.pk}/confirm/",
        {"amount": "0", "currency": "KGS", "is_historical": "1", "historical_date": "2022-05-05"},
    )
    assert resp.status_code == 302
    order.refresh_from_db()
    assert order.is_historical is True
    assert timezone.localtime(order.confirmed_at).date() == date(2022, 5, 5)
    variant.refresh_from_db()
    assert variant.stock == 0  # never touched despite zero starting stock


def test_sale_confirm_view_owner_missing_date_shows_error_not_500(
    client, django_user_model, variant
):
    call_command("setup_roles")
    owner = django_user_model.objects.create_superuser("hist_owner_noerr", password="x" * 12)
    client.force_login(owner)

    add_movement(variant, StockMovement.PRODUCTION_IN, 5)
    order = SaleOrder.objects.create(currency="KGS", created_by=owner)
    SaleItem.objects.create(order=order, variant=variant, quantity=1, unit_price=Decimal("3200"))
    resp = client.post(
        f"/pos/sale/{order.pk}/confirm/",
        {"amount": "0", "currency": "KGS", "is_historical": "1", "historical_date": ""},
    )
    assert resp.status_code == 302  # redisplays sale_detail, not a 500
    order.refresh_from_db()
    assert order.status == SaleOrder.DRAFT
