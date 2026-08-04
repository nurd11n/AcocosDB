"""Per-client transaction history + debt Excel export.

Reachable from /panel/ only: a button on the Client change page (one client)
and a bulk action on the Client changelist (several clients at once). Bulk
export choice, noted per the spec: ONE WORKBOOK, one sheet per client —
simpler than a zip of files, and every client's sheet needs its own
variable-length summary block below the table anyway, which a single sheet
handles naturally.

Every figure here is the SAME derivation used everywhere else this data is
shown (client_detail, receipts, the daily report) — CLAUDE.md's core rule
that debt/balance is never a second, independently-computed number:
SaleOrder.total/paid_amount/balance for the per-order columns, and
apps.clients.services.client_debt for the one authoritative OUTSTANDING DEBT
figure (which alone applies the sub-сом rounding-tolerance clamp — naively
subtracting "purchased minus paid" here instead would occasionally disagree
with what the rest of the app calls "fully paid" by a few kopecks).
"""

from decimal import Decimal

from django.utils import timezone

from django.conf import settings
from django.utils.translation import gettext as _
from openpyxl.styles import Font, PatternFill

from apps.clients.services import client_debt
from apps.reports.export import DATE, INT, MONEY, TEXT, Column, MONEY_FORMAT, write_sheet

_COLUMNS = [
    Column(_("Дата"), DATE),
    Column(_("Продажа №"), TEXT),
    Column(_("Товары"), TEXT),
    Column(_("Количество"), INT),
    Column(_("Подытог"), MONEY),
    Column(_("Скидка"), MONEY),
    Column(_("Итого"), MONEY),
    Column(_("Оплачено"), MONEY),
    Column(_("Остаток"), MONEY),
    Column(_("Валюта"), TEXT),
    Column(_("Статус"), TEXT),
]

# SaleOrder.payment_status values -> the Russian shown throughout /pos/
# (badge--paid/partial/debt) — same labels, not a second vocabulary.
_STATUS_RU = {"unpaid": _("долг"), "partial": _("частично"), "paid": _("оплачено")}

_SUMMARY_FILL = PatternFill("solid", fgColor="FFF3CD")
_SUMMARY_FONT = Font(bold=True)


def _order_row(order, items) -> list:
    subtotal = sum((i.subtotal for i in items), Decimal("0"))
    discount = sum((i.discount_amount for i in items), Decimal("0"))
    quantity = sum((i.quantity for i in items), 0)
    products = ", ".join(f"{i.variant} ×{i.quantity}" for i in items)
    return [
        # Bishkek-local date, not UTC — order.confirmed_at is a UTC-aware
        # datetime (USE_TZ=True); a bare .date() reads the wrong calendar
        # date for anything confirmed in the early-morning hours locally.
        timezone.localtime(order.confirmed_at).date() if order.confirmed_at else None,
        f"№{order.pk}",
        products,
        quantity,
        subtotal,
        discount,
        order.total,
        order.paid_amount,
        order.balance,
        order.currency,
        _STATUS_RU.get(order.payment_status, order.payment_status),
    ]


def write_client_sheet(ws, client) -> None:
    """Writes the Transactions table + the per-currency summary block for
    ONE client onto an already-created worksheet."""
    from apps.sales.models import SaleOrder

    orders = list(
        SaleOrder.objects.filter(client=client, status=SaleOrder.CONFIRMED)
        .prefetch_related("items__variant__product")
        .order_by("confirmed_at")
    )
    rows = [_order_row(o, list(o.items.all())) for o in orders]
    write_sheet(ws, _COLUMNS, rows)

    # Purchased/paid, summed per currency straight from the same rows just
    # written (never re-queried) — never summed ACROSS currencies.
    totals: dict[str, dict[str, Decimal]] = {}
    for o in orders:
        t = totals.setdefault(o.currency, {"purchased": Decimal("0"), "paid": Decimal("0")})
        t["purchased"] += o.total
        t["paid"] += o.paid_amount
    debts = client_debt(client)  # {currency: positive debt} — the one authoritative source

    currencies = sorted(set(totals) | set(debts)) or [settings.CURRENCY]

    row = ws.max_row + 2
    for currency in currencies:
        t = totals.get(currency, {"purchased": Decimal("0"), "paid": Decimal("0")})
        debt = debts.get(currency, Decimal("0"))
        for label, value, highlight in (
            (f"{_('ВСЕГО КУПЛЕНО')} ({currency})", t["purchased"], False),
            (f"{_('ВСЕГО ОПЛАЧЕНО')} ({currency})", t["paid"], False),
            (f"{_('ЗАДОЛЖЕННОСТЬ')} ({currency})", debt, True),
        ):
            label_cell = ws.cell(row=row, column=1, value=label)
            value_cell = ws.cell(row=row, column=2, value=value)
            label_cell.font = _SUMMARY_FONT
            value_cell.font = _SUMMARY_FONT
            value_cell.number_format = MONEY_FORMAT
            if highlight:
                label_cell.fill = _SUMMARY_FILL
                value_cell.fill = _SUMMARY_FILL
            row += 1
        row += 1  # blank line between currencies


def _sheet_title(client, used: set) -> str:
    # Excel: 31-char cap, no \ / * ? : [ ]. Staff-only descriptor deliberately
    # left OUT (see the filename note below) — first name (+ id if that
    # collides with another selected client) is enough to tell sheets apart.
    base = "".join(ch for ch in client.first_name if ch not in "\\/*?:[]")[:25] or "client"
    title = base
    n = 2
    while title in used:
        title = f"{base}-{n}"[:31]
        n += 1
    used.add(title)
    return title


def client_workbook_bytes(clients) -> bytes:
    """One workbook, one sheet per client in `clients` (any iterable)."""
    from openpyxl import Workbook

    wb = Workbook()
    used_titles: set = set()
    clients = list(clients)
    for i, client in enumerate(clients):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = _sheet_title(client, used_titles)
        write_client_sheet(ws, client)
    if not clients:
        wb.active.title = "empty"
    import io

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def client_export_filename(client) -> str:
    """{client_name}_{YYYY-MM-DD}.xlsx — first name ONLY, never client.name
    (which parenthesises the staff-only descriptor, e.g. "сестра Розы"): a
    downloaded file is far more likely to get forwarded/shared outside
    /panel/ than an on-screen label ever is, so it gets the same
    client-facing-text discipline as a receipt or WhatsApp message."""
    from django.utils import timezone

    safe_name = "".join(ch for ch in client.first_name if ch not in '\\/*?:"<>|') or "client"
    return f"{safe_name}_{timezone.localdate():%Y-%m-%d}.xlsx"
