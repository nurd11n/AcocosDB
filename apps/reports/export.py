"""Spreadsheet serialization shared by every export — the daily report, the
storage report, and the analytics dashboard.

One place that turns a typed column spec + rows into a formatted .xlsx
workbook, or a single sheet into UTF-8-BOM .csv (BOM so Excel renders
Cyrillic correctly). Each report renders from the same aggregates it shows
on screen, then hands the rows here — so a download can never disagree with
the page.

WHY TYPED COLUMNS: every value used to be run through str() before it reached
openpyxl, so Excel stored the whole sheet as TEXT. "3200.00" in a text cell
can't be summed, sorted numerically, or filtered by range — the owner had to
retype figures to total a column, and sorting put 1000 before 900. Declaring
a kind per column lets the writer put a real number/date in the cell and
attach a display format, so the file behaves like a spreadsheet instead of a
screenshot of one.
"""

import csv
import datetime as dt
import io
from dataclasses import dataclass, field
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TEXT, INT, MONEY, DATE, DATETIME, TIME = "text", "int", "money", "date", "datetime", "time"

# Space-grouped thousands: readable next to a comma decimal separator, which
# is what a Russian-locale Excel shows. Renders 3 200,00.
MONEY_FORMAT = "# ##0.00"
INT_FORMAT = "# ##0"
DATE_FORMAT = "DD.MM.YYYY"
DATETIME_FORMAT = "DD.MM.YYYY HH:MM"
# A full timestamp displayed as just the clock time — every row in a daily
# report is the same day, so the date would be noise, but keeping the real
# datetime underneath means the column still sorts chronologically.
TIME_FORMAT = "HH:MM"

_NUMERIC_KINDS = {INT, MONEY}
_NUMBER_FORMATS = {
    INT: INT_FORMAT,
    MONEY: MONEY_FORMAT,
    DATE: DATE_FORMAT,
    DATETIME: DATETIME_FORMAT,
    TIME: TIME_FORMAT,
}

_HEADER_FILL = PatternFill("solid", fgColor="1F3B54")  # deep slate, prints legibly
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TOTAL_FONT = Font(bold=True)
_TOTAL_BORDER = Border(top=Side(style="thin", color="1F3B54"))
_MIN_WIDTH = 9
_MAX_WIDTH = 46  # keep one long product name from stretching the sheet off-screen


@dataclass(frozen=True)
class Column:
    """A column's header text and how its values render.

    `kind` drives the Excel cell type and number format. Anything not
    numeric/temporal stays TEXT — including deliberate placeholders like
    "н/д", which must never silently become 0.
    """

    header: str
    kind: str = TEXT


@dataclass
class Sheet:
    """A named sheet's column spec, rows, and optional bold totals row."""

    columns: list[Column]
    rows: list[list] = field(default_factory=list)
    totals: list | None = None

    def as_flat_rows(self) -> list[list]:
        """Header + body (+ totals) as plain lists — the CSV shape."""
        out = [[c.header for c in self.columns]] + [list(r) for r in self.rows]
        if self.totals:
            out.append(list(self.totals))
        return out


def _is_number(value) -> bool:
    # bool is an int subclass; it is never a figure we want number-formatted.
    return isinstance(value, (int, float, Decimal)) and not isinstance(value, bool)


def _cell(value):
    """Excel has no timezone concept and openpyxl refuses an aware datetime
    outright. Callers hand us `timezone.localtime(...)`, i.e. already in
    Asia/Bishkek — dropping tzinfo here keeps the displayed clock time
    correct while satisfying the format. Done centrally so no report author
    has to remember it."""
    if isinstance(value, dt.datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def write_sheet(ws, columns: list[Column], rows, totals: list | None = None) -> None:
    """Write one formatted sheet: styled frozen header, typed body cells,
    autofilter, sized columns, and an optional bold totals row."""
    ws.append([c.header for c in columns])
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    widths = [len(c.header) + 2 for c in columns]
    for row in rows:
        ws.append([_cell(v) for v in row])
        for i, value in enumerate(row):
            if i < len(widths) and value is not None:
                widths[i] = max(widths[i], len(str(value)) + 2)

    if totals:
        ws.append([_cell(v) for v in totals])
        for i, value in enumerate(totals):
            if i < len(widths) and value is not None:
                widths[i] = max(widths[i], len(str(value)) + 2)
        for cell in ws[ws.max_row]:
            cell.font = _TOTAL_FONT
            cell.border = _TOTAL_BORDER

    # Number formats per column, over the body (and totals) but never the header.
    for i, col in enumerate(columns, start=1):
        fmt = _NUMBER_FORMATS.get(col.kind)
        if not fmt:
            continue
        for cell in ws[get_column_letter(i)][1:]:
            if _is_number(cell.value) or isinstance(cell.value, (dt.date, dt.datetime)):
                cell.number_format = fmt

    # Freeze the header and let the owner filter/sort in place — the whole
    # point of shipping a real spreadsheet rather than a static table.
    ws.freeze_panes = "A2"
    if columns:
        # Applied even to a header-only sheet (a quiet day's «Долги», nobody
        # owing anything): Excel accepts the range, and a sheet that gains
        # rows tomorrow is then already filterable rather than inconsistently not.
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{ws.max_row}"
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(width, _MIN_WIDTH), _MAX_WIDTH)


def to_xlsx(sheets: dict) -> bytes:
    """{name: Sheet} -> a formatted workbook. A bare list-of-lists (header row
    first) is still accepted so an un-migrated caller keeps working — it gets
    the header styling and widths, just no per-column number formats."""
    wb = Workbook()
    for i, (name, sheet) in enumerate(sheets.items()):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = name[:31]  # Excel caps sheet names at 31 chars
        if isinstance(sheet, Sheet):
            write_sheet(ws, sheet.columns, sheet.rows, sheet.totals)
        else:
            rows = list(sheet)
            columns = [Column(str(h)) for h in (rows[0] if rows else [])]
            write_sheet(ws, columns, rows[1:])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_csv_bytes(rows_or_sheet) -> bytes:
    rows = rows_or_sheet.as_flat_rows() if isinstance(rows_or_sheet, Sheet) else list(rows_or_sheet)
    buf = io.StringIO()
    csv.writer(buf).writerows(rows)
    # UTF-8 BOM — plain UTF-8 CSV shows broken Cyrillic when opened in Excel.
    return ("﻿" + buf.getvalue()).encode("utf-8")
